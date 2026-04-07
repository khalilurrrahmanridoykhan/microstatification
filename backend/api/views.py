from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework import status, viewsets
from rest_framework.exceptions import PermissionDenied, ValidationError
from .models import Template, Patient, TrashBin, Form, Submission, UserProfile, AppVersion
from .serializers import (
    TemplateSerializer,
    PatientSerializer,
    TrashBinListSerializer,
    TrashBinSerializer,
    FormSerializer,
    FormWithoutSubmissionSerializer,
    AppVersionSerializer,
)

# API endpoint to get a template by template ID

from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.core.cache import cache
from django.db import transaction, IntegrityError, connection
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import os
import re
import random
import string
import json
import logging
import uuid
from collections import Counter
from datetime import timedelta, datetime, date, time
from django.utils import timezone
from django.utils.text import slugify
from django.db.models import Count, Q, Prefetch
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek
from django.utils.dateparse import parse_datetime, parse_date
from django.conf import settings


# Define the exact column order as requested
DATA_COLLECTION_FIELD_ORDER = [
    'row_number',
    'date',
    'name_of_staff',
    'organization',
    'division',
    'district',
    'upazila',
    'union',
    'ward',
    'area',
    'location',
    'hh_id',
    'hh_head_name',
    'suspected_in_the_disease',
    'name_of_the_person_with_suspected_case',
    'mobile_number',
    'patient_id_type',
    'age',
    'sex',
    'suspected_disease',
    'referred',
    'bed_net_use_practice_during_sleep',
    'handwashing_practice_with_soap__water',
    'type_latrine_use',
    'presence_of_stagnant_water_mosquito_breeding_sites',
    'presence_of_mosquito_larvae',
    'did_any_disaster_occur_in_last_7_days_',
    'what_types',
    'remarks',
    'Validation',
    '_id',
    '_version',
    'group_name',
    'formType',
    'has_followup',
    'followup_count',
    'followup_forms',
    'followup_follow_up_date',
    'followup_start',
    'followup_end',
    'followup_follow_up_status_2472_h',
    'followup_suspected_in_the_disease',
    'followup_if_yes__name_of_the_diseases_1',
    'followup_hh_head_name',
    'followup_hh_id',
    'instanceID',
    'submitted_by',
    'start',
    'end'
]

BASE_ROW_FIELDS = [
    'row_number',
    'Validation',
    '_id',
    '_version',
    'group_name',
    'formType',
    'has_followup',
    'followup_count',
    'followup_forms',
    'followup_follow_up_date',
    'followup_start',
    'followup_end',
    'followup_follow_up_status_2472_h',
    'followup_suspected_in_the_disease',
    'followup_if_yes__name_of_the_diseases_1',
    'followup_hh_head_name',
    'followup_hh_id',
]

BASE_FIELD_DEFAULTS = {
    'row_number': None,
    'Validation': '',
    '_id': '',
    '_version': '',
    'group_name': '',
    'formType': '',
    'has_followup': False,
    'followup_count': 0,
    'followup_forms': '',
    'followup_follow_up_date': '',
    'followup_start': '',
    'followup_end': '',
    'followup_follow_up_status_2472_h': '',
    'followup_suspected_in_the_disease': '',
    'followup_if_yes__name_of_the_diseases_1': '',
    'followup_hh_head_name': '',
    'followup_hh_id': '',
}

BRAC_WHAT_TYPES_SPLIT_FIELDS = [
    "what_types_flood",
    "what_types_cyclone",
    "what_types_heavy_rainfall",
    "what_types_drought",
    "what_types_landslides",
    "what_types_others",
]

BRAC_FOLLOWUP_DISEASE_SPLIT_FIELDS = [
    "followup_disease_dengue",
    "followup_disease_awd",
    "followup_disease_malaria",
    "followup_disease_others",
    "followup_disease_not_diagonosd",
]

BRAC_DERIVED_FIELD_LABELS = {
    "total_reporting_sites": "total reporting sites",
    "age_month": "Age Month",
    "year_age": "Year age",
    "suspected_disease_dengue": "Suspected Disease: Dengue",
    "suspected_disease_awd": "Suspected Disease: AWD",
    "suspected_disease_malaria": "Suspected Disease: Malaria",
    "what_types_flood": "What Types: Flood",
    "what_types_cyclone": "What Types: Cyclone",
    "what_types_heavy_rainfall": "What Types: Heavy rainfall",
    "what_types_drought": "What Types: Drought",
    "what_types_landslides": "What Types: Landslides",
    "what_types_others": "What Types: Others",
    "followup_disease_dengue": "Follow-up Disease: Dengue",
    "followup_disease_awd": "Follow-up Disease: AWD",
    "followup_disease_malaria": "Follow-up Disease: Malaria",
    "followup_disease_others": "Follow-up Disease: Others",
    "followup_disease_not_diagonosd": "Follow-up Disease: Not diagnosed",
    "user_identification_11_9943_01976848561": "user identification",
}

BRAC_DOWNLOAD_DEFAULT_FIELDS = [
    "row_number",
    "date",
    "start",
    "end",
    "name_of_staff",
    "organization",
    "designation_1",
    "reporting_sites",
    "total_reporting_sites",
    "division",
    "district",
    "upazila",
    "union",
    "city_corporation",
    "ward",
    "ward_1",
    "area",
    "location",
    "hh_id",
    "hh_head_name",
    "suspected_in_the_disease",
    "name_of_the_person_with_suspected_case",
    "mobile_number",
    "user_identification_11_9943_01976848561",
    "age_month",
    "year_age",
    "sex",
    "pregnent",
    "suspected_disease_dengue",
    "suspected_disease_awd",
    "suspected_disease_malaria",
    "referred",
    "referral_place",
    "if_referred_to_govt",
    "bed_net_use_practice_during_sleep",
    "handwashing_practice_with_soap__water",
    "type_latrine_use",
    "no._of_already_diagnosed_cases_of_dengue_in_the_hh_1",
    "no._of_already_diagnosed_cases_of_malaria_in_the_hh",
    "no._of_already_diagnosed_cases_of_awd_in_the_hh",
    "presence_of_stagnant_water_mosquito_breeding_sites",
    "presence_of_mosquito_larvae",
    "did_any_disaster_occur_in_last_7_days_",
    *BRAC_WHAT_TYPES_SPLIT_FIELDS,
    "remarks",
    "followup_follow_up_date",
    "followup_start",
    "followup_end",
    "followup_follow_up_status_2472_h",
    "followup_suspected_in_the_disease",
    *BRAC_FOLLOWUP_DISEASE_SPLIT_FIELDS,
    "followup_hh_head_name",
    "followup_hh_id",
    "_id",
    "_version",
    "group_name",
    "formType",
    "submitted_by",
]

def get_ordered_field_list(all_fields):
    """
    Return fields in the specified order: data collection fields first, then followup fields
    """
    ordered_fields = []

    # Add data collection fields in the specified order
    for field in DATA_COLLECTION_FIELD_ORDER:
        if field in all_fields:
            ordered_fields.append(field)

    # Add any remaining data collection fields that weren't in the predefined list
    remaining_dc_fields = [f for f in all_fields if f not in ordered_fields and not f.startswith('followup_')]
    ordered_fields.extend(sorted(remaining_dc_fields))

    # Add followup fields at the end, sorted by form name
    followup_fields = [f for f in all_fields if f.startswith('followup_') and f not in ordered_fields]
    ordered_fields.extend(sorted(followup_fields))

    return ordered_fields


logger = logging.getLogger(__name__)

MICROSTATIFICATION_ADMIN_ROLE = 7
MICROSTATIFICATION_MANAGED_ROLES = {4, 8, 9}
MICROSTATIFICATION_GLOBAL_VISIBLE_ROLES = {8, 9}
MICROSTATIFICATION_CREATOR_SCOPED_ROLE = 4
MICROSTATIFICATION_ALLOWED_PROFILE_FIELDS = {
    'data_collection_type',
    'micro_role',
    'micro_division',
    'micro_district',
    'micro_upazila',
    'micro_union',
    'micro_village',
    'micro_villages',
    'micro_ward_no',
    'micro_sk_shw_name',
    'micro_designation',
    'micro_ss_name',
}


def _get_client_ip(request):
    forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if forwarded_for:
        return forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', 'unknown')


def _normalize_instance_id(instance_id):
    if instance_id is None:
        return ''
    value = str(instance_id).strip()
    if not value:
        return ''
    if value.startswith('uuid:'):
        return value
    return f'uuid:{value}'


def _extract_submission_data(payload):
    if not isinstance(payload, dict):
        return {}
    if isinstance(payload.get('data'), dict):
        return payload['data']
    return payload


def _find_field_in_payload(payload, field_name):
    """Recursively find the first matching field in nested submission payloads."""
    if not field_name:
        return None

    if isinstance(payload, dict):
        if field_name in payload:
            return payload[field_name]

        for value in payload.values():
            found = _find_field_in_payload(value, field_name)
            if found is not None:
                return found

    if isinstance(payload, list):
        for item in payload:
            found = _find_field_in_payload(item, field_name)
            if found is not None:
                return found

    return None


def _ensure_instance_id(submission_data):
    if not isinstance(submission_data, dict):
        submission_data = {}

    meta = submission_data.get('meta')
    if not isinstance(meta, dict):
        meta = {}

    instance_id = _normalize_instance_id(
        meta.get('instanceID') or submission_data.get('instanceID')
    )
    if not instance_id:
        instance_id = f'uuid:{uuid.uuid4()}'

    meta['instanceID'] = instance_id
    submission_data['meta'] = meta
    return instance_id, submission_data


def _submission_entry_from_record(record):
    submission_data = _extract_submission_data(record.data or {})
    if record.instance_id:
        meta = submission_data.get('meta')
        if not isinstance(meta, dict):
            meta = {}
        meta.setdefault('instanceID', record.instance_id)
        submission_data['meta'] = meta
    return {'data': submission_data}


def _extract_instance_id(submission_data):
    if not isinstance(submission_data, dict):
        return ''
    meta = submission_data.get('meta')
    if isinstance(meta, dict) and meta.get('instanceID'):
        return _normalize_instance_id(meta.get('instanceID'))
    if submission_data.get('instanceID'):
        return _normalize_instance_id(submission_data.get('instanceID'))
    return ''


def _get_form_submission_entries(form):
    records_qs = (
        Submission.objects.filter(form=form)
        .only('data', 'instance_id', 'is_deleted')
        .order_by('created_at')
    )
    records_entries = []
    deleted_instance_ids = set()
    for record in records_qs:
        normalized_instance = _normalize_instance_id(record.instance_id)
        if record.is_deleted:
            if normalized_instance:
                deleted_instance_ids.add(normalized_instance)
            continue
        records_entries.append(_submission_entry_from_record(record))
    legacy_entries = form.submission if isinstance(form.submission, list) else []

    if not legacy_entries:
        return records_entries

    merged_entries = []
    seen_instance_ids = set()

    def append_unique(entry):
        if not isinstance(entry, dict):
            return
        payload = entry.get('data')
        normalized_instance = _extract_instance_id(payload)
        if normalized_instance and normalized_instance in deleted_instance_ids:
            return
        if normalized_instance:
            if normalized_instance in seen_instance_ids:
                return
            seen_instance_ids.add(normalized_instance)
        merged_entries.append(entry)

    # Keep legacy order first, then append new row-based submissions.
    for entry in legacy_entries:
        append_unique(entry)

    for entry in records_entries:
        append_unique(entry)

    return merged_entries


def _collect_followup_lookup_for_forms(followup_form_ids):
    """
    Build follow-up rows mapped by data_collection_form_uuid using both:
    - Submission table rows (new storage)
    - Form.submission JSON rows (legacy storage)
    """
    ids = [int(form_id) for form_id in followup_form_ids if form_id]
    if not ids:
        return {}, {}, {}

    form_map = {
        form.id: form
        for form in Form.objects.only('id', 'name', 'submission', 'questions').filter(id__in=ids)
    }
    form_name_map = {form_id: (form.name or 'Follow Up') for form_id, form in form_map.items()}
    form_dc_uuid_fallback = {}

    for form_id, form in form_map.items():
        fallback_uuid = ''
        questions = form.questions if isinstance(form.questions, list) else []
        for question in questions:
            if not isinstance(question, dict):
                continue
            if question.get('name') != 'data_collection_form_uuid':
                continue
            fallback_uuid = _normalize_instance_id(question.get('default'))
            if fallback_uuid:
                break
        form_dc_uuid_fallback[int(form_id)] = fallback_uuid

    followup_lookup = {}
    linked_form_ids = set()
    seen_keys = set()
    sequence_counter = 0

    def append_followup(form_id, payload, instance_id=''):
        nonlocal sequence_counter
        if not isinstance(payload, dict):
            return
        dc_uuid = payload.get('data_collection_form_uuid')
        if not dc_uuid:
            dc_uuid = form_dc_uuid_fallback.get(int(form_id), '')
            if dc_uuid:
                payload = payload.copy()
                payload['data_collection_form_uuid'] = dc_uuid
        dc_uuid = _normalize_instance_id(dc_uuid)
        if not dc_uuid:
            return

        normalized_instance = _normalize_instance_id(instance_id) if instance_id else _extract_instance_id(payload)
        if normalized_instance:
            dedupe_key = (int(form_id), normalized_instance)
            if dedupe_key in seen_keys:
                return
            seen_keys.add(dedupe_key)

        linked_form_ids.add(int(form_id))
        followup_lookup.setdefault(dc_uuid, []).append({
            'form_id': int(form_id),
            'submission_data': payload,
            'submission_date': _extract_submission_date(payload),
            'sequence': sequence_counter,
        })
        sequence_counter += 1

    # Legacy rows first
    for form_id, form in form_map.items():
        for entry in (form.submission or []):
            payload = _extract_submission_data(entry)
            append_followup(form_id, payload)

    # Then normalized Submission rows
    for row in (
        Submission.objects.filter(form_id__in=ids, is_deleted=False)
        .only('form_id', 'data', 'instance_id', 'created_at')
        .order_by('created_at', 'id')
        .iterator(chunk_size=1000)
    ):
        payload = _extract_submission_data(row.data or {})
        append_followup(row.form_id, payload, row.instance_id or '')

    # Deterministic oldest-first ordering per data_collection_form_uuid.
    for dc_uuid in followup_lookup.keys():
        followup_lookup[dc_uuid].sort(
            key=lambda item: (
                item.get('submission_date') is None,
                item.get('submission_date') or date.max,
                item.get('sequence', 0),
            )
        )

    followup_forms_catalog = {
        form_id: form_name_map.get(form_id, 'Follow Up')
        for form_id in linked_form_ids
    }

    return followup_lookup, form_name_map, followup_forms_catalog


def _shift_year_month(year, month, delta):
    total = (year * 12 + (month - 1)) + delta
    return total // 12, (total % 12) + 1


def _coerce_to_local_date(value):
    if value in (None, ''):
        return None

    if isinstance(value, datetime):
        parsed_dt = value
    elif isinstance(value, str):
        parsed_dt = parse_datetime(value)
        if parsed_dt is None:
            parsed_date = parse_date(value)
            if parsed_date is not None:
                return parsed_date
            try:
                parsed_dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
            except Exception:
                return None
    else:
        return None

    if timezone.is_naive(parsed_dt):
        parsed_dt = timezone.make_aware(parsed_dt, timezone.get_current_timezone())
    return timezone.localdate(parsed_dt)


def _extract_submission_date(payload):
    submission_data = _extract_submission_data(payload)
    if not isinstance(submission_data, dict):
        return None

    meta = submission_data.get('meta')
    if not isinstance(meta, dict):
        meta = {}

    candidates = [
        submission_data.get('end'),
        submission_data.get('_submission_time'),
        submission_data.get('submitted_at'),
        submission_data.get('created_at'),
        meta.get('end'),
        meta.get('timeEnd'),
    ]

    for value in candidates:
        parsed_date = _coerce_to_local_date(value)
        if parsed_date is not None:
            return parsed_date
    return None


def _build_submission_bar_data(now_date, date_counts, month_counts):
    past7 = [0] * 7
    for diff in range(7):
        target_date = now_date - timedelta(days=diff)
        past7[6 - diff] = int(date_counts.get(target_date, 0))

    past31 = [0] * 5
    for diff in range(31):
        target_date = now_date - timedelta(days=diff)
        count = int(date_counts.get(target_date, 0))
        if diff < 7:
            past31[4] += count
        elif diff < 14:
            past31[3] += count
        elif diff < 21:
            past31[2] += count
        elif diff < 28:
            past31[1] += count
        else:
            past31[0] += count

    past3m = []
    past12m = []
    for delta in range(-2, 1):
        year, month = _shift_year_month(now_date.year, now_date.month, delta)
        past3m.append(int(month_counts.get((year, month), 0)))

    for delta in range(-11, 1):
        year, month = _shift_year_month(now_date.year, now_date.month, delta)
        past12m.append(int(month_counts.get((year, month), 0)))

    return {
        'past7': past7,
        'past31': past31,
        'past3m': past3m,
        'past12m': past12m,
    }


def _find_submission_record(form, instance_id, include_deleted=False):
    normalized_instance_id = _normalize_instance_id(instance_id)
    if not normalized_instance_id:
        return None
    queryset = Submission.objects.filter(
        form=form,
        instance_id=normalized_instance_id,
    )
    if not include_deleted:
        queryset = queryset.filter(is_deleted=False)
    return queryset.order_by('-created_at').first()


def _check_submit_rate_limit(request, form_id, user=None):
    window_seconds = max(1, int(getattr(settings, 'SUBMISSION_RATE_LIMIT_WINDOW_SECONDS', 60)))
    auth_limit = max(1, int(getattr(settings, 'SUBMISSION_RATE_LIMIT_AUTH_LIMIT', 60)))
    anon_limit = max(1, int(getattr(settings, 'SUBMISSION_RATE_LIMIT_ANON_LIMIT', 30)))

    if user and user.is_authenticated:
        identity = f"user:{user.id}"
        limit = auth_limit
    else:
        identity = f"ip:{_get_client_ip(request)}"
        limit = anon_limit

    rate_key = f"submit-rate:v1:form:{form_id}:{identity}"
    current_count = cache.get(rate_key)

    if current_count is None:
        cache.set(rate_key, 1, timeout=window_seconds)
        return True

    try:
        current_count = int(current_count)
    except (TypeError, ValueError):
        current_count = 0

    if current_count >= limit:
        return False

    try:
        cache.incr(rate_key)
    except ValueError:
        cache.set(rate_key, current_count + 1, timeout=window_seconds)

    return True


def _should_force_inline_processing(form, submission_data):
    if not bool(getattr(settings, "SUBMISSION_INLINE_PRIORITY_ENABLED", True)):
        return False

    form_ids = getattr(settings, "SUBMISSION_INLINE_PRIORITY_FORM_IDS", [1079]) or []
    try:
        allowed_form_ids = {int(form_id) for form_id in form_ids}
    except (TypeError, ValueError):
        allowed_form_ids = {1079}

    if not form or form.id not in allowed_form_ids:
        return False

    priority_field = (
        str(getattr(settings, "SUBMISSION_INLINE_PRIORITY_FIELD", "suspected_in_the_disease")).strip()
        or "suspected_in_the_disease"
    )
    priority_value = _find_field_in_payload(submission_data, priority_field)
    if not _is_affirmative(priority_value):
        return False

    try:
        queue_threshold = int(getattr(settings, "SUBMISSION_INLINE_PRIORITY_QUEUE_THRESHOLD", 500))
    except (TypeError, ValueError):
        queue_threshold = 500
    queue_threshold = max(0, queue_threshold)

    queued_backlog = Submission.objects.filter(processing_status="queued", is_deleted=False).count()
    return queued_backlog >= queue_threshold


def _enqueue_submission_processing(submission_id, force_inline=False):
    from .tasks import process_submission as process_submission_task

    async_enabled = bool(getattr(settings, 'SUBMISSION_ASYNC_ENABLED', True))
    inline_fallback = bool(getattr(settings, 'SUBMISSION_PROCESS_INLINE_FALLBACK', True))
    submit_queue = getattr(settings, 'SUBMISSION_CELERY_QUEUE', 'submission_processing')

    if async_enabled and not force_inline:
        try:
            process_submission_task.apply_async(args=[submission_id], queue=submit_queue)
            return True
        except Exception:
            logger.exception("Failed to enqueue submission %s", submission_id)
            if not inline_fallback:
                return False

    try:
        process_submission_task(submission_id)
        return True
    except Exception:
        logger.exception("Synchronous post-processing failed for submission %s", submission_id)
        return False


def _is_soft_submission_trash_entry(trash_entry):
    if not trash_entry or getattr(trash_entry, 'item_type', None) != 'submission':
        return False
    item_data = getattr(trash_entry, 'item_data', None)
    return (
        isinstance(item_data, dict)
        and item_data.get('soft_delete_mode') == 'submission_is_deleted'
    )


def _mark_trash_entry_restored(trash_entry, user=None):
    if hasattr(trash_entry, 'restored'):
        trash_entry.restored = True
        trash_entry.restored_at = timezone.now()
        trash_entry.restored_by = user
        trash_entry.save(update_fields=['restored', 'restored_at', 'restored_by'])
    else:
        trash_entry.delete()


def _delete_soft_submission_for_trash_entry(trash_entry):
    if not _is_soft_submission_trash_entry(trash_entry):
        return

    item_data = trash_entry.item_data or {}
    submission_id = item_data.get('submission_id') or trash_entry.item_id
    if submission_id is None:
        return

    Submission.objects.filter(id=submission_id).delete()


def move_submission_to_soft_trash(submission, user=None):
    """
    Soft-delete a submission row and create a TrashBin entry for restore.
    """
    if not submission:
        return None

    if getattr(submission, 'is_deleted', False):
        return None

    auto_delete_date = timezone.now() + timedelta(days=30)
    instance_label = submission.instance_id or f"id:{submission.id}"
    form_name = submission.form.name if getattr(submission, 'form_id', None) and submission.form else 'Submission'
    display_name = f"{form_name} | {instance_label}"

    item_data = {
        'soft_delete_mode': 'submission_is_deleted',
        'submission_id': submission.id,
        'form_id': submission.form_id,
        'instance_id': submission.instance_id,
        'source': submission.source,
        'created_at': submission.created_at.isoformat() if submission.created_at else None,
    }

    with transaction.atomic():
        trash_entry = TrashBin.objects.create(
            item_type='submission',
            item_id=submission.id,
            item_data=item_data,
            item_name=display_name,
            deleted_by=user,
            auto_delete_at=auto_delete_date,
        )

        submission.is_deleted = True
        submission.save(update_fields=['is_deleted'])

    return trash_entry


def _create_submission_tombstone(form, instance_id, user=None):
    """
    Ensure a Submission row exists so legacy Form.submission entries
    can be soft-deleted via is_deleted markers.
    """
    normalized_instance_id = _normalize_instance_id(instance_id)
    if not normalized_instance_id:
        return None

    placeholder_payload = {'data': {'meta': {'instanceID': normalized_instance_id}}}
    submission, created = Submission.objects.get_or_create(
        form=form,
        instance_id=normalized_instance_id,
        defaults={
            'data': placeholder_payload,
            'user': user,
            'source': 'api',
        },
    )

    if not created and user and not submission.user_id:
        submission.user = user
        submission.save(update_fields=['user'])

    return submission

# Trash Bin Utility Functions
def move_to_trash(item, item_type, user=None):
    """
    Move an item to trash bin instead of permanently deleting it
    Also handles related items that would be cascade deleted
    """
    from django.core import serializers
    import logging

    logger = logging.getLogger(__name__)

    if item_type == 'submission':
        return move_submission_to_soft_trash(item, user)

    # Serialize the item data
    serialized_data = serializers.serialize('json', [item])
    item_data = json.loads(serialized_data)[0]

    # Determine display name
    display_name = getattr(item, 'name', None) or getattr(item, 'title', None) or str(item)

    # Calculate auto-delete date (30 days from now)
    auto_delete_date = timezone.now() + timedelta(days=30)

    # Store related items that would be cascade deleted
    related_items = []

    if item_type == 'project':
        # Handle project's related items
        logger.info(f"📦 Finding related items for project {item.id}")

        # Get all templates in this project
        templates = Template.objects.filter(project=item)
        for template in templates:
            template_data = serializers.serialize('json', [template])
            template_dict = json.loads(template_data)[0]

            # Manually add many-to-many field data
            template_dict['fields']['lookup_forms_ids'] = list(template.lookup_forms.values_list('id', flat=True))
            template_dict['fields']['generated_lookup_forms_ids'] = list(template.generated_lookup_forms.values_list('id', flat=True))

            logger.error(f"📄 FORCE LOG: Found template: {template.name} (M2M: lookup_forms={template.lookup_forms.all().count()}, generated={template.generated_lookup_forms.all().count()})")
            logger.error(f"📄 FORCE LOG: Template M2M IDs - lookup_forms={list(template.lookup_forms.values_list('id', flat=True))}, generated={list(template.generated_lookup_forms.values_list('id', flat=True))}")

            related_items.append({
                'type': 'template',
                'id': template.id,
                'name': template.name,
                'data': template_dict
            })
            logger.info(f"📄 Found template: {template.name} (M2M: lookup_forms={len(template.lookup_forms.all())}, generated={len(template.generated_lookup_forms.all())})")

        # Get all forms in this project
        forms = Form.objects.filter(project=item)
        for form in forms:
            form_data = serializers.serialize('json', [form])
            form_dict = json.loads(form_data)[0]

            # Manually add many-to-many field data for forms
            form_dict['fields']['other_languages_ids'] = list(form.other_languages.values_list('id', flat=True))

            related_items.append({
                'type': 'form',
                'id': form.id,
                'name': form.name,
                'data': form_dict
            })
            logger.info(f"📋 Found form: {form.name} (M2M: other_languages={len(form.other_languages.all())})")

    elif item_type == 'template':
        # Handle template's related forms that would be cascade deleted
        logger.info(f"📄 Finding related forms for template {item.id}")

        # We need to capture the template's M2M relationships before it's deleted
        template_lookup_forms_ids = list(item.lookup_forms.values_list('id', flat=True))
        template_generated_forms_ids = list(item.generated_lookup_forms.values_list('id', flat=True))

        logger.error(f"📄 FORCE LOG: Template M2M before deletion - lookup_forms={template_lookup_forms_ids}, generated={template_generated_forms_ids}")

        # Get data collection form
        if item.data_collection_form:
            form_data = serializers.serialize('json', [item.data_collection_form])
            form_dict = json.loads(form_data)[0]
            form_dict['fields']['other_languages_ids'] = list(item.data_collection_form.other_languages.values_list('id', flat=True))

            related_items.append({
                'type': 'form',
                'id': item.data_collection_form.id,
                'name': item.data_collection_form.name,
                'data': form_dict
            })
            logger.info(f"📋 Found data collection form: {item.data_collection_form.name}")

        # Get lookup forms
        for form in item.lookup_forms.all():
            form_data = serializers.serialize('json', [form])
            form_dict = json.loads(form_data)[0]
            form_dict['fields']['other_languages_ids'] = list(form.other_languages.values_list('id', flat=True))

            related_items.append({
                'type': 'form',
                'id': form.id,
                'name': form.name,
                'data': form_dict
            })
            logger.info(f"📋 Found lookup form: {form.name}")

        # Get generated lookup forms
        for form in item.generated_lookup_forms.all():
            form_data = serializers.serialize('json', [form])
            form_dict = json.loads(form_data)[0]
            form_dict['fields']['other_languages_ids'] = list(form.other_languages.values_list('id', flat=True))

            related_items.append({
                'type': 'form',
                'id': form.id,
                'name': form.name,
                'data': form_dict
            })
            logger.info(f"📋 Found generated lookup form: {form.name}")

        # Store the M2M relationships in the main template data so they can be restored
        # Add the M2M IDs to the existing item_data
        item_data['fields']['lookup_forms_ids'] = template_lookup_forms_ids
        item_data['fields']['generated_lookup_forms_ids'] = template_generated_forms_ids

        logger.error(f"📄 FORCE LOG: Added M2M to template data - lookup_forms_ids={template_lookup_forms_ids}, generated_lookup_forms_ids={template_generated_forms_ids}")    # For individual forms, no related items to capture
    # (forms don't have cascade relationships to other items)

    # Add related items to the main item data
    item_data['related_items'] = related_items
    logger.info(f"💾 Stored {len(related_items)} related items")

    # Create trash bin entry
    trash_entry = TrashBin.objects.create(
        item_type=item_type,
        item_id=item.id,
        item_data=item_data,
        item_name=display_name,
        deleted_by=user,
        auto_delete_at=auto_delete_date
    )

    # Delete the original item (this will cascade delete related items)
    item.delete()

    logger.info(f"🗑️ Moved {item_type} '{display_name}' to trash with {len(related_items)} related items")
    return trash_entry
    item.delete()

    return trash_entry

def restore_from_trash(trash_entry, user=None):
    """
    Restore an item from trash bin
    """
    from django.core import serializers
    from django.apps import apps
    import logging

    logger = logging.getLogger(__name__)
    logger.info(f"🔄 Starting restore for item {trash_entry.id}: {trash_entry.item_name}")

    if hasattr(trash_entry, 'restored') and trash_entry.restored:
        logger.error("❌ Item already restored")
        raise ValueError("Item has already been restored")

    if _is_soft_submission_trash_entry(trash_entry):
        item_data = trash_entry.item_data or {}
        submission_id = item_data.get('submission_id') or trash_entry.item_id
        submission = Submission.objects.filter(id=submission_id).first()
        if not submission:
            raise ValueError("Original submission no longer exists and cannot be restored")

        with transaction.atomic():
            submission.is_deleted = False
            submission.save(update_fields=['is_deleted'])
            _mark_trash_entry_restored(trash_entry, user)
        logger.info(f"✅ Restored soft-deleted submission {submission.id}")
        return submission

    # Get the model class
    model_map = {
        'project': 'Project',
        'form': 'Form',
        'submission': 'Submission',
        'patient': 'Patient',
        'organization': 'Organization',
        'template': 'Template',
    }

    model_name = model_map.get(trash_entry.item_type)
    if not model_name:
        logger.error(f"❌ Unknown item type: {trash_entry.item_type}")
        raise ValueError(f"Unknown item type: {trash_entry.item_type}")

    logger.info(f"📦 Restoring {model_name} with ID {trash_entry.item_id}")

    try:
        model_class = apps.get_model('api', model_name)
    except Exception as e:
        logger.error(f"❌ Failed to get model class: {e}")
        raise ValueError(f"Failed to get model class: {str(e)}")

    # Restore the item from serialized data
    item_data = trash_entry.item_data
    logger.info(f"📄 Item data type: {type(item_data)}")

    # Handle both old and new serialization formats
    if isinstance(item_data, list) and len(item_data) > 0:
        # Old format: Django serializer format
        fields = item_data[0]['fields']
        original_pk = item_data[0]['pk']
        logger.info("📋 Using old Django serializer format")
    elif isinstance(item_data, dict) and 'fields' in item_data:
        # Old format: Single item
        fields = item_data['fields']
        original_pk = item_data.get('pk', trash_entry.item_id)
        logger.info("📋 Using old single item format")
    else:
        # New format: Direct fields
        fields = item_data
        original_pk = trash_entry.item_id
        logger.info("📋 Using direct fields format")

    logger.info(f"🔧 Attempting to restore with ID {original_pk}")
    logger.info(f"🔧 Fields to restore: {list(fields.keys()) if isinstance(fields, dict) else 'Invalid fields'}")

    try:
        # Fix foreign key relationships before creating the object
        fixed_fields = {}
        m2m_fields = {}  # Store M2M fields for later processing

        for field_name, field_value in fields.items():
            # Check for our custom M2M IDs fields first (before trying to get field_obj)
            if field_name.endswith('_ids'):
                m2m_field_name = field_name[:-4]
                m2m_fields[m2m_field_name] = field_value
                logger.error(f"🔗 FORCE LOG: Found main item M2M field {m2m_field_name}: {field_value}")
                continue

            field_obj = None
            try:
                field_obj = model_class._meta.get_field(field_name)
            except:
                # Field doesn't exist, skip it
                continue

            # Check if this is a many-to-many field (actual Django M2M fields)
            if hasattr(field_obj, 'many_to_many') and field_obj.many_to_many:
                # Skip actual M2M fields since Django serializer includes them but we can't restore them directly
                continue

            if hasattr(field_obj, 'related_model') and field_obj.related_model:
                # This is a foreign key field
                if field_value is not None:
                    try:
                        # Get the related object by ID
                        related_obj = field_obj.related_model.objects.get(id=field_value)
                        fixed_fields[field_name] = related_obj
                        logger.info(f"🔗 Fixed FK {field_name}: {field_value} -> {related_obj}")
                    except field_obj.related_model.DoesNotExist:
                        logger.warning(f"⚠️ Related object {field_obj.related_model.__name__} with ID {field_value} not found")
                        fixed_fields[field_name] = None
                else:
                    fixed_fields[field_name] = None
            else:
                # Regular field, use as-is
                fixed_fields[field_name] = field_value

        # Ensure is_deleted field is set to False for all models that have it
        if hasattr(model_class, '_meta'):
            try:
                is_deleted_field = model_class._meta.get_field('is_deleted')
                fixed_fields['is_deleted'] = False
                logger.info(f"🔧 Set is_deleted=False for {model_class.__name__}")
            except:
                # Model doesn't have is_deleted field, which is fine
                pass

        logger.info(f"🔧 Fixed fields: {list(fixed_fields.keys())}")

        # Try to restore with original ID if possible
        restored_item = model_class.objects.create(id=original_pk, **fixed_fields)
        logger.info(f"✅ Restored with original ID {original_pk}")
    except Exception as e1:
        logger.warning(f"⚠️ Failed to restore with original ID: {e1}")
        # If original ID fails (conflict), create without ID
        try:
            # Remove problematic fields that might cause conflicts
            clean_fields = {k: v for k, v in fixed_fields.items()
                          if k not in ['id', 'created_at', 'updated_at']}
            # Ensure is_deleted is included in clean_fields if it exists
            if hasattr(model_class, '_meta'):
                try:
                    is_deleted_field = model_class._meta.get_field('is_deleted')
                    clean_fields['is_deleted'] = False
                    logger.info(f"🔧 Set is_deleted=False in clean_fields for {model_class.__name__}")
                except:
                    pass
            logger.info(f"🧹 Cleaned fields: {list(clean_fields.keys())}")
            restored_item = model_class.objects.create(**clean_fields)
            logger.info(f"✅ Restored with new ID {restored_item.id}")
        except Exception as e2:
            logger.error(f"❌ Failed to restore item: {e2}")
            raise ValueError(f"Failed to restore item: {str(e2)}")

    # Process M2M fields for the main restored item
    if m2m_fields:
        for m2m_field_name, m2m_ids in m2m_fields.items():
            if m2m_ids and isinstance(m2m_ids, list):
                try:
                    m2m_manager = getattr(restored_item, m2m_field_name)
                    # Look for IDs in existing objects (they should exist since this is the main item)
                    existing_objects = []

                    logger.error(f"🔍 FORCE LOG: Processing main item M2M {m2m_field_name}: {m2m_ids}")

                    for m2m_id in m2m_ids:
                        try:
                            field_obj = model_class._meta.get_field(m2m_field_name)
                            existing_obj = field_obj.related_model.objects.get(id=m2m_id)
                            existing_objects.append(existing_obj)
                            logger.info(f"✅ Main item M2M found: {m2m_id}")
                        except Exception as e:
                            logger.warning(f"⚠️ Main item M2M object not found: {m2m_id} - {e}")

                    if existing_objects:
                        m2m_manager.set(existing_objects)
                        logger.info(f"🔗 Main item: Set M2M {m2m_field_name}: {[obj.id for obj in existing_objects]}")
                    else:
                        logger.warning(f"⚠️ No valid M2M objects for main item {m2m_field_name}")

                except Exception as m2m_error:
                    logger.error(f"❌ Failed to set main item M2M {m2m_field_name}: {m2m_error}")

    # Restore related items if they exist
    related_items = item_data.get('related_items', [])
    if related_items:
        logger.error(f"🔄 FORCE LOG: Restoring {len(related_items)} related items in two phases")

        # Phase 1: Restore all items without cross-references to other deleted items
        restored_items_map = {}  # Map original_id -> restored_item
        items_to_update = []  # Items that need FK updates in phase 2

        for related_item in related_items:
            try:
                related_type = related_item['type']
                related_data = related_item['data']
                related_name = related_item['name']
                related_id = related_item['id']

                logger.error(f"📋 FORCE LOG: Phase 1: Restoring {related_type}: {related_name}")

                # Get the model class for the related item
                related_model_name = model_map.get(related_type)
                if not related_model_name:
                    logger.warning(f"⚠️ Unknown related item type: {related_type}")
                    continue

                related_model_class = apps.get_model('api', related_model_name)
                related_fields = related_data['fields']

                logger.error(f"🔍 FORCE LOG: Available fields in {related_type} {related_name}: {list(related_fields.keys())}")

                # Separate fields for phase 1 (safe to restore) and phase 2 (cross-references)
                phase1_fields = {}
                phase2_fields = {}
                m2m_fields = {}

                for field_name, field_value in related_fields.items():
                    # Check for our custom M2M IDs fields first (before trying to get field_obj)
                    if field_name.endswith('_ids'):
                        m2m_field_name = field_name[:-4]
                        m2m_fields[m2m_field_name] = field_value
                        logger.error(f"🔗 FORCE LOG: Found stored M2M field {m2m_field_name}: {field_value}")
                        continue

                    field_obj = None
                    try:
                        field_obj = related_model_class._meta.get_field(field_name)
                    except:
                        # Field doesn't exist in model (could be our custom fields), skip it
                        continue

                    # Check if this is a many-to-many field (actual Django M2M fields)
                    if hasattr(field_obj, 'many_to_many') and field_obj.many_to_many:
                        # Skip actual M2M fields since Django serializer includes them but we can't restore them directly
                        continue

                    # Handle foreign key fields
                    if hasattr(field_obj, 'related_model') and field_obj.related_model:
                        if field_value is not None:
                            # Special handling for project reference - use the restored item
                            if field_name == 'project' and related_type in ['template', 'form']:
                                phase1_fields[field_name] = restored_item
                                logger.info(f"🔗 Phase 1: Linked {related_type} to restored project")
                            else:
                                # Check if this references another item in our related_items list
                                cross_reference = any(item['id'] == field_value for item in related_items)
                                if cross_reference:
                                    # This is a cross-reference to another deleted item, handle in phase 2
                                    phase2_fields[field_name] = field_value
                                    logger.info(f"🔄 Phase 2 needed for {field_name}: {field_value}")
                                else:
                                    # This references an existing item, safe to restore now
                                    try:
                                        related_obj = field_obj.related_model.objects.get(id=field_value)
                                        phase1_fields[field_name] = related_obj
                                        logger.info(f"🔗 Phase 1: Found existing {field_name}: {field_value}")
                                    except field_obj.related_model.DoesNotExist:
                                        logger.warning(f"⚠️ Phase 1: Related object not found for {field_name}: {field_value}")
                                        phase1_fields[field_name] = None
                        else:
                            phase1_fields[field_name] = None
                    else:
                        phase1_fields[field_name] = field_value

                # Ensure is_deleted field is set to False for all models that have it
                if hasattr(related_model_class, '_meta'):
                    try:
                        is_deleted_field = related_model_class._meta.get_field('is_deleted')
                        phase1_fields['is_deleted'] = False
                        logger.info(f"🔧 Set is_deleted=False for related {related_model_class.__name__}")
                    except:
                        # Model doesn't have is_deleted field, which is fine
                        pass

                # Restore the item with phase 1 fields
                restored_related_item = None
                try:
                    restored_related_item = related_model_class.objects.create(id=related_id, **phase1_fields)
                    logger.info(f"✅ Phase 1: Restored {related_type} '{related_name}' with original ID {related_id}")
                except:
                    clean_fields = {k: v for k, v in phase1_fields.items()
                                  if k not in ['id', 'created_at', 'updated_at']}
                    # Ensure is_deleted is included in clean_fields if it exists
                    if hasattr(related_model_class, '_meta'):
                        try:
                            is_deleted_field = related_model_class._meta.get_field('is_deleted')
                            clean_fields['is_deleted'] = False
                        except:
                            pass
                    restored_related_item = related_model_class.objects.create(**clean_fields)
                    logger.info(f"✅ Phase 1: Restored {related_type} '{related_name}' with new ID {restored_related_item.id}")

                # Store for phase 2 updates
                if restored_related_item:
                    restored_items_map[related_id] = restored_related_item
                    if phase2_fields or m2m_fields:
                        items_to_update.append({
                            'item': restored_related_item,
                            'phase2_fields': phase2_fields,
                            'm2m_fields': m2m_fields,
                            'model_class': related_model_class,
                            'name': related_name,
                            'type': related_type
                        })

            except Exception as e:
                logger.error(f"❌ Phase 1: Failed to restore related item {related_name}: {e}")
                continue

        # Phase 2: Update cross-references and many-to-many fields
        logger.error(f"🔄 FORCE LOG: Phase 2: Updating {len(items_to_update)} items with cross-references")

        for update_info in items_to_update:
            try:
                item = update_info['item']
                phase2_fields = update_info['phase2_fields']
                m2m_fields = update_info['m2m_fields']
                model_class = update_info['model_class']
                name = update_info['name']
                item_type = update_info['type']

                # Update foreign key cross-references
                for field_name, field_value in phase2_fields.items():
                    if field_value in restored_items_map:
                        setattr(item, field_name, restored_items_map[field_value])
                        logger.info(f"🔗 Phase 2: Updated {item_type} {name}.{field_name} -> {restored_items_map[field_value].id}")
                    else:
                        logger.warning(f"⚠️ Phase 2: Cross-reference not found: {field_name} -> {field_value}")

                # Save the updated item
                if phase2_fields:
                    item.save()
                    logger.info(f"💾 Phase 2: Saved {item_type} {name} with updated cross-references")

                # Handle many-to-many fields
                for m2m_field_name, m2m_ids in m2m_fields.items():
                    if m2m_ids and isinstance(m2m_ids, list):
                        try:
                            m2m_manager = getattr(item, m2m_field_name)
                            # Look for IDs in both restored items and existing objects
                            existing_objects = []

                            logger.error(f"🔍 FORCE LOG: Processing M2M {m2m_field_name} for {item_type} {name}: {m2m_ids}")

                            for m2m_id in m2m_ids:
                                if m2m_id in restored_items_map:
                                    # This M2M references a restored item
                                    existing_objects.append(restored_items_map[m2m_id])
                                    logger.info(f"✅ M2M found in restored items: {m2m_id} -> {restored_items_map[m2m_id].id}")
                                else:
                                    # This M2M references an existing item
                                    try:
                                        field_obj = model_class._meta.get_field(m2m_field_name)
                                        existing_obj = field_obj.related_model.objects.get(id=m2m_id)
                                        existing_objects.append(existing_obj)
                                        logger.info(f"✅ M2M found in existing objects: {m2m_id}")
                                    except Exception as e:
                                        logger.warning(f"⚠️ M2M object not found: {m2m_id} - {e}")

                            if existing_objects:
                                m2m_manager.set(existing_objects)
                                logger.info(f"🔗 Phase 2: Set M2M {m2m_field_name}: {[obj.id for obj in existing_objects]}")
                            else:
                                logger.warning(f"⚠️ No valid M2M objects for {m2m_field_name}")

                        except Exception as m2m_error:
                            logger.error(f"❌ Failed to set M2M {m2m_field_name}: {m2m_error}")
                            import traceback
                            logger.error(traceback.format_exc())

            except Exception as e:
                logger.error(f"❌ Phase 2: Failed to update item {update_info['name']}: {e}")
                continue

        logger.info(f"🎉 Two-phase restoration complete: {len(restored_items_map)} items restored")

    _mark_trash_entry_restored(trash_entry, user)
    logger.info("✅ Marked trash entry as restored")

    logger.info(f"🎉 Successfully restored item {restored_item.id}")
    return restored_item

def cleanup_expired_trash():
    """
    Permanently delete expired items from trash bin
    """
    try:
        # Try to filter with restored field
        expired_entries = TrashBin.objects.filter(
            auto_delete_at__lte=timezone.now(),
            restored=False
        )
    except Exception:
        # If restored field doesn't exist, just filter by auto_delete_at
        expired_entries = TrashBin.objects.filter(
            auto_delete_at__lte=timezone.now()
        )

    expired_entry_list = list(expired_entries)
    count = len(expired_entry_list)

    for entry in expired_entry_list:
        _delete_soft_submission_for_trash_entry(entry)

    expired_entries.delete()

    return count

def clean_numeric_field(value):
    """
    Clean numeric fields by removing non-numeric characters
    """
    if not value:
        return None

    # Convert to string if not already
    str_value = str(value).strip()

    # Remove common non-numeric characters but keep decimal points
    cleaned = re.sub(r'[^\d.]', '', str_value)

    # Handle empty string after cleaning
    if not cleaned:
        return None

    try:
        # Try to convert to integer first (for age, etc.)
        if '.' not in cleaned:
            return int(cleaned)
        else:
            return float(cleaned)
    except (ValueError, TypeError):
        return None

def clean_text_field(value):
    """
    Clean text fields by removing extra whitespace and ensuring proper format
    """
    if not value:
        return None

    # Convert to string and clean up
    cleaned = str(value).strip()

    # Remove excessive whitespace
    cleaned = re.sub(r'\s+', ' ', cleaned)

    return cleaned if cleaned else None

def create_or_update_patient_from_submission(submission_data, user=None):
    """
    Check if submission contains user_identification_11_9943_01976848561 field.
    If yes, create or update patient record.
    Handles both direct fields and grouped fields.
    """
    PATIENT_ID_FIELD = 'user_identification_11_9943_01976848561'

    def find_field_in_data(data, field_name):
        """
        Recursively search for a field in submission data.
        Handles both direct fields and grouped/nested fields.
        """
        # Direct field access
        if field_name in data:
            print(f"[DEBUG] Found field '{field_name}' directly: {data[field_name]}")
            return data[field_name]

        # Search in grouped fields (group_name/field_name format)
        for key, value in data.items():
            if isinstance(value, dict):
                # Recursively search in nested dictionaries
                result = find_field_in_data(value, field_name)
                if result:
                    print(f"[DEBUG] Found field '{field_name}' in nested group '{key}': {result}")
                    return result
            elif '/' in key and key.split('/')[-1] == field_name:
                # Handle grouped field format: "group_name/field_name"
                print(f"[DEBUG] Found field '{field_name}' in grouped format '{key}': {value}")
                return value

        return None

    def find_patient_field(data, field_names):
        """Find patient field from a list of possible field names"""
        for field_name in field_names:
            value = find_field_in_data(data, field_name)
            if value:
                return value
        return None

    # Find patient ID in submission data
    patient_id = find_field_in_data(submission_data, PATIENT_ID_FIELD)
    print(f"[DEBUG] Looking for patient ID field '{PATIENT_ID_FIELD}' in submission: {patient_id}")

    if not patient_id:
        print(f"[DEBUG] Patient ID field not found in submission data keys: {list(submission_data.keys())}")
        return None

    print(f"[DEBUG] Found patient ID: {patient_id}, creating/updating patient record")

    try:
        # Clean and prepare patient data
        name = clean_text_field(find_patient_field(submission_data, ['name', 'patient_name', 'full_name']))
        email = clean_text_field(find_patient_field(submission_data, ['email', 'patient_email']))
        phone = clean_text_field(find_patient_field(submission_data, ['phone', 'patient_phone', 'contact', 'phone_number']))
        age = clean_numeric_field(find_patient_field(submission_data, ['age', 'patient_age']))
        gender = clean_text_field(find_patient_field(submission_data, ['gender', 'patient_gender']))
        address = clean_text_field(find_patient_field(submission_data, ['address', 'patient_address']))

        print(f"[DEBUG] Cleaned patient data - name: {name}, age: {age}, gender: {gender}, phone: {phone}")

        # Try to get existing patient
        patient, created = Patient.objects.get_or_create(
            patient_id=patient_id,
            defaults={
                'name': name,
                'email': email,
                'phone': phone,
                'age': age,
                'gender': gender,
                'address': address,
                'submission_data': submission_data,
                'created_by': user,
            }
        )

        if not created:
            # Update existing patient with latest submission data
            name = clean_text_field(find_patient_field(submission_data, ['name', 'patient_name', 'full_name']))
            if name:
                patient.name = name

            email = clean_text_field(find_patient_field(submission_data, ['email', 'patient_email']))
            if email:
                patient.email = email

            phone = clean_text_field(find_patient_field(submission_data, ['phone', 'patient_phone', 'contact', 'phone_number']))
            if phone:
                patient.phone = phone

            age = clean_numeric_field(find_patient_field(submission_data, ['age', 'patient_age']))
            if age is not None:
                patient.age = age

            gender = clean_text_field(find_patient_field(submission_data, ['gender', 'patient_gender']))
            if gender:
                patient.gender = gender

            address = clean_text_field(find_patient_field(submission_data, ['address', 'patient_address']))
            if address:
                patient.address = address

            patient.submission_data = submission_data
            patient.save()
            print(f"[DEBUG] Updated existing patient {patient_id}: {patient.name}")
        else:
            print(f"[DEBUG] Created new patient {patient_id}: {patient.name}")

        return patient

    except Exception as e:
        print(f"[ERROR] Failed to create/update patient with ID '{patient_id}': {e}")
        print(f"[ERROR] Submission data: {submission_data}")
        import traceback
        print(f"[ERROR] Full traceback: {traceback.format_exc()}")
        return None

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_template_by_id(request, template_id):
    cache_ttl = int(getattr(settings, 'SUBMISSION_TEMPLATE_CACHE_TTL_SECONDS', 30))
    cache_key = f"template-by-id:v1:{template_id}"
    cached = cache.get(cache_key)
    if cached is not None:
        return Response(cached, status=status.HTTP_200_OK)

    try:
        template = Template.objects.get(id=template_id)
        serializer = TemplateSerializer(template)
        payload = serializer.data
        cache.set(cache_key, payload, timeout=cache_ttl)
        return Response(payload, status=status.HTTP_200_OK)
    except Template.DoesNotExist:
        return Response({'error': 'Template not found'}, status=status.HTTP_404_NOT_FOUND)


def _extract_form_submission_uuid_from_questions(question_list):
    """Read data-collection instance UUID from generated follow-up form questions."""
    if not isinstance(question_list, list):
        return ''

    for question in question_list:
        if not isinstance(question, dict):
            continue
        if question.get('name') != 'data_collection_form_uuid':
            continue
        return _normalize_instance_id(question.get('default'))
    return ''


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_template_followup_forms(request, template_id):
    """
    Lightweight endpoint for mobile apps.
    Returns filter questions + user-accessible generated follow-up forms only.
    """
    user = request.user
    user_role = getattr(user, 'role', 4)
    cache_ttl = int(getattr(settings, 'SUBMISSION_TEMPLATE_CACHE_TTL_SECONDS', 30))
    cache_key = f"template-followup-forms:v1:template:{template_id}:user:{user.id}:role:{user_role}"
    cached = cache.get(cache_key)
    if cached is not None:
        return Response(cached, status=status.HTTP_200_OK)

    try:
        template = (
            Template.objects
            .select_related('data_collection_form')
            .only(
                'id',
                'name',
                'project_id',
                'data_collection_form_id',
                'data_collection_form__id',
                'data_collection_form__questions',
            )
            .get(id=template_id)
        )
    except Template.DoesNotExist:
        return Response({'error': 'Template not found'}, status=status.HTTP_404_NOT_FOUND)

    dc_questions = []
    if template.data_collection_form and isinstance(template.data_collection_form.questions, list):
        dc_questions = template.data_collection_form.questions

    filter_questions = []
    for question in dc_questions:
        if not isinstance(question, dict):
            continue

        q_type = (question.get('type') or '').lower()
        if q_type in {'begin_group', 'end_group', 'begin_repeat', 'end_repeat'}:
            continue
        if question.get('make_mandatory') is not True:
            continue

        q_name = question.get('name') or question.get('id')
        if not q_name:
            continue

        raw_label = question.get('label')
        if isinstance(raw_label, str):
            q_label = raw_label.strip() or q_name
        elif isinstance(raw_label, dict):
            q_label = next((str(v).strip() for v in raw_label.values() if str(v).strip()), q_name)
        else:
            q_label = q_name

        filter_questions.append({
            'name': q_name,
            'label': q_label,
            'type': question.get('type') or '',
            'make_mandatory': True,
        })

    generated_forms_qs = template.generated_lookup_forms.all()

    # Role-aware access filtering to avoid additional get-user-forms calls from app.
    if user_role in [4, 5]:
        if hasattr(user, 'profile'):
            generated_forms_qs = generated_forms_qs.filter(
                id__in=user.profile.forms.values_list('id', flat=True)
            )
        else:
            generated_forms_qs = generated_forms_qs.none()
    elif user_role == 6:
        project_ids = _get_assigned_project_ids(user)
        if project_ids and template.project_id in project_ids:
            generated_forms_qs = generated_forms_qs.filter(project_id__in=project_ids)
        else:
            generated_forms_qs = generated_forms_qs.none()

    forms = (
        generated_forms_qs
        .only('id', 'name', 'project_id', 'questions')
        .order_by('-id')
    )

    forms_payload = []
    for form in forms.iterator(chunk_size=500):
        forms_payload.append({
            'uid': form.id,
            'project': form.project_id,
            'template': template.id,
            'display_name': form.name,
            'submission_uuid': _extract_form_submission_uuid_from_questions(form.questions),
        })

    payload = {
        'template_id': template.id,
        'template_name': template.name,
        'filter_questions': filter_questions,
        'forms': forms_payload,
    }
    cache.set(cache_key, payload, timeout=cache_ttl)
    return Response(payload, status=status.HTTP_200_OK)
from rest_framework.decorators import api_view, permission_classes
def assign_template_to_user(user, template):
    """
    Helper function to assign a template to a user through their UserProfile.
    Creates UserProfile if it doesn't exist.
    """
    try:
        from .models import UserProfile

        # Get or create user profile
        profile, created = UserProfile.objects.get_or_create(user=user)

        # Check if template is already assigned
        if not profile.templates.filter(id=template.id).exists():
            profile.templates.add(template)
            profile.save()
            print(f"[DEBUG] Assigned template '{template.name}' (ID: {template.id}) to user '{user.username}'")
            return True
        else:
            print(f"[DEBUG] Template '{template.name}' (ID: {template.id}) already assigned to user '{user.username}'")
            return False
    except Exception as e:
        print(f"[ERROR] Failed to assign template {template.id} to user {user.username}: {e}")
        return False

def check_and_assign_template_for_submission(form, user):
    """
    Check if the submitted form is a data collection form of a template.
    If yes, assign that template to the user.
    """
    try:
        # Check if this form is a data collection form of any template
        if hasattr(form, 'template') and form.template:
            template = form.template
            # Check if this form is the data collection form of the template
            if template.data_collection_form and template.data_collection_form.id == form.id:
                print(f"[DEBUG] Form '{form.name}' (ID: {form.id}) is a data collection form of template '{template.name}' (ID: {template.id})")
                if user and user.is_authenticated:
                    assign_template_to_user(user, template)
                    return True
                else:
                    print(f"[DEBUG] User is not authenticated, skipping template assignment")
            else:
                print(f"[DEBUG] Form '{form.name}' (ID: {form.id}) is linked to template '{template.name}' but is not the data collection form")
        else:
            print(f"[DEBUG] Form '{form.name}' (ID: {form.id}) is not linked to any template")
    except Exception as e:
        print(f"[ERROR] Error checking template assignment for form {form.id}: {e}")

    return False


def assign_generated_forms_to_profiles(source_form, template, generated_lookup_form_ids, user=None):
    """
    Assign generated lookup forms to users who should see them.

    - Authenticated submitter: assign only to submitter profile (existing behavior).
    - Anonymous submitter: assign to profiles already assigned to the source form/template.
    """
    if not generated_lookup_form_ids:
        return 0

    profile_ids = set()

    if user and user.is_authenticated and hasattr(user, 'profile'):
        profile_ids.add(user.profile.id)
    else:
        profile_ids.update(
            UserProfile.objects.filter(forms=source_form).values_list('id', flat=True)
        )
        if template:
            profile_ids.update(
                UserProfile.objects.filter(templates=template).values_list('id', flat=True)
            )

    if not profile_ids:
        return 0

    assigned_profiles = 0
    for profile in UserProfile.objects.filter(id__in=profile_ids):
        existing_ids = set(
            profile.forms.filter(id__in=generated_lookup_form_ids).values_list('id', flat=True)
        )
        missing_ids = [form_id for form_id in generated_lookup_form_ids if form_id not in existing_ids]
        if missing_ids:
            profile.forms.add(*missing_ids)
            assigned_profiles += 1

    return assigned_profiles


def process_submission_side_effects(submission_id):
    """
    Run expensive, non-ack-critical submission side effects.
    """
    submission = Submission.objects.select_related('form', 'user').filter(id=submission_id).first()
    if not submission or not submission.form:
        return

    form = submission.form
    user = submission.user
    data = _extract_submission_data(submission.data or {})
    instance_id, data = _ensure_instance_id(data)

    meta = data.get('meta')
    if not isinstance(meta, dict):
        meta = {}
    if user and user.is_authenticated:
        meta.setdefault('submitted_by', user.username)
        meta.setdefault('submission_type', 'authenticated')
    else:
        meta.setdefault('submitted_by', 'anonymous')
        meta.setdefault('submission_type', 'anonymous')
    data['meta'] = meta

    submission.instance_id = instance_id
    submission.data = {'data': data}
    submission.save(update_fields=['instance_id', 'data'])

    # Template assignment and generated lookup-form flow.
    check_and_assign_template_for_submission(form, user)

    GENERATED_LOOKUP_SYSTEM_GROUP_LABEL_EN = "Go ahead for follow-up questions."
    GENERATED_LOOKUP_SYSTEM_GROUP_LABEL_BN = "ফলো-আপ প্রশ্নগুলির জন্য এগিয়ে যান।"
    GENERATED_LOOKUP_SYSTEM_GROUP_TRANSLATIONS = {
        "English (en)": GENERATED_LOOKUP_SYSTEM_GROUP_LABEL_EN,
        "Bangla (bn)": GENERATED_LOOKUP_SYSTEM_GROUP_LABEL_BN,
    }

    try:
        uuid_val = meta.get('instanceID', '')
        template = getattr(form, 'template', None)
        template_lookup_forms = list(template.lookup_forms.all()) if template else []
        generated_lookup_form_ids = []

        if uuid_val and template_lookup_forms:
            for template_lookup_form in template_lookup_forms:
                criteria = getattr(template_lookup_form, "criteria", None)

                if criteria:
                    question_name = criteria.get("question_name")
                    expected_value = criteria.get("expected_value")
                    submission_value = _find_field_in_payload(data, question_name)
                    criteria_met = False
                    if isinstance(expected_value, list):
                        if isinstance(submission_value, list):
                            criteria_met = all(opt in submission_value for opt in expected_value)
                        elif isinstance(submission_value, str):
                            submission_list = submission_value.split()
                            criteria_met = all(opt in submission_list for opt in expected_value)
                    else:
                        criteria_met = (str(submission_value) == str(expected_value)) if submission_value is not None else False

                    if not criteria_met:
                        print(
                            f"[DEBUG] Skipping lookup form template {template_lookup_form.id} for submission {submission.id}: "
                            f"criteria {question_name} expected={expected_value} got={submission_value}"
                        )
                        continue

                system_questions = [{
                    'type': 'text',
                    'name': 'data_collection_form_uuid',
                    'default': uuid_val,
                    'read_only': 'yes',
                    'label': 'Data Collection Form UUID',
                }]

                data_form = template.data_collection_form if template else None
                if data_form:
                    for q in (data_form.questions or []):
                        if q.get('make_mandatory') is True:
                            qname = q.get('name')
                            value = _find_field_in_payload(data, qname)
                            newq = q.copy()
                            newq['read_only'] = 'yes'
                            newq['default'] = value if value is not None else ''
                            system_questions.append(newq)

                template_questions = [q.copy() for q in template_lookup_form.questions] if template_lookup_form.questions else []
                existing_names = {
                    str(q.get('name')).strip()
                    for q in (system_questions + template_questions)
                    if q.get('name')
                }
                group_name = "prefilled_system_questions"
                suffix = 2
                while group_name in existing_names:
                    group_name = f"prefilled_system_questions_{suffix}"
                    suffix += 1

                questions = [{
                    'type': 'begin_group',
                    'name': group_name,
                    'label': GENERATED_LOOKUP_SYSTEM_GROUP_LABEL_EN,
                    'appearance': 'field-list',
                    'read_only': 'yes',
                    'translations': GENERATED_LOOKUP_SYSTEM_GROUP_TRANSLATIONS.copy(),
                }]
                questions.extend(system_questions)
                questions.append({
                    'type': 'end_group',
                    'name': '',
                    'label': '',
                })

                for q in template_questions:
                    qname = q.get('name')
                    if not any(qq.get('name') == qname for qq in questions):
                        questions.append(q)

                data_collection_form_name = template.data_collection_form.name if template and template.data_collection_form else "Data Collection Form"
                hh_id = _find_field_in_payload(data, "hh_id")
                hh_head_name = _find_field_in_payload(data, "hh_head_name")
                submission_date = _extract_submission_date({'data': data})
                if submission_date is None and submission.created_at:
                    submission_date = timezone.localdate(submission.created_at)
                submission_date_text = submission_date.strftime('%Y-%m-%d') if submission_date else ''

                if hh_id and hh_head_name:
                    lookup_form_name = f"{data_collection_form_name} - {hh_id} - {hh_head_name}"
                else:
                    lookup_form_name = f"{data_collection_form_name} {uuid_val}"

                if submission_date_text:
                    lookup_form_name = f"{lookup_form_name} - {submission_date_text}"

                form_translations = {}
                if template_lookup_form and hasattr(template_lookup_form, 'translations') and template_lookup_form.translations:
                    form_translations = template_lookup_form.translations.copy()

                if not form_translations and template_lookup_form and hasattr(template_lookup_form, 'questions') and template_lookup_form.questions:
                    for question in template_lookup_form.questions:
                        question_name = question.get('name')
                        question_translations = question.get('translations', {})
                        if question_name and question_translations:
                            form_translations[question_name] = question_translations
                    for question in template_lookup_form.questions:
                        question_name = question.get('name')
                        options = question.get('options', [])
                        if question_name and options:
                            for option in options:
                                option_name = option.get('name')
                                option_translations = option.get('translations', {})
                                if option_name and option_translations:
                                    option_key = f"{question_name}_{option_name}"
                                    form_translations[option_key] = option_translations

                if not isinstance(form_translations, dict):
                    form_translations = {}
                existing_group_translations = form_translations.get(group_name, {})
                if not isinstance(existing_group_translations, dict):
                    existing_group_translations = {}
                form_translations[group_name] = {
                    **existing_group_translations,
                    **GENERATED_LOOKUP_SYSTEM_GROUP_TRANSLATIONS,
                }

                lookup_form = Form.objects.create(
                    project=template.project,
                    name=lookup_form_name,
                    description=template_lookup_form.description,
                    questions=questions,
                    form_style='pages',
                    translations=form_translations,
                    submission=[],
                    enketo_id=None,
                    allow_anonymous_submissions=False,
                    template=None
                )

                if template and template.data_collection_form:
                    data_form = template.data_collection_form
                    lookup_form.default_language = data_form.default_language
                    lookup_form.other_languages.set(data_form.other_languages.all())
                    lookup_form.save()

                template.generated_lookup_forms.add(lookup_form)
                generated_lookup_form_ids.append(lookup_form.id)
                print(
                    f"[DEBUG] Generated lookup form {lookup_form.id} from template lookup form {template_lookup_form.id} "
                    f"for submission {submission.id}"
                )

                try:
                    data_form = template.data_collection_form if template else None
                    default_language = data_form.default_language if data_form else None
                    other_languages = data_form.other_languages.all() if data_form else []

                    if form_translations:
                        for question in questions:
                            question_name = question.get('name')
                            if question_name and question_name in form_translations:
                                if 'translations' not in question:
                                    question['translations'] = {}
                                question_translations = form_translations[question_name]
                                for lang_key, translation_text in question_translations.items():
                                    question['translations'][lang_key] = translation_text

                    generate_xlsx_file(lookup_form, questions, default_language, other_languages, lookup_form_name)
                except Exception as exc:
                    print(f"[DEBUG] Error updating XLSX for lookup form {lookup_form.id}: {exc}")

            template.save()
            try:
                assigned_profiles = assign_generated_forms_to_profiles(
                    source_form=form,
                    template=template,
                    generated_lookup_form_ids=generated_lookup_form_ids,
                    user=user,
                )
                if assigned_profiles:
                    print(
                        f"[DEBUG] Assigned generated lookup forms {generated_lookup_form_ids} "
                        f"to {assigned_profiles} profile(s)"
                    )
            except Exception as exc:
                print(f"[DEBUG] Error updating profile.forms for generated lookup forms: {exc}")
    except Exception as exc:
        print(f"Error generating lookup forms for submission: {exc}")

    try:
        if submission.xml_file:
            with submission.xml_file.open('rb') as xml_stream:
                save_submission_to_xlsx(xml_stream)
    except Exception as exc:
        print(f"Error writing submission XLSX: {exc}")

    try:
        patient = create_or_update_patient_from_submission(data, user)
        if patient:
            patient.submissions.add(submission)
    except Exception as exc:
        print(f"Error creating/updating patient record: {exc}")

def _flatten_form_payload(payload, prefix=""):
    """Flatten nested dictionaries while preserving prefix structure."""
    flat = {}
    if not isinstance(payload, dict):
        return flat

    for key, value in payload.items():
        if key == 'meta':
            continue

        new_key = f"{prefix}{key}" if not prefix else f"{prefix}_{key}"

        if isinstance(value, dict):
            flat.update(_flatten_form_payload(value, new_key))
        elif isinstance(value, list):
            flat[new_key] = json.dumps(value)
        else:
            flat[new_key] = value

    return flat

def _extract_question_sequence(question_list):
    """Return flattened field order following the XLSForm question sequence."""
    if not isinstance(question_list, list):
        return []

    order = []
    group_stack = []

    for question in question_list:
        if not isinstance(question, dict):
            continue

        q_type = (question.get('type') or '').lower()
        name = question.get('name') or question.get('id')

        if q_type in {'begin_group', 'begin_repeat'}:
            if name:
                group_stack.append(name)
            continue

        if q_type in {'end_group', 'end_repeat'}:
            if group_stack:
                group_stack.pop()
            continue

        if not name:
            continue

        prefix_parts = group_stack + [name]
        field_name = "_".join(part for part in prefix_parts if part)

        if field_name and field_name not in order:
            order.append(field_name)

    return order

def _extract_question_labels(question_list):
    """Return mapping from field names to labels following the XLSForm question sequence."""
    if not isinstance(question_list, list):
        return {}

    field_to_label = {}
    group_stack = []

    for question in question_list:
        if not isinstance(question, dict):
            continue

        q_type = (question.get('type') or '').lower()
        name = question.get('name') or question.get('id')

        if q_type in {'begin_group', 'begin_repeat'}:
            if name:
                group_stack.append(name)
            continue

        if q_type in {'end_group', 'end_repeat'}:
            if group_stack:
                group_stack.pop()
            continue

        if not name:
            continue

        prefix_parts = group_stack + [name]
        field_name = "_".join(part for part in prefix_parts if part)

        field_label = _resolve_field_label(question, name)

        if field_name:
            field_to_label[field_name] = field_label

        # Also handle sub-questions if they exist
        if 'subQuestions' in question and isinstance(question['subQuestions'], list):
            for sub_question in question['subQuestions']:
                if isinstance(sub_question, dict):
                    sub_name = sub_question.get('name') or sub_question.get('id')
                    if sub_name:
                        sub_prefix_parts = group_stack + [name, sub_name]
                        sub_field_name = "_".join(part for part in sub_prefix_parts if part)

                        sub_field_label = _resolve_field_label(sub_question, sub_name)

                        if sub_field_name:
                            field_to_label[sub_field_name] = sub_field_label

    return field_to_label

def _contains_bangla_text(value):
    if not isinstance(value, str):
        return False
    return bool(re.search(r'[\u0980-\u09FF]', value))

def _is_english_language_key(raw_key):
    key = str(raw_key or '').strip().lower()
    if not key:
        return False

    if 'english' in key:
        return True

    normalized = key.replace('_', '-')
    if normalized == 'en' or normalized.startswith('en-'):
        return True

    for code in re.findall(r'\(([a-zA-Z_-]+)\)', key):
        normalized_code = code.replace('_', '-').strip().lower()
        if normalized_code == 'en' or normalized_code.startswith('en-'):
            return True

    return False

def _resolve_label_from_map(label_map, english_only=False):
    if not isinstance(label_map, dict):
        return ''

    for key, candidate in label_map.items():
        if _is_english_language_key(key) and isinstance(candidate, str) and candidate.strip():
            return candidate.strip()

    if english_only:
        return ''

    for candidate in label_map.values():
        if isinstance(candidate, str) and candidate.strip():
            return candidate.strip()

    return ''

def _resolve_choice_label(label_value, fallback='', english_only=False):
    """Resolve choice label text from string/dict translation structures."""
    if isinstance(label_value, str):
        cleaned = label_value.strip()
        if not cleaned:
            return fallback
        if english_only and _contains_bangla_text(cleaned):
            return fallback
        return cleaned

    if isinstance(label_value, dict):
        resolved = _resolve_label_from_map(label_value, english_only=english_only)
        if resolved:
            return resolved

    return fallback

def _resolve_field_label(question, fallback_name):
    """
    Resolve question/sub-question label with English-first preference.
    Falls back safely when no English value exists.
    """
    label_value = question.get('label', '')
    translation_value = question.get('translations')

    # 1) Prefer explicit English variants from label maps like "English (en)".
    label_english = _resolve_choice_label(label_value, '', english_only=True)
    if label_english:
        return label_english

    # 2) If base label is non-English (e.g., Bangla), prefer English translation when available.
    translation_english = _resolve_choice_label(
        translation_value,
        '',
        english_only=True,
    )
    if translation_english:
        return translation_english

    # 3) Fall back to base label, then any translation, then the field name.
    base_label = _resolve_choice_label(label_value, '')
    if base_label:
        return base_label

    any_translation = _resolve_choice_label(translation_value, '')
    if any_translation:
        return any_translation

    return fallback_name

def _normalize_option_alias(value):
    if value is None:
        return ''
    cleaned = re.sub(r'\s+', ' ', str(value)).strip()
    return cleaned.lower() if cleaned else ''

def _build_select_field_option_maps(question_list):
    """
    Build a map for select questions so submitted option codes can be converted
    to display labels in All Rows APIs.
    """
    if not isinstance(question_list, list):
        return {}

    select_maps = {}
    group_stack = []

    def register_select_question(field_name, question):
        if not field_name or not isinstance(question, dict):
            return

        q_type = (question.get('type') or '').lower()
        if 'select_one' not in q_type and 'select_multiple' not in q_type:
            return

        options = question.get('options') or question.get('choices') or []
        option_map = {}
        option_alias_map = {}

        for option in options:
            if not isinstance(option, dict):
                continue

            option_name = option.get('name')
            if option_name in (None, ''):
                continue

            option_name = str(option_name).strip()
            if not option_name:
                continue

            option_label_from_label = _resolve_choice_label(
                option.get('label'),
                ''
            )
            option_label_from_translation_english = _resolve_choice_label(
                option.get('translations'),
                '',
                english_only=True,
            )
            option_label_from_translation_any = _resolve_choice_label(
                option.get('translations'),
                ''
            )

            option_label = (
                option_label_from_translation_english
                or option_label_from_label
                or option_label_from_translation_any
                or option_name
            )

            option_map[option_name] = option_label or option_name

            alias_candidates = [
                option_name,
                option.get('value'),
                option_label,
                option_label_from_label,
                option_label_from_translation_any,
            ]

            for source in (option.get('label'), option.get('translations')):
                if isinstance(source, dict):
                    alias_candidates.extend(
                        value
                        for value in source.values()
                        if isinstance(value, str) and value.strip()
                    )
                elif isinstance(source, str):
                    alias_candidates.append(source)

            for alias_candidate in alias_candidates:
                normalized_alias = _normalize_option_alias(alias_candidate)
                if normalized_alias and normalized_alias not in option_alias_map:
                    option_alias_map[normalized_alias] = option_label

        select_maps[field_name] = {
            'type': q_type,
            'options': option_map,
            'aliases': option_alias_map,
        }

    for question in question_list:
        if not isinstance(question, dict):
            continue

        q_type = (question.get('type') or '').lower()
        name = question.get('name') or question.get('id')

        if q_type in {'begin_group', 'begin_repeat'}:
            if name:
                group_stack.append(name)
            continue

        if q_type in {'end_group', 'end_repeat'}:
            if group_stack:
                group_stack.pop()
            continue

        if not name:
            continue

        prefix_parts = group_stack + [name]
        field_name = "_".join(part for part in prefix_parts if part)
        register_select_question(field_name, question)

        if 'subQuestions' in question and isinstance(question['subQuestions'], list):
            for sub_question in question['subQuestions']:
                if not isinstance(sub_question, dict):
                    continue
                sub_name = sub_question.get('name') or sub_question.get('id')
                if not sub_name:
                    continue

                sub_prefix_parts = group_stack + [name, sub_name]
                sub_field_name = "_".join(
                    part for part in sub_prefix_parts if part
                )
                register_select_question(sub_field_name, sub_question)

    return select_maps

def _convert_select_value_to_label(value, question_type, option_map, option_alias_map=None):
    """Convert select_one/select_multiple stored values to readable labels."""
    if value in (None, '') or not isinstance(option_map, dict) or not option_map:
        return value

    q_type = (question_type or '').lower()

    option_alias_map = option_alias_map if isinstance(option_alias_map, dict) else {}

    def _convert_single_option(raw_option):
        option_name = str(raw_option or '').strip()
        if not option_name:
            return option_name

        if option_name in option_map:
            return option_map[option_name]

        normalized_alias = _normalize_option_alias(option_name)
        if normalized_alias and normalized_alias in option_alias_map:
            return option_alias_map[normalized_alias]

        return option_name

    if 'select_one' in q_type:
        option_name = value.strip() if isinstance(value, str) else str(value).strip()
        if not option_name:
            return value
        converted = _convert_single_option(option_name)
        return converted if converted != '' else value

    if 'select_multiple' in q_type:
        option_names = []

        if isinstance(value, list):
            option_names = [str(item).strip() for item in value if str(item).strip()]
        elif isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return value

            direct_converted = _convert_single_option(cleaned)
            if direct_converted != cleaned:
                return direct_converted

            parsed = None
            if cleaned.startswith('[') and cleaned.endswith(']'):
                try:
                    loaded = json.loads(cleaned)
                    if isinstance(loaded, list):
                        parsed = loaded
                except (TypeError, ValueError, json.JSONDecodeError):
                    parsed = None

            if parsed is not None:
                option_names = [str(item).strip() for item in parsed if str(item).strip()]
            else:
                option_names = [name for name in re.split(r'[\s,]+', cleaned) if name]
        else:
            option_name = str(value).strip()
            option_names = [option_name] if option_name else []

        converted = [_convert_single_option(name) for name in option_names]
        return ', '.join(converted)

    return value

def _apply_select_option_labels(flat_data, select_field_option_maps):
    """Mutate flattened row values to show labels for select question answers."""
    if not isinstance(flat_data, dict) or not isinstance(select_field_option_maps, dict):
        return flat_data

    for row_key, row_value in list(flat_data.items()):
        if row_value in (None, ''):
            continue

        # Primary match: exact key.
        meta = select_field_option_maps.get(row_key)

        # Fallback: question keys may include group prefixes while row keys do not.
        # Example: `suspected_patient-related_information_reporting_sites` -> `reporting_sites`
        if not meta:
            matches = [
                (field_key, field_meta)
                for field_key, field_meta in select_field_option_maps.items()
                if isinstance(field_key, str)
                and field_key.endswith(f"_{row_key}")
            ]
            if len(matches) == 1:
                meta = matches[0][1]

        if not meta:
            continue

        flat_data[row_key] = _convert_select_value_to_label(
            row_value,
            meta.get('type', ''),
            meta.get('options', {}),
            meta.get('aliases', {}),
        )

    return flat_data

def _inject_meta_fields(source_payload, target, prefix=""):
    """Copy meta fields into the flattened dictionary to aid filtering."""
    if not isinstance(source_payload, dict):
        return target

    meta = source_payload.get('meta', {}) if isinstance(source_payload.get('meta'), dict) else {}
    mapping = {
        'instanceID': 'instanceID',
        'submitted_by': 'submitted_by',
        'submission_type': 'submission_type',
    }

    for meta_key, suffix in mapping.items():
        if meta.get(meta_key):
            meta_value = meta.get(meta_key)
            if prefix:
                target[f"{prefix}_{suffix}"] = meta_value
            else:
                target[suffix] = meta_value

    return target

def _format_followup_summary(form_name, flat_data):
    """Return a readable summary string for a single follow-up submission."""
    if not isinstance(flat_data, dict):
        return form_name or 'Follow Up'

    preferred_keys = [
        'follow_up_date',
        'follow_up_status_2472_h',
        'suspected_in_the_disease',
        'if_yes__name_of_the_diseases_1',
        'remarks',
        'hh_head_name',
        'hh_id',
        'submitAs',
        'submitted_by',
        'submission_type',
        'start',
        'end',
    ]

    summary_lines = [f"Form: {form_name}"]

    for key in preferred_keys:
        value = flat_data.get(key)
        if value not in (None, ''):
            summary_lines.append(f"- {key}: {value}")

    remaining = {
        key: value for key, value in flat_data.items()
        if key not in preferred_keys and value not in (None, '')
    }

    if remaining:
        remaining_str = ', '.join(f"{key}: {value}" for key, value in remaining.items())
        if remaining_str:
            summary_lines.append(f"- other: {remaining_str}")

    return '\n'.join(summary_lines)

def _build_combined_template_rows(templates, filter_followup='all', followup_form_ids=None):
    """Create flattened rows that combine data collection + follow-up submissions."""
    followup_form_ids = followup_form_ids or set()
    combined_rows = []
    field_order = list(BASE_ROW_FIELDS)
    followup_forms_catalog = {}
    field_labels = {}

    # Create default labels for base fields
    default_labels = {
        'row_number': 'Row Number',
        'Validation': 'Validation',
        '_id': 'ID',
        '_version': 'Version',
        'group_name': 'Group Name',
        'formType': 'Form Type',
        'has_followup': 'Has Follow-up',
        'followup_count': 'Follow-up Count',
        'followup_forms': 'Follow-up Forms',
        'followup_follow_up_date': 'Follow-up Date',
        'followup_start': 'Follow-up start',
        'followup_end': 'Follow-up end',
        'followup_follow_up_status_2472_h': 'Follow-up Status (24-72h)',
        'followup_suspected_in_the_disease': 'Follow-up: Suspected Disease',
        'followup_if_yes__name_of_the_diseases_1': 'Follow-up: Disease Name',
        'followup_hh_head_name': 'Follow-up: Household Head Name',
        'followup_hh_id': 'Follow-up: Household ID',
        'instanceID': 'Instance ID',
        'submitted_by': 'Submitted By',
        'start': 'Start Time',
        'end': 'End Time'
    }
    field_labels.update(default_labels)

    for template in templates:
        dc_form_id = getattr(template, 'data_collection_form_id', None)
        if not dc_form_id:
            continue

        dc_form = Form.objects.only('id', 'name', 'questions').filter(id=dc_form_id).first()
        if not dc_form:
            continue

        dc_questions = dc_form.questions or []
        question_sequence = _extract_question_sequence(dc_questions)
        dc_labels = _extract_question_labels(dc_questions)
        dc_select_field_option_maps = _build_select_field_option_maps(dc_questions)
        field_labels.update(dc_labels)

        for field_name in question_sequence:
            if field_name not in field_order:
                field_order.append(field_name)

        dc_submissions = _get_form_submission_entries(dc_form)
        if not dc_submissions:
            continue

        lookup_form_ids = list(template.lookup_forms.values_list('id', flat=True))
        generated_form_ids = list(template.generated_lookup_forms.values_list('id', flat=True))
        followup_form_ids_all = lookup_form_ids + generated_form_ids

        followup_lookup = {}
        followup_form_name_map = {}
        followup_select_option_maps_by_form = {}
        if followup_form_ids_all:
            (
                followup_lookup,
                followup_form_name_map,
                linked_followup_forms,
            ) = _collect_followup_lookup_for_forms(followup_form_ids_all)
            followup_forms_catalog.update(linked_followup_forms)

            followup_forms = Form.objects.only('id', 'questions').filter(
                id__in=followup_form_ids_all
            )
            followup_select_option_maps_by_form = {
                followup_form.id: _build_select_field_option_maps(
                    followup_form.questions or []
                )
                for followup_form in followup_forms
            }

        for submission in dc_submissions:
            if not isinstance(submission, dict):
                continue

            submission_data = submission.get('data', {})
            if not isinstance(submission_data, dict) or not submission_data:
                continue

            uuid = submission_data.get('meta', {}).get('instanceID') or submission_data.get('instanceID')
            if not uuid:
                continue

            base_row = _flatten_form_payload(submission_data)
            _inject_meta_fields(submission_data, base_row)
            _apply_select_option_labels(base_row, dc_select_field_option_maps)

            # Prepare placeholders for consistent ordering
            if 'row_number' not in base_row:
                base_row['row_number'] = None
            if 'followup_forms' not in base_row:
                base_row['followup_forms'] = ''
            if 'followup_count' not in base_row:
                base_row['followup_count'] = 0
            if 'has_followup' not in base_row:
                base_row['has_followup'] = False

            patient_name = (
                submission_data.get('patient_name')
                or submission_data.get('name')
                or submission_data.get('patient')
                or submission_data.get('administrative', {}).get('name_of_the_person_with_suspected_case')
                or submission_data.get('hh_head_name')
                or "Unknown Patient"
            )

            base_row['group_name'] = f"{dc_form.name} - {patient_name}"
            base_row['formType'] = 'Data Collection'

            associated_followups = followup_lookup.get(uuid, [])
            followup_ids_present = set()
            followup_names = []
            followup_primary_summary = ''

            # Initialize individual followup fields to empty values
            followup_fields = [
                'follow_up_date',
                'start',
                'end',
                'follow_up_status_2472_h',
                'suspected_in_the_disease',
                'if_yes__name_of_the_diseases_1',
                'hh_head_name',
                'hh_id'
            ]

            # Initialize followup columns with empty values
            for field in followup_fields:
                base_row[f'followup_{field}'] = ''

            for followup_entry in associated_followups:
                followup_form_id = followup_entry.get('form_id')
                form_name = followup_form_name_map.get(followup_form_id) or 'Follow Up'
                followup_names.append(form_name)
                if followup_form_id:
                    try:
                        followup_ids_present.add(int(followup_form_id))
                    except (TypeError, ValueError):
                        pass

            # Row-level follow-up fields come from the oldest follow-up only.
            if associated_followups:
                primary_followup_entry = associated_followups[0]
                followup_form_id = primary_followup_entry.get('form_id')
                followup_data = primary_followup_entry.get('submission_data') or {}
                form_name = followup_form_name_map.get(followup_form_id) or 'Follow Up'

                flat_followup = _flatten_form_payload(followup_data)
                _inject_meta_fields(followup_data, flat_followup)
                followup_form_id_int = None
                if followup_form_id:
                    try:
                        followup_form_id_int = int(followup_form_id)
                    except (TypeError, ValueError):
                        followup_form_id_int = None
                _apply_select_option_labels(
                    flat_followup,
                    followup_select_option_maps_by_form.get(followup_form_id_int, {})
                )
                followup_primary_summary = _format_followup_summary(form_name, flat_followup)

                for field in followup_fields:
                    field_value = flat_followup.get(field)
                    if field_value not in (None, ''):
                        base_row[f'followup_{field}'] = field_value

            has_followup = len(associated_followups) > 0

            if filter_followup == 'with' and not has_followup:
                continue
            if filter_followup == 'without' and has_followup:
                continue
            if followup_form_ids and not followup_ids_present.intersection(followup_form_ids):
                continue

            base_row['has_followup'] = has_followup
            base_row['followup_count'] = len(associated_followups)
            base_row['followup_forms'] = ', '.join(sorted(set(followup_names))) if followup_names else ''
            base_row['followup_details'] = followup_primary_summary

            ordered_keys = []
            for key in BASE_ROW_FIELDS:
                if key in base_row and key not in ordered_keys:
                    ordered_keys.append(key)
            for key in question_sequence:
                if key in base_row and key not in ordered_keys:
                    ordered_keys.append(key)
            for key in base_row.keys():
                if key not in ordered_keys:
                    ordered_keys.append(key)

            ordered_row = {key: base_row[key] for key in ordered_keys}

            for key in ordered_row.keys():
                if key not in field_order:
                    field_order.append(key)

            combined_rows.append(ordered_row)

    combined_rows.sort(key=lambda row: (
        row.get('end') or '',
        row.get('start') or '',
        row.get('instanceID') or ''
    ), reverse=True)

    total_rows = len(combined_rows)
    for idx, row in enumerate(combined_rows):
        row['row_number'] = total_rows - idx
        for key in row.keys():
            if key not in field_order:
                field_order.append(key)

    # Apply the proper field ordering with data collection fields first, then followup fields
    ordered_field_list = get_ordered_field_list(field_order)

    field_labels.update(
        {
            "total_reporting_sites": "total reporting sites",
            "age_month": "Age Month",
            "year_age": "Year age",
            "followup_suspected_in_the_disease": "Follow-up: Suspected Disease",
            "followup_hh_head_name": "Follow-up: Household Head Name",
            "followup_hh_id": "Follow-up: Household ID",
        }
    )

    return combined_rows, ordered_field_list, followup_forms_catalog, field_labels


def _is_brac_download_user(user):
    username = getattr(settings, "BRAC_DOWNLOAD_USERNAME", "").strip()
    return bool(username) and bool(user and user.is_authenticated and user.username == username)


def _has_project_scope_access(user, project_id):
    user_role = getattr(user, "role", 4)
    if user_role == 6:
        return bool(
            hasattr(user, "profile")
            and user.profile.projects.filter(id=project_id).exists()
        )
    return True


def _normalize_binary_flag(value):
    normalized = str(value or "").strip().lower()
    if normalized in {"yes", "true", "1", "y", "হ্যাঁ"}:
        return "yes"
    if normalized in {"no", "false", "0", "n", "না"}:
        return "no"
    return ""


def _is_affirmative(value):
    return _normalize_binary_flag(value) == "yes"


def _contains_any_marker(text, markers):
    if not text:
        return False
    return any(marker in text for marker in markers)


def _resolve_first_non_empty(source, candidate_keys):
    for key in candidate_keys:
        value = source.get(key)
        if str(value or "").strip():
            return value
    return ""


def _apply_brac_export_fields(row):
    # Keep export behavior aligned with AllRowsDataTable transformations.
    next_row = dict(row)
    division_value = str(next_row.get("division") or "").strip()
    city_corporation_value = str(next_row.get("city_corporation") or "").strip()
    next_row["total_reporting_sites"] = " / ".join(
        [value for value in [division_value, city_corporation_value] if value]
    )

    disease_source_candidates = [
        "suspected_disease",
        "suspected_patient-related_information_suspected_disease",
        "if_yes__name_of_the_diseases_1",
        "administrative_suspected_disease",
    ]
    dynamic_disease_key = next(
        (
            key
            for key, value in next_row.items()
            if key != "suspected_disease"
            and key.endswith("_suspected_disease")
            and str(value or "").strip()
        ),
        None,
    )
    if dynamic_disease_key:
        disease_source_candidates.append(dynamic_disease_key)

    disease_value = _resolve_first_non_empty(next_row, disease_source_candidates)
    disease_text = str(disease_value).lower()

    current_flags = {
        "dengue": _normalize_binary_flag(next_row.get("suspected_disease_dengue")),
        "awd": _normalize_binary_flag(next_row.get("suspected_disease_awd")),
        "malaria": _normalize_binary_flag(next_row.get("suspected_disease_malaria")),
    }
    inferred_flags = {"dengue": "", "awd": "", "malaria": ""}
    if disease_text:
        inferred_flags["dengue"] = (
            "yes"
            if _contains_any_marker(disease_text, ["dengue", "ডেঙ্গু"])
            else "no"
        )
        inferred_flags["awd"] = (
            "yes"
            if _contains_any_marker(
                disease_text,
                [
                    "awd",
                    "acute watery diarrhea",
                    "diarrhea",
                    "diarrhoea",
                    "ডায়রিয়া",
                    "ডায়রিয়া",
                    "ডায়রিয়া",
                    "ডায়রিয়া",
                ],
            )
            else "no"
        )
        inferred_flags["malaria"] = (
            "yes"
            if _contains_any_marker(disease_text, ["malaria", "ম্যালেরিয়া", "ম্যালেরিয়া"])
            else "no"
        )

    suspected_yes = _is_affirmative(next_row.get("suspected_in_the_disease"))
    resolved_dengue = inferred_flags["dengue"] or current_flags["dengue"]
    resolved_awd = inferred_flags["awd"] or current_flags["awd"]
    resolved_malaria = inferred_flags["malaria"] or current_flags["malaria"]

    next_row["suspected_disease_dengue"] = resolved_dengue or ("" if suspected_yes else "no")
    next_row["suspected_disease_awd"] = resolved_awd or ("" if suspected_yes else "no")
    next_row["suspected_disease_malaria"] = resolved_malaria or ("" if suspected_yes else "no")

    what_types_source_candidates = [
        "what_types",
        "administrative_what_types",
    ]
    for key, value in next_row.items():
        if (
            key != "what_types"
            and key.endswith("_what_types")
            and str(value or "").strip()
        ):
            what_types_source_candidates.append(key)

    what_types_value = _resolve_first_non_empty(next_row, what_types_source_candidates)
    what_types_text = str(what_types_value).lower()
    current_what_type_flags = {
        "flood": _normalize_binary_flag(next_row.get("what_types_flood")),
        "cyclone": _normalize_binary_flag(next_row.get("what_types_cyclone")),
        "heavy_rainfall": _normalize_binary_flag(next_row.get("what_types_heavy_rainfall")),
        "drought": _normalize_binary_flag(next_row.get("what_types_drought")),
        "landslides": _normalize_binary_flag(next_row.get("what_types_landslides")),
        "others": _normalize_binary_flag(next_row.get("what_types_others")),
    }
    inferred_what_type_flags = dict(current_what_type_flags)
    if what_types_text:
        inferred_what_type_flags["flood"] = (
            "yes" if _contains_any_marker(what_types_text, ["flood", "বন্যা"]) else "no"
        )
        inferred_what_type_flags["cyclone"] = (
            "yes" if _contains_any_marker(what_types_text, ["cyclone", "সাইক্লোন"]) else "no"
        )
        inferred_what_type_flags["heavy_rainfall"] = (
            "yes"
            if _contains_any_marker(
                what_types_text,
                ["heavy_rainfall", "heavy rainfall", "অতিবৃষ্টি"],
            )
            else "no"
        )
        inferred_what_type_flags["drought"] = (
            "yes" if _contains_any_marker(what_types_text, ["drought", "খরা"]) else "no"
        )
        inferred_what_type_flags["landslides"] = (
            "yes"
            if _contains_any_marker(what_types_text, ["landslides", "landslide", "ভূমিধ্বস"])
            else "no"
        )
        inferred_what_type_flags["others"] = (
            "yes" if _contains_any_marker(what_types_text, ["others", "other", "অন্যান্য"]) else "no"
        )

    disaster_yes = _is_affirmative(next_row.get("did_any_disaster_occur_in_last_7_days_"))
    next_row["what_types_flood"] = inferred_what_type_flags["flood"] or ("" if disaster_yes else "no")
    next_row["what_types_cyclone"] = inferred_what_type_flags["cyclone"] or ("" if disaster_yes else "no")
    next_row["what_types_heavy_rainfall"] = (
        inferred_what_type_flags["heavy_rainfall"] or ("" if disaster_yes else "no")
    )
    next_row["what_types_drought"] = inferred_what_type_flags["drought"] or ("" if disaster_yes else "no")
    next_row["what_types_landslides"] = inferred_what_type_flags["landslides"] or ("" if disaster_yes else "no")
    next_row["what_types_others"] = inferred_what_type_flags["others"] or ("" if disaster_yes else "no")

    followup_disease_source_candidates = [
        "followup_if_yes__name_of_the_diseases_1",
        "if_yes__name_of_the_diseases_1",
    ]
    for key, value in next_row.items():
        if (
            key != "followup_if_yes__name_of_the_diseases_1"
            and key.startswith("followup_")
            and key.endswith("_if_yes__name_of_the_diseases_1")
            and str(value or "").strip()
        ):
            followup_disease_source_candidates.append(key)

    followup_disease_value = _resolve_first_non_empty(
        next_row,
        followup_disease_source_candidates,
    )
    followup_disease_text = str(followup_disease_value).lower()
    current_followup_disease_flags = {
        "dengue": _normalize_binary_flag(next_row.get("followup_disease_dengue")),
        "awd": _normalize_binary_flag(next_row.get("followup_disease_awd")),
        "malaria": _normalize_binary_flag(next_row.get("followup_disease_malaria")),
        "others": _normalize_binary_flag(next_row.get("followup_disease_others")),
        "not_diagonosd": _normalize_binary_flag(next_row.get("followup_disease_not_diagonosd")),
    }
    inferred_followup_disease_flags = dict(current_followup_disease_flags)
    if followup_disease_text:
        inferred_followup_disease_flags["dengue"] = (
            "yes"
            if _contains_any_marker(followup_disease_text, ["dengue", "ডেঙ্গু"])
            else "no"
        )
        inferred_followup_disease_flags["awd"] = (
            "yes"
            if _contains_any_marker(
                followup_disease_text,
                [
                    "awd",
                    "acute watery diarrhea",
                    "diarrhea",
                    "diarrhoea",
                    "ডায়রিয়া",
                    "ডায়রিয়া",
                    "ডায়রিয়া",
                    "ডায়রিয়া",
                ],
            )
            else "no"
        )
        inferred_followup_disease_flags["malaria"] = (
            "yes"
            if _contains_any_marker(
                followup_disease_text,
                ["malaria", "ম্যালেরিয়া", "ম্যালেরিয়া"],
            )
            else "no"
        )
        inferred_followup_disease_flags["others"] = (
            "yes" if _contains_any_marker(followup_disease_text, ["others", "other", "অন্যান্য"]) else "no"
        )
        inferred_followup_disease_flags["not_diagonosd"] = (
            "yes"
            if _contains_any_marker(
                followup_disease_text,
                ["not_diagonosd", "not diagonosd", "not diagnosed", "নির্ণয়"],
            )
            else "no"
        )

    followup_suspected_yes = _is_affirmative(next_row.get("followup_suspected_in_the_disease"))
    next_row["followup_disease_dengue"] = (
        inferred_followup_disease_flags["dengue"] or ("" if followup_suspected_yes else "no")
    )
    next_row["followup_disease_awd"] = (
        inferred_followup_disease_flags["awd"] or ("" if followup_suspected_yes else "no")
    )
    next_row["followup_disease_malaria"] = (
        inferred_followup_disease_flags["malaria"] or ("" if followup_suspected_yes else "no")
    )
    next_row["followup_disease_others"] = (
        inferred_followup_disease_flags["others"] or ("" if followup_suspected_yes else "no")
    )
    next_row["followup_disease_not_diagonosd"] = (
        inferred_followup_disease_flags["not_diagonosd"] or ("" if followup_suspected_yes else "no")
    )

    # Merge age_year and age into a single "Year age" column.
    age_year_val = str(next_row.get("age_year") or "").strip()
    age_val = str(next_row.get("age") or "").strip()
    next_row["year_age"] = age_year_val if age_year_val else age_val

    return next_row


def _to_excel_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        try:
            value = json.dumps(value, ensure_ascii=False)
        except Exception:
            value = str(value)
    if isinstance(value, str):
        cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)
        # Excel cell text max length is 32767 characters.
        return cleaned[:32767]
    return value


def _parse_brac_followup_filter(raw_filter):
    normalized = str(raw_filter or "all").strip().lower()
    if normalized not in {"all", "with", "without"}:
        return "all"
    return normalized


def _parse_query_bool(raw_value):
    return str(raw_value or "").strip().lower() in {"1", "true", "yes", "on"}


def _parse_brac_followup_form_ids(raw_value):
    followup_form_ids = set()
    if not raw_value:
        return followup_form_ids

    for value in str(raw_value).split(","):
        value = value.strip()
        if value.isdigit():
            followup_form_ids.add(int(value))
    return followup_form_ids


def _parse_brac_selected_fields(raw_value):
    if not raw_value:
        return list(BRAC_DOWNLOAD_DEFAULT_FIELDS), False

    selected_fields = [field.strip() for field in str(raw_value).split(",") if field.strip()]
    if not selected_fields:
        return list(BRAC_DOWNLOAD_DEFAULT_FIELDS), False

    return selected_fields, True


def _resolve_brac_export_header_label(field_name, field_labels):
    """
    Resolve export column header labels with explicit overrides for known fields.
    """
    if str(field_name).endswith("user_identification_11_9943_01976848561"):
        return "user identification"

    label = field_labels.get(field_name)
    if label:
        normalized_label = re.sub(r"\s+", " ", str(label).strip().lower())
        if normalized_label == "user identification 11 9943 01976848561":
            return "user identification"
        return label

    return BRAC_DERIVED_FIELD_LABELS.get(field_name, str(field_name).replace("_", " "))


def _get_brac_snapshot_paths(project_id, followup_filter="all"):
    configured_subdir = str(
        getattr(settings, "BRAC_SNAPSHOT_SUBDIR", "downloads/brac_snapshots")
    ).strip() or "downloads/brac_snapshots"
    snapshot_dir = (
        configured_subdir
        if os.path.isabs(configured_subdir)
        else os.path.join(settings.MEDIA_ROOT, configured_subdir)
    )
    safe_followup_filter = re.sub(r"[^a-z0-9_-]", "", str(followup_filter or "all").lower()) or "all"
    snapshot_name = f"project_{project_id}_followup_{safe_followup_filter}_latest.xlsx"
    meta_name = f"project_{project_id}_followup_{safe_followup_filter}_latest.json"
    return (
        os.path.join(snapshot_dir, snapshot_name),
        os.path.join(snapshot_dir, meta_name),
    )


def _read_brac_snapshot_meta(meta_path):
    if not os.path.exists(meta_path):
        return {}
    try:
        with open(meta_path, "r", encoding="utf-8") as meta_file:
            payload = json.load(meta_file)
            return payload if isinstance(payload, dict) else {}
    except Exception:
        logger.warning("Failed to read BRAC snapshot metadata: %s", meta_path, exc_info=True)
        return {}


def _write_brac_snapshot_meta(meta_path, payload):
    meta_dir = os.path.dirname(meta_path)
    os.makedirs(meta_dir, exist_ok=True)
    temp_meta_path = f"{meta_path}.tmp.{uuid.uuid4().hex}"
    with open(temp_meta_path, "w", encoding="utf-8") as meta_file:
        json.dump(payload, meta_file, ensure_ascii=False, indent=2)
    os.replace(temp_meta_path, meta_path)


def _build_brac_xlsx_file(project_id, selected_fields, followup_filter, followup_form_ids, persist_snapshot=False):
    from .models import Project, Template

    project = Project.objects.get(id=project_id)

    templates = (
        Template.objects.filter(project=project)
        .select_related("data_collection_form")
        .prefetch_related("lookup_forms", "generated_lookup_forms")
    )
    if not templates.exists():
        raise ValueError("No templates found for this project")

    combined_rows, _, _, field_labels = _build_combined_template_rows(
        templates,
        filter_followup=followup_filter,
        followup_form_ids=followup_form_ids,
    )

    workbook = openpyxl.Workbook(write_only=True)
    worksheet = workbook.create_sheet(title="BRAC_Data")

    header_row = ["row_index"] + [
        _resolve_brac_export_header_label(field, field_labels)
        for field in selected_fields
    ]
    worksheet.append(header_row)

    _FILTER_EMPTY_FIELDS = ("total_reporting_sites", "location", "hh_id", "hh_head_name")
    output_rows = []
    for row in combined_rows:
        processed_row = _apply_brac_export_fields(row)
        if any(not str(processed_row.get(f, "") or "").strip() for f in _FILTER_EMPTY_FIELDS):
            continue
        output_rows.append(processed_row)

    for index, processed_row in enumerate(output_rows, start=1):
        worksheet.append(
            [index] + [_to_excel_cell_value(processed_row.get(field, "")) for field in selected_fields]
        )

    generated_at = timezone.now()
    timestamp = generated_at.strftime("%Y%m%d_%H%M%S")
    file_name = f"project_{project_id}_brac_data_{timestamp}.xlsx"
    row_count = len(output_rows)

    if persist_snapshot:
        snapshot_path, meta_path = _get_brac_snapshot_paths(project_id, followup_filter=followup_filter)
        os.makedirs(os.path.dirname(snapshot_path), exist_ok=True)
        temp_snapshot_path = f"{snapshot_path}.tmp.{uuid.uuid4().hex}"
        try:
            workbook.save(temp_snapshot_path)
            os.replace(temp_snapshot_path, snapshot_path)
        finally:
            if os.path.exists(temp_snapshot_path):
                os.remove(temp_snapshot_path)

        snapshot_size_bytes = os.path.getsize(snapshot_path) if os.path.exists(snapshot_path) else 0
        snapshot_meta = {
            "project_id": int(project_id),
            "followup_filter": str(followup_filter),
            "selected_fields": list(selected_fields),
            "followup_form_ids": sorted(int(item) for item in followup_form_ids),
            "generated_at": generated_at.isoformat(),
            "row_count": int(row_count),
            "file_size_bytes": int(snapshot_size_bytes),
        }
        _write_brac_snapshot_meta(meta_path, snapshot_meta)
        return {
            "file_path": snapshot_path,
            "file_name": file_name,
            "row_count": row_count,
            "generated_at": snapshot_meta["generated_at"],
            "snapshot_size_bytes": snapshot_size_bytes,
            "from_snapshot": True,
        }

    download_dir = os.path.join(settings.MEDIA_ROOT, "downloads")
    os.makedirs(download_dir, exist_ok=True)
    file_path = os.path.join(download_dir, file_name)
    workbook.save(file_path)
    return {
        "file_path": file_path,
        "file_name": file_name,
        "row_count": row_count,
        "generated_at": generated_at.isoformat(),
        "snapshot_size_bytes": os.path.getsize(file_path) if os.path.exists(file_path) else 0,
        "from_snapshot": False,
    }


def _load_prepared_brac_snapshot(project_id, followup_filter="all"):
    snapshot_path, meta_path = _get_brac_snapshot_paths(project_id, followup_filter=followup_filter)
    if not os.path.exists(snapshot_path):
        return None

    snapshot_meta = _read_brac_snapshot_meta(meta_path)
    generated_at_value = snapshot_meta.get("generated_at")
    generated_at_dt = parse_datetime(generated_at_value) if generated_at_value else None
    if generated_at_dt and timezone.is_naive(generated_at_dt):
        generated_at_dt = timezone.make_aware(generated_at_dt, timezone.get_current_timezone())
    if not generated_at_dt:
        modified_ts = os.path.getmtime(snapshot_path)
        generated_at_dt = datetime.fromtimestamp(modified_ts, tz=timezone.get_current_timezone())

    generated_stamp = generated_at_dt.strftime("%Y%m%d_%H%M%S")
    return {
        "file_path": snapshot_path,
        "file_name": f"project_{project_id}_brac_data_{generated_stamp}.xlsx",
        "row_count": snapshot_meta.get("row_count"),
        "generated_at": generated_at_dt.isoformat(),
        "snapshot_size_bytes": snapshot_meta.get("file_size_bytes")
        or (os.path.getsize(snapshot_path) if os.path.exists(snapshot_path) else 0),
        "from_snapshot": True,
    }


def _build_brac_file_response(download_payload):
    from django.http import FileResponse

    response = FileResponse(
        open(download_payload["file_path"], "rb"),
        as_attachment=True,
        filename=download_payload["file_name"],
    )
    row_count = download_payload.get("row_count")
    if row_count is not None:
        response["X-Total-Count"] = str(row_count)

    generated_at = download_payload.get("generated_at")
    if generated_at:
        response["X-Snapshot-Generated-At"] = str(generated_at)

    snapshot_size_bytes = download_payload.get("snapshot_size_bytes")
    if snapshot_size_bytes is not None:
        response["X-Snapshot-Size-Bytes"] = str(snapshot_size_bytes)

    response["X-Export-Source"] = "prepared_snapshot" if download_payload.get("from_snapshot") else "on_demand"
    return response


# Get all templates for a specific project
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_project_templates(request, project_id):
    from .models import Project, Template
    from .serializers import FormSerializer

    if _is_brac_download_user(request.user):
        allowed_project_id = int(getattr(settings, "BRAC_DOWNLOAD_PROJECT_ID", 55))
        if project_id != allowed_project_id:
            return Response(
                {'error': f'Access denied. This account can only access project {allowed_project_id}.'},
                status=status.HTTP_403_FORBIDDEN,
            )

    include_submissions = request.GET.get('include_submissions', 'true').strip().lower() not in {
        '0', 'false', 'no'
    }

    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return Response({'error': 'Project not found'}, status=404)

    if getattr(request.user, 'role', 4) == 6:
        if not hasattr(request.user, 'profile') or not request.user.profile.projects.filter(id=project_id).exists():
            return Response({'error': 'You do not have access to this project'}, status=403)

    templates = Template.objects.filter(project=project).only('id', 'data_collection_form_id')
    result = []

    def serialize_form_compact(form, include_questions=False):
        if not form:
            return None
        data = {
            'id': form.id,
            'project': form.project_id,
            'name': form.name,
            'description': form.description,
            'criteria': form.criteria,
            'created_at': form.created_at,
            'updated_at': form.updated_at,
        }
        if include_questions:
            data['questions'] = form.questions
        return data

    for template in templates:
        template_data = {
            'id': template.id,
            'name': template.name,
            'description': template.description,
            'project': template.project_id,
            'created_at': template.created_at,
            'updated_at': template.updated_at,
        }
        # Attach data_collection_form details
        if template.data_collection_form:
            if include_submissions:
                dc_form_data = FormSerializer(template.data_collection_form).data
                dc_form_data['submission'] = _get_form_submission_entries(template.data_collection_form)
            else:
                dc_form_data = serialize_form_compact(template.data_collection_form, include_questions=True)
            template_data['data_collection_form'] = dc_form_data
        else:
            template_data['data_collection_form'] = None
        # Attach lookup_forms details
        template_data['lookup_forms'] = []
        for lf in template.lookup_forms.all():
            if include_submissions:
                lf_data = FormSerializer(lf).data
                lf_data['submission'] = _get_form_submission_entries(lf)
            else:
                lf_data = serialize_form_compact(lf, include_questions=False)
            template_data['lookup_forms'].append(lf_data)
        # Attach generated_lookup_forms details
        template_data['generated_lookup_forms'] = []
        for generated_form in template.generated_lookup_forms.all():
            if include_submissions:
                generated_data = FormSerializer(generated_form).data
                generated_data['submission'] = _get_form_submission_entries(generated_form)
            else:
                generated_data = serialize_form_compact(generated_form, include_questions=False)
            template_data['generated_lookup_forms'].append(generated_data)
        result.append(template_data)
    return Response(result)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_project_templates_paginated(request, project_id):
    """Return paginated combined rows for data collection + follow-ups."""
    from .models import Project, Template
    from django.core.paginator import Paginator, EmptyPage

    if _is_brac_download_user(request.user):
        allowed_project_id = int(getattr(settings, "BRAC_DOWNLOAD_PROJECT_ID", 55))
        return Response(
            {
                'error': (
                    f'This account can only download Excel from '
                    f'/api/get-project-templates-full/{allowed_project_id}/'
                )
            },
            status=status.HTTP_403_FORBIDDEN,
        )

    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return Response({'error': 'Project not found'}, status=404)

    page = max(int(request.GET.get('page', 1)), 1)
    page_size = max(int(request.GET.get('page_size', 20)), 1)

    followup_filter = request.GET.get('followup_filter', 'all').lower()
    if followup_filter not in {'all', 'with', 'without'}:
        followup_filter = 'all'

    followup_forms_param = request.GET.get('followup_forms', '')
    followup_form_ids = set()
    if followup_forms_param:
        for value in followup_forms_param.split(','):
            value = value.strip()
            if value.isdigit():
                followup_form_ids.add(int(value))

    templates = list(
        Template.objects.filter(project=project).only('id', 'data_collection_form_id')
    )

    if not templates:
        return Response({
            'data': [],
            'total_count': 0,
            'page': page,
            'page_size': page_size,
            'total_pages': 0,
            'visible_fields': BASE_ROW_FIELDS,
            'followup_forms': []
        })

    # Fast path for large projects: page data collection submissions directly while
    # still attaching follow-up data in the same row (Download Full Data format).
    if followup_filter == 'all' and not followup_form_ids and len(templates) == 1:
        template = templates[0]
        dc_form_id = getattr(template, 'data_collection_form_id', None)
        dc_form = (
            Form.objects.only('id', 'name', 'questions', 'submission')
            .filter(id=dc_form_id)
            .first()
            if dc_form_id
            else None
        )

        if dc_form:
            dc_questions = dc_form.questions or []
            question_sequence = _extract_question_sequence(dc_questions)
            dc_select_field_option_maps = _build_select_field_option_maps(dc_questions)
            field_labels = {
                'row_number': 'Row Number',
                'Validation': 'Validation',
                '_id': 'ID',
                '_version': 'Version',
                'group_name': 'Group Name',
                'formType': 'Form Type',
                'has_followup': 'Has Follow-up',
                'followup_count': 'Follow-up Count',
                'followup_forms': 'Follow-up Forms',
                'followup_follow_up_date': 'Follow-up Date',
                'followup_start': 'Follow-up start',
                'followup_end': 'Follow-up end',
                'followup_follow_up_status_2472_h': 'Follow-up Status (24-72h)',
                'followup_suspected_in_the_disease': 'Follow-up: Suspected Disease',
                'followup_if_yes__name_of_the_diseases_1': 'Follow-up: Disease Name',
                'followup_hh_head_name': 'Follow-up: Household Head Name',
                'followup_hh_id': 'Follow-up: Household ID',
                'instanceID': 'Instance ID',
                'submitted_by': 'Submitted By',
                'start': 'Start Time',
                'end': 'End Time',
            }
            field_labels.update(_extract_question_labels(dc_questions))

            dc_submissions = _get_form_submission_entries(dc_form)
            total_count = len(dc_submissions)
            total_pages = (total_count + page_size - 1) // page_size if total_count else 0

            lookup_form_ids = list(template.lookup_forms.values_list('id', flat=True))
            generated_form_ids = list(template.generated_lookup_forms.values_list('id', flat=True))
            followup_form_ids_all = lookup_form_ids + generated_form_ids
            followup_lookup, followup_form_name_map, followup_forms_catalog = _collect_followup_lookup_for_forms(
                followup_form_ids_all
            )
            followup_select_option_maps_by_form = {}
            if followup_form_ids_all:
                followup_forms = Form.objects.only('id', 'questions').filter(
                    id__in=followup_form_ids_all
                )
                followup_select_option_maps_by_form = {
                    followup_form.id: _build_select_field_option_maps(
                        followup_form.questions or []
                    )
                    for followup_form in followup_forms
                }

            if total_pages > 0 and page > total_pages:
                page = total_pages

            if total_count == 0:
                return Response({
                    'data': [],
                    'total_count': 0,
                    'page': page,
                    'page_size': page_size,
                    'total_pages': 0,
                    'visible_fields': list(get_ordered_field_list(BASE_ROW_FIELDS + question_sequence)),
                    'field_labels': field_labels,
                    'followup_forms': [
                        {'id': form_id, 'name': name}
                        for form_id, name in sorted(
                            followup_forms_catalog.items(), key=lambda item: item[1].lower()
                        )
                    ],
                })

            start_idx = (page - 1) * page_size
            end_idx = min(start_idx + page_size, total_count)
            slice_start = max(total_count - end_idx, 0)
            slice_end = total_count - start_idx
            page_entries = list(reversed(dc_submissions[slice_start:slice_end]))

            rows = []
            field_order = list(BASE_ROW_FIELDS)
            followup_fields = [
                'follow_up_date',
                'start',
                'end',
                'follow_up_status_2472_h',
                'suspected_in_the_disease',
                'if_yes__name_of_the_diseases_1',
                'hh_head_name',
                'hh_id'
            ]

            for offset, submission in enumerate(page_entries):
                submission_data = submission.get('data', {}) if isinstance(submission, dict) else {}
                if not isinstance(submission_data, dict) or not submission_data:
                    continue

                uuid = submission_data.get('meta', {}).get('instanceID') or submission_data.get('instanceID')

                base_row = _flatten_form_payload(submission_data)
                _inject_meta_fields(submission_data, base_row)
                _apply_select_option_labels(base_row, dc_select_field_option_maps)

                for key, default_value in BASE_FIELD_DEFAULTS.items():
                    base_row.setdefault(key, default_value)
                base_row.setdefault('followup_details', '')
                for field in followup_fields:
                    base_row[f'followup_{field}'] = ''

                patient_name = (
                    submission_data.get('patient_name')
                    or submission_data.get('name')
                    or submission_data.get('patient')
                    or submission_data.get('administrative', {}).get('name_of_the_person_with_suspected_case')
                    or submission_data.get('hh_head_name')
                    or "Unknown Patient"
                )
                base_row['group_name'] = f"{dc_form.name} - {patient_name}"
                base_row['formType'] = 'Data Collection'

                associated_followups = followup_lookup.get(uuid, []) if uuid else []
                followup_names = []
                followup_primary_summary = ''

                for followup_entry in associated_followups:
                    followup_form_id = followup_entry.get('form_id')
                    form_name = followup_form_name_map.get(followup_form_id) or 'Follow Up'
                    followup_names.append(form_name)

                # Row-level follow-up fields come from the oldest follow-up only.
                if associated_followups:
                    primary_followup_entry = associated_followups[0]
                    followup_form_id = primary_followup_entry.get('form_id')
                    followup_data = primary_followup_entry.get('submission_data') or {}
                    form_name = followup_form_name_map.get(followup_form_id) or 'Follow Up'

                    flat_followup = _flatten_form_payload(followup_data)
                    _inject_meta_fields(followup_data, flat_followup)
                    followup_form_id_int = None
                    if followup_form_id:
                        try:
                            followup_form_id_int = int(followup_form_id)
                        except (TypeError, ValueError):
                            followup_form_id_int = None
                    _apply_select_option_labels(
                        flat_followup,
                        followup_select_option_maps_by_form.get(followup_form_id_int, {})
                    )
                    followup_primary_summary = _format_followup_summary(form_name, flat_followup)

                    for field in followup_fields:
                        field_value = flat_followup.get(field)
                        if field_value not in (None, ''):
                            base_row[f'followup_{field}'] = field_value

                base_row['has_followup'] = len(associated_followups) > 0
                base_row['followup_count'] = len(associated_followups)
                base_row['followup_forms'] = ', '.join(sorted(set(followup_names))) if followup_names else ''
                base_row['followup_details'] = followup_primary_summary

                global_idx = start_idx + offset
                base_row['row_number'] = total_count - global_idx

                ordered_keys = []
                for key in BASE_ROW_FIELDS:
                    if key in base_row and key not in ordered_keys:
                        ordered_keys.append(key)
                for key in question_sequence:
                    if key in base_row and key not in ordered_keys:
                        ordered_keys.append(key)
                for key in base_row.keys():
                    if key not in ordered_keys:
                        ordered_keys.append(key)

                ordered_row = {key: base_row.get(key, '') for key in ordered_keys}
                rows.append(ordered_row)

                for key in ordered_row.keys():
                    if key not in field_order:
                        field_order.append(key)

            return Response({
                'data': rows,
                'total_count': total_count,
                'page': page,
                'page_size': page_size,
                'total_pages': total_pages,
                'visible_fields': list(get_ordered_field_list(field_order)),
                'field_labels': field_labels,
                'followup_forms': [
                    {'id': form_id, 'name': name}
                    for form_id, name in sorted(
                        followup_forms_catalog.items(), key=lambda item: item[1].lower()
                    )
                ],
            })

    combined_rows, field_names, followup_forms_catalog, field_labels = _build_combined_template_rows(
        templates,
        filter_followup=followup_filter,
        followup_form_ids=followup_form_ids
    )

    if not combined_rows:
        return Response({
            'data': [],
            'total_count': 0,
            'page': page,
            'page_size': page_size,
            'total_pages': 0,
            'visible_fields': list(field_names),
            'followup_forms': [
                {'id': form_id, 'name': name}
                for form_id, name in sorted(followup_forms_catalog.items(), key=lambda item: item[1].lower())
            ]
        })

    paginator = Paginator(combined_rows, page_size)

    try:
        page_obj = paginator.page(page)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    return Response({
        'data': list(page_obj.object_list),
        'total_count': paginator.count,
        'page': page_obj.number,
        'page_size': page_size,
        'total_pages': paginator.num_pages,
        'visible_fields': list(field_names),
        'field_labels': field_labels,
        'followup_forms': [
            {'id': form_id, 'name': name}
            for form_id, name in sorted(followup_forms_catalog.items(), key=lambda item: item[1].lower())
        ]
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_project_templates_full(request, project_id):
    """Return the entire combined dataset for export (no pagination)."""
    from .models import Project, Template

    if _is_brac_download_user(request.user):
        allowed_project_id = int(getattr(settings, "BRAC_DOWNLOAD_PROJECT_ID", 55))
        if project_id != allowed_project_id:
            return Response(
                {'error': f'Access denied. This account can only access project {allowed_project_id}.'},
                status=status.HTTP_403_FORBIDDEN,
            )

    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return Response({'error': 'Project not found'}, status=404)

    templates = (
        Template.objects.filter(project=project)
        .select_related('data_collection_form')
        .prefetch_related('lookup_forms', 'generated_lookup_forms')
    )

    if not templates.exists():
        return Response({
            'data': [],
            'total_count': 0,
            'message': 'No templates found for this project'
        })

    selected_fields_param = request.GET.get('selected_fields', '')
    selected_fields = [f.strip() for f in selected_fields_param.split(',') if f.strip()] if selected_fields_param else None

    followup_filter = request.GET.get('followup_filter', 'all').lower()
    if followup_filter not in {'all', 'with', 'without'}:
        followup_filter = 'all'

    followup_forms_param = request.GET.get('followup_forms', '')
    followup_form_ids = set()
    if followup_forms_param:
        for value in followup_forms_param.split(','):
            value = value.strip()
            if value.isdigit():
                followup_form_ids.add(int(value))

    combined_rows, _, followup_forms_catalog, field_labels = _build_combined_template_rows(
        templates,
        filter_followup=followup_filter,
        followup_form_ids=followup_form_ids
    )

    if selected_fields:
        filtered_rows = []
        for row in combined_rows:
            filtered_row = {}
            for field in selected_fields:
                filtered_row[field] = row.get(field, '')
            filtered_rows.append(filtered_row)
        combined_rows = filtered_rows

    return Response({
        'data': combined_rows,
        'total_count': len(combined_rows),
        'message': f'Successfully retrieved {len(combined_rows)} records',
        'field_labels': field_labels,
        'followup_forms': [
            {'id': form_id, 'name': name}
            for form_id, name in sorted(followup_forms_catalog.items(), key=lambda item: item[1].lower())
        ]
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_project_templates_full_xlsx(request, project_id):
    from .models import Project

    if _is_brac_download_user(request.user):
        allowed_project_id = int(getattr(settings, "BRAC_DOWNLOAD_PROJECT_ID", 55))
        if project_id != allowed_project_id:
            return Response(
                {'error': f'Access denied. This account can only access project {allowed_project_id}.'},
                status=status.HTTP_403_FORBIDDEN,
            )
    elif not _has_project_scope_access(request.user, project_id):
        return Response(
            {'error': 'You do not have access to this project.'},
            status=status.HTTP_403_FORBIDDEN,
        )

    selected_fields, selected_fields_explicit = _parse_brac_selected_fields(
        request.GET.get("selected_fields", "")
    )
    followup_filter = _parse_brac_followup_filter(request.GET.get("followup_filter", "all"))
    followup_form_ids = _parse_brac_followup_form_ids(request.GET.get("followup_forms", ""))
    force_fresh = _parse_query_bool(request.GET.get("force_fresh"))

    snapshot_enabled = bool(getattr(settings, "BRAC_SNAPSHOT_ENABLED", True))
    is_default_snapshot_request = (
        snapshot_enabled
        and not force_fresh
        and not selected_fields_explicit
        and followup_filter == "all"
        and not followup_form_ids
    )

    if is_default_snapshot_request:
        prepared_snapshot = _load_prepared_brac_snapshot(project_id, followup_filter=followup_filter)
        if prepared_snapshot:
            return _build_brac_file_response(prepared_snapshot)

    try:
        download_payload = _build_brac_xlsx_file(
            project_id=project_id,
            selected_fields=selected_fields,
            followup_filter=followup_filter,
            followup_form_ids=followup_form_ids,
            persist_snapshot=is_default_snapshot_request,
        )
    except Project.DoesNotExist:
        return Response({'error': 'Project not found'}, status=status.HTTP_404_NOT_FOUND)
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_404_NOT_FOUND)
    except Exception:
        logger.exception(
            "Failed to generate BRAC export XLSX for project=%s followup_filter=%s",
            project_id,
            followup_filter,
        )
        return Response(
            {'error': 'Failed to generate BRAC export file.'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    return _build_brac_file_response(download_payload)

# Delete a template and all its forms
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def delete_template(request, template_id):
    """Move template to trash instead of permanent deletion"""
    from .models import Template, Form
    import logging

    logger = logging.getLogger(__name__)
    logger.info(f"🗑️ delete_template called for template {template_id}")

    try:
        template = Template.objects.get(id=template_id)
        logger.info(f"📄 Template found: {template.name}")
    except Template.DoesNotExist:
        return Response({'error': 'Template not found'}, status=status.HTTP_404_NOT_FOUND)

    try:
        # Move template to trash
        move_to_trash(template, 'template', request.user)
        logger.info(f"✅ Template {template_id} moved to trash successfully")
        return Response({'message': 'Template moved to trash successfully'}, status=status.HTTP_204_NO_CONTENT)
    except Exception as e:
        logger.error(f"❌ Failed to delete template {template_id}: {e}")
        return Response({'error': f'Failed to delete template: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Update a template (name, description, etc.)
@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_template(request, template_id):
    """Update template fields like name and description"""
    from .models import Template
    import logging

    logger = logging.getLogger(__name__)
    logger.info(f"📝 update_template called for template {template_id}")

    try:
        template = Template.objects.get(id=template_id)
        logger.info(f"📄 Template found: {template.name}")
    except Template.DoesNotExist:
        return Response({'error': 'Template not found'}, status=status.HTTP_404_NOT_FOUND)

    try:
        # Update template fields
        if 'name' in request.data:
            old_name = template.name
            template.name = request.data['name']
            logger.info(f"📝 Updated template name from '{old_name}' to '{template.name}'")

        if 'description' in request.data:
            template.description = request.data['description']
            logger.info(f"📝 Updated template description")

        template.save()

        serializer = TemplateSerializer(template)
        logger.info(f"✅ Template {template_id} updated successfully")
        return Response(serializer.data, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"❌ Failed to update template {template_id}: {e}")
        return Response({'error': f'Failed to update template: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Clone a template to another project
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def clone_template(request, template_id):
    """Clone a template and all its forms to another project"""
    from .models import Template, Form, Project
    import copy

    logger = logging.getLogger(__name__)
    logger.info(f"📋 clone_template called for template {template_id}")

    target_project_id = request.data.get('target_project_id')
    new_template_name = request.data.get('template_name', None)

    if not target_project_id:
        return Response({'error': 'target_project_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Get source template
        source_template = Template.objects.get(id=template_id)
        logger.info(f"📄 Source template found: {source_template.name}")

        # Get target project
        target_project = Project.objects.get(id=target_project_id)
        logger.info(f"📦 Target project found: {target_project.name}")

    except Template.DoesNotExist:
        return Response({'error': 'Source template not found'}, status=status.HTTP_404_NOT_FOUND)
    except Project.DoesNotExist:
        return Response({'error': 'Target project not found'}, status=status.HTTP_404_NOT_FOUND)

    try:
        # Create new template
        cloned_template = Template.objects.create(
            name=new_template_name if new_template_name else f"{source_template.name} (Copy)",
            description=f"Cloned from template {source_template.id}: {source_template.description}",
            project=target_project
        )
        logger.info(f"✅ Created cloned template: {cloned_template.name}")

        # Clone data collection form if exists
        cloned_data_form = None
        if source_template.data_collection_form:
            source_data_form = source_template.data_collection_form
            cloned_data_form = Form.objects.create(
                project=target_project,
                name=f"{source_data_form.name} (Copy)",
                description=f"Cloned from form {source_data_form.id}: {source_data_form.description}",
                questions=copy.deepcopy(source_data_form.questions),
                translations=copy.deepcopy(source_data_form.translations),
                submission=[],  # Don't copy submissions
                enketo_id=None,  # Will be generated when needed
                allow_anonymous_submissions=source_data_form.allow_anonymous_submissions,
                template=cloned_template,
                form_style=getattr(source_data_form, 'form_style', 'default')
            )

            # Copy many-to-many relationships
            if hasattr(source_data_form, 'other_languages'):
                cloned_data_form.other_languages.set(source_data_form.other_languages.all())

            cloned_template.data_collection_form = cloned_data_form
            logger.info(f"✅ Cloned data collection form: {cloned_data_form.name}")

        # Clone lookup forms
        cloned_lookup_forms = []
        for source_lookup_form in source_template.lookup_forms.all():
            cloned_lookup_form = Form.objects.create(
                project=target_project,
                name=f"{source_lookup_form.name} (Copy)",
                description=f"Cloned from form {source_lookup_form.id}: {source_lookup_form.description}",
                questions=copy.deepcopy(source_lookup_form.questions),
                translations=copy.deepcopy(source_lookup_form.translations),
                submission=[],  # Don't copy submissions
                enketo_id=None,  # Will be generated when needed
                allow_anonymous_submissions=source_lookup_form.allow_anonymous_submissions,
                template=cloned_template,
                form_style=getattr(source_lookup_form, 'form_style', 'default'),
                criteria=copy.deepcopy(getattr(source_lookup_form, 'criteria', {}))
            )

            # Copy many-to-many relationships
            if hasattr(source_lookup_form, 'other_languages'):
                cloned_lookup_form.other_languages.set(source_lookup_form.other_languages.all())

            cloned_lookup_forms.append(cloned_lookup_form)
            logger.info(f"✅ Cloned lookup form: {cloned_lookup_form.name}")

        # Set lookup forms to template
        cloned_template.lookup_forms.set(cloned_lookup_forms)

        # Clone generated lookup forms
        cloned_generated_forms = []
        for source_generated_form in source_template.generated_lookup_forms.all():
            cloned_generated_form = Form.objects.create(
                project=target_project,
                name=f"{source_generated_form.name} (Copy)",
                description=f"Cloned from form {source_generated_form.id}: {source_generated_form.description}",
                questions=copy.deepcopy(source_generated_form.questions),
                translations=copy.deepcopy(source_generated_form.translations),
                submission=[],  # Don't copy submissions
                enketo_id=None,  # Will be generated when needed
                allow_anonymous_submissions=source_generated_form.allow_anonymous_submissions,
                template=cloned_template,
                form_style=getattr(source_generated_form, 'form_style', 'default'),
                criteria=copy.deepcopy(getattr(source_generated_form, 'criteria', {}))
            )

            # Copy many-to-many relationships
            if hasattr(source_generated_form, 'other_languages'):
                cloned_generated_form.other_languages.set(source_generated_form.other_languages.all())

            cloned_generated_forms.append(cloned_generated_form)
            logger.info(f"✅ Cloned generated form: {cloned_generated_form.name}")

        # Set generated lookup forms to template
        cloned_template.generated_lookup_forms.set(cloned_generated_forms)

        # Save the template with all relationships
        cloned_template.save()

        # Prepare response data
        serializer = TemplateSerializer(cloned_template)
        response_data = serializer.data

        # Add form details to response
        if cloned_data_form:
            from .serializers import FormSerializer
            response_data['data_collection_form'] = FormSerializer(cloned_data_form).data

        response_data['lookup_forms'] = [
            {'id': form.id, 'name': form.name} for form in cloned_lookup_forms
        ]
        response_data['generated_lookup_forms'] = [
            {'id': form.id, 'name': form.name} for form in cloned_generated_forms
        ]

        logger.info(f"🎉 Successfully cloned template {template_id} to project {target_project_id}")
        return Response({
            'message': 'Template cloned successfully',
            'cloned_template': response_data
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        logger.error(f"❌ Failed to clone template {template_id}: {e}")
        return Response({'error': f'Failed to clone template: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Get all projects for template cloning selection
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_all_projects(request):
    """Get all projects that the user can access for template cloning"""
    from .models import Project
    from .serializers import ProjectSerializer

    try:
        # Get all projects (you can add user filtering here if needed)
        projects = Project.objects.all().order_by('name')
        if getattr(request.user, 'role', 4) == 6:
            project_ids = _get_assigned_project_ids(request.user)
            projects = projects.filter(id__in=project_ids)
        serializer = ProjectSerializer(projects, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': f'Failed to get projects: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db import transaction


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from .models import Form

# ...existing code...

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def lookup_form_set_criteria(request):
    lookup_form_id = request.data.get("lookup_form_id")
    question_name = request.data.get("question_name")
    expected_value = request.data.get("expected_value")
    if not lookup_form_id or not question_name:
        return Response({"error": "Missing required fields."}, status=400)
    try:
        lf = Form.objects.get(id=lookup_form_id)
        lf.criteria = {"question_name": question_name, "expected_value": expected_value}
        lf.save()
        return Response({"success": True})
    except Form.DoesNotExist:
        return Response({"error": "LookupForm not found."}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)

# Create a new lookup form for a given data collection form and project
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def create_lookup_form(request):
    """
    Creates a new lookup form for the given data collection form and project.
    Adds the new lookup form's ID to the related_lookup_forms field of the data collection form.
    """
    user = request.user
    data_collection_form_id = request.data.get('data_collection_form_id')
    project_id = request.data.get('project_id')
    from .models import Project, Form
    try:
        project = Project.objects.get(id=project_id)
        data_form = Form.objects.get(id=data_collection_form_id, project=project)
    except (Project.DoesNotExist, Form.DoesNotExist):
        return Response({'error': 'Project or Data Collection Form not found'}, status=404)

    # Count existing lookup forms for this template
    template = data_form.template
    lookup_count = template.lookup_forms.count() if template else 0

    # Create a new lookup form and link to template
    lookup_form = Form.objects.create(
        project=project,
        name=f"Lookup Form {lookup_count + 1}",
        description=f"Automatically generated lookup form {lookup_count + 1}",
        questions=[],
        translations={},
        submission=[],
        enketo_id=None,
        allow_anonymous_submissions=False,
        template=template
    )
    if template:
        template.lookup_forms.add(lookup_form)
        template.save()

    # Do NOT generate any generated lookup forms at creation time
    return Response({
        "lookup_form_id": lookup_form.id,
        "data_collection_form_id": data_form.id,
        "template_id": template.id if template else None
    }, status=201)
# Template creation endpoint: creates a data collection form and two lookup forms, all connected
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db import transaction

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def create_template(request):
    """
    Creates a template with:
    - 1 data collection form
    - 2 lookup forms (each referencing the data collection form)
    All forms are created using the Form model and can be managed via FormViewSet.
    """
    user = request.user
    project_id = request.data.get('project_id')
    from .models import Project, Form
    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return Response({'error': 'Project not found'}, status=404)

    # 1. Create the Template object (without forms yet)
    template = Template.objects.create(
        name=f"Template for Project {project.id}",
        description="Automatically generated template",
        project=project
    )

    # 2. Create the data collection form and link to template
    data_form = Form.objects.create(
        project=project,
        name="Data Collection Form",
        description="Automatically generated data collection form",
        questions=[],
        translations={},
        submission=[],
        enketo_id=None,
        allow_anonymous_submissions=False,
        template=template
    )

    # 3. Create two lookup forms, referencing the data collection form and link to template
    lookup_forms = []
    for i in range(2):
        lookup_form = Form.objects.create(
            project=project,
            name=f"Lookup Form {i+1}",
            description=f"Automatically generated lookup form {i+1}",
            questions=[],
            translations={},
            submission=[],
            enketo_id=None,
            allow_anonymous_submissions=False,
            template=template
        )
        lookup_forms.append(lookup_form)

    # 4. Link forms to template
    template.data_collection_form = data_form
    template.save()
    template.lookup_forms.set(lookup_forms)

    # 5. Return the template structure
    return Response({
        "template_id": template.id,
        "data_collection_form": data_form.id,
        "lookup_forms": [lf.id for lf in lookup_forms]
    }, status=201)
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from django.views.decorators.csrf import csrf_exempt

# Media file submission endpoint
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
@csrf_exempt
def submit_form_media(request, form_id):
    import os
    from django.conf import settings
    from django.http import JsonResponse

    # Authenticate user (Token or Basic Auth)
    user = None
    auth_header = request.META.get('HTTP_AUTHORIZATION', '')
    if auth_header.startswith('Token ') or auth_header.startswith('Bearer '):
        token_key = auth_header.split(' ')[1]
        from rest_framework.authtoken.models import Token
        try:
            token_obj = Token.objects.get(key=token_key)
            user = token_obj.user
        except Token.DoesNotExist:
            return JsonResponse({'error': 'Invalid or expired token'}, status=401)
    elif auth_header.startswith('Basic '):
        import base64
        try:
            encoded_credentials = auth_header.split(' ')[1]
            decoded_credentials = base64.b64decode(encoded_credentials).decode('utf-8')
            username, password = decoded_credentials.split(':', 1)
            from django.contrib.auth import authenticate
            user = authenticate(username=username, password=password)
            if not user:
                return JsonResponse({'error': 'Invalid username or password'}, status=401)
        except Exception:
            return JsonResponse({'error': 'Invalid Basic Auth header'}, status=401)
    else:
        return JsonResponse({'error': 'Authentication required (Token or Basic Auth)'}, status=401)

    # Validate form exists
    from .models import Form
    try:
        form = Form.objects.get(id=form_id)
    except Form.DoesNotExist:
        return JsonResponse({'error': 'Form not found'}, status=404)

    # Save uploaded files in form_<form_id>/<next_submission_id>/
    saved_files = []
    from .models import Submission
    # Get the last submission for this form_id by checking the xml_file path or by counting all submissions
    last_submission = Submission.objects.order_by('-id').first()
    next_submission_id = (last_submission.id + 1) if last_submission else 1
    form_media_dir = os.path.join(settings.MEDIA_ROOT, 'submissions', f'form_{form_id}', str(next_submission_id))
    os.makedirs(form_media_dir, exist_ok=True)

    for key, uploaded_file in request.FILES.items():
        file_path = os.path.join(form_media_dir, uploaded_file.name)
        with open(file_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)
        saved_files.append({
            'filename': uploaded_file.name,
            'path': file_path,
            'url': f"{settings.MEDIA_URL}submissions/form_{form_id}/{next_submission_id}/{uploaded_file.name}"
        })

    return JsonResponse({'message': 'Media files saved', 'files': saved_files}, status=201)
from django.shortcuts import render
from rest_framework import viewsets, status
from .models import Form, Submission, Project, Language, Organization, DownloadLog, Template
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.authtoken.models import Token
from rest_framework.response import Response
from rest_framework.views import APIView
from django.contrib.auth import authenticate
from rest_framework.decorators import action, api_view, permission_classes, parser_classes
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt  # Add this import
from django.utils.decorators import method_decorator
from urllib.parse import quote  # Add this import if not already present
import openpyxl
import os
import random
import string
from django.conf import settings
import re
from django.db import transaction
from django.contrib.auth.models import User
from .serializers import UserSerializer, OrganizationSerializer
from django.http import FileResponse, Http404, JsonResponse  # Add JsonResponse if not present
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import AllowAny
import tempfile
from pyxform.xls2xform import xls2xform_convert
from django.urls import reverse
from django.utils.html import escape
import xml.etree.ElementTree as ET
import csv
import zipfile
import io
import requests  # Add this import
import base64    # Add this import
from .auth_views import generate_auth_token

def generate_random_id(length=7):
    return ''.join(random.choices(string.ascii_letters + string.digits, k=length))

def get_ordinal_suffix(n):
    if 10 <= n % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
    return f'_{n}{suffix}_choice'

def sanitize_name(name):
    return re.sub(r'[^a-zA-Z0-9_]', '', name)


def _language_description(language):
    return f"{language.description} ({language.subtag})"


def _normalize_language_selection(default_language, other_languages):
    """
    Normalize language inputs used during XLSX generation.
    - Always return a valid default language description/subtag fallback.
    - Remove duplicate additional languages.
    - Exclude the default language from additional languages.
    """
    if default_language:
        default_language_description = _language_description(default_language)
        default_language_subtag = (default_language.subtag or '').strip().lower()
    else:
        default_language_description = 'English (en)'
        default_language_subtag = 'en'

    normalized_other_languages = []
    seen_ids = set()
    for language in list(other_languages or []):
        if not language:
            continue
        if default_language and language.id == default_language.id:
            continue
        if language.id in seen_ids:
            continue
        seen_ids.add(language.id)
        normalized_other_languages.append(language)

    return default_language_description, default_language_subtag, normalized_other_languages


def _get_language_value(value_map, language=None, default_description='', default_subtag=''):
    """
    Resolve translation/hint text using:
    1) Exact language description key, e.g. 'English (en)'
    2) Language subtag key, e.g. 'en'
    3) Any key ending with '(en)'
    """
    if not isinstance(value_map, dict):
        return ''

    preferred_description = default_description
    preferred_subtag = default_subtag
    if language is not None:
        preferred_description = _language_description(language)
        preferred_subtag = (language.subtag or '').strip().lower()

    if preferred_description:
        value = value_map.get(preferred_description, '')
        if value:
            return value

    if preferred_subtag:
        direct = value_map.get(preferred_subtag, '')
        if direct:
            return direct
        for key, value in value_map.items():
            if isinstance(key, str) and key.strip().lower().endswith(f'({preferred_subtag})') and value:
                return value

    return ''


def _get_default_text(item, field_name, default_language_description, default_language_subtag):
    """
    Return text for the current default language.
    Falls back to the base field when no translation/hint exists for default language.
    """
    base_value = item.get(field_name, '') if isinstance(item, dict) else ''
    if field_name == 'label':
        map_name = 'translations'
    elif field_name == 'hint':
        map_name = 'hints'
    elif field_name == 'guidance_hint':
        map_name = 'guidance_hints'
    else:
        map_name = None
    mapped_value = _get_language_value(
        item.get(map_name, {}) if (isinstance(item, dict) and map_name) else {},
        language=None,
        default_description=default_language_description,
        default_subtag=default_language_subtag,
    )
    return mapped_value or base_value


def _default_language_identity(language):
    if language:
        return _language_description(language), (language.subtag or '').strip().lower()
    return 'English (en)', 'en'


def _sync_node_for_default_language_switch(node, old_description, old_subtag, new_description, new_subtag):
    """
    Keep source-language text safe when default language changes and
    switch base labels/hints to the new default language when available.
    """
    if not isinstance(node, dict):
        return False

    changed = False

    if not isinstance(node.get('translations'), dict):
        node['translations'] = {}
        changed = True
    translations = node['translations']

    base_label = node.get('label', '')
    existing_old_label = _get_language_value(
        translations,
        language=None,
        default_description=old_description,
        default_subtag=old_subtag,
    )
    if base_label and not existing_old_label:
        translations[old_description] = base_label
        changed = True

    new_default_label = _get_language_value(
        translations,
        language=None,
        default_description=new_description,
        default_subtag=new_subtag,
    )
    if new_default_label and new_default_label != base_label:
        node['label'] = new_default_label
        changed = True

    base_hint = node.get('hint', '')
    if base_hint:
        if not isinstance(node.get('hints'), dict):
            node['hints'] = {}
            changed = True
        hints = node['hints']

        existing_old_hint = _get_language_value(
            hints,
            language=None,
            default_description=old_description,
            default_subtag=old_subtag,
        )
        if base_hint and not existing_old_hint:
            hints[old_description] = base_hint
            changed = True

        new_default_hint = _get_language_value(
            hints,
            language=None,
            default_description=new_description,
            default_subtag=new_subtag,
        )
        if new_default_hint and new_default_hint != base_hint:
            node['hint'] = new_default_hint
            changed = True

    return changed


def _sync_questions_for_default_language_switch(questions, old_default_language, new_default_language):
    if not isinstance(questions, list):
        return False

    old_description, old_subtag = _default_language_identity(old_default_language)
    new_description, new_subtag = _default_language_identity(new_default_language)

    if old_description == new_description and old_subtag == new_subtag:
        return False

    changed = False
    for question in questions:
        if _sync_node_for_default_language_switch(
            question,
            old_description,
            old_subtag,
            new_description,
            new_subtag,
        ):
            changed = True

        options = question.get('options', []) if isinstance(question, dict) else []
        if isinstance(options, list):
            for option in options:
                if _sync_node_for_default_language_switch(
                    option,
                    old_description,
                    old_subtag,
                    new_description,
                    new_subtag,
                ):
                    changed = True

        sub_questions = question.get('subQuestions', []) if isinstance(question, dict) else []
        if isinstance(sub_questions, list):
            for sub_question in sub_questions:
                if _sync_node_for_default_language_switch(
                    sub_question,
                    old_description,
                    old_subtag,
                    new_description,
                    new_subtag,
                ):
                    changed = True

    return changed


def generate_xlsx_file(form, questions, default_language, other_languages, form_name, output_path=None):
    """
    Shared function to generate XLSX file with consistent choice_filter logic
    across create_form, update, and update_translation functions
    """
    import openpyxl
    from django.conf import settings
    import os

    if not isinstance(questions, list):
        questions = []

    default_language_description, default_language_subtag, other_languages = _normalize_language_selection(
        default_language,
        other_languages,
    )

    if output_path is None:
        output_dir = os.path.join(settings.MEDIA_ROOT, 'update')
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f'{form.id}.xlsx')
    else:
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    survey_ws = wb.active
    survey_ws.title = 'survey'
    settings_ws = wb.create_sheet(title='settings')
    choices_ws = wb.create_sheet(title='choices')

    # --- Survey headers ---
    survey_headers = [
        'type',
        'name',
        'label',
        'required',
        'appearance',
        'parameters',
        'repeat_count',
        'hint',
        'default',
        'guidance_hint',
        'hxl',
        'constraint_message',
        'constraint',
        'relevant',
        'trigger',
        'calculation',
        'read_only',
        'choice_filter',
    ]
    for language in other_languages:
        language_description = _language_description(language)
        survey_headers.append(f'label::{language_description}')
        survey_headers.append(f'hint::{language_description}')
    survey_ws.append(survey_headers)

    # Add start and end rows after headers
    survey_ws.append(['start', 'start'] + [''] * (len(survey_headers) - 2))
    survey_ws.append(['end', 'end'] + [''] * (len(survey_headers) - 2))

    # --- Collect all unique filter columns from optionFilterMap ---
    choice_filter_columns = set()
    for question in questions:
        option_filter_map = question.get('optionFilterMap', {})
        if option_filter_map and isinstance(option_filter_map, dict) and len(option_filter_map) > 0:
            # Find parent question using filterQuestionId or question index
            filter_question_id = question.get('filterQuestionId')
            if filter_question_id is not None:
                parent_question_name = None

                # Try to find parent question by index first
                try:
                    parent_question_index = int(filter_question_id)
                    if 0 <= parent_question_index < len(questions):
                        parent_question_name = questions[parent_question_index].get('name')
                except (ValueError, TypeError):
                    # If filterQuestionId is not a valid index, try to find by UUID
                    for q in questions:
                        if q.get('_uuid') == filter_question_id:
                            parent_question_name = q.get('name')
                            break

                if parent_question_name:
                    choice_filter_columns.add(f'{parent_question_name}filter')

    # Convert to sorted list for consistent order
    choice_filter_columns = sorted(list(choice_filter_columns))

    # --- Choices headers ---
    choices_headers = ['list_name', 'name', 'label']

    # Add multi-language choice label columns
    for language in other_languages:
        language_description = _language_description(language)
        choices_headers.append(f'label::{language_description}')

    # Add filter columns at the end
    choices_headers.extend(choice_filter_columns)
    choices_ws.append(choices_headers)

    # --- Settings sheet setup ---
    settings_ws.append(['form_title', 'form_id', 'version', 'default_language', 'style'])
    settings_ws.append([
        form_name,
        f"form_{form.id}",
        "20240601",
        default_language_description,
        form.form_style if hasattr(form, 'form_style') and form.form_style else 'default'
    ])

    # --- Add questions to survey sheet ---
    for idx, question in enumerate(questions):
        question_type = question.get('type', 'text')
        question_name = question.get('name', '')
        question_label = _get_default_text(
            question,
            'label',
            default_language_description,
            default_language_subtag,
        )
        question_required = question.get('required', False)
        question_parameters = question.get('parameters', '')
        question_repeat_count = question.get('repeat_count', '')
        question_hint = _get_default_text(
            question,
            'hint',
            default_language_description,
            default_language_subtag,
        )
        question_default = question.get('default', '')
        question_appearance = question.get('appearance', '')
        question_guidance_hint = _get_default_text(
            question,
            'guidance_hint',
            default_language_description,
            default_language_subtag,
        )
        question_hxl = question.get('hxl', '')
        question_constraint_message = question.get('constraint_message', '')
        question_constraint = question.get('constraint', '')
        question_relevant = question.get('relevant', '')
        question_trigger = question.get('trigger', '')
        question_calculation = question.get('calculation', '')
        question_read_only = question.get('read_only', '')

        # Skip empty begin/end block pairs as pyxform fails on empty containers.
        if question_type in ('begin_group', 'begin_repeat'):
            next_type = questions[idx + 1].get('type', '') if idx + 1 < len(questions) else ''
            expected_end = 'end_group' if question_type == 'begin_group' else 'end_repeat'
            if next_type == expected_end:
                continue
        elif question_type in ('end_group', 'end_repeat'):
            prev_type = questions[idx - 1].get('type', '') if idx > 0 else ''
            expected_begin = 'begin_group' if question_type == 'end_group' else 'begin_repeat'
            if prev_type == expected_begin:
                continue

        # --- Build choice_filter value ---
        choice_filter_value = ''
        option_filter_map = question.get('optionFilterMap', {})
        if option_filter_map and isinstance(option_filter_map, dict) and len(option_filter_map) > 0:
            # Find parent question using filterQuestionId
            filter_question_id = question.get('filterQuestionId')
            if filter_question_id is not None:
                parent_question_name = None

                # Try to find parent question by index first
                try:
                    parent_question_index = int(filter_question_id)
                    if 0 <= parent_question_index < len(questions):
                        parent_question_name = questions[parent_question_index].get('name')
                except (ValueError, TypeError):
                    # If filterQuestionId is not a valid index, try to find by UUID
                    for q in questions:
                        if q.get('_uuid') == filter_question_id:
                            parent_question_name = q.get('name')
                            break

                if parent_question_name:
                    choice_filter_value = f'{parent_question_name}filter=${{{parent_question_name}}}'

        if question_type == 'rating':
            list_id = question.get('list_id', generate_random_id())
            survey_ws.append(
                [
                    'begin_group',
                    question_name,
                    question_label,
                    question_required,
                    question_appearance,
                    question_parameters,
                    '',
                    question_hint,
                    question_default,
                    question_guidance_hint,
                    question_hxl,
                    question_constraint_message,
                    question_constraint,
                    question_relevant,
                    question_trigger,
                    question_calculation,
                    question_read_only,
                    '',
                ]
                + [''] * (len(survey_headers) - 18)
            )
            for sub_question in question.get('subQuestions', []):
                sub_question_type = sub_question.get('type', 'text')
                sub_question_name = sub_question.get('name', '')
                sub_question_label = _get_default_text(
                    sub_question,
                    'label',
                    default_language_description,
                    default_language_subtag,
                )
                sub_question_required = sub_question.get('required', False)
                sub_question_parameters = sub_question.get('parameters', '')
                sub_question_hint = _get_default_text(
                    sub_question,
                    'hint',
                    default_language_description,
                    default_language_subtag,
                )
                sub_question_default = sub_question.get('default', '')
                sub_question_appearance = sub_question.get('appearance', '')
                sub_question_guidance_hint = _get_default_text(
                    sub_question,
                    'guidance_hint',
                    default_language_description,
                    default_language_subtag,
                )
                sub_question_hxl = sub_question.get('hxl', '')
                sub_question_constraint_message = sub_question.get('constraint_message', '')
                sub_question_constraint = sub_question.get('constraint', '')
                sub_question_relevant = sub_question.get('relevant', '')
                sub_question_trigger = sub_question.get('trigger', '')
                sub_question_calculation = sub_question.get('calculation', '')
                sub_question_read_only = sub_question.get('read_only', '')

                sub_row = [
                    f'select_one {list_id}',
                    sub_question_name,
                    sub_question_label,
                    sub_question_required,
                    sub_question_appearance,
                    sub_question_parameters,
                    '',
                    sub_question_hint,
                    sub_question_default,
                    sub_question_guidance_hint,
                    sub_question_hxl,
                    sub_question_constraint_message,
                    sub_question_constraint,
                    sub_question_relevant,
                    sub_question_trigger,
                    sub_question_calculation,
                    sub_question_read_only,
                    '',
                ] + [''] * (len(survey_headers) - 18)
                for language in other_languages:
                    sub_row.append(_get_language_value(sub_question.get('translations', {}), language=language))
                    sub_row.append(_get_language_value(sub_question.get('hints', {}), language=language))
                survey_ws.append(sub_row)
            survey_ws.append(
                [
                    'end_group',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                ]
                + [''] * (len(survey_headers) - 18)
            )
            for option in question.get('options', []):
                # Build the row with basic info
                row = [list_id, option['name'], option['label']]

                # Add multi-language translations for this option
                for language in other_languages:
                    translation = _get_language_value(option.get('translations', {}), language=language)
                    row.append(translation)

                # Add filter columns (no filter columns for rating options)
                row.extend([''] * len(choice_filter_columns))
                choices_ws.append(row)
        elif question_type in ('begin_group', 'begin_repeat'):
            # Handle begin_group / begin_repeat block starts
            row = [
                question_type,
                question_name,
                question_label,
                question_required,
                question_appearance,
                question_parameters,
                question_repeat_count if question_type == 'begin_repeat' else '',
                question_hint,
                question_default,
                question_guidance_hint,
                question_hxl,
                question_constraint_message,
                question_constraint,
                question_relevant,
                question_trigger,
                question_calculation,
                question_read_only,
                '',
            ]
            for language in other_languages:
                row.append(_get_language_value(question.get('translations', {}), language=language))
                row.append(_get_language_value(question.get('hints', {}), language=language))
            survey_ws.append(row + [''] * (len(survey_headers) - len(row)))
        elif question_type in ('end_group', 'end_repeat'):
            # Handle end_group / end_repeat block ends
            row = [
                question_type,
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
            ]
            for language in other_languages:
                row.append('')  # No translations for end_group
                row.append('')  # No hints for end_group
            survey_ws.append(row + [''] * (len(survey_headers) - len(row)))
        else:
            row = [
                question_type,
                question_name,
                question_label,
                question_required,
                question_appearance,
                question_parameters,
                '',
                question_hint,
                question_default,
                question_guidance_hint,
                question_hxl,
                question_constraint_message,
                question_constraint,
                question_relevant,
                question_trigger,
                question_calculation,
                question_read_only,
                choice_filter_value,
            ]
            for language in other_languages:
                row.append(_get_language_value(question.get('translations', {}), language=language))
                row.append(_get_language_value(question.get('hints', {}), language=language))
            survey_ws.append(row + [''] * (len(survey_headers) - len(row)))

            if question_type.startswith('select_one') or question_type.startswith('select_multiple'):
                list_id = question_type.split(' ')[1]
                options = question.get('options', [])
                option_filter_map = question.get('optionFilterMap', {})

                for option in options:
                    # Build filter column values for this option
                    filter_col_dict = {col: '' for col in choice_filter_columns}
                    if option_filter_map:
                        # Find parent question using filterQuestionId
                        filter_question_id = question.get('filterQuestionId')
                        if filter_question_id is not None:
                            parent_question_name = None
                            parent_question = None

                            # Try to find parent question by index first
                            try:
                                parent_question_index = int(filter_question_id)
                                if 0 <= parent_question_index < len(questions):
                                    parent_question = questions[parent_question_index]
                                    parent_question_name = parent_question.get('name')
                            except (ValueError, TypeError):
                                # If filterQuestionId is not a valid index, try to find by UUID
                                for q in questions:
                                    if q.get('_uuid') == filter_question_id:
                                        parent_question = q
                                        parent_question_name = q.get('name')
                                        break

                            if parent_question_name and parent_question:
                                col_name = f'{parent_question_name}filter'
                                # Find which parent option this child option belongs to
                                for filter_option, self_vals in option_filter_map.items():
                                    # Check if current option name matches any in self_vals
                                    if option['name'] in self_vals:
                                        # filter_option is the parent option label, find the corresponding name
                                        parent_option_name = filter_option  # Default fallback
                                        for parent_opt in parent_question.get('options', []):
                                            if parent_opt.get('label') == filter_option:
                                                parent_option_name = parent_opt.get('name', filter_option)
                                                break
                                        filter_col_dict[col_name] = parent_option_name
                                        break

                    # Write the row in the order of choices_headers
                    row = [list_id, option['name'], option['label']]

                    # Add multi-language translations for this option
                    for language in other_languages:
                        translation = _get_language_value(option.get('translations', {}), language=language)
                        row.append(translation)

                    # Add filter column values
                    row.extend([filter_col_dict.get(col, '') for col in choice_filter_columns])
                    choices_ws.append(row)

            for sub_question in question.get('subQuestions', []):
                sub_question_type = sub_question.get('type', 'text')
                sub_question_name = sub_question.get('name', '')
                sub_question_label = _get_default_text(
                    sub_question,
                    'label',
                    default_language_description,
                    default_language_subtag,
                )
                sub_question_required = sub_question.get('required', False)
                sub_question_parameters = sub_question.get('parameters', '')
                sub_question_hint = _get_default_text(
                    sub_question,
                    'hint',
                    default_language_description,
                    default_language_subtag,
                )
                sub_question_default = sub_question.get('default', '')
                sub_question_appearance = sub_question.get('appearance', '')
                sub_question_guidance_hint = _get_default_text(
                    sub_question,
                    'guidance_hint',
                    default_language_description,
                    default_language_subtag,
                )
                sub_question_hxl = sub_question.get('hxl', '')
                sub_question_constraint_message = sub_question.get('constraint_message', '')
                sub_question_constraint = sub_question.get('constraint', '')
                sub_question_relevant = sub_question.get('relevant', '')
                sub_question_trigger = sub_question.get('trigger', '')
                sub_question_calculation = sub_question.get('calculation', '')
                sub_question_read_only = sub_question.get('read_only', '')

                sub_row = [
                    sub_question_type,
                    sub_question_name,
                    sub_question_label,
                    sub_question_required,
                    sub_question_appearance,
                    sub_question_parameters,
                    '',
                    sub_question_hint,
                    sub_question_default,
                    sub_question_guidance_hint,
                    sub_question_hxl,
                    sub_question_constraint_message,
                    sub_question_constraint,
                    sub_question_relevant,
                    sub_question_trigger,
                    sub_question_calculation,
                    sub_question_read_only,
                    '',
                ]
                for language in other_languages:
                    sub_row.append(_get_language_value(sub_question.get('translations', {}), language=language))
                    sub_row.append(_get_language_value(sub_question.get('hints', {}), language=language))
                survey_ws.append(sub_row + [''] * (len(survey_headers) - len(sub_row)))

    wb.save(output_path)
    return output_path


QUESTION_VALIDATION_FIELDS = (
    'relevant',
    'constraint',
    'calculation',
    'trigger',
    'default',
    'repeat_count',
)


def _collect_question_names_for_validation(questions):
    names = set()
    if not isinstance(questions, list):
        return names

    for question in questions:
        if not isinstance(question, dict):
            continue
        q_name = question.get('name')
        if isinstance(q_name, str) and q_name.strip():
            names.add(q_name.strip())

        sub_questions = question.get('subQuestions', [])
        if isinstance(sub_questions, list):
            for sub_question in sub_questions:
                if not isinstance(sub_question, dict):
                    continue
                sub_name = sub_question.get('name')
                if isinstance(sub_name, str) and sub_name.strip():
                    names.add(sub_name.strip())

    return names


def _iter_question_expression_entries(questions):
    if not isinstance(questions, list):
        return

    for question_index, question in enumerate(questions):
        if not isinstance(question, dict):
            continue

        question_name = question.get('name', '')
        question_label = question.get('label', '')
        for field_name in QUESTION_VALIDATION_FIELDS:
            expression = question.get(field_name)
            if not isinstance(expression, str) or '${' not in expression:
                continue
            references = re.findall(r'\$\{([^}]+)\}', expression)
            if not references:
                continue
            yield {
                'question_index': question_index,
                'question_number': question_index + 1,
                'question_name': question_name,
                'question_label': question_label,
                'subquestion_index': None,
                'subquestion_name': '',
                'subquestion_label': '',
                'field': field_name,
                'expression': expression,
                'references': references,
            }

        sub_questions = question.get('subQuestions', [])
        if not isinstance(sub_questions, list):
            continue

        for subquestion_index, sub_question in enumerate(sub_questions):
            if not isinstance(sub_question, dict):
                continue

            subquestion_name = sub_question.get('name', '')
            subquestion_label = sub_question.get('label', '')
            for field_name in QUESTION_VALIDATION_FIELDS:
                expression = sub_question.get(field_name)
                if not isinstance(expression, str) or '${' not in expression:
                    continue
                references = re.findall(r'\$\{([^}]+)\}', expression)
                if not references:
                    continue
                yield {
                    'question_index': question_index,
                    'question_number': question_index + 1,
                    'question_name': question_name,
                    'question_label': question_label,
                    'subquestion_index': subquestion_index,
                    'subquestion_name': subquestion_name,
                    'subquestion_label': subquestion_label,
                    'field': field_name,
                    'expression': expression,
                    'references': references,
                }


def _collect_missing_reference_question_errors(questions):
    from difflib import get_close_matches

    known_names = _collect_question_names_for_validation(questions)
    expression_entries = list(_iter_question_expression_entries(questions))
    errors = []
    seen_keys = set()
    sorted_names = sorted(known_names)

    for entry in expression_entries:
        for reference in entry['references']:
            if not reference or reference in known_names:
                continue

            dedupe_key = (
                entry['question_index'],
                entry['subquestion_index'],
                entry['field'],
                reference,
            )
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)

            suggestions = get_close_matches(reference, sorted_names, n=3, cutoff=0.6)
            subject_name = entry['subquestion_name'] or entry['question_name']
            subject_label = entry['subquestion_label'] or entry['question_label']
            location_label = (
                f"Sub-question {entry['subquestion_index'] + 1}"
                if entry['subquestion_index'] is not None
                else f"Question {entry['question_number']}"
            )
            message = (
                f"{location_label}: `{entry['field']}` references `${{{reference}}}` "
                f"but that field does not exist."
            )

            errors.append({
                'question_index': entry['question_index'],
                'question_number': entry['question_number'],
                'question_name': entry['question_name'],
                'question_label': entry['question_label'],
                'subquestion_index': entry['subquestion_index'],
                'subquestion_name': entry['subquestion_name'],
                'subquestion_label': entry['subquestion_label'],
                'field': entry['field'],
                'expression': entry['expression'],
                'missing_reference': reference,
                'suggestions': suggestions,
                'subject_name': subject_name,
                'subject_label': subject_label,
                'message': message,
            })

    return errors


def _collect_block_structure_question_errors(questions):
    if not isinstance(questions, list):
        return []

    opening_to_closing = {
        'begin_group': 'end_group',
        'begin_repeat': 'end_repeat',
    }
    closing_to_opening = {closing: opening for opening, closing in opening_to_closing.items()}

    errors = []
    stack = []

    for question_index, question in enumerate(questions):
        if not isinstance(question, dict):
            continue

        question_type = str(question.get('type') or '').strip().lower()
        question_name = question.get('name', '')
        question_label = question.get('label', '')

        if question_type in opening_to_closing:
            stack.append(
                {
                    'type': question_type,
                    'index': question_index,
                    'name': question_name,
                    'label': question_label,
                }
            )
            continue

        if question_type not in closing_to_opening:
            continue

        question_number = question_index + 1

        if not stack:
            expected_open = closing_to_opening[question_type]
            errors.append(
                {
                    'question_index': question_index,
                    'question_number': question_number,
                    'question_name': question_name,
                    'question_label': question_label,
                    'subquestion_index': None,
                    'subquestion_name': '',
                    'subquestion_label': '',
                    'field': 'type',
                    'message': (
                        f"Question {question_number}: `{question_type}` has no matching "
                        f"`{expected_open}`."
                    ),
                    'suggestions': [],
                }
            )
            continue

        opening = stack.pop()
        expected_close = opening_to_closing[opening['type']]
        if question_type == expected_close:
            continue

        opening_number = opening['index'] + 1
        errors.append(
            {
                'question_index': question_index,
                'question_number': question_number,
                'question_name': question_name,
                'question_label': question_label,
                'subquestion_index': None,
                'subquestion_name': '',
                'subquestion_label': '',
                'field': 'type',
                'message': (
                    f"Question {question_number}: `{question_type}` closes `{opening['type']}` from "
                    f"Question {opening_number}. Expected `{expected_close}`."
                ),
                'suggestions': [expected_close],
            }
        )

    for opening in stack:
        opening_number = opening['index'] + 1
        expected_close = opening_to_closing[opening['type']]
        errors.append(
            {
                'question_index': opening['index'],
                'question_number': opening_number,
                'question_name': opening.get('name', ''),
                'question_label': opening.get('label', ''),
                'subquestion_index': None,
                'subquestion_name': '',
                'subquestion_label': '',
                'field': 'type',
                'message': (
                    f"Question {opening_number}: `{opening['type']}` has no matching "
                    f"`{expected_close}`."
                ),
                'suggestions': [expected_close],
            }
        )

    return errors


def _collect_prevalidation_question_errors(questions):
    missing_reference_errors = _collect_missing_reference_question_errors(questions)
    structure_errors = _collect_block_structure_question_errors(questions)
    return missing_reference_errors + structure_errors


def _build_form_validation_error_payload(questions, exc):
    raw_error = str(exc).strip() if exc else 'Unknown form validation error'
    question_errors = _collect_prevalidation_question_errors(questions)
    global_errors = []

    if not question_errors:
        global_errors.append({'message': raw_error})

    return {
        'valid': False,
        'error_type': 'form_validation_error',
        'message': 'Form validation failed. Fix the listed question errors and try again.',
        'validation_error': raw_error,
        'question_errors': question_errors,
        'global_errors': global_errors,
    }


def _validate_form_definition_for_enketo(form, questions, default_language, other_languages, form_name):
    prevalidation_errors = _collect_prevalidation_question_errors(questions)
    if prevalidation_errors:
        return {
            'valid': False,
            'error_type': 'form_validation_error',
            'message': 'Form validation failed. Fix the listed question errors and try again.',
            'validation_error': 'Form pre-validation failed before XLSForm conversion.',
            'question_errors': prevalidation_errors,
            'global_errors': [],
        }

    temp_fd, temp_xlsx_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(temp_fd)

    try:
        generate_xlsx_file(
            form=form,
            questions=questions,
            default_language=default_language,
            other_languages=other_languages,
            form_name=form_name,
            output_path=temp_xlsx_path,
        )
        with tempfile.NamedTemporaryFile(suffix='.xml') as temp_xml:
            xls2xform_convert(temp_xlsx_path, temp_xml.name, validate=False)
        return {
            'valid': True,
            'message': 'Form validation passed.',
            'question_errors': [],
            'global_errors': [],
            'validation_error': '',
        }
    except Exception as exc:
        return _build_form_validation_error_payload(questions, exc)
    finally:
        try:
            if os.path.exists(temp_xlsx_path):
                os.remove(temp_xlsx_path)
        except Exception:
            pass

class FormViewSet(viewsets.ModelViewSet):
    queryset = Form.objects.all()

    def _should_include_submissions_for_detail(self):
        request = getattr(self, 'request', None)
        if not request:
            return True
        include_param = request.query_params.get('include_submissions')
        if include_param is None:
            return True
        return str(include_param).lower() in ('1', 'true', 'yes', 'on')

    def _build_detail_payload(self, instance, include_submissions=None):
        if include_submissions is None:
            include_submissions = self._should_include_submissions_for_detail()

        serializer_class = (
            FormSerializer if include_submissions else FormWithoutSubmissionSerializer
        )
        serializer = serializer_class(instance, context=self.get_serializer_context())
        data = dict(serializer.data)

        default_language = instance.default_language
        other_languages = list(instance.other_languages.all())
        data['default_language_meta'] = (
            {
                'id': default_language.id,
                'description': default_language.description,
                'subtag': default_language.subtag,
            }
            if default_language
            else None
        )
        data['other_languages_meta'] = [
            {
                'id': language.id,
                'description': language.description,
                'subtag': language.subtag,
            }
            for language in other_languages
        ]

        if include_submissions:
            data['submission'] = _get_form_submission_entries(instance)

        return data

    def get_serializer_class(self):
        from .serializers import FormSerializer, FormWithoutSubmissionSerializer
        if getattr(self, 'action', None) == 'retrieve':
            if self._should_include_submissions_for_detail():
                return FormSerializer
            return FormWithoutSubmissionSerializer
        return FormSerializer

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        data = self._build_detail_payload(instance)
        return Response(data)

    @action(detail=True, methods=['get'], url_path='questions-json')
    def questions_json(self, request, pk=None):
        form = self.get_object()
        payload = self._build_detail_payload(form, include_submissions=False)
        file_slug = slugify(form.name or '') or f'form-{form.id}'
        filename = f'{file_slug}-questions.json'
        response = HttpResponse(
            json.dumps(payload, ensure_ascii=False, indent=2),
            content_type='application/json; charset=utf-8',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def get_queryset(self):
        """
        Filter forms based on user role:
        - Role 4 and 5: Only see forms they have access to (via profile.forms)
        - Other roles: See all forms
        """
        user = self.request.user
        user_role = getattr(user, 'role', 4)
        queryset = Form.objects.none()

        # For role 4 and 5, filter by profile.forms
        if user_role in [4, 5]:
            if hasattr(user, 'profile'):
                user_form_ids = user.profile.forms.values_list('id', flat=True)
                queryset = Form.objects.filter(id__in=user_form_ids)
            else:
                # User has no profile, return empty queryset
                queryset = Form.objects.none()

        # For officer role, restrict to assigned projects only
        elif user_role == 6:
            project_ids = _get_assigned_project_ids(user)
            if not project_ids:
                queryset = Form.objects.none()
            else:
                queryset = Form.objects.filter(project_id__in=project_ids)

        # For other roles (1, 2, 3), return all forms
        else:
            queryset = Form.objects.all()

        if getattr(self, 'action', None) == 'retrieve':
            queryset = queryset.select_related('default_language').prefetch_related('other_languages')

        if getattr(self, 'action', None) == 'retrieve' and not self._should_include_submissions_for_detail():
            queryset = queryset.defer('submission')
        elif getattr(self, 'action', None) == 'summary_stats':
            queryset = queryset.defer('questions', 'translations', 'description')
        return queryset

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        old_default_language = instance.default_language
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        refresh_xlsx_fields = {'questions', 'name', 'default_language', 'other_languages'}
        should_refresh_xlsx = any(field in request.data for field in refresh_xlsx_fields)

        if should_refresh_xlsx:
            import copy

            preview_questions = serializer.validated_data.get('questions', instance.questions)
            if not isinstance(preview_questions, list):
                preview_questions = []
            else:
                preview_questions = copy.deepcopy(preview_questions)

            preview_default_language = serializer.validated_data.get(
                'default_language',
                instance.default_language,
            )
            preview_other_languages = serializer.validated_data.get('other_languages')
            if preview_other_languages is None:
                preview_other_languages = list(instance.other_languages.all())
            else:
                preview_other_languages = list(preview_other_languages)

            preview_form_name = serializer.validated_data.get('name', instance.name) or 'Form'

            _sync_questions_for_default_language_switch(
                preview_questions,
                old_default_language,
                preview_default_language,
            )

            validation_result = _validate_form_definition_for_enketo(
                form=instance,
                questions=preview_questions,
                default_language=preview_default_language,
                other_languages=preview_other_languages,
                form_name=preview_form_name,
            )
            if not validation_result.get('valid', False):
                return Response(validation_result, status=status.HTTP_400_BAD_REQUEST)

        self.perform_update(serializer)

        # Update the associated XLSX file using shared function
        form = serializer.instance

        if should_refresh_xlsx:
            if form.default_language_id and form.other_languages.filter(id=form.default_language_id).exists():
                form.other_languages.remove(form.default_language_id)

            if _sync_questions_for_default_language_switch(
                form.questions,
                old_default_language,
                form.default_language,
            ):
                form.save(update_fields=['questions', 'updated_at'])

            questions = form.questions if isinstance(form.questions, list) else []

            default_language = form.default_language
            other_languages = form.other_languages.all()
            form_name = form.name if hasattr(form, 'name') else 'Form'

            # Generate XLSX with consistent choice_filter logic
            generate_xlsx_file(form, questions, default_language, other_languages, form_name)

        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='validate')
    def validate(self, request, pk=None):
        form = self.get_object()
        questions = form.questions if isinstance(form.questions, list) else []
        default_language = form.default_language
        other_languages = form.other_languages.all()
        form_name = form.name if hasattr(form, 'name') else 'Form'

        validation_result = _validate_form_definition_for_enketo(
            form=form,
            questions=questions,
            default_language=default_language,
            other_languages=other_languages,
            form_name=form_name,
        )
        return Response(validation_result, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    @transaction.atomic
    def create_form(self, request, pk=None):
        print(f"🔍 create_form called - user: {request.user.username}, project_id: {pk}")
        print(f"🔍 user role: {getattr(request.user, 'role', 'NO_ROLE')}")
        print(f"🔍 user is_superuser: {request.user.is_superuser}")
        print(f"🔍 request data: {request.data}")

        try:
            project = Project.objects.get(pk=pk)
            print(f"🔍 project found: {project.name}, created_by: {project.created_by.username}")

            # Check if user has permission to create forms for this project
            user = request.user
            user_role = getattr(user, 'role', 4)  # Default to role 4 (user)

            # Allow admins (role 1), organization users (role 2), and project users (role 3)
            if user_role == 6:
                has_access = hasattr(user, 'profile') and user.profile.projects.filter(id=project.id).exists()
                if not has_access:
                    return Response(
                        {'error': 'You do not have permission to create forms for this project'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            elif user_role not in [1, 2, 3] and project.created_by != user:
                return Response(
                    {'error': 'You do not have permission to create forms for this project'},
                    status=status.HTTP_403_FORBIDDEN
                )

            form_name = request.data.get('name')
            questions = request.data.get('questions', [])
            form_style = request.data.get('formStyle', 'default')  # Add form style parameter
            default_language_id = request.data.get('default_language')
            other_languages = request.data.get('other_languages', [])

            form = Form.objects.create(
                project=project,
                name=form_name,
                questions=questions,
                form_style=form_style,  # Add form style to creation
                default_language_id=default_language_id
            )

            if other_languages:
                form.other_languages.set(other_languages)

            # Generate XLSX with consistent choice_filter logic using shared function
            default_language = form.default_language
            other_languages_qs = form.other_languages.all()
            generate_xlsx_file(form, questions, default_language, other_languages_qs, form_name)

            return Response({'message': 'Form created and files generated successfully', 'form_id': form.id}, status=status.HTTP_201_CREATED)
        except Project.DoesNotExist:
            print(f"❌ Project {pk} not found")
            return Response({'error': 'Project not found'}, status=status.HTTP_404_NOT_FOUND)

    def destroy(self, request, *args, **kwargs):
        """Move form to trash instead of permanent deletion"""
        form = self.get_object()

        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"🗑️ FormViewSet.destroy called for form {form.id}: {form.name}")

        try:
            # Move to trash
            move_to_trash(form, 'form', request.user)
            logger.info(f"✅ Form {form.id} moved to trash successfully")
            return Response({'message': 'Form moved to trash successfully'}, status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            logger.error(f"❌ Failed to delete form {form.id}: {e}")
            return Response({'error': f'Failed to delete form: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='summary-stats')
    def summary_stats(self, request, pk=None):
        form = self.get_object()
        now_date = timezone.localdate()

        submission_queryset = Submission.objects.filter(form=form, is_deleted=False)
        total_submissions = submission_queryset.count()

        if total_submissions > 0:
            current_tz = timezone.get_current_timezone()
            start_31 = now_date - timedelta(days=30)
            start_12_year, start_12_month = _shift_year_month(now_date.year, now_date.month, -11)
            start_12 = date(start_12_year, start_12_month, 1)
            start_31_dt = timezone.make_aware(datetime.combine(start_31, time.min), current_tz)
            start_12_dt = timezone.make_aware(datetime.combine(start_12, time.min), current_tz)

            date_counts = {}
            for row in (
                submission_queryset
                .filter(created_at__gte=start_31_dt)
                .annotate(bucket=TruncDate('created_at'))
                .values('bucket')
                .annotate(total=Count('id'))
            ):
                bucket = row.get('bucket')
                if bucket is None:
                    continue
                if isinstance(bucket, datetime):
                    if timezone.is_naive(bucket):
                        bucket = timezone.make_aware(bucket, timezone.get_current_timezone())
                    bucket = timezone.localdate(bucket)
                date_counts[bucket] = int(row.get('total', 0))

            month_counts = {}
            for row in (
                submission_queryset
                .filter(created_at__gte=start_12_dt)
                .annotate(bucket=TruncMonth('created_at'))
                .values('bucket')
                .annotate(total=Count('id'))
            ):
                bucket = row.get('bucket')
                if bucket is None:
                    continue
                if isinstance(bucket, datetime):
                    if timezone.is_naive(bucket):
                        bucket = timezone.make_aware(bucket, timezone.get_current_timezone())
                    bucket = timezone.localtime(bucket)
                month_counts[(bucket.year, bucket.month)] = int(row.get('total', 0))

            today_submissions = int(date_counts.get(now_date, 0))
            bar_data = _build_submission_bar_data(now_date, date_counts, month_counts)
        else:
            legacy_submissions = form.submission if isinstance(form.submission, list) else []
            date_counts = Counter()
            month_counts = Counter()
            for entry in legacy_submissions:
                submission_date = _extract_submission_date(entry)
                if submission_date is None:
                    continue
                date_counts[submission_date] += 1
                month_counts[(submission_date.year, submission_date.month)] += 1

            total_submissions = len(legacy_submissions)
            today_submissions = int(date_counts.get(now_date, 0))
            bar_data = _build_submission_bar_data(now_date, date_counts, month_counts)

        return Response({
            'form_id': form.id,
            'total_submissions': int(total_submissions),
            'today_submissions': int(today_submissions),
            'bar_data': bar_data,
        })

    @action(detail=True, methods=['get'], url_path='xlsx')
    def xlsx(self, request, pk=None):
        form_id = pk
        file_path = os.path.join(settings.MEDIA_ROOT, 'update', f'{form_id}.xlsx')
        if not os.path.exists(file_path):
            raise Http404("XLSX file not found.")
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=f'{form_id}.xlsx')

class SubmissionViewSet(viewsets.ModelViewSet):
    queryset = Submission.objects.all()

    def get_serializer_class(self):
        from .serializers import SubmissionSerializer
        return SubmissionSerializer

    def destroy(self, request, *args, **kwargs):
        """Soft-delete submission and place it in trash for restore."""
        submission = self.get_object()

        try:
            move_submission_to_soft_trash(submission, request.user)
            return Response({'message': 'Submission moved to trash successfully'}, status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            return Response({'error': f'Failed to delete submission: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all()

    def _should_include_forms(self):
        request = getattr(self, 'request', None)
        if not request:
            return False
        include_param = request.query_params.get('include_forms')
        if include_param is None:
            return False
        return str(include_param).lower() in ('1', 'true', 'yes', 'on')

    def _should_include_forms_for_detail(self):
        request = getattr(self, 'request', None)
        if not request:
            return True
        include_param = request.query_params.get('include_forms')
        if include_param is None:
            return True
        return str(include_param).lower() in ('1', 'true', 'yes', 'on')

    def _should_filter_has_templates(self):
        request = getattr(self, 'request', None)
        if not request:
            return False
        include_param = request.query_params.get('has_templates')
        if include_param is None:
            return False
        return str(include_param).lower() in ('1', 'true', 'yes', 'on')

    def _filter_queryset(self, queryset):
        request = getattr(self, 'request', None)
        if not request:
            return queryset
        if getattr(self, 'action', None) in ('list', 'user_projects'):
            organization_id = request.query_params.get('organization')
            if organization_id:
                queryset = queryset.filter(organization_id=organization_id)
        return queryset

    def get_serializer_class(self):
        from .serializers import (
            ProjectSerializer,
            ProjectListSerializer,
            ProjectWithFormListSerializer,
        )

        if getattr(self, 'action', None) == 'retrieve':
            if self._should_include_forms_for_detail():
                return ProjectSerializer
            return ProjectListSerializer

        if getattr(self, 'action', None) in ('list', 'user_projects'):
            if self._should_include_forms():
                return ProjectWithFormListSerializer
            return ProjectListSerializer
        return ProjectSerializer

    def get_queryset(self):
        from .models import Form, FormAccess, Setting

        request = getattr(self, 'request', None)
        user = request.user if request else None
        user_role = getattr(user, 'role', 4) if user else 4
        assigned_project_ids = None
        if user and user_role == 6:
            assigned_project_ids = _get_assigned_project_ids(user)
            if not assigned_project_ids:
                return Project.objects.none()

        if getattr(self, 'action', None) in ('list', 'user_projects'):
            if not self._should_include_forms():
                queryset = (
                    Project.objects.select_related('organization', 'created_by')
                    .annotate(
                        forms_count=Count('forms', distinct=True),
                        templates_count=Count('templates', distinct=True),
                    )
                    .order_by('name')
                )
                queryset = self._filter_queryset(queryset)
                if self._should_filter_has_templates():
                    queryset = queryset.filter(templates_count__gt=0)
                if assigned_project_ids is not None:
                    queryset = queryset.filter(id__in=assigned_project_ids)
                return queryset

            optimized_forms = Form.objects.only(
                'id', 'project_id', 'name', 'description',
                'created_at', 'updated_at', 'default_language_id',
                'enketo_id', 'allow_anonymous_submissions',
                'template_id', 'form_style', 'criteria'
            ).order_by('id')

            optimized_access = FormAccess.objects.select_related('user').only(
                'id', 'project_id', 'user_id', 'access_level'
            )

            optimized_settings = Setting.objects.only(
                'id', 'project_id', 'key', 'value'
            )

            queryset = (
                Project.objects.select_related('organization', 'created_by')
                .prefetch_related(
                    Prefetch('forms', queryset=optimized_forms),
                    Prefetch('form_access', queryset=optimized_access),
                    Prefetch('project_settings', queryset=optimized_settings),
                )
                .order_by('name')
            )

            queryset = self._filter_queryset(queryset)
            if assigned_project_ids is not None:
                queryset = queryset.filter(id__in=assigned_project_ids)
            return queryset

        if getattr(self, 'action', None) == 'retrieve':
            if self._should_include_forms_for_detail():
                queryset = (
                    Project.objects.select_related('organization', 'created_by')
                    .prefetch_related('forms', 'form_access', 'project_settings')
                    .order_by('name')
                )
            else:
                queryset = (
                    Project.objects.select_related('organization', 'created_by')
                    .annotate(
                        forms_count=Count('forms', distinct=True),
                        templates_count=Count('templates', distinct=True),
                    )
                    .order_by('name')
                )
        elif getattr(self, 'action', None) in ('stats', 'forms'):
            queryset = (
                Project.objects.select_related('organization', 'created_by')
                .order_by('name')
            )
        else:
            queryset = (
                Project.objects.select_related('organization', 'created_by')
                .prefetch_related('forms', 'form_access', 'project_settings')
                .order_by('name')
            )

        if assigned_project_ids is not None:
            queryset = queryset.filter(id__in=assigned_project_ids)
        return queryset

    @action(detail=False, methods=['get'], url_path='user-projects')
    def user_projects(self, request):
        """
        Optimized endpoint that returns projects filtered by:
        - Projects created by the user
        - Projects assigned to the user
        Returns merged list without duplicates
        """
        from django.db.models import Q
        from .models import Form

        user = request.user

        if user.is_superuser or getattr(user, 'role', 4) == 1:
            queryset = self.get_queryset().distinct().order_by('name')
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)

        assigned_project_ids = []
        assigned_form_ids = []
        assigned_org_ids = []

        if hasattr(user, 'profile'):
            if hasattr(user.profile, 'projects'):
                try:
                    assigned_project_ids = list(
                        user.profile.projects.all().values_list('id', flat=True)
                    )
                except Exception as exc:
                    print(f"Error getting profile projects: {exc}")
            if hasattr(user.profile, 'organizations'):
                try:
                    assigned_org_ids = list(
                        user.profile.organizations.all().values_list('id', flat=True)
                    )
                except Exception as exc:
                    print(f"Error getting profile organizations: {exc}")
            if hasattr(user.profile, 'forms'):
                try:
                    assigned_form_ids = list(
                        user.profile.forms.all().values_list('id', flat=True)
                    )
                except Exception as exc:
                    print(f"Error getting profile forms: {exc}")

        form_project_ids = []
        if assigned_form_ids:
            form_project_ids = list(
                Form.objects.filter(id__in=assigned_form_ids).values_list('project_id', flat=True)
            )

        if getattr(user, 'role', 4) == 6:
            queryset = self.get_queryset().filter(
                Q(id__in=assigned_project_ids)
            ).distinct().order_by('name')
        else:
            queryset = self.get_queryset().filter(
                Q(created_by=user) |
                Q(id__in=assigned_project_ids) |
                Q(id__in=form_project_ids) |
                Q(organization_id__in=assigned_org_ids)
            ).distinct().order_by('name')

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def perform_create(self, serializer):
        project = serializer.save(created_by=self.request.user)
        if getattr(self.request.user, 'role', 4) == 6:
            try:
                self.request.user.profile.projects.add(project)
            except Exception:
                pass

    def destroy(self, request, *args, **kwargs):
        """Move project to trash instead of permanent deletion"""
        project = self.get_object()

        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"🗑️ ProjectViewSet.destroy called for project {project.id}: {project.name}")

        try:
            # Move to trash
            move_to_trash(project, 'project', request.user)
            logger.info(f"✅ Project {project.id} moved to trash successfully")
            return Response({'message': 'Project moved to trash successfully'}, status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            logger.error(f"❌ Failed to delete project {project.id}: {e}")
            return Response({'error': f'Failed to delete project: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def forms(self, request, pk=None):
        from .serializers import FormSerializer, FormWithoutSubmissionSerializer, FormListSerializer

        include_submissions = str(
            request.query_params.get('include_submissions', 'false')
        ).lower() in ('1', 'true', 'yes', 'on')
        compact = str(
            request.query_params.get('compact', 'false')
        ).lower() in ('1', 'true', 'yes', 'on')

        project = self.get_object()
        forms = Form.objects.filter(project=project)
        if include_submissions:
            serializer_class = FormSerializer
        elif compact:
            serializer_class = FormListSerializer
        else:
            serializer_class = FormWithoutSubmissionSerializer
        serializer = serializer_class(forms, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def stats(self, request, pk=None):
        project = self.get_object()

        forms_count = Form.objects.filter(project=project).count()
        submissions_count = Submission.objects.filter(form__project=project, is_deleted=False).count()

        # Backward compatibility: older projects may still store submissions
        # in Form.submission JSON instead of Submission rows.
        if submissions_count == 0:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT COALESCE(
                        SUM(
                            CASE
                                WHEN jsonb_typeof(submission) = 'array'
                                THEN jsonb_array_length(submission)
                                ELSE 0
                            END
                        ),
                        0
                    )
                    FROM api_form
                    WHERE project_id = %s
                    """,
                    [project.id],
                )
                row = cursor.fetchone()
                submissions_count = int(row[0] if row and row[0] is not None else 0)

        return Response({
            'project_id': project.id,
            'forms_count': forms_count,
            'submissions_count': submissions_count,
        })

from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator

@method_decorator(csrf_exempt, name="dispatch")
class CustomAuthToken(APIView):
    permission_classes = [AllowAny]
    authentication_classes = []  # Skip all authentication for login

    def options(self, request, *args, **kwargs):
        """Handle OPTIONS requests for CORS preflight"""
        response = Response()
        response["Access-Control-Allow-Origin"] = request.META.get("HTTP_ORIGIN", "*")
        response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        response["Access-Control-Allow-Headers"] = "Content-Type, X-CSRFToken, X-Requested-With"
        response["Access-Control-Allow-Credentials"] = "true"
        return response

    def post(self, request, *args, **kwargs):
        username = request.data.get('username') or request.data.get('email')
        password = request.data.get('password')

        print(f"Login attempt: {username}")

        user = authenticate(username=username, password=password)
        if not user and username:
            fallback_user = User.objects.filter(email__iexact=username).first()
            if fallback_user:
                user = authenticate(username=fallback_user.username, password=password)
        if user:
            token, created = Token.objects.get_or_create(user=user)
            
            # Use optimized serializer for login response
            from .serializers import UserListSerializer
            user_data = UserListSerializer(user).data

            response_data = {'token': token.key, 'user': user_data}
            response = Response(response_data)

            # Set explicit CORS headers for login responses
            response["Access-Control-Allow-Origin"] = request.META.get("HTTP_ORIGIN", "http://localhost:5173")
            response["Access-Control-Allow-Credentials"] = "true"

            # Also set the Enketo cookie during regular login
            auth_token = generate_auth_token(user)

            # For local development
            if settings.DEBUG:
                response.set_cookie(
                    'commicplan_auth',
                    auth_token,
                    domain=None,  # No domain restriction for local dev
                    httponly=False,  # Allow JS to access the cookie in dev
                    secure=False,  # Allow HTTP for local dev
                    samesite='Lax',
                    max_age=7*24*60*60
                )
            else:
                # Production settings
                secure_cookie = request.is_secure() or settings.FRONTEND_URL.startswith("https://")
                cookie_kwargs = {
                    'httponly': True,
                    'secure': secure_cookie,
                    'samesite': 'None' if secure_cookie else 'Lax',
                    'max_age': 7 * 24 * 60 * 60,
                }
                if settings.AUTH_COOKIE_DOMAIN:
                    cookie_kwargs['domain'] = settings.AUTH_COOKIE_DOMAIN
                response.set_cookie('commicplan_auth', auth_token, **cookie_kwargs)

            return response
        return Response({'error': 'Invalid credentials'}, status=status.HTTP_400_BAD_REQUEST)

class RegisterUser(APIView):
    permission_classes = [IsAuthenticated]  # Changed from AllowAny to require authentication

    def post(self, request):
        if _is_brac_download_user(request.user):
            return Response(
                {'error': 'This account is restricted to BRAC data download only.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        from .serializers import UserSerializer
        data = request.data.copy()
        # Set default role to 4 (user)
        data['role'] = 4
        # Pass the creating user through context
        context = {'created_by': request.user}
        print(f"🔧 Setting created_by to: {request.user.username} (ID: {request.user.id})")

        serializer = UserSerializer(data=data, context=context)
        if serializer.is_valid():
            user = serializer.save()
            print(f"✅ User created: {user.username}, created_by: {user.created_by}")
            return Response({'message': 'User registered successfully'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LanguageViewSet(viewsets.ModelViewSet):
    queryset = Language.objects.filter(
        type__iexact='language',
        deprecated__isnull=True,
    ).order_by('description', 'subtag')

    def get_serializer_class(self):
        from .serializers import LanguageSerializer
        return LanguageSerializer
    permission_classes = [IsAuthenticated]  # Ensure only authenticated users can access this view

def _get_accessible_ids_for_user(user):
    from .models import Project, Form

    org_ids = set()
    project_ids = set()
    form_ids = set()

    profile = None
    try:
        profile = user.profile
    except Exception:
        profile = None

    if profile is not None:
        try:
            org_ids.update(profile.organizations.values_list('id', flat=True))
        except Exception:
            pass
        try:
            project_ids.update(profile.projects.values_list('id', flat=True))
        except Exception:
            pass
        try:
            form_ids.update(profile.forms.values_list('id', flat=True))
        except Exception:
            pass

    try:
        project_ids.update(Project.objects.filter(created_by=user).values_list('id', flat=True))
    except Exception:
        pass

    if org_ids:
        project_ids.update(
            Project.objects.filter(organization_id__in=org_ids).values_list('id', flat=True)
        )

    if form_ids:
        project_ids.update(
            Form.objects.filter(id__in=form_ids).values_list('project_id', flat=True)
        )

    if project_ids:
        org_ids.update(
            Project.objects.filter(id__in=project_ids).values_list('organization_id', flat=True)
        )
        form_ids.update(
            Form.objects.filter(project_id__in=project_ids).values_list('id', flat=True)
        )

    org_ids.discard(None)
    return org_ids, project_ids, form_ids

def _get_assigned_project_ids(user):
    try:
        return set(user.profile.projects.values_list('id', flat=True))
    except Exception:
        return set()

def _is_microstatification_admin(user):
    return getattr(user, 'role', 4) == MICROSTATIFICATION_ADMIN_ROLE

def _get_microstatification_managed_users_filter(user):
    return (
        Q(role__in=MICROSTATIFICATION_GLOBAL_VISIBLE_ROLES)
        | Q(role=MICROSTATIFICATION_CREATOR_SCOPED_ROLE, created_by=user)
    )

def _get_microstatification_user_filter(user):
    return Q(id=user.id) | _get_microstatification_managed_users_filter(user)

def _get_user_access_filter_for_user(user, org_ids=None, project_ids=None, form_ids=None):
    if org_ids is None or project_ids is None or form_ids is None:
        org_ids, project_ids, form_ids = _get_accessible_ids_for_user(user)

    queryset = Q(created_by=user) | Q(id=user.id)

    if org_ids:
        queryset |= Q(profile__organizations__id__in=org_ids)
    if project_ids:
        queryset |= Q(profile__projects__id__in=project_ids)
    if form_ids:
        queryset |= Q(profile__forms__id__in=form_ids)

    return queryset

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.select_related('profile').prefetch_related(
        'profile__organizations',
        'profile__projects',
        'profile__forms',
        'profile__templates'
    )
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'username'
    pagination_class = None  # Disable pagination for now to return all users

    def get_queryset(self):
        """Filter users based on logged-in user's access rights"""
        user = self.request.user
        base_queryset = self.queryset

        if user.is_superuser or getattr(user, 'role', 4) == 1:
            queryset = base_queryset
        elif _is_microstatification_admin(user):
            queryset = base_queryset.filter(
                _get_microstatification_user_filter(user)
            ).distinct()
        else:
            queryset = base_queryset.filter(_get_user_access_filter_for_user(user)).distinct()

        # Apply the filter
        if self.action == 'list':
            if _is_microstatification_admin(user):
                return queryset.filter(
                    _get_microstatification_managed_users_filter(user)
                ).distinct()
            return queryset.exclude(role=1).distinct()

        return queryset

    def get_serializer_class(self):
        """Use lightweight serializer for list view"""
        if self.action == 'list':
            from .serializers import UserListSerializer
            return UserListSerializer
        return UserSerializer

    @action(detail=False, methods=['get'], url_path='list-view')
    def list_view(self, request):
        """
        Paginated, server-filtered user list for the user/all page.
        """
        from django.db.models import Q
        from django.utils.dateparse import parse_date
        from .serializers import UserListViewSerializer

        user = request.user
        queryset = User.objects.select_related('created_by')

        if not (user.is_superuser or getattr(user, 'role', 4) == 1):
            if _is_microstatification_admin(user):
                queryset = queryset.filter(
                    _get_microstatification_managed_users_filter(user)
                )
            else:
                queryset = queryset.filter(
                    _get_user_access_filter_for_user(user)
                ).distinct()
                queryset = queryset.exclude(role=1)
        else:
            queryset = queryset.exclude(role=1)

        organization_id = request.query_params.get('organization')
        if organization_id:
            queryset = queryset.filter(
                Q(profile__organizations__id=organization_id) |
                Q(profile__projects__organization_id=organization_id) |
                Q(profile__forms__project__organization_id=organization_id)
            )

        project_id = request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(
                Q(profile__projects__id=project_id) |
                Q(profile__forms__project_id=project_id)
            )

        form_id = request.query_params.get('form')
        if form_id:
            queryset = queryset.filter(profile__forms__id=form_id)

        created_from = parse_date(request.query_params.get('created_from') or "")
        if created_from:
            queryset = queryset.filter(date_joined__date__gte=created_from)

        created_to = parse_date(request.query_params.get('created_to') or "")
        if created_to:
            queryset = queryset.filter(date_joined__date__lte=created_to)

        created_by_filter = (request.query_params.get('created_by') or "all").lower()
        if created_by_filter == "me":
            queryset = queryset.filter(created_by=user)
        elif created_by_filter == "others":
            queryset = queryset.exclude(created_by=user)

        search = (request.query_params.get('search') or "").strip()
        if search:
            search_q = (
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(username__icontains=search) |
                Q(email__icontains=search) |
                Q(created_by__username__icontains=search)
            )

            role_map = {
                "admin": 1,
                "administrator": 1,
                "organizer": 2,
                "organization": 2,
                "project": 3,
                "project manager": 3,
                "user": 4,
                "data collector": 5,
                "datacollector": 5,
                "officer": 6,
                "microstatification admin": 7,
                "micro admin": 7,
                "sk": 8,
                "shw": 9,
            }
            normalized = search.lower()
            role_matches = []
            if normalized in role_map:
                role_matches.append(role_map[normalized])
            elif normalized in {"pm", "project manager"}:
                role_matches.append(3)
            if normalized.isdigit():
                try:
                    role_matches.append(int(normalized))
                except ValueError:
                    pass
            if role_matches:
                search_q |= Q(role__in=role_matches)

            queryset = queryset.filter(search_q)

        queryset = queryset.distinct().order_by('username')

        try:
            page = int(request.query_params.get('page', 1))
        except ValueError:
            page = 1
        try:
            page_size = int(request.query_params.get('page_size', 10))
        except ValueError:
            page_size = 10

        if page < 1:
            page = 1
        if page_size < 1:
            page_size = 10
        if page_size > 200:
            page_size = 200

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        page_queryset = queryset[start:end]

        serializer = UserListViewSerializer(page_queryset, many=True)
        return Response({
            "count": total,
            "page": page,
            "page_size": page_size,
            "results": serializer.data,
        })

    @action(detail=False, methods=['get'], url_path='basic-list')
    def basic_list(self, request):
        """
        Optimized endpoint that returns only basic user info (id, username, email)
        for dropdown/selection purposes. Much faster than full user serialization.
        Respects same access control as main list view.
        """
        user = request.user
        if user.is_superuser or getattr(user, 'role', 4) == 1:
            users = (
                User.objects.exclude(role=1)
                .distinct()
                .only('id', 'username', 'email')
                .order_by('username')
            )
        elif _is_microstatification_admin(user):
            users = (
                User.objects.filter(_get_microstatification_managed_users_filter(user))
                .distinct()
                .only('id', 'username', 'email')
                .order_by('username')
            )
        else:
            queryset = _get_user_access_filter_for_user(user)
            users = (
                User.objects.filter(queryset)
                .exclude(role=1)
                .distinct()
                .only('id', 'username', 'email')
                .order_by('username')
            )

        data = [
            {
                'id': user.id,
                'username': user.username,
                'email': user.email or ''
            }
            for user in users
        ]
        return Response(data)

    @action(detail=False, methods=['get'], url_path='user-list')
    def user_list(self, request):
        """
        Optimized endpoint that returns user info with role (id, username, email, role)
        for assign user dropdown. Much faster than full user serialization.
        Respects same access control as main list view.
        """
        user = request.user

        if user.is_superuser or getattr(user, 'role', 4) == 1:
            users = (
                User.objects.exclude(role=1)
                .distinct()
                .only('id', 'username', 'email', 'role')
                .order_by('username')
            )
        elif _is_microstatification_admin(user):
            users = (
                User.objects.filter(_get_microstatification_managed_users_filter(user))
                .distinct()
                .only('id', 'username', 'email', 'role')
                .order_by('username')
            )
        else:
            queryset = _get_user_access_filter_for_user(user)
            users = (
                User.objects.filter(queryset)
                .exclude(role=1)
                .distinct()
                .only('id', 'username', 'email', 'role')
                .order_by('username')
            )

        data = [
            {
                'id': user.id,
                'username': user.username,
                'email': user.email or '',
                'role': user.role if hasattr(user, 'role') else 4
            }
            for user in users
        ]
        return Response(data)

    def _sanitize_microstatification_admin_update(self, request, instance):
        if getattr(instance, 'role', None) not in MICROSTATIFICATION_MANAGED_ROLES:
            raise PermissionDenied('Microstatification admins can update only User, SK, and SHW users.')

        raw_profile = request.data.get('profile', {})
        if raw_profile in (None, ''):
            raw_profile = {}
        if not isinstance(raw_profile, dict):
            raise ValidationError({'profile': 'Invalid profile payload.'})

        sanitized_data = {}

        if 'role' in request.data:
            try:
                requested_role = int(request.data.get('role'))
            except (TypeError, ValueError):
                raise ValidationError({'role': 'Role must be a valid number.'})

            if requested_role not in MICROSTATIFICATION_MANAGED_ROLES:
                raise PermissionDenied('Microstatification admins can assign only User, SK, or SHW roles.')

            sanitized_data['role'] = requested_role

        sanitized_profile = {
            key: value
            for key, value in raw_profile.items()
            if key in MICROSTATIFICATION_ALLOWED_PROFILE_FIELDS
        }
        if sanitized_profile:
            sanitized_data['profile'] = sanitized_profile

        return sanitized_data

    def _sanitize_microstatification_admin_create(self, request):
        raw_profile = request.data.get('profile', {})
        if raw_profile in (None, ''):
            raw_profile = {}
        if not isinstance(raw_profile, dict):
            raise ValidationError({'profile': 'Invalid profile payload.'})

        raw_role = request.data.get('role', 4)
        try:
            requested_role = int(raw_role)
        except (TypeError, ValueError):
            raise ValidationError({'role': 'Role must be a valid number.'})

        if requested_role not in MICROSTATIFICATION_MANAGED_ROLES:
            raise PermissionDenied('Microstatification admins can create only User, SK, or SHW users.')

        sanitized_data = {
            'username': request.data.get('username'),
            'email': request.data.get('email'),
            'password': request.data.get('password'),
            'first_name': request.data.get('first_name', ''),
            'last_name': request.data.get('last_name', ''),
            'role': requested_role,
            'is_staff': False,
        }

        sanitized_profile = {
            key: value
            for key, value in raw_profile.items()
            if key in MICROSTATIFICATION_ALLOWED_PROFILE_FIELDS
        }

        username = (request.data.get('username') or '').strip()
        if requested_role == 8:
            sanitized_profile.setdefault('data_collection_type', 'microstatification')
            sanitized_profile.setdefault('micro_role', 'sk')
            sanitized_profile.setdefault('micro_designation', 'SK')
            sanitized_profile.setdefault('micro_sk_shw_name', username)
        elif requested_role == 9:
            sanitized_profile.setdefault('data_collection_type', 'microstatification')
            sanitized_profile.setdefault('micro_role', 'shw')
            sanitized_profile.setdefault('micro_designation', 'SHW')
            sanitized_profile.setdefault('micro_sk_shw_name', username)
        else:
            sanitized_profile.setdefault('data_collection_type', 'normal')

        if sanitized_profile:
            sanitized_data['profile'] = sanitized_profile

        return sanitized_data

    def create(self, request, *args, **kwargs):
        """Override create to set created_by field"""
        serializer_data = request.data
        if _is_microstatification_admin(request.user):
            serializer_data = self._sanitize_microstatification_admin_create(request)

        serializer = self.get_serializer(data=serializer_data, context={'created_by': request.user})
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer_data = request.data
        serializer_partial = partial

        if _is_microstatification_admin(request.user):
            serializer_data = self._sanitize_microstatification_admin_update(
                request,
                instance,
            )
            serializer_partial = True

        serializer = self.get_serializer(
            instance,
            data=serializer_data,
            partial=serializer_partial,
        )
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        # Check for role and set superuser/staff status
        role = serializer_data.get('role')
        if role is not None:
            if str(role) == "1":
                instance.is_superuser = True
                instance.is_staff = True
            else:
                instance.is_superuser = False
                # Optionally, set is_staff based on your logic
            instance.save()

        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        if _is_microstatification_admin(request.user):
            raise PermissionDenied('Microstatification admins cannot delete users.')

        return super().destroy(request, *args, **kwargs)

class DashboardSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        warnings = set()

        totals = {
            "users": 0,
            "organizations": 0,
            "projects": 0,
            "forms": 0,
        }
        org_stats = []
        chart = {"week": [], "month": []}

        user_queryset = User.objects.none()
        try:
            user_filter = _get_user_access_filter_for_user(user)
            user_queryset = User.objects.filter(user_filter).exclude(role=1).distinct()
            totals["users"] = user_queryset.count()
        except Exception as exc:
            logging.getLogger(__name__).warning(
                "DashboardSummaryView: failed to count users: %s", exc
            )
            warnings.add("user count")

        org_ids = set()
        project_ids = set()
        form_ids = set()

        try:
            org_ids, project_ids, form_ids = _get_accessible_ids_for_user(user)
            totals["projects"] = len(project_ids)
            totals["organizations"] = len(org_ids)
        except Exception as exc:
            logging.getLogger(__name__).warning(
                "DashboardSummaryView: failed to fetch access ids: %s", exc
            )
            warnings.add("access ids")

        try:
            from django.db.models import Q, F

            forms_qs = Form.objects.filter(
                Q(project_id__in=project_ids) | Q(id__in=form_ids)
            )
            forms_qs = forms_qs.exclude(
                template_generated_lookup_forms__isnull=False
            )
            forms_qs = forms_qs.filter(
                Q(template__isnull=True) | Q(template__data_collection_form_id=F("id"))
            )
            totals["forms"] = forms_qs.distinct().count()
        except Exception as exc:
            logging.getLogger(__name__).warning(
                "DashboardSummaryView: failed to count forms: %s", exc
            )
            warnings.add("form count")

        try:
            if org_ids:
                org_counts = (
                    Project.objects.filter(organization_id__in=org_ids)
                    .values("organization_id")
                    .annotate(project_count=Count("id"))
                )

                organization_lookup = {
                    entry["id"]: entry["name"]
                    for entry in Organization.objects.filter(id__in=org_ids).values(
                        "id", "name"
                    )
                }

                org_stats = [
                    {
                        "id": entry["organization_id"],
                        "name": organization_lookup.get(entry["organization_id"], ""),
                        "project_count": entry["project_count"],
                    }
                    for entry in org_counts
                    if entry["organization_id"] in organization_lookup
                ]
                org_stats.sort(key=lambda item: item["project_count"], reverse=True)
                org_stats = org_stats[:6]
        except Exception as exc:
            logging.getLogger(__name__).warning(
                "DashboardSummaryView: failed to build organization stats: %s", exc
            )
            warnings.add("organization stats")

        try:
            reference_date = timezone.now() - timedelta(days=365)
            recent_users = user_queryset.filter(date_joined__gte=reference_date)

            weekly = (
                recent_users.annotate(period=TruncWeek("date_joined"))
                .values("period")
                .annotate(count=Count("id"))
                .order_by("period")
            )
            weekly_list = [
                {
                    "period": entry["period"].isoformat(),
                    "count": entry["count"],
                }
                for entry in weekly
                if entry["period"] is not None
            ]
            if len(weekly_list) > 12:
                weekly_list = weekly_list[-12:]
            chart["week"] = weekly_list

            monthly = (
                recent_users.annotate(period=TruncMonth("date_joined"))
                .values("period")
                .annotate(count=Count("id"))
                .order_by("period")
            )
            monthly_list = [
                {
                    "period": entry["period"].isoformat(),
                    "count": entry["count"],
                }
                for entry in monthly
                if entry["period"] is not None
            ]
            if len(monthly_list) > 12:
                monthly_list = monthly_list[-12:]
            chart["month"] = monthly_list
        except Exception as exc:
            logging.getLogger(__name__).warning(
                "DashboardSummaryView: failed to build user trend chart: %s", exc
            )
            warnings.add("registration chart")

        payload = {
            "totals": totals,
            "org_stats": org_stats,
            "chart": chart,
            "errors": sorted(warnings),
            "last_updated": timezone.now().isoformat(),
        }
        return Response(payload)

class OrganizationViewSet(viewsets.ModelViewSet):
    queryset = Organization.objects.all()
    serializer_class = OrganizationSerializer
    permission_classes = [IsAuthenticated]  # This should be sufficient

    def get_permissions(self):
        """Allow authenticated users with admin role full access"""
        user = self.request.user

        # Debug logging
        print(f"🔍 OrganizationViewSet permission check - user: {user.username}")
        print(f"🔍 User role: {getattr(user, 'role', 'NO_ROLE')}")
        print(f"🔍 Is superuser: {user.is_superuser}")
        print(f"🔍 Action: {self.action}")

        # Always allow superusers or role 1 users
        if user.is_superuser or getattr(user, 'role', 4) == 1:
            return [IsAuthenticated()]

        # For other users, use default permissions
        return [IsAuthenticated()]

    def dispatch(self, request, *args, **kwargs):
        """Add debugging to see what's happening"""
        print(f"🔍 OrganizationViewSet.dispatch - Method: {request.method}")
        print(f"🔍 User: {request.user.username}, Role: {getattr(request.user, 'role', 'NO_ROLE')}")
        print(f"🔍 Is superuser: {request.user.is_superuser}")
        print(f"🔍 Is authenticated: {request.user.is_authenticated}")
        return super().dispatch(request, *args, **kwargs)

    def perform_create(self, serializer):
        # Save the organization first
        organization = serializer.save(created_by=self.request.user, updated_by=self.request.user)

        # Automatically assign this organization to the user's profile
        try:
            user_profile = self.request.user.profile
            user_profile.organizations.add(organization)
            print(f"✅ Auto-assigned organization '{organization.name}' to user '{self.request.user.username}' profile")
        except Exception as e:
            print(f"❌ Error auto-assigning organization to user profile: {e}")
            # Don't fail the organization creation if profile assignment fails
            pass

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        """Move organization to trash instead of permanent deletion"""
        organization = self.get_object()

        try:
            # Move to trash
            move_to_trash(organization, 'organization', request.user)
            return Response({'message': 'Organization moved to trash successfully'}, status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            return Response({'error': f'Failed to delete organization: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['PUT'])
def update_translations(request, form_id):
    try:
        form = Form.objects.get(id=form_id)
    except Form.DoesNotExist:
        return Response({'error': 'Form not found'}, status=status.HTTP_404_NOT_FOUND)

    translations = request.data.get('translations', {})
    hints = request.data.get('hints', {})
    guidance_hints = request.data.get('guidance_hints', {})
    language_subtag = (request.data.get('language_subtag', '') or '').strip().lower()

    # Validate input
    if not isinstance(translations, dict):
        return Response({'error': 'Invalid translations payload'}, status=status.HTTP_400_BAD_REQUEST)

    if not isinstance(hints, dict):
        return Response({'error': 'Invalid hints payload'}, status=status.HTTP_400_BAD_REQUEST)

    if not isinstance(guidance_hints, dict):
        return Response({'error': 'Invalid guidance_hints payload'}, status=status.HTTP_400_BAD_REQUEST)

    if not translations and not hints and not guidance_hints:
        return Response({'error': 'No translations, hints, or guidance_hints provided'}, status=status.HTTP_400_BAD_REQUEST)

    if not language_subtag:
        return Response({'error': 'Language subtag is required'}, status=status.HTTP_400_BAD_REQUEST)

    # Ensure the translations field is initialized
    if not form.translations:
        form.translations = {}

    # Ensure form.questions is a list
    if not isinstance(form.questions, list):
        form.questions = []

    # Get the actual default language description and subtag
    default_language = form.default_language
    other_languages = form.other_languages.all()

    if default_language:
        default_language_description = f"{default_language.description} ({default_language.subtag})"
        default_language_subtag = (default_language.subtag or '').strip().lower()
    else:
        default_language_description = 'English (en)'
        default_language_subtag = 'en'  # Default fallback

    def _get_payload_value(node, payload_map, field_name):
        """
        Find translated text from payload for the requested field.
        Supports both current base text and existing default-language mapped text.
        """
        if not isinstance(node, dict):
            return None

        candidates = []
        base_text = node.get(field_name)
        if base_text is not None and str(base_text) != '':
            candidates.append(str(base_text))

        if field_name == 'label':
            map_name = 'translations'
        elif field_name == 'hint':
            map_name = 'hints'
        elif field_name == 'guidance_hint':
            map_name = 'guidance_hints'
        else:
            map_name = None

        existing_default_text = _get_language_value(
            node.get(map_name, {}) if map_name else {},
            language=None,
            default_description=default_language_description,
            default_subtag=default_language_subtag,
        )
        if existing_default_text:
            candidates.append(str(existing_default_text))

        for candidate in candidates:
            if candidate in payload_map:
                value = payload_map.get(candidate)
                if value is not None:
                    return str(value)
        return None

    # Update the translations in the form's questions
    if language_subtag == default_language_subtag:
        for question in form.questions:
            try:
                updated_question_label = _get_payload_value(question, translations, 'label')
                if updated_question_label is not None:
                    if not isinstance(question.get('translations'), dict):
                        question['translations'] = {}
                    question['translations'][default_language_description] = updated_question_label
                    # Keep base label synchronized with the active default language
                    question['label'] = updated_question_label

                updated_question_hint = _get_payload_value(question, hints, 'hint')
                if updated_question_hint is not None:
                    if not isinstance(question.get('hints'), dict):
                        question['hints'] = {}
                    question['hints'][default_language_description] = updated_question_hint
                    # Keep base hint synchronized with the active default language
                    question['hint'] = updated_question_hint

                updated_question_guidance_hint = _get_payload_value(
                    question,
                    guidance_hints,
                    'guidance_hint',
                )
                if updated_question_guidance_hint is not None:
                    if not isinstance(question.get('guidance_hints'), dict):
                        question['guidance_hints'] = {}
                    question['guidance_hints'][default_language_description] = updated_question_guidance_hint
                    question['guidance_hint'] = updated_question_guidance_hint

                for option in question.get('options', []):
                    updated_option_label = _get_payload_value(option, translations, 'label')
                    if updated_option_label is not None:
                        if not isinstance(option.get('translations'), dict):
                            option['translations'] = {}
                        option['translations'][default_language_description] = updated_option_label
                        option['label'] = updated_option_label

                    updated_option_hint = _get_payload_value(option, hints, 'hint')
                    if updated_option_hint is not None:
                        if not isinstance(option.get('hints'), dict):
                            option['hints'] = {}
                        option['hints'][default_language_description] = updated_option_hint
                        option['hint'] = updated_option_hint

                for sub_question in question.get('subQuestions', []):
                    updated_sub_label = _get_payload_value(sub_question, translations, 'label')
                    if updated_sub_label is not None:
                        if not isinstance(sub_question.get('translations'), dict):
                            sub_question['translations'] = {}
                        sub_question['translations'][default_language_description] = updated_sub_label
                        sub_question['label'] = updated_sub_label

                    updated_sub_hint = _get_payload_value(sub_question, hints, 'hint')
                    if updated_sub_hint is not None:
                        if not isinstance(sub_question.get('hints'), dict):
                            sub_question['hints'] = {}
                        sub_question['hints'][default_language_description] = updated_sub_hint
                        sub_question['hint'] = updated_sub_hint

                    updated_sub_guidance_hint = _get_payload_value(
                        sub_question,
                        guidance_hints,
                        'guidance_hint',
                    )
                    if updated_sub_guidance_hint is not None:
                        if not isinstance(sub_question.get('guidance_hints'), dict):
                            sub_question['guidance_hints'] = {}
                        sub_question['guidance_hints'][default_language_description] = updated_sub_guidance_hint
                        sub_question['guidance_hint'] = updated_sub_guidance_hint
            except (KeyError, TypeError) as e:
                print(f"Error processing default language translations: {e}")
                continue
    else:
        # Add translations for other languages in XLSForm format
        language_description = None
        for language in other_languages:
            if (language.subtag or '').strip().lower() == language_subtag:
                language_description = _language_description(language)
                break

        if not language_description:
            return Response({'error': 'Language is not configured for this form'}, status=status.HTTP_400_BAD_REQUEST)

        for question in form.questions:
            try:
                updated_question_label = _get_payload_value(question, translations, 'label')
                if updated_question_label is not None:
                    if not isinstance(question.get('translations'), dict):
                        question['translations'] = {}
                    question['translations'][language_description] = updated_question_label

                updated_question_hint = _get_payload_value(question, hints, 'hint')
                if updated_question_hint is not None:
                    if not isinstance(question.get('hints'), dict):
                        question['hints'] = {}
                    question['hints'][language_description] = updated_question_hint

                updated_question_guidance_hint = _get_payload_value(
                    question,
                    guidance_hints,
                    'guidance_hint',
                )
                if updated_question_guidance_hint is not None:
                    if not isinstance(question.get('guidance_hints'), dict):
                        question['guidance_hints'] = {}
                    question['guidance_hints'][language_description] = updated_question_guidance_hint

                for option in question.get('options', []):
                    updated_option_label = _get_payload_value(option, translations, 'label')
                    if updated_option_label is not None:
                        if not isinstance(option.get('translations'), dict):
                            option['translations'] = {}
                        option['translations'][language_description] = updated_option_label

                    updated_option_hint = _get_payload_value(option, hints, 'hint')
                    if updated_option_hint is not None:
                        if not isinstance(option.get('hints'), dict):
                            option['hints'] = {}
                        option['hints'][language_description] = updated_option_hint

                for sub_question in question.get('subQuestions', []):
                    updated_sub_label = _get_payload_value(sub_question, translations, 'label')
                    if updated_sub_label is not None:
                        if not isinstance(sub_question.get('translations'), dict):
                            sub_question['translations'] = {}
                        sub_question['translations'][language_description] = updated_sub_label

                    updated_sub_hint = _get_payload_value(sub_question, hints, 'hint')
                    if updated_sub_hint is not None:
                        if not isinstance(sub_question.get('hints'), dict):
                            sub_question['hints'] = {}
                        sub_question['hints'][language_description] = updated_sub_hint

                    updated_sub_guidance_hint = _get_payload_value(
                        sub_question,
                        guidance_hints,
                        'guidance_hint',
                    )
                    if updated_sub_guidance_hint is not None:
                        if not isinstance(sub_question.get('guidance_hints'), dict):
                            sub_question['guidance_hints'] = {}
                        sub_question['guidance_hints'][language_description] = updated_sub_guidance_hint
            except (KeyError, TypeError) as e:
                print(f"Error processing translations for {language_description}: {e}")
                continue

    # Save the translations in the form
    try:
        form.save()
    except Exception as e:
        return Response({'error': f'Failed to save form: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Update the associated XLSX file using shared function with consistent choice_filter logic
    try:
        generate_xlsx_file(form, form.questions, default_language, other_languages, form.name)
    except Exception as e:
        # Don't fail the entire request if XLSX generation fails
        print(f"Warning: Failed to generate XLSX file: {e}")

    return Response({'message': 'Translations updated successfully'}, status=status.HTTP_200_OK)

@api_view(['POST', 'HEAD'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser])
def convert_xlsform(request):
    # Save uploaded file to temp location
    xlsfile = request.FILES.get('file')
    if not xlsfile:
        return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp:
        for chunk in xlsfile.chunks():
            temp.write(chunk)
        temp_path = temp.name

    try:
        # Prepare output temp file for XForm XML
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xml') as xml_temp:
            xml_output_path = xml_temp.name

        # Convert XLSX to XForm XML using pyxform
        xls2xform_convert(
            xlsform_path=temp_path,
            xform_path=xml_output_path,
            validate=False  # Set to True if you want validation
        )

        # Read the generated XML
        with open(xml_output_path, 'r', encoding='utf-8') as f:
            xform_xml = f.read()

        # Clean up temp files
        os.remove(temp_path)
        os.remove(xml_output_path)

        return Response({'xform': xform_xml}, status=status.HTTP_200_OK)
    except Exception as e:
        # Clean up temp files in case of error
        if os.path.exists(temp_path):
            os.remove(temp_path)
        if 'xml_output_path' in locals() and os.path.exists(xml_output_path):
            os.remove(xml_output_path)
        return Response({'error': f'Conversion failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

from django.views.decorators.csrf import ensure_csrf_cookie

@ensure_csrf_cookie
@api_view(['GET'])
@permission_classes([AllowAny])
def openrosa_xform(request, form_id):
    import os
    from django.conf import settings
    xlsx_path = os.path.join(settings.MEDIA_ROOT, 'update', f'{form_id}.xlsx')
    if not os.path.exists(xlsx_path):
        return HttpResponse("Form not found", status=404)
    from pyxform.xls2xform import xls2xform_convert
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xml") as temp_xml:
        xls2xform_convert(xlsx_path, temp_xml.name, validate=False)
        temp_xml.seek(0)
        xml_content = temp_xml.read()
    return HttpResponse(xml_content, content_type="text/xml")

from .models import Form  # Make sure Form is imported

def xml_to_json(xml_file):
    import xml.etree.ElementTree as ET

    def _xml_element_to_data(element):
        children = list(element)
        if not children:
            return (element.text or '').strip()

        parsed = {}
        for child in children:
            child_tag = child.tag.split('}')[-1]
            child_value = _xml_element_to_data(child)

            if child_tag in parsed:
                if not isinstance(parsed[child_tag], list):
                    parsed[child_tag] = [parsed[child_tag]]
                parsed[child_tag].append(child_value)
            else:
                parsed[child_tag] = child_value

        return parsed

    xml_file.seek(0)
    tree = ET.parse(xml_file)
    root = tree.getroot()
    data = {}
    # Add attributes as _id, _version, _xmlns:jr, etc.
    for attr, value in root.attrib.items():
        if attr == "id":
            data["_id"] = value
        elif attr == "version":
            data["_version"] = value
        else:
            data[f"_{attr}"] = value
    # Add namespaces
    for k, v in root.nsmap.items() if hasattr(root, "nsmap") else []:
        data[f"_xmlns:{k}"] = v
    # Parse children recursively while preserving repeated tags as arrays.
    for child in root:
        tag = child.tag.split('}')[-1]
        value = _xml_element_to_data(child)
        if tag in data:
            if not isinstance(data[tag], list):
                data[tag] = [data[tag]]
            data[tag].append(value)
        else:
            data[tag] = value

    # --- BEGIN: Geopoint Extraction Enhancement ---
    # Try to extract geopoint fields from the form definition if possible
    try:
        from .models import Form
        form_id = data.get('_id') or data.get('id')
        if form_id and str(form_id).startswith('form_'):
            form_id_num = str(form_id).replace('form_', '')
            form = Form.objects.filter(id=form_id_num).first()
            if form and hasattr(form, 'questions'):
                for q in form.questions:
                    if q.get('type', '').startswith('geopoint'):
                        qname = q.get('name')
                        val = data.get(qname)
                        if val and isinstance(val, str):
                            parts = val.strip().split()
                            if len(parts) == 4:
                                data[f'{qname}_latitude'] = parts[0]
                                data[f'{qname}_longitude'] = parts[1]
                                data[f'{qname}_altitude'] = parts[2]
                                data[f'{qname}_accuracy'] = parts[3]
    except Exception as e:
        # Log but do not fail
        import sys
        print(f"[xml_to_json geopoint extraction error]: {e}", file=sys.stderr)
    # --- END: Geopoint Extraction Enhancement ---

    return {"data": data}


from rest_framework.decorators import api_view, permission_classes, parser_classes
from django.views.decorators.csrf import csrf_exempt

@api_view(['POST', 'HEAD'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser])
@csrf_exempt
def openrosa_submission(request):
    if request.method == 'HEAD':
        return HttpResponse(status=200)

    xml_file = request.FILES.get('xml_submission_file')
    if not xml_file:
        return HttpResponse("No XML submission file", status=400)

    try:
        json_data = xml_to_json(xml_file)
    except Exception as exc:
        print(f"Failed to parse XML submission: {exc}")
        return HttpResponse("Invalid XML submission", status=400)

    payload_data = json_data.get("data", {}) if isinstance(json_data, dict) else {}
    if not isinstance(payload_data, dict):
        payload_data = {}

    form_identifier = payload_data.get("_id") or payload_data.get("id")
    if not form_identifier or not str(form_identifier).startswith("form_"):
        return HttpResponse("Invalid form identifier", status=400)

    form_id_num = str(form_identifier).replace("form_", "")
    form = Form.objects.filter(id=form_id_num).first()
    if not form:
        return HttpResponse("Form not found", status=404)

    if not form.allow_anonymous_submissions:
        if not (hasattr(request, 'user') and request.user.is_authenticated):
            current_url = request.build_absolute_uri()
            login_base = getattr(settings, "FRONTEND_URL", "").rstrip("/")
            login_url = f"{login_base}/auth/enketo-login?return={quote(current_url)}"
            response = HttpResponse(
                f"Authentication required. Please authenticate here: {login_url}",
                status=401
            )
            response['WWW-Authenticate'] = f'Bearer realm="CommicPlan", auth_url="{login_url}"'
            return response
        user = request.user
    else:
        user = request.user if (hasattr(request, 'user') and request.user.is_authenticated) else None

    if not _check_submit_rate_limit(request, form.id, user):
        response = HttpResponse("Submission rate limit exceeded", status=429)
        response['Retry-After'] = str(getattr(settings, 'SUBMISSION_RATE_LIMIT_WINDOW_SECONDS', 60))
        return response

    instance_id, payload_data = _ensure_instance_id(payload_data)
    submission_meta = payload_data.get("meta", {})
    if not isinstance(submission_meta, dict):
        submission_meta = {}

    if user:
        submission_meta["submitted_by"] = user.username
        submission_meta["submission_type"] = "authenticated"
    else:
        submission_meta["submitted_by"] = "anonymous"
        submission_meta["submission_type"] = "anonymous"
    payload_data["meta"] = submission_meta
    force_inline = _should_force_inline_processing(form, payload_data)

    deprecated_id = _normalize_instance_id(submission_meta.get("deprecatedID"))
    if deprecated_id and deprecated_id != instance_id:
        Submission.objects.filter(form=form, instance_id=deprecated_id).update(is_deleted=True)
        if form.submission:
            original_count = len(form.submission)
            form.submission = [
                sub for sub in form.submission
                if sub.get('data', {}).get('meta', {}).get('instanceID') != deprecated_id
            ]
            if len(form.submission) != original_count:
                form.save(update_fields=['submission'])

    existing_submission = _find_submission_record(form, instance_id, include_deleted=True)
    if existing_submission:
        if existing_submission.processing_status == 'failed':
            existing_submission.processing_status = 'queued'
            existing_submission.processing_error = ''
            existing_submission.data = {'data': payload_data}
            should_update_user = bool(user and not existing_submission.user_id)
            if should_update_user:
                existing_submission.user = user
            update_fields = ['processing_status', 'processing_error', 'data']
            if should_update_user:
                update_fields.append('user')
            existing_submission.save(update_fields=update_fields)
            _enqueue_submission_processing(existing_submission.id, force_inline=force_inline)

        response = HttpResponse(status=201)
        response['X-Submission-Deduplicated'] = 'true'
        return response

    try:
        xml_file.seek(0)
    except Exception:
        pass

    try:
        submission = Submission.objects.create(
            form=form,
            instance_id=instance_id,
            data={'data': payload_data},
            xml_file=xml_file,
            user=user,
            source='openrosa',
            processing_status='queued',
        )
    except IntegrityError:
        response = HttpResponse(status=201)
        response['X-Submission-Deduplicated'] = 'true'
        return response

    media_form_id = f"form_{form_id_num}"
    for key, uploaded_file in request.FILES.items():
        if key == 'xml_submission_file':
            continue
        media_dir = os.path.join(settings.MEDIA_ROOT, 'submissions', media_form_id, str(submission.id))
        os.makedirs(media_dir, exist_ok=True)
        file_path = os.path.join(media_dir, uploaded_file.name)
        with open(file_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

    queued = _enqueue_submission_processing(submission.id, force_inline=force_inline)
    if not queued:
        Submission.objects.filter(id=submission.id).update(
            processing_status='failed',
            processing_error='Failed to enqueue submission processing task',
        )

    return HttpResponse(status=201)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_enketo_survey(request):
    """
    Proxy Enketo survey creation to avoid CORS and keep API key server-side.
    """
    form_id = request.data.get("form_id")
    server_url = request.data.get("server_url")

    if not form_id or not server_url:
        return Response(
            {"error": "form_id and server_url are required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not settings.ENKETO_URL:
        return Response(
            {"error": "ENKETO_URL is not configured"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
    if not settings.ENKETO_API_KEY:
        return Response(
            {"error": "ENKETO_API_KEY is not configured"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    enketo_url = settings.ENKETO_URL.rstrip("/") + "/api/v2/survey"
    payload = {"form_id": form_id, "server_url": server_url}

    auth_scheme = (getattr(settings, "ENKETO_AUTH_SCHEME", "") or "basic").lower()
    if auth_scheme == "basic":
        import base64
        auth_header = "Basic " + base64.b64encode(f"{settings.ENKETO_API_KEY}:".encode()).decode()
    elif auth_scheme == "bearer":
        auth_header = f"Bearer {settings.ENKETO_API_KEY}"
    elif auth_scheme == "token":
        auth_header = f"Token {settings.ENKETO_API_KEY}"
    else:
        import base64
        auth_header = "Basic " + base64.b64encode(f"{settings.ENKETO_API_KEY}:".encode()).decode()

    print(f"[Enketo] POST {enketo_url} | payload={payload} | auth_scheme={auth_scheme}")
    try:
        response = requests.post(
            enketo_url,
            json=payload,
            headers={"Authorization": auth_header},
            timeout=20,
        )
    except Exception as e:
        print(f"[Enketo] Network error reaching {enketo_url}: {e}")
        return Response(
            {"error": f"Failed to reach Enketo: {e}"},
            status=status.HTTP_502_BAD_GATEWAY,
        )

    print(f"[Enketo] Response status={response.status_code} body={response.text[:500]}")

    if response.status_code >= 400:
        return Response(
            {
                "error": "Enketo API error",
                "status_code": response.status_code,
                "details": response.text,
            },
            status=status.HTTP_502_BAD_GATEWAY,
        )

    try:
        return Response(response.json(), status=status.HTTP_200_OK)
    except ValueError:
        print(f"[Enketo] Non-JSON response from Enketo: {response.text[:300]}")
        return Response(
            {"error": "Invalid JSON response from Enketo"},
            status=status.HTTP_502_BAD_GATEWAY,
        )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def set_enketo_cookie(request):
    """Set authentication cookie for Enketo access"""
    try:
        user = request.user
        auth_token = generate_auth_token(user)

        print(f"Setting cookie for user: {user.username}")
        print(f"Auth token: {auth_token[:20]}...")

        response = JsonResponse({'message': 'Cookie set successfully'})

        secure_cookie = request.is_secure() or settings.FRONTEND_URL.startswith("https://")
        cookie_kwargs = {
            'httponly': True,
            'secure': secure_cookie,
            'samesite': 'None' if secure_cookie else 'Lax',
            'max_age': 7 * 24 * 60 * 60,
        }
        if settings.AUTH_COOKIE_DOMAIN:
            cookie_kwargs['domain'] = settings.AUTH_COOKIE_DOMAIN

        response.set_cookie('commicplan_auth', auth_token, **cookie_kwargs)

        # Set CORS headers
        response['Access-Control-Allow-Credentials'] = 'true'
        origin = request.META.get('HTTP_ORIGIN')
        allowed_origins = getattr(settings, "ENKETO_ALLOWED_ORIGINS", [])
        if origin and origin in allowed_origins:
            response['Access-Control-Allow-Origin'] = origin
        else:
            response['Access-Control-Allow-Origin'] = settings.FRONTEND_URL

        return response
    except Exception as e:
        print(f"Error setting cookie: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@ensure_csrf_cookie
@api_view(['GET', 'POST', 'HEAD'])
@permission_classes([AllowAny])
def openrosa_form_list(request):
    import hashlib

    # Fast-path for HEAD requests (no body required)
    if request.method == 'HEAD':
        return HttpResponse(status=200)

    # Handle both GET and POST requests
    form_id = request.GET.get('formID') or request.POST.get('formID')

    print(f"Received formID parameter: {form_id}")  # Debug logging

    if form_id:
        # If formID is provided, look for the specific form
        if form_id.startswith("form_"):
            # Extract numeric ID from "form_30" -> "30"
            numeric_id = form_id.replace("form_", "")
            try:
                forms = Form.objects.filter(id=int(numeric_id))
            except ValueError:
                forms = Form.objects.none()
        else:
            # Try to find by direct ID
            try:
                forms = Form.objects.filter(id=int(form_id))
            except ValueError:
                forms = Form.objects.none()
    else:
        # Return all forms (you might want to filter by user permissions here)
        forms = Form.objects.all()

    print(f"Found {forms.count()} forms")  # Debug logging

    xml = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml += '<xforms xmlns="http://openrosa.org/xforms/xformsList">\n'

    # Hashes are required by Enketo for cache checks. Compute hashes for single-form
    # requests (formID provided), or when explicitly requested via ?hashes=1.
    compute_hashes = request.GET.get('hashes') == '1' or bool(form_id)

    for form in forms:
        name = escape(form.name)
        form_id_string = f"form_{form.id}"  # Always return form_XX format
        download_url = request.build_absolute_uri(reverse('openrosa_xform', args=[form.id]))
        xlsx_path = os.path.join(settings.MEDIA_ROOT, 'update', f'{form.id}.xlsx')
        version = "20240601"
        hash_value = ""

        print(f"Processing form {form.id}: {name}")  # Debug logging

        if compute_hashes and os.path.exists(xlsx_path):
            try:
                with tempfile.NamedTemporaryFile(suffix=".xml") as temp_xml:
                    xls2xform_convert(xlsx_path, temp_xml.name, validate=False)
                    temp_xml.seek(0)
                    xml_content = temp_xml.read()
                    hash_value = hashlib.md5(xml_content).hexdigest()
                    print(f"Generated hash for form {form.id}: {hash_value}")  # Debug logging
            except Exception as e:
                print(f"Error processing form {form.id}: {e}")
                continue  # Skip this form if there's an error

        xml += f'  <xform>\n'
        xml += f'    <formID>{form_id_string}</formID>\n'  # Always use form_XX format
        xml += f'    <name>{name}</name>\n'
        xml += f'    <version>{version}</version>\n'
        if hash_value:
            xml += f'    <hash>md5:{hash_value}</hash>\n'
        xml += f'    <downloadUrl>{download_url}</downloadUrl>\n'
        xml += f'  </xform>\n'

    xml += '</xforms>\n'

    print(f"Generated XML response:\n{xml}")  # Debug logging

    return HttpResponse(xml, content_type="text/xml; charset=utf-8")

from django.http import FileResponse

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_submission_xml(request, submission_id):
    try:
        submission = Submission.objects.get(id=submission_id)
    except Submission.DoesNotExist:
        return HttpResponse("Submission not found", status=404)
    return FileResponse(submission.xml_file, as_attachment=True, filename=f'submission_{submission_id}.xml')

def save_submission_to_xlsx(xml_file):
    import openpyxl
    import os
    import xml.etree.ElementTree as ET

    # Parse XML
    xml_file.seek(0)
    tree = ET.parse(xml_file)
    root = tree.getroot()
    form_id = root.attrib.get('id')
    if not form_id:
        print("No form_id in XML")
        return

    # Extract data fields (excluding meta)
    data = {}
    for child in root:
        tag = child.tag.split('}')[-1]  # Remove namespace if present
        if tag == 'meta':
            continue
        data[tag] = child.text

    # Prepare XLSX file path in submissions/XLSX/
    xlsx_dir = os.path.join(settings.MEDIA_ROOT, 'submissions', 'XLSX')
    os.makedirs(xlsx_dir, exist_ok=True)
    xlsx_path = os.path.join(xlsx_dir, f'{form_id}.xlsx')

    # Print the path where XLSX will be saved
    print(f"XLSX will be saved to: {xlsx_path}")

    # Write or append to XLSX
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        # Add new columns if new fields are present
        for key in data.keys():
            if key not in headers:
                headers.append(key)
                ws.cell(row=1, column=len(headers)).value = key
        # Prepare row in header order
        row = [data.get(h, "") for h in headers]
        ws.append(row)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = list(data.keys())
        ws.append(headers)
        ws.append([data[h] for h in headers])

    wb.save(xlsx_path)
    print(f"Saved submission to {xlsx_path}")

def flatten_submission_data(data):
    """
    Flatten nested submission payloads, including repeat arrays.
    Examples:
      {meta: {ward: "x"}} -> {"meta.ward": "x"}
      {rep: [{q1: "a"}, {q1: "b"}]} -> {"rep[0].q1": "a", "rep[1].q1": "b"}
    """
    flattened = {}

    def walk(node, prefix=""):
        if isinstance(node, dict):
            for key, value in node.items():
                current_key = f"{prefix}.{key}" if prefix else key

                if isinstance(value, dict):
                    walk(value, current_key)
                elif isinstance(value, list):
                    if not value:
                        flattened[current_key] = []
                        continue

                    # Keep scalar arrays at the direct key (select_multiple list values, etc.)
                    if all(not isinstance(item, (dict, list)) for item in value):
                        flattened[current_key] = value

                    for idx, item in enumerate(value):
                        indexed_key = f"{current_key}[{idx}]"
                        if isinstance(item, (dict, list)):
                            walk(item, indexed_key)
                        else:
                            flattened[indexed_key] = item
                else:
                    flattened[current_key] = value
        elif isinstance(node, list):
            for idx, item in enumerate(node):
                indexed_key = f"{prefix}[{idx}]" if prefix else f"[{idx}]"
                if isinstance(item, (dict, list)):
                    walk(item, indexed_key)
                else:
                    flattened[indexed_key] = item

    if isinstance(data, dict):
        walk(data)

    return flattened


def index_submission_values_by_field(data):
    """
    Build a field -> values index across nested groups/repeats.
    Used as a fallback when flattened path matching cannot resolve a field.
    """
    field_index = {}

    def add_value(field_name, value):
        if field_name in (None, ""):
            return
        if value is None:
            return
        if isinstance(value, str) and not value.strip():
            return
        field_index.setdefault(field_name, []).append(value)

    def walk(node):
        if isinstance(node, dict):
            for key, value in node.items():
                if isinstance(value, dict):
                    walk(value)
                elif isinstance(value, list):
                    if not value:
                        continue

                    if all(not isinstance(item, (dict, list)) for item in value):
                        add_value(key, value)
                    else:
                        for item in value:
                            walk(item)
                else:
                    add_value(key, value)
        elif isinstance(node, list):
            for item in node:
                walk(item)

    if isinstance(data, dict):
        walk(data)

    return field_index


def _field_key_base(key):
    base = key.split(".")[-1] if "." in key else key
    return re.sub(r"\[\d+\]", "", base)


def resolve_submission_column_values(flattened_data, field_value_index, column_name):
    """
    Resolve raw values for a single logical question column.
    Supports direct fields, grouped paths, and repeated entries.
    """
    values = []

    def add_value(value):
        if value is None:
            return
        if isinstance(value, str) and not value.strip():
            return
        values.append(value)

    # 1) Exact direct key match.
    if flattened_data.get(column_name) is not None:
        add_value(flattened_data.get(column_name))

    # 2) Backward-compatible grouped paths.
    if not values:
        possible_paths = [
            f"_.{column_name}",
            f"_._1.{column_name}",
            f"_._2.{column_name}",
            f"_._3.{column_name}",
            f"address.{column_name}",
        ]
        for path in possible_paths:
            if flattened_data.get(path) is not None:
                add_value(flattened_data.get(path))

    # 3) Partial key matching (including repeat-indexed paths like repeat[0].field).
    matching_keys = [
        key for key in flattened_data.keys()
        if (
            _field_key_base(key) == column_name
            or key.endswith(f".{column_name}")
            or (
                column_name
                and "_" in column_name
                and _field_key_base(key)
                and "_" in _field_key_base(key)
                and _field_key_base(key).lower() == column_name.lower()
            )
        )
    ]
    for key in sorted(matching_keys):
        add_value(flattened_data.get(key))

    # 4) Deep field-name index fallback.
    if not values:
        for indexed_value in field_value_index.get(column_name, []):
            add_value(indexed_value)

    return values


def collapse_export_values(values):
    """Collapse a list of values into a de-duplicated display string."""
    if values is None:
        return ''

    if not isinstance(values, list):
        values = [values]

    def normalize(value):
        if value is None:
            return ''
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, list):
            parts = [normalize(item) for item in value]
            parts = [part for part in parts if part not in ('', None)]
            return ', '.join(parts)
        if isinstance(value, dict):
            try:
                return json.dumps(value, ensure_ascii=False, sort_keys=True)
            except Exception:
                return str(value)
        return str(value)

    normalized_values = []
    seen = set()
    for raw_value in values:
        normalized = normalize(raw_value)
        if not normalized:
            continue
        if normalized in seen:
            continue
        seen.add(normalized)
        normalized_values.append(normalized)

    if not normalized_values:
        return ''
    if len(normalized_values) == 1:
        return normalized_values[0]
    return ' | '.join(normalized_values)

def get_option_label(question, option_name, default_language='english'):
    """
    Convert option name to option label for select_one/select_multiple questions
    """
    # Handle both 'choices' (old format) and 'options' (new format)
    options_list = question.get('options') or question.get('choices', [])

    if not options_list:
        return option_name

    # Find the option with matching name
    for option in options_list:
        if option.get('name') == option_name:
            # Return the label directly
            return option.get('label', option_name)

    return option_name

def process_submission_value(question, value, default_language='english'):
    """
    Process submission value based on question type.
    For select_one/select_multiple, convert option names to labels.
    """
    if not question:
        return value

    question_type = question.get('type', '')

    if 'select_one' in question_type:
        if isinstance(value, list):
            option_labels = []
            seen = set()
            for item in value:
                if item in (None, ''):
                    continue
                label = get_option_label(question, item, default_language)
                if label not in seen:
                    seen.add(label)
                    option_labels.append(label)
            if not option_labels:
                return ''
            if len(option_labels) == 1:
                return option_labels[0]
            return ' | '.join(option_labels)
        return get_option_label(question, value, default_language)
    elif 'select_multiple' in question_type:
        if isinstance(value, str):
            # Handle space-separated values
            option_names = value.split(' ') if value else []
            option_labels = [get_option_label(question, name, default_language) for name in option_names if name.strip()]
            return ', '.join(option_labels)  # Use comma separation like frontend
        elif isinstance(value, list):
            if all(not isinstance(item, (dict, list)) for item in value):
                # A list here usually means one select_multiple answer stored as array.
                option_labels = [get_option_label(question, name, default_language) for name in value if name]
                return ', '.join(option_labels)

            # Fallback for unusual nested values.
            option_labels = []
            for item in value:
                if item in (None, ''):
                    continue
                if isinstance(item, str):
                    split_names = [name for name in item.split(' ') if name]
                    if split_names:
                        option_labels.append(
                            ', '.join(get_option_label(question, name, default_language) for name in split_names)
                        )
                    else:
                        option_labels.append(get_option_label(question, item, default_language))
                else:
                    option_labels.append(get_option_label(question, item, default_language))
            return ' | '.join(option_labels)

    return value

@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser])
def save_xlsx_submission(request):
    xml_file = request.FILES.get('xml_submission_file')
    if not xml_file:
        return HttpResponse("No XML submission file", status=400)
    save_submission_to_xlsx(xml_file)
    return HttpResponse("Saved to XLSX", status=201)

@api_view(['GET'])
@permission_classes([AllowAny])
def download_form_csv(request, form_id):
    import openpyxl
    import os
    import csv
    from django.http import HttpResponse
    from django.utils import timezone

    try:
        form = Form.objects.get(id=form_id)
    except Form.DoesNotExist:
        return HttpResponse("Form not found", status=404)

    # Get question-based columns (matching frontend approach)
    all_columns = []

    # Add standard fields
    all_columns.extend(['start', 'end'])

    # Get all actual questions (non-group questions) from the form definition
    form_questions = [q for q in form.questions if
                     q.get('type') not in ['begin_group', 'end_group', 'begin_repeat', 'end_repeat'] and
                     q.get('name') and q.get('name').strip()]

    # Add all question field names
    for question in form_questions:
        all_columns.append(question['name'])

    # Add meta fields
    all_columns.extend(['meta.submitted_by', 'meta.submission_type', 'uuid'])

    # Create headers using question labels (matching frontend)
    headers = []
    for col in all_columns:
        if col == 'start':
            headers.append('Start Time')
        elif col == 'end':
            headers.append('End Time')
        elif col == 'meta.submitted_by':
            headers.append('Submitted By')
        elif col == 'meta.submission_type':
            headers.append('Submission Type')
        elif col == 'uuid':
            headers.append('UUID')
        else:
            # Find question label
            question = next((q for q in form_questions if q['name'] == col), None)
            if question:
                # Handle special designation cases (matching frontend)
                if col == 'designation_1':
                    headers.append('Designation (BRAC)')
                elif col == 'designation':
                    headers.append('Designation (Govt)')
                else:
                    headers.append(question.get('label', col))
            else:
                headers.append(col)

    # Save CSV to media/downloads/ with timestamp
    download_dir = os.path.join(settings.MEDIA_ROOT, 'downloads')
    os.makedirs(download_dir, exist_ok=True)
    timestamp = timezone.now().strftime("%Y%m%d%H%M%S")
    file_name = f'form_{form_id}_{timestamp}.csv'
    file_path = os.path.join(download_dir, file_name)

    # Create question lookup dictionary
    question_lookup = {q['name']: q for q in form_questions}

    with open(file_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for sub in _get_form_submission_entries(form):
            data = sub.get('data', {})
            flattened_data = flatten_submission_data(data)
            field_value_index = index_submission_values_by_field(data)

            row = []
            for col in all_columns:
                value = None

                # Handle special fields
                if col == 'start':
                    value = data.get('start')
                elif col == 'end':
                    value = data.get('end')
                elif col == 'meta.submitted_by':
                    value = data.get('meta', {}).get('submitted_by')
                elif col == 'meta.submission_type':
                    value = data.get('meta', {}).get('submission_type')
                elif col == 'uuid':
                    value = data.get('meta', {}).get('instanceID')
                else:
                    raw_values = resolve_submission_column_values(
                        flattened_data,
                        field_value_index,
                        col,
                    )

                    if col in question_lookup and raw_values:
                        question = question_lookup[col]
                        raw_values = [
                            process_submission_value(question, raw_value)
                            for raw_value in raw_values
                        ]

                    value = collapse_export_values(raw_values)

                if value in (None, ''):
                    row.append('')
                else:
                    row.append(value)

            writer.writerow(row)

    # Log download
    DownloadLog.objects.create(
        user=request.user,
        form_id=form_id,
        type='CSV',
        file_path=file_path,
    )

    # Stream the file as response
    response = HttpResponse(open(file_path, 'rb'), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={file_name}'
    return response

import os
from django.conf import settings
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from .models import Form, Submission
from .serializers import SubmissionSerializer

from django.views.decorators.csrf import csrf_exempt

from django.views.decorators.csrf import csrf_exempt



from rest_framework.authentication import TokenAuthentication
from rest_framework.exceptions import AuthenticationFailed
from django.contrib.auth import authenticate


@api_view(['POST'])
@csrf_exempt
def submit_form(request, form_id):
    from django.utils import timezone

    user = None
    auth_header = request.META.get('HTTP_AUTHORIZATION', '')

    if auth_header.startswith('Token ') or auth_header.startswith('Bearer '):
        token_key = auth_header.split(' ')[1]
        from rest_framework.authtoken.models import Token
        try:
            token_obj = Token.objects.get(key=token_key)
            user = token_obj.user
        except Token.DoesNotExist:
            raise AuthenticationFailed('Invalid or expired token')
    elif auth_header.startswith('Basic '):
        import base64
        try:
            encoded_credentials = auth_header.split(' ')[1]
            decoded_credentials = base64.b64decode(encoded_credentials).decode('utf-8')
            username, password = decoded_credentials.split(':', 1)
            user = authenticate(username=username, password=password)
            if not user:
                raise AuthenticationFailed('Invalid username or password')
        except Exception:
            raise AuthenticationFailed('Invalid Basic Auth header')
    else:
        return Response({'error': 'Authentication required (Token or Basic Auth)'}, status=401)

    try:
        form = Form.objects.get(id=form_id)
    except Form.DoesNotExist:
        return Response({'error': 'Form not found'}, status=status.HTTP_404_NOT_FOUND)

    if not _check_submit_rate_limit(request, form.id, user):
        return Response(
            {
                'error': 'Submission rate limit exceeded',
                'retry_after_seconds': int(getattr(settings, 'SUBMISSION_RATE_LIMIT_WINDOW_SECONDS', 60)),
            },
            status=429,
        )

    data = request.data.copy()
    data['_id'] = f'form_{form.id}'
    data['_version'] = '20240601'
    data['start'] = data.get('start', timezone.now().isoformat())
    data['end'] = data.get('end', timezone.now().isoformat())

    instance_id, data = _ensure_instance_id(data)
    meta = data.get('meta', {})
    if not isinstance(meta, dict):
        meta = {}

    meta['submitted_by'] = user.username
    meta['submission_type'] = 'authenticated'
    data['meta'] = meta
    force_inline = _should_force_inline_processing(form, data)

    deprecated_id = _normalize_instance_id(meta.get('deprecatedID'))
    if deprecated_id and deprecated_id != instance_id:
        Submission.objects.filter(form=form, instance_id=deprecated_id).update(is_deleted=True)
        if form.submission:
            original_count = len(form.submission)
            form.submission = [
                sub for sub in form.submission
                if sub.get('data', {}).get('meta', {}).get('instanceID') != deprecated_id
            ]
            if len(form.submission) != original_count:
                form.save(update_fields=['submission'])

    existing_submission = _find_submission_record(form, instance_id, include_deleted=True)
    if existing_submission:
        if existing_submission.processing_status == 'failed':
            existing_submission.processing_status = 'queued'
            existing_submission.processing_error = ''
            existing_submission.data = {'data': data}
            existing_submission.user = user
            existing_submission.source = 'api'
            existing_submission.save(update_fields=['processing_status', 'processing_error', 'data', 'user', 'source'])
            _enqueue_submission_processing(existing_submission.id, force_inline=force_inline)
        current_status = Submission.objects.filter(id=existing_submission.id).values_list(
            'processing_status',
            flat=True,
        ).first() or existing_submission.processing_status
        return Response(
            {
                'message': 'Duplicate submission ignored',
                'instance_id': instance_id,
                'submission_id': existing_submission.id,
                'processing_status': current_status,
                'status_endpoint': f'/api/submissions/{instance_id}/status/',
            },
            status=status.HTTP_200_OK,
        )

    try:
        submission = Submission.objects.create(
            form=form,
            instance_id=instance_id,
            data={'data': data},
            user=user,
            source='api',
            processing_status='queued',
        )
    except IntegrityError:
        return Response({'message': 'Duplicate submission ignored', 'instance_id': instance_id}, status=status.HTTP_200_OK)

    queued = _enqueue_submission_processing(submission.id, force_inline=force_inline)
    if not queued:
        Submission.objects.filter(id=submission.id).update(
            processing_status='failed',
            processing_error='Failed to enqueue submission processing task',
        )
    current_status = Submission.objects.filter(id=submission.id).values_list(
        'processing_status',
        flat=True,
    ).first() or ('queued' if queued else 'failed')

    return Response(
        {
            'message': 'Submission accepted',
            'instance_id': instance_id,
            'submission_id': submission.id,
            'processing_status': current_status,
            'status_endpoint': f'/api/submissions/{instance_id}/status/',
        },
        status=status.HTTP_201_CREATED,
    )

@api_view(['GET'])
@permission_classes([AllowAny])
def list_form_media(request, form_id):
    """
    Returns a list of all media files for a given form_id.
    """
    media_root = settings.MEDIA_ROOT
    media_url = settings.MEDIA_URL
    form_dir = os.path.join(media_root, 'submissions', f'form_{form_id}')
    media_files = []

    if os.path.exists(form_dir):
        for submission_id in os.listdir(form_dir):
            submission_dir = os.path.join(form_dir, submission_id)
            if os.path.isdir(submission_dir):
                for fname in os.listdir(submission_dir):
                    file_url = f"{media_url}submissions/form_{form_id}/{submission_id}/{fname}"
                    media_files.append({
                        "url": file_url,
                        "filename": fname,
                        "submission_id": submission_id,
                    })

    return JsonResponse(media_files, safe=False)

@api_view(['GET'])
@permission_classes([AllowAny])
def download_form_zip(request, form_id):
    import openpyxl
    import zipfile
    import io
    import csv
    from django.utils import timezone

    try:
        form = Form.objects.get(id=form_id)
    except Form.DoesNotExist:
        return HttpResponse("Form not found", status=404)

    media_root = settings.MEDIA_ROOT
    download_dir = os.path.join(media_root, 'downloads')
    os.makedirs(download_dir, exist_ok=True)
    timestamp = timezone.now().strftime("%Y%m%d%H%M%S")

    # Prepare headers: start, end, all question names, geopoint columns, Validation, uuid
    question_names = [q['name'] for q in form.questions]
    geopoint_questions = [q['name'] for q in form.questions if q.get('type', '').startswith('geopoint')]
    geopoint_headers = []
    for qn in geopoint_questions:
        geopoint_headers.extend([
            f"{qn}_latitude",
            f"{qn}_longitude",
            f"{qn}_altitude",
            f"{qn}_accuracy"
        ])
    headers = ['start', 'end'] + question_names + geopoint_headers + ['Validation', 'uuid']

    # --- Generate XLSX with Validation column ---
    xlsx_path = os.path.join(download_dir, f'form_{form_id}_{timestamp}.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for sub in _get_form_submission_entries(form):
        data = sub.get('data', {})
        row = [
            data.get('start', ''),
            data.get('end', ''),
        ]
        for qn in question_names:
            row.append(data.get(qn, ''))
        # Add geopoint columns if present
        for qn in geopoint_questions:
            row.append(data.get(f'{qn}_latitude', ''))
            row.append(data.get(f'{qn}_longitude', ''))
            row.append(data.get(f'{qn}_altitude', ''))
            row.append(data.get(f'{qn}_accuracy', ''))
        row.append(data.get('Validation', ''))
        uuid = data.get('meta', {}).get('instanceID', '')
        row.append(uuid)
        ws.append(row)
    wb.save(xlsx_path)

    # --- Generate CSV from the same data ---
    csv_path = os.path.join(download_dir, f'form_{form_id}_{timestamp}.csv')
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for sub in _get_form_submission_entries(form):
            data = sub.get('data', {})
            row = [
                data.get('start', ''),
                data.get('end', ''),
            ]
            for qn in question_names:
                row.append(data.get(qn, ''))
            # Add geopoint columns if present
            for qn in geopoint_questions:
                row.append(data.get(f'{qn}_latitude', ''))
                row.append(data.get(f'{qn}_longitude', ''))
                row.append(data.get(f'{qn}_altitude', ''))
                row.append(data.get(f'{qn}_accuracy', ''))
            row.append(data.get('Validation', ''))
            uuid = data.get('meta', {}).get('instanceID', '')
            row.append(uuid)
            writer.writerow(row)

    # --- Prepare ZIP in memory ---
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        # Add XLSX
        if os.path.exists(xlsx_path):
            zip_file.write(xlsx_path, arcname=os.path.basename(xlsx_path))
        # Add CSV
        if os.path.exists(csv_path):
            zip_file.write(csv_path, arcname=os.path.basename(csv_path))
        # Add media files
        form_media_dir = os.path.join(media_root, 'submissions', f'form_{form_id}')
        if os.path.exists(form_media_dir):
            for submission_id in os.listdir(form_media_dir):
                submission_dir = os.path.join(form_media_dir, submission_id)
                if os.path.isdir(submission_dir):
                    for fname in os.listdir(submission_dir):
                        fpath = os.path.join(submission_dir, fname)
                        arcname = f"media/{submission_id}/{fname}"
                        zip_file.write(fpath, arcname=arcname)

    zip_buffer.seek(0)
    zip_filename = f'form_{form_id}_{timestamp}_all_with_media.zip'
    zip_path = os.path.join(download_dir, zip_filename)
    with open(zip_path, 'wb') as f:
        f.write(zip_buffer.getvalue())

    # Log download
    DownloadLog.objects.create(
        user=request.user,
        form_id=form_id,
        type='ZIP',
        file_path=zip_path,
    )

    # Serve the ZIP file
    response = HttpResponse(open(zip_path, 'rb'), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename={zip_filename}'
    return response

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_submission_validation(request, instance_id):
    validation = request.data.get('validation')
    if validation is None:
        return Response({'error': 'Validation value required'}, status=400)

    normalized_instance_id = _normalize_instance_id(instance_id)

    submission_record = Submission.objects.filter(
        instance_id=normalized_instance_id,
        is_deleted=False,
    ).order_by('-created_at').first()
    if not submission_record and normalized_instance_id.startswith('uuid:'):
        submission_record = Submission.objects.filter(
            instance_id=normalized_instance_id.replace('uuid:', '', 1),
            is_deleted=False,
        ).order_by('-created_at').first()

    if submission_record:
        payload = _extract_submission_data(submission_record.data or {})
        payload['Validation'] = validation
        submission_record.data = {'data': payload}
        submission_record.save(update_fields=['data'])
        return Response({'success': True})

    # Legacy fallback (JSON blob on Form)
    found = False
    for form in Form.objects.filter(submission__isnull=False):
        updated_submissions = []
        changed = False
        for submission in form.submission or []:
            meta_instance = submission.get('data', {}).get('meta', {}).get('instanceID')
            if _normalize_instance_id(meta_instance) == normalized_instance_id:
                submission.setdefault('data', {})['Validation'] = validation
                changed = True
                found = True
            updated_submissions.append(submission)
        if changed:
            form.submission = updated_submissions
            form.save(update_fields=['submission'])

    if not found:
        return Response({'error': 'Submission not found'}, status=404)

    return Response({'success': True})

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_submission(request, form_id, instance_id):
    """
    Soft-delete a specific submission from a form by instanceID.
    Marks row-based Submission records with is_deleted=True.
    """
    try:
        try:
            form = Form.objects.get(id=form_id)
        except Form.DoesNotExist:
            return Response({'error': 'Form not found'}, status=404)

        normalized_instance_id = _normalize_instance_id(instance_id)
        if not normalized_instance_id:
            return Response({'error': 'Invalid instance ID'}, status=400)

        candidate_instance_ids = {normalized_instance_id}
        if normalized_instance_id.startswith('uuid:'):
            candidate_instance_ids.add(normalized_instance_id.replace('uuid:', '', 1))
        else:
            candidate_instance_ids.add(f'uuid:{normalized_instance_id}')

        # Soft delete row-based submissions and add to trash.
        submission_rows = Submission.objects.filter(
            form=form,
            instance_id__in=candidate_instance_ids,
            is_deleted=False,
        ).select_related('form')

        deleted_count = 0
        with transaction.atomic():
            submission_list = list(submission_rows)
            for submission_row in submission_list:
                trash_entry = move_submission_to_soft_trash(submission_row, request.user)
                if trash_entry:
                    deleted_count += 1

            # Legacy fallback: create a tombstone marker when data exists only in Form.submission.
            if deleted_count == 0 and isinstance(form.submission, list):
                matched_legacy_instance_id = ''
                for legacy_entry in form.submission:
                    legacy_payload = _extract_submission_data(legacy_entry)
                    legacy_instance_id = _extract_instance_id(legacy_payload)
                    if legacy_instance_id and legacy_instance_id in candidate_instance_ids:
                        matched_legacy_instance_id = legacy_instance_id
                        break

                if matched_legacy_instance_id:
                    tombstone = _create_submission_tombstone(
                        form,
                        matched_legacy_instance_id,
                        request.user,
                    )
                    trash_entry = move_submission_to_soft_trash(tombstone, request.user)
                    if trash_entry:
                        deleted_count += 1

        if deleted_count == 0:
            return Response({'error': 'Submission not found'}, status=404)

        remaining_rows = Submission.objects.filter(form=form, is_deleted=False).count()
        remaining_submissions = remaining_rows if remaining_rows else len(form.submission or [])

        return Response({
            'success': True,
            'message': 'Submission soft deleted successfully',
            'remaining_submissions': remaining_submissions,
            'deleted_count': deleted_count,
        }, status=200)

    except Exception as e:
        print(f"Error deleting submission: {e}")
        return Response({'error': 'Failed to delete submission'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_soft_delete_project_submissions(request, project_id):
    """
    Soft-delete multiple submissions inside a project by instance IDs.
    """
    try:
        from .models import Project

        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return Response({'error': 'Project not found'}, status=404)

        if getattr(request.user, 'role', 4) == 6:
            has_access = (
                hasattr(request.user, 'profile')
                and request.user.profile.projects.filter(id=project_id).exists()
            )
            if not has_access:
                return Response({'error': 'You do not have access to this project'}, status=403)

        raw_instance_ids = request.data.get('instance_ids', [])
        if not isinstance(raw_instance_ids, list) or not raw_instance_ids:
            return Response({'error': 'instance_ids (non-empty list) is required'}, status=400)

        candidate_instance_ids = set()
        for raw_value in raw_instance_ids:
            normalized = _normalize_instance_id(raw_value)
            if not normalized:
                continue
            candidate_instance_ids.add(normalized)
            if normalized.startswith('uuid:'):
                candidate_instance_ids.add(normalized.replace('uuid:', '', 1))
            else:
                candidate_instance_ids.add(f'uuid:{normalized}')

        if not candidate_instance_ids:
            return Response({'error': 'No valid instance IDs provided'}, status=400)

        project_form_ids = list(Form.objects.filter(project=project).values_list('id', flat=True))
        if not project_form_ids:
            return Response({
                'success': True,
                'deleted_count': 0,
                'requested_count': len(raw_instance_ids),
                'remaining_submissions': 0,
            }, status=200)

        queryset = Submission.objects.filter(
            form_id__in=project_form_ids,
            instance_id__in=candidate_instance_ids,
            is_deleted=False,
        ).select_related('form')

        deleted_count = 0
        normalized_requested_ids = {
            _normalize_instance_id(value)
            for value in raw_instance_ids
            if _normalize_instance_id(value)
        }
        matched_requested_ids = set()

        with transaction.atomic():
            submissions_to_delete = list(queryset)
            for submission in submissions_to_delete:
                trash_entry = move_submission_to_soft_trash(submission, request.user)
                if trash_entry:
                    deleted_count += 1
                    normalized_instance = _normalize_instance_id(submission.instance_id)
                    if normalized_instance:
                        matched_requested_ids.add(normalized_instance)

            # Legacy fallback: scan Form.submission and create tombstone markers
            # for requested instance IDs that do not yet exist in Submission rows.
            unresolved_ids = normalized_requested_ids - matched_requested_ids
            if unresolved_ids:
                for project_form in Form.objects.filter(id__in=project_form_ids).only('id', 'submission').iterator():
                    legacy_entries = project_form.submission if isinstance(project_form.submission, list) else []
                    if not legacy_entries:
                        continue

                    for legacy_entry in legacy_entries:
                        legacy_payload = _extract_submission_data(legacy_entry)
                        legacy_instance_id = _extract_instance_id(legacy_payload)
                        if not legacy_instance_id or legacy_instance_id not in unresolved_ids:
                            continue

                        tombstone = _create_submission_tombstone(
                            project_form,
                            legacy_instance_id,
                            request.user,
                        )
                        trash_entry = move_submission_to_soft_trash(tombstone, request.user)
                        if trash_entry:
                            deleted_count += 1

                        unresolved_ids.discard(legacy_instance_id)
                        if not unresolved_ids:
                            break

                    if not unresolved_ids:
                        break

        remaining_submissions = Submission.objects.filter(
            form_id__in=project_form_ids,
            is_deleted=False,
        ).count()

        return Response({
            'success': True,
            'deleted_count': deleted_count,
            'requested_count': len(raw_instance_ids),
            'remaining_submissions': remaining_submissions,
        }, status=200)
    except Exception as e:
        print(f"Error bulk soft deleting submissions for project {project_id}: {e}")
        return Response({'error': 'Failed to bulk delete submissions'}, status=500)

from .models import DownloadLog
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from django.utils import timezone

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_xls_download(request, form_id):
    import openpyxl
    import os
    from django.utils import timezone
    from django.conf import settings
    from django.http import FileResponse

    try:
        form = Form.objects.get(id=form_id)
    except Form.DoesNotExist:
        return Response({'error': 'Form not found'}, status=404)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Get question-based columns (matching frontend approach)
    all_columns = []

    # Add standard fields
    all_columns.extend(['start', 'end'])

    # Get all actual questions (non-group questions) from the form definition
    form_questions = [q for q in form.questions if
                     q.get('type') not in ['begin_group', 'end_group', 'begin_repeat', 'end_repeat'] and
                     q.get('name') and q.get('name').strip()]

    # Add all question field names
    for question in form_questions:
        all_columns.append(question['name'])

    # Add meta fields
    all_columns.extend(['meta.submitted_by', 'meta.submission_type', 'uuid'])

    # Create headers using question labels (matching frontend)
    headers = []
    for col in all_columns:
        if col == 'start':
            headers.append('Start Time')
        elif col == 'end':
            headers.append('End Time')
        elif col == 'meta.submitted_by':
            headers.append('Submitted By')
        elif col == 'meta.submission_type':
            headers.append('Submission Type')
        elif col == 'uuid':
            headers.append('UUID')
        else:
            # Find question label
            question = next((q for q in form_questions if q['name'] == col), None)
            if question:
                # Handle special designation cases (matching frontend)
                if col == 'designation_1':
                    headers.append('Designation (BRAC)')
                elif col == 'designation':
                    headers.append('Designation (Govt)')
                else:
                    headers.append(question.get('label', col))
            else:
                headers.append(col)

    ws.append(headers)

    # Create question lookup dictionary
    question_lookup = {q['name']: q for q in form_questions}

    # Data rows
    for sub in _get_form_submission_entries(form):
        data = sub.get('data', {})
        flattened_data = flatten_submission_data(data)
        field_value_index = index_submission_values_by_field(data)

        row = []
        for col in all_columns:
            value = None

            # Handle special fields
            if col == 'start':
                value = data.get('start')
            elif col == 'end':
                value = data.get('end')
            elif col == 'meta.submitted_by':
                value = data.get('meta', {}).get('submitted_by')
            elif col == 'meta.submission_type':
                value = data.get('meta', {}).get('submission_type')
            elif col == 'uuid':
                value = data.get('meta', {}).get('instanceID')
            else:
                raw_values = resolve_submission_column_values(
                    flattened_data,
                    field_value_index,
                    col,
                )

                if col in question_lookup and raw_values:
                    question = question_lookup[col]
                    raw_values = [
                        process_submission_value(question, raw_value)
                        for raw_value in raw_values
                    ]

                value = collapse_export_values(raw_values)

            if value in (None, ''):
                row.append('')
            else:
                row.append(value)

        ws.append(row)

    # Save XLSX to media/downloads/
    download_dir = os.path.join(settings.MEDIA_ROOT, 'downloads')
    os.makedirs(download_dir, exist_ok=True)
    file_name = f'form_{form_id}_{timezone.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    file_path = os.path.join(download_dir, file_name)
    wb.save(file_path)

    # Log download
    DownloadLog.objects.create(
        user=request.user,
        form=form,
        type='XLS',
        file_path=file_path,
    )

    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_name)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_download_logs(request, form_id):
    logs = DownloadLog.objects.filter(form_id=form_id).order_by('-created')
    data = [
        {
            'id': log.id,
            'type': log.type,
            'created': log.created,
            'file_url': f"{settings.MEDIA_URL}downloads/{os.path.basename(log.file_path)}"
        }
        for log in logs
    ]
    return Response(data)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_download_log(request, log_id):
    try:
        log = DownloadLog.objects.get(id=log_id)
        if os.path.exists(log.file_path):
            os.remove(log.file_path)
        log.delete()
        return Response({'success': True})
    except DownloadLog.DoesNotExist:
        return Response({'error': 'Log not found'}, status=404)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_enketo_edit_url(request, form_id, instance_id):
    """
    Generate Enketo edit URL for a specific submission.
    Similar to KoboToolbox API pattern with instance caching.
    """
    try:
        # Get the form to retrieve enketo_id
        form = Form.objects.get(id=form_id)
        enketo_id = getattr(form, 'enketo_id', None)

        if not enketo_id:
            return Response(
                {'error': 'Enketo ID not available. Please fetch Enketo data first.'},
                status=400
            )

        normalized_instance_id = _normalize_instance_id(instance_id)
        clean_instance_id = normalized_instance_id[5:] if normalized_instance_id.startswith("uuid:") else normalized_instance_id

        submission_data = None
        submission_record = _find_submission_record(form, normalized_instance_id)
        if submission_record:
            submission_data = _extract_submission_data(submission_record.data or {})
        else:
            for submission in _get_form_submission_entries(form):
                candidate = submission.get('data', {}).get('meta', {}).get('instanceID')
                if _normalize_instance_id(candidate) == normalized_instance_id:
                    submission_data = submission.get('data', {})
                    break

        if not submission_data:
            return Response({'error': 'Submission not found'}, status=404)

        # Convert submission data to XML format for Enketo
        xml_data = convert_submission_to_xml(submission_data, form_id)

        # Step 1: Cache the instance with Enketo API
        enketo_base_url = settings.ENKETO_URL.rstrip('/') if settings.ENKETO_URL else ''
        openrosa_server_url = settings.OPENROSA_SERVER_URL.rstrip('/') if settings.OPENROSA_SERVER_URL else ''

        if not enketo_base_url or not openrosa_server_url:
            return Response(
                {'error': 'Enketo configuration is incomplete. Please set ENKETO_URL and OPENROSA_SERVER_URL.'},
                status=500
            )
        if not settings.ENKETO_API_KEY:
            return Response(
                {'error': 'Enketo API key is not configured.'},
                status=500
            )

        cache_url = f"{enketo_base_url}/api/v2/instance"

        cache_payload = {
            "server_url": openrosa_server_url,
            "form_id": f"form_{form_id}",
            "instance": xml_data,
            "instance_id": clean_instance_id
        }


        print(f"🔍 Enketo cache_payload: {cache_payload}")
        # Actually POST to Enketo and assign response
        import requests, base64
        auth_scheme = (getattr(settings, "ENKETO_AUTH_SCHEME", "token") or "token").strip().lower()
        if auth_scheme == "basic":
            auth_header = "Basic " + base64.b64encode(f"{settings.ENKETO_API_KEY}:".encode()).decode()
        elif auth_scheme == "bearer":
            auth_header = f"Bearer {settings.ENKETO_API_KEY}"
        else:
            auth_header = f"Token {settings.ENKETO_API_KEY}"
        cache_response = requests.post(
            cache_url,
            headers={
                "Authorization": auth_header,
                "Content-Type": "application/json",
            },
            json=cache_payload
        )

        # Log Enketo POST response status and body
        print(f"🔍 Enketo POST status: {cache_response.status_code}")
        try:
            print(f"🔍 Enketo POST response: {cache_response.text}")
        except Exception as e:
            print(f"🔍 Error reading Enketo POST response: {e}")

        if cache_response.status_code == 201:
            # Step 2: Get the edit URL from the response
            try:
                cache_data = cache_response.json()
            except Exception as e:
                print(f"🔍 Error parsing Enketo response JSON: {e}")
                return Response({'error': 'Invalid response from Enketo'}, status=500)
            edit_url = cache_data.get('edit_url')

            if edit_url:
                return Response({
                    'url': edit_url,
                    'version_uid': f"v{form.id}_{clean_instance_id[:8]}"
                })
            else:
                print(f"🔍 No edit_url in Enketo response: {cache_data}")
                return Response({'error': 'No edit URL returned from Enketo'}, status=500)
        else:
            print(f"🔍 Enketo POST failed: {cache_response.status_code} {cache_response.text}")
            return Response({'error': f'Failed to cache record in Enketo: {cache_response.status_code}', 'details': cache_response.text}, status=500)

    except Form.DoesNotExist:
        return Response({'error': 'Form not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)


def convert_submission_to_xml(submission_data, form_id):
    """
    Convert submission data back to XML format for Enketo caching.
    """
    # Get form details for proper XML structure
    try:
        form = Form.objects.get(id=form_id)
    except Form.DoesNotExist:
        return ""

    # Build XML string
    xml_parts = []
    xml_parts.append(f'<data xmlns:jr="http://openrosa.org/javarosa" xmlns:orx="http://openrosa.org/xforms" id="form_{form_id}" version="20240601">')

    # Add form fields
    for key, value in submission_data.items():
        if key == 'meta':
            # Handle meta data specially
            xml_parts.append('<meta>')
            for meta_key, meta_value in value.items():
                xml_parts.append(f'<{meta_key}>{meta_value}</{meta_key}>')
            xml_parts.append('</meta>')
        elif isinstance(value, dict):
            # Handle nested objects
            xml_parts.append(f'<{key}>')
            for sub_key, sub_value in value.items():
                xml_parts.append(f'<{sub_key}>{sub_value}</{sub_key}>')
            xml_parts.append(f'</{key}>')
        else:
            # Handle simple fields
            xml_parts.append(f'<{key}>{value}</{key}>')

    xml_parts.append('</data>')

    return ''.join(xml_parts)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_submission_by_instance_id(request, instance_id):
    """
    Get submission data by instance_id (UUID).
    This endpoint is used by the mobile app to fetch submission data for lookup form filtering.
    """
    try:
        normalized_instance_id = _normalize_instance_id(instance_id)
        submission_record = Submission.objects.filter(
            instance_id=normalized_instance_id,
            is_deleted=False,
        ).order_by('-created_at').first()

        if not submission_record and normalized_instance_id.startswith('uuid:'):
            submission_record = Submission.objects.filter(
                instance_id=normalized_instance_id.replace('uuid:', '', 1),
                is_deleted=False,
            ).order_by('-created_at').first()

        if not submission_record:
            return Response(
                {'error': f'Submission with instance_id {instance_id} not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        payload = _extract_submission_data(submission_record.data or {})
        if payload:
            return Response({'data': payload}, status=status.HTTP_200_OK)

        # Legacy fallback for very old rows that only stored XML.
        if submission_record.xml_file and submission_record.xml_file.name:
            try:
                json_payload = xml_to_json(submission_record.xml_file)
                submission_data = _extract_submission_data(json_payload)
                return Response({'data': submission_data}, status=status.HTTP_200_OK)
            except Exception as xml_error:
                print(f"DEBUG: Failed to parse XML for submission {submission_record.id}: {xml_error}")

        return Response(
            {'error': f'Submission with instance_id {instance_id} has no readable payload'},
            status=status.HTTP_404_NOT_FOUND
        )

    except Exception as e:
        print(f"DEBUG: Error in get_submission_by_instance_id: {str(e)}")
        return Response(
            {'error': f'Error retrieving submission: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_submission_processing_status(request, instance_id):
    """
    Return async processing status for a submission instance.
    """
    normalized_instance_id = _normalize_instance_id(instance_id)
    if not normalized_instance_id:
        return Response(
            {'error': f'Invalid instance_id: {instance_id}'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    instance_candidates = {normalized_instance_id}
    if normalized_instance_id.startswith('uuid:'):
        instance_candidates.add(normalized_instance_id.replace('uuid:', '', 1))
    else:
        instance_candidates.add(f'uuid:{normalized_instance_id}')

    submission_record = (
        Submission.objects.filter(
            instance_id__in=list(instance_candidates),
            is_deleted=False,
        )
        .order_by('-created_at')
        .first()
    )
    if not submission_record:
        return Response(
            {'error': f'Submission with instance_id {instance_id} not found'},
            status=status.HTTP_404_NOT_FOUND,
        )

    return Response(
        {
            'submission_id': submission_record.id,
            'instance_id': submission_record.instance_id,
            'processing_status': submission_record.processing_status,
            'is_processed': submission_record.processing_status == 'completed',
            'created_at': submission_record.created_at.isoformat() if submission_record.created_at else None,
            'processed_at': submission_record.processed_at.isoformat() if submission_record.processed_at else None,
            'processing_error': submission_record.processing_error or '',
        },
        status=status.HTTP_200_OK,
    )


# Add this API endpoint to update anonymous submission setting
@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_form_anonymous_setting(request, form_id):
    try:
        form = Form.objects.get(id=form_id)
        allow_anonymous = request.data.get('allow_anonymous_submissions', False)
        form.allow_anonymous_submissions = allow_anonymous
        form.save()

        return JsonResponse({
            'message': 'Anonymous submission setting updated successfully',
            'allow_anonymous_submissions': allow_anonymous
        })
    except Form.DoesNotExist:
        return JsonResponse({'error': 'Form not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def generate_auth_token(user):
    """Generate JWT token for user authentication"""
    import jwt
    from django.conf import settings
    from datetime import datetime, timedelta

    payload = {
        'user_id': user.id,
        'username': user.username,
        'exp': datetime.utcnow() + timedelta(days=7)  # 7 days expiration
    }

    token = jwt.encode(payload, settings.SECRET_KEY, algorithm='HS256')
    return token

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def set_enketo_cookie(request):
    """Set authentication cookie for Enketo access"""
    try:
        user = request.user
        auth_token = generate_auth_token(user)

        print(f"Setting cookie for user: {user.username}")
        print(f"Auth token: {auth_token[:20]}...")

        response = JsonResponse({'message': 'Cookie set successfully'})

        secure_cookie = request.is_secure() or settings.FRONTEND_URL.startswith("https://")
        cookie_kwargs = {
            'httponly': True,
            'secure': secure_cookie,
            'samesite': 'None' if secure_cookie else 'Lax',
            'max_age': 7 * 24 * 60 * 60,
        }
        if settings.AUTH_COOKIE_DOMAIN:
            cookie_kwargs['domain'] = settings.AUTH_COOKIE_DOMAIN

        response.set_cookie('commicplan_auth', auth_token, **cookie_kwargs)

        # Set CORS headers
        response['Access-Control-Allow-Credentials'] = 'true'
        origin = request.META.get('HTTP_ORIGIN')
        allowed_origins = getattr(settings, "ENKETO_ALLOWED_ORIGINS", [])
        if origin and origin in allowed_origins:
            response['Access-Control-Allow-Origin'] = origin
        else:
            response['Access-Control-Allow-Origin'] = settings.FRONTEND_URL

        return response
    except Exception as e:
        print(f"Error setting cookie: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([AllowAny])
def debug_auth(request):
    return JsonResponse({
        'path': request.path,
        'method': request.method,
        'get_params': dict(request.GET),
        'post_params': dict(request.POST),
        'authenticated': request.user.is_authenticated,
        'user': request.user.username if request.user.is_authenticated else None,
        'cookies': list(request.COOKIES.keys()),
        'return_url': request.GET.get('return', 'NOT_FOUND')
    })

@api_view(['GET'])
@permission_classes([AllowAny])
def debug_enketo_auth(request):
    """Debug endpoint to check Enketo authentication flow"""
    return JsonResponse({
        'path': request.path,
        'full_url': request.build_absolute_uri(),
        'method': request.method,
        'get_params': dict(request.GET),
        'headers': dict(request.META),
        'cookies': dict(request.COOKIES),
        'authenticated': request.user.is_authenticated,
        'user': request.user.username if request.user.is_authenticated else None,
    })

@api_view(['GET'])
@permission_classes([AllowAny])
def test_cookie(request):
    """Test endpoint to check cookie setting and reading"""

    # Check if cookie exists
    auth_cookie = request.COOKIES.get('commicplan_auth')

    response_data = {
        'cookies_received': dict(request.COOKIES),
        'commicplan_auth_present': bool(auth_cookie),
        'domain': request.get_host(),
        'origin': request.META.get('HTTP_ORIGIN'),
    }

    response = JsonResponse(response_data)

    secure_cookie = request.is_secure() or settings.FRONTEND_URL.startswith("https://")
    cookie_kwargs = {
        'httponly': False,
        'secure': secure_cookie,
        'samesite': 'None' if secure_cookie else 'Lax',
        'max_age': 300,
    }
    if settings.AUTH_COOKIE_DOMAIN:
        cookie_kwargs['domain'] = settings.AUTH_COOKIE_DOMAIN

    # Set a test cookie
    response.set_cookie('test_cookie', 'test_value', **cookie_kwargs)

    return response


@api_view(['GET'])
@permission_classes([AllowAny])
def get_latest_app_version(request):
    latest = (
        AppVersion.objects
        .only('id', 'version', 'update_info', 'created_at')
        .order_by('-created_at', '-id')
        .first()
    )

    if latest is None:
        return Response(
            {'detail': 'No app version information found.'},
            status=status.HTTP_404_NOT_FOUND,
        )

    serializer = AppVersionSerializer(latest)
    return Response(serializer.data, status=status.HTTP_200_OK)

from django.middleware.csrf import get_token
from django.views.decorators.csrf import ensure_csrf_cookie
from django.http import JsonResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
import os

 # Custom CSRF token endpoint for Enketo (__csrf cookie)
@ensure_csrf_cookie
@api_view(['GET', 'POST', 'HEAD'])
@permission_classes([AllowAny])
def get_csrf_token(request):
    from django.middleware.csrf import get_token
    import os
    csrf_token = get_token(request)
    response = JsonResponse({'csrfToken': csrf_token})

    secure_cookie = request.is_secure() or settings.FRONTEND_URL.startswith("https://")
    cookie_kwargs = {
        'httponly': False,
        'secure': secure_cookie,
        'samesite': 'None' if secure_cookie else 'Lax',
        'max_age': 60 * 60 * 24 * 7,
    }
    if settings.AUTH_COOKIE_DOMAIN:
        cookie_kwargs['domain'] = settings.AUTH_COOKIE_DOMAIN

    # Always set the __csrf cookie for Enketo compatibility
    response.set_cookie(key='__csrf', value=csrf_token, **cookie_kwargs)

    # Set CORS headers for browser compatibility
    origin = request.META.get('HTTP_ORIGIN')
    allowed_origins = getattr(settings, "ENKETO_ALLOWED_ORIGINS", [])
    if origin and origin in allowed_origins:
        response['Access-Control-Allow-Origin'] = origin
    else:
        response['Access-Control-Allow-Origin'] = settings.FRONTEND_URL
    response['Access-Control-Allow-Credentials'] = 'true'
    response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
    response['Access-Control-Allow-Methods'] = 'GET, HEAD, OPTIONS'
    return response

# New optimized API endpoints for mobile app performance
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_forms(request):
    """
    Get only forms assigned to the current user (via profile.forms)
    Returns forms with template and project information in one request
    Optimized for role 4/5 users who only need to see their assigned forms
    """
    try:
        user = request.user
        user_role = getattr(user, 'role', 4)

        # For role 4 and 5, only return forms they have access to
        if user_role in [4, 5]:
            if not hasattr(user, 'profile'):
                return Response([], status=status.HTTP_200_OK)

            user_form_ids = user.profile.forms.values_list('id', flat=True)

            # Fetch forms with related data in one query
            forms = Form.objects.filter(
                id__in=user_form_ids
            ).select_related('template', 'project').only(
                'id', 'name', 'template__id', 'project__id', 'questions'
            )

            result = []
            for form in forms:
                result.append({
                    'uid': form.id,
                    'name': form.name,
                    'template': form.template.id if form.template else None,
                    'project': form.project.id if form.project else None,
                    'questions': form.questions or []
                })

            return Response(result, status=status.HTTP_200_OK)

        # For officer role, return forms only from assigned projects
        if user_role == 6:
            project_ids = _get_assigned_project_ids(user)
            if not project_ids:
                return Response([], status=status.HTTP_200_OK)

            forms = Form.objects.filter(
                project_id__in=project_ids
            ).select_related('template', 'project').only(
                'id', 'name', 'template__id', 'project__id', 'questions'
            )

            result = []
            for form in forms:
                result.append({
                    'uid': form.id,
                    'name': form.name,
                    'template': form.template.id if form.template else None,
                    'project': form.project.id if form.project else None,
                    'questions': form.questions or []
                })

            return Response(result, status=status.HTTP_200_OK)

        # For other roles, return all forms
        else:
            forms = Form.objects.all().select_related('template', 'project').only(
                'id', 'name', 'template__id', 'project__id', 'questions'
            )

            result = []
            for form in forms:
                result.append({
                    'uid': form.id,
                    'name': form.name,
                    'template': form.template.id if form.template else None,
                    'project': form.project.id if form.project else None,
                    'questions': form.questions or []
                })

            return Response(result, status=status.HTTP_200_OK)

    except Exception as e:
        print(f"Error in get_user_forms: {e}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_forms_without_submission(request):
    """
    Lightweight form list endpoint that excludes the submission payload.

    Optional query params:
      - project: filter by project ID
      - ids: comma-separated list of form IDs to include
    """
    try:
        # Role-based filtering: users with role 4/5 only see their assigned forms
        user = request.user
        if user.role in [4, 5]:
            forms = Form.objects.filter(id__in=user.profile.forms.all())
        else:
            forms = Form.objects.filter(project__in=user.profile.projects.all())

        project_id = request.query_params.get('project')
        if project_id:
            forms = forms.filter(project_id=project_id)

        ids_param = request.query_params.get('ids')
        if ids_param:
            try:
                id_list = [
                    int(form_id.strip()) for form_id in ids_param.split(',')
                    if form_id.strip()
                ]
            except ValueError:
                return Response(
                    {'error': 'Invalid ids parameter, expected comma-separated integers.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if id_list:
                forms = forms.filter(id__in=id_list)

        forms = forms.select_related('project', 'template', 'default_language').prefetch_related('other_languages')
        from .serializers import FormWithoutSubmissionSerializer
        serializer = FormWithoutSubmissionSerializer(forms, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as exc:
        return Response(
            {'error': str(exc)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def get_templates_bulk(request):
    """
    Get multiple templates and their data in one request
    Accepts a list of template IDs and returns all template data
    """
    try:
        template_ids = request.data.get('template_ids', [])
        if not template_ids:
            return Response([], status=status.HTTP_200_OK)

        normalized_ids = sorted({
            int(template_id) for template_id in template_ids
            if str(template_id).isdigit()
        })
        if not normalized_ids:
            return Response([], status=status.HTTP_200_OK)

        cache_ttl = int(getattr(settings, 'SUBMISSION_TEMPLATE_CACHE_TTL_SECONDS', 30))
        cache_key = f"templates-bulk:v1:{','.join(str(i) for i in normalized_ids)}"
        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached, status=status.HTTP_200_OK)

        # Fetch templates with related data
        templates = Template.objects.filter(
            id__in=normalized_ids
        ).prefetch_related(
            'generated_lookup_forms',
            'lookup_forms',
            'data_collection_form'
        )

        result = []
        for template in templates:
            template_data = {
                'id': template.id,
                'name': template.name,
                'data_collection_form': None,
                'generated_lookup_forms': []
            }

            # Add data collection form
            if template.data_collection_form:
                template_data['data_collection_form'] = {
                    'id': template.data_collection_form.id,
                    'name': template.data_collection_form.name,
                    'uid': template.data_collection_form.id
                }

            # Add generated lookup forms
            for lookup_form in template.generated_lookup_forms.all():
                template_data['generated_lookup_forms'].append({
                    'id': lookup_form.id,
                    'name': lookup_form.name,
                    'uid': lookup_form.id,
                    'questions': lookup_form.questions or []
                })

            result.append(template_data)

        cache.set(cache_key, result, timeout=cache_ttl)
        return Response(result, status=status.HTTP_200_OK)

    except Exception as e:
        print(f"Error in get_templates_bulk: {e}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

# Patient management endpoints

class PatientViewSet(viewsets.ModelViewSet):
    queryset = Patient.objects.all()
    serializer_class = PatientSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Patient.objects.all().order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def update(self, request, *args, **kwargs):
        """
        Custom update method with debugging for patient updates
        """
        partial = kwargs.pop('partial', False)
        instance = self.get_object()

        print(f"[DEBUG] Patient Update Request:")
        print(f"[DEBUG] Patient ID: {instance.id} ({instance.patient_id})")
        print(f"[DEBUG] Request Data: {request.data}")
        print(f"[DEBUG] Request User: {request.user}")

        # Filter out read-only fields from request data
        allowed_fields = ['name', 'email', 'phone', 'age', 'gender', 'address']
        filtered_data = {k: v for k, v in request.data.items() if k in allowed_fields}

        print(f"[DEBUG] Filtered Data: {filtered_data}")

        serializer = self.get_serializer(instance, data=filtered_data, partial=partial)

        try:
            serializer.is_valid(raise_exception=True)
            print(f"[DEBUG] Serializer validation passed")
        except Exception as e:
            print(f"[DEBUG] Serializer validation failed: {e}")
            print(f"[DEBUG] Serializer errors: {serializer.errors}")
            raise

        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            instance._prefetched_objects_cache = {}

        print(f"[DEBUG] Patient updated successfully: {serializer.data}")
        return Response(serializer.data)

    def perform_update(self, serializer):
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        """Move patient to trash instead of permanent deletion"""
        patient = self.get_object()
        try:
            trash_entry = enhanced_delete_patient(patient, request.user)
            return Response({
                'message': f'Patient {patient.patient_id} moved to trash',
                'trash_id': trash_entry.id,
                'auto_delete_date': trash_entry.auto_delete_at.isoformat()
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to move patient to trash: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def search(self, request):
        from django.db import models
        query = request.query_params.get('q', '')
        if query:
            patients = Patient.objects.filter(
                models.Q(patient_id__icontains=query) |
                models.Q(name__icontains=query) |
                models.Q(email__icontains=query) |
                models.Q(phone__icontains=query)
            ).order_by('-created_at')
        else:
            patients = Patient.objects.all().order_by('-created_at')

        serializer = self.get_serializer(patients, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def submissions(self, request, pk=None):
        """Get all submissions for a specific patient"""
        try:
            patient = self.get_object()
            submissions = patient.submissions.all().order_by('-created_at')

            # Convert submissions to a format similar to AllRowsDataTable
            submission_data = []
            for submission in submissions:
                try:
                    # Basic submission data
                    flat_data = {
                        'submission_id': str(submission.id),
                        'submitted_by': submission.user.username if submission.user else 'Unknown',
                        'created_at': submission.created_at.isoformat() if submission.created_at else '',
                        'patient_id': patient.patient_id,
                        'patient_name': patient.name or '',
                        'xml_file': submission.xml_file.url if submission.xml_file else '',
                    }

                    # Try to parse XML data if available
                    if submission.xml_file:
                        try:
                            import xml.etree.ElementTree as ET
                            # Read the XML file and parse it for form data
                            xml_content = submission.xml_file.read()
                            root = ET.fromstring(xml_content)

                            # Extract data from XML (this is a basic implementation)
                            for elem in root.iter():
                                if elem.text and elem.tag:
                                    # Clean tag name and add to flat_data
                                    tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                                    if tag_name not in ['meta', 'instanceID']:
                                        flat_data[tag_name] = elem.text
                        except Exception as e:
                            print(f"Error parsing XML for submission {submission.id}: {e}")
                            flat_data['xml_parse_error'] = str(e)

                    submission_data.append(flat_data)
                except Exception as e:
                    print(f"Error processing submission {submission.id}: {e}")
                    continue

            return Response({
                'patient': {
                    'id': patient.id,
                    'patient_id': patient.patient_id,
                    'name': patient.name,
                    'email': patient.email,
                    'phone': patient.phone,
                    'age': patient.age,
                    'gender': patient.gender,
                    'address': patient.address,
                    'created_at': patient.created_at.isoformat() if patient.created_at else None,
                },
                'submissions': submission_data,
                'total_submissions': len(submission_data)
            })
        except Patient.DoesNotExist:
            return Response({'error': 'Patient not found'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)    @action(detail=True, methods=['get'])
    def submissions_csv(self, request, pk=None):
        """Download CSV of all submissions for a specific patient"""
        import csv
        from django.http import HttpResponse

        try:
            patient = self.get_object()
            submissions = patient.submissions.all().order_by('-created_at')

            # Prepare CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="patient_{patient.patient_id}_submissions.csv"'

            writer = csv.writer(response)

            # Collect all unique field names
            all_fields = set(['submission_id', 'submitted_by', 'created_at', 'patient_id', 'patient_name', 'xml_file'])
            submission_data = []

            for submission in submissions:
                try:
                    flat_data = {
                        'submission_id': str(submission.id),
                        'submitted_by': submission.user.username if submission.user else 'Unknown',
                        'created_at': submission.created_at.strftime('%Y-%m-%d %H:%M:%S') if submission.created_at else '',
                        'patient_id': patient.patient_id,
                        'patient_name': patient.name or '',
                        'xml_file': submission.xml_file.url if submission.xml_file else '',
                    }

                    # Try to parse XML data if available
                    if submission.xml_file:
                        try:
                            import xml.etree.ElementTree as ET
                            xml_content = submission.xml_file.read()
                            root = ET.fromstring(xml_content)

                            for elem in root.iter():
                                if elem.text and elem.tag:
                                    tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                                    if tag_name not in ['meta', 'instanceID']:
                                        all_fields.add(tag_name)
                                        flat_data[tag_name] = elem.text
                        except Exception as e:
                            print(f"Error parsing XML for submission {submission.id}: {e}")
                            all_fields.add('xml_parse_error')
                            flat_data['xml_parse_error'] = str(e)

                    submission_data.append(flat_data)
                except Exception as e:
                    print(f"Error processing submission {submission.id}: {e}")
                    continue

            # Write header
            sorted_fields = sorted(all_fields)
            writer.writerow(sorted_fields)

            # Write data rows
            for row_data in submission_data:
                row = [row_data.get(field, '') for field in sorted_fields]
                writer.writerow(row)

            return response

        except Patient.DoesNotExist:
            return Response({'error': 'Patient not found'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_patients_csv(request):
    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="patients.csv"'

    writer = csv.writer(response)
    writer.writerow(['Patient ID', 'Name', 'Email', 'Phone', 'Age', 'Gender', 'Address', 'Submission Count', 'Created At'])

    patients = Patient.objects.all().order_by('-created_at')
    for patient in patients:
        writer.writerow([
            patient.patient_id,
            patient.name or '',
            patient.email or '',
            patient.phone or '',
            patient.age or '',
            patient.gender or '',
            patient.address or '',
            patient.submissions.count(),
            patient.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])

    return response


# Trash Bin Management ViewSet
class TrashBinViewSet(viewsets.ModelViewSet):
    queryset = TrashBin.objects.all()
    serializer_class = TrashBinSerializer
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return TrashBinListSerializer
        return TrashBinSerializer

    def get_queryset(self):
        """Filter trash items by type if specified"""
        queryset = TrashBin.objects.all().order_by('-deleted_at')

        # Filter out restored items if the field exists
        try:
            # Try to filter by restored field
            queryset = queryset.filter(restored=False)
        except Exception:
            # Field doesn't exist yet, show all items
            pass

        item_type = self.request.query_params.get('type', None)
        if item_type:
            queryset = queryset.filter(item_type=item_type)
        return queryset

    def create(self, request, *args, **kwargs):
        """Prevent direct creation through API"""
        return Response(
            {'error': 'Items are moved to trash through delete operations'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )

    def update(self, request, *args, **kwargs):
        """Prevent updating trash entries"""
        return Response(
            {'error': 'Trash entries cannot be updated'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )

    def destroy(self, request, *args, **kwargs):
        """Permanently delete item from trash"""
        trash_entry = self.get_object()

        # Check if item is restored (if field exists)
        if hasattr(trash_entry, 'restored') and trash_entry.restored:
            return Response(
                {'error': 'Cannot delete restored item'},
                status=status.HTTP_400_BAD_REQUEST
            )

        _delete_soft_submission_for_trash_entry(trash_entry)
        trash_entry.delete()
        return Response({'message': 'Item permanently deleted'}, status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        """Restore item from trash"""
        import logging
        logger = logging.getLogger(__name__)

        logger.info(f"🔄 Restore request for trash item {pk}")
        trash_entry = self.get_object()
        logger.info(f"📋 Trash entry: {trash_entry.item_type} - {trash_entry.item_name}")

        try:
            restored_item = restore_from_trash(trash_entry, request.user)
            logger.info(f"✅ Successfully restored item {restored_item.id}")
            return Response({
                'message': 'Item restored successfully',
                'restored_item_id': restored_item.id
            }, status=status.HTTP_200_OK)
        except ValueError as e:
            logger.error(f"❌ ValueError during restore: {e}")
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"❌ Exception during restore: {e}")
            return Response({'error': f'Failed to restore item: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def cleanup_expired(self, request):
        """Cleanup expired items (admin only)"""
        if not request.user.is_staff:
            return Response(
                {'error': 'Only admin users can cleanup expired items'},
                status=status.HTTP_403_FORBIDDEN
            )

        count = cleanup_expired_trash()
        return Response({
            'message': f'Cleaned up {count} expired items',
            'count': count
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get trash bin statistics"""
        total_items = TrashBin.objects.filter(restored=False).count()
        expired_items = TrashBin.objects.filter(
            auto_delete_at__lte=timezone.now(),
            restored=False
        ).count()

        # Items by type
        items_by_type = {}
        for item_type, _ in TrashBin.ITEM_TYPES:
            items_by_type[item_type] = TrashBin.objects.filter(
                item_type=item_type,
                restored=False
            ).count()

        return Response({
            'total_items': total_items,
            'expired_items': expired_items,
            'items_by_type': items_by_type
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_current_user_profile(request):
    """Get current user's profile with assigned organizations"""
    try:
        user_profile = request.user.profile
        organizations = user_profile.organizations.all()

        organizations_data = []
        for org in organizations:
            organizations_data.append({
                'id': org.id,
                'name': org.name,
                'type': org.type,
                'email': org.email,
                'website': org.website,
                'location': org.location,
                'active_user': org.active_user,
                'receive_updates': org.receive_updates,
                'created_at': org.created_at,
                'updated_at': org.updated_at,
            })

        return Response({
            'user_id': request.user.id,
            'username': request.user.username,
            'email': request.user.email,
            'role': getattr(request.user, 'role', None),
            'organizations': organizations_data
        })

    except Exception as e:
        return Response(
            {'error': f'Error getting user profile: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# Enhanced delete methods for existing ViewSets
def enhanced_delete_project(project, user=None):
    """Move project to trash instead of permanent deletion"""
    return move_to_trash(project, 'project', user)

def enhanced_delete_patient(patient, user=None):
    """Move patient to trash instead of permanent deletion"""
    return move_to_trash(patient, 'patient', user)

def enhanced_delete_form(form, user=None):
    """Move form to trash instead of permanent deletion"""
    return move_to_trash(form, 'form', user)

def enhanced_delete_submission(submission, user=None):
    """Move submission to trash instead of permanent deletion"""
    return move_to_trash(submission, 'submission', user)

def enhanced_delete_organization(organization, user=None):
    """Move organization to trash instead of permanent deletion"""
    return move_to_trash(organization, 'organization', user)

def enhanced_delete_template(template, user=None):
    """Move template to trash instead of permanent deletion"""
    return move_to_trash(template, 'template', user)
