from calendar import monthrange
import csv
from collections import defaultdict
from copy import copy
from datetime import date, datetime
import hashlib
from io import BytesIO, StringIO
import json
import math
from pathlib import Path
import re
from tempfile import TemporaryFile
from urllib.parse import urlencode
from zoneinfo import ZoneInfo
import zipfile
from xml.etree import ElementTree as ET

from django.contrib.auth.models import User
from django.core import signing
from django.core.paginator import EmptyPage, Paginator
from django.core.signing import BadSignature, SignatureExpired
from django.http import FileResponse, HttpResponse
from django.db.models import Count, Max, Prefetch, Q, Sum, Value
from django.db.models.functions import Coalesce, Upper
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from rest_framework import renderers, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .metadata_approval import (
    apply_local_metadata_pending_to_db,
    apply_nonlocal_metadata_pending_to_db,
    filter_nonlocal_metadata_submission,
    local_metadata_submission_nonempty,
    merge_local_metadata_pending,
    merge_nonlocal_metadata_pending,
    parse_local_metadata_from_request,
    reject_local_metadata,
    reject_nonlocal_metadata,
)
from .models import (
    District,
    LocalRecord,
    MalariaGridColumnLayout,
    MalariaUserRole,
    METADATA_APPROVAL_APPROVED,
    METADATA_APPROVAL_PENDING,
    METADATA_APPROVAL_REJECTED,
    MicrostatificationDataUpload,
    MonthAccessSetting,
    MonthlyApproval,
    NonLocalRecord,
    Union,
    Upazila,
    Village,
)
from .microstatification_sync import sync_microstatification_workbook
from .permissions import (
    HasMalariaAccess,
    IsMalariaAdmin,
    IsMalariaGlobalAdmin,
    IsMalariaPrivileged,
    get_dm_district_id,
    get_malaria_role,
    has_malaria_access,
    has_malaria_privileged_access,
    is_malaria_dm,
    is_malaria_global_admin,
)
from .serializers import (
    DistrictSerializer,
    LocalRecordSerializer,
    MalariaGridColumnLayoutPayloadSerializer,
    MalariaGridColumnLayoutResponseSerializer,
    MalariaSessionSerializer,
    MalariaUserCreateSerializer,
    MalariaUserRoleSerializer,
    MicrostatificationDataUploadSerializer,
    MicrostatificationDataUploadFileSerializer,
    MonthAccessSettingSerializer,
    MonthlyApprovalSerializer,
    NonLocalRecordSerializer,
    ProfileSerializer,
    UnionSerializer,
    UpazilaSerializer,
    VillageSerializer,
    build_full_name,
    get_local_record_display_details,
)


MONTH_COLUMNS = [
    "jan_cases",
    "feb_cases",
    "mar_cases",
    "apr_cases",
    "may_cases",
    "jun_cases",
    "jul_cases",
    "aug_cases",
    "sep_cases",
    "oct_cases",
    "nov_cases",
    "dec_cases",
]
MONTH_COLUMN_LABELS = [
    ("jan_cases", "January", "Jan"),
    ("feb_cases", "February", "Feb"),
    ("mar_cases", "March", "Mar"),
    ("apr_cases", "April", "Apr"),
    ("may_cases", "May", "May"),
    ("jun_cases", "June", "Jun"),
    ("jul_cases", "July", "Jul"),
    ("aug_cases", "August", "Aug"),
    ("sep_cases", "September", "Sep"),
    ("oct_cases", "October", "Oct"),
    ("nov_cases", "November", "Nov"),
    ("dec_cases", "December", "Dec"),
]
ITN_FIELDS = ["itn_2023", "itn_2024", "itn_2025", "itn_2026"]
COUNT_QUERY_VALUES = {"1", "true", "yes", "on", "exact"}
REQUESTED_FIELD_PATTERN = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
MICROSTATIFICATION_TEMPLATE_DIR = (
    Path(__file__).resolve().parents[2]
    / "Malaria-Reporting-System"
    / "public"
    / "data"
)
MICROSTATIFICATION_EXPORT_DEBUG_DIR = (
    MICROSTATIFICATION_TEMPLATE_DIR / "generated_microstatification_exports"
)
MICROSTATIFICATION_EXPORT_CACHE_TTL_SECONDS = 0
MICROSTATIFICATION_TEMPLATE_FILES = {
    "Bandarban": "01_Microstatification Format _Bandarban_District_2026.xlsx",
    "Khagrachhari": "02_Microstatification Format _Khagrachhari_District_2026.xlsx",
    "Cox's Bazar": "03_Microstatification Format _Cox's_Bazar_District_2026.xlsx",
    "Rangamati": "04_Microstatification Format _Rangamati_District_2026.xlsx",
    "Chattogram": "05_Microstatification Format _Chattogram_District_2026.xlsx",
}
MICROSTATIFICATION_TEMPLATE_HEADER_ROW = 6
MICROSTATIFICATION_TEMPLATE_DATA_START_ROW = 7
MICROSTATIFICATION_ALL_DISTRICTS_TOKEN = "__all__"
MICROSTATIFICATION_ALL_DISTRICTS_VALUES = {"all", "all-districts", "all_districts", "*"}
MICROSTATIFICATION_EXPORT_FORMATS = {"xlsx", "csv"}
MICROSTATIFICATION_DIVISION_BY_DISTRICT = {
    "Bandarban": "Chattogram",
    "Khagrachhari": "Chattogram",
    "Cox's Bazar": "Chattogram",
    "Rangamati": "Chattogram",
    "Chattogram": "Chattogram",
}
MICROSTATIFICATION_CSV_HEADERS = [
    "SL",
    "Country",
    "Division",
    "District",
    "Upazila",
    "Union",
    "Ward No",
    "Name of SK/SHW",
    "Desig.",
    "Name of SS",
    "Village Name (English)",
    "Village Name \n(Bangla)",
    "Village \nCode",
    "Latitude",
    "Longitute",
    "Population",
    "HH Number",
    "2026 \n(Active LLINs)",
    "2025\n(Active LLINs)",
    "2024\n(Active LLINs)",
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
    "Name of\n MMW, \nHealth \npost &\nCHW(C)",
    "Village \nDistance\n from \nupazila office (KM)",
    "Name of Border\n with others\n country",
    "Others\n Activities\n(TDA/Dev care)",
]
MICRO_DASHBOARD_ROLE_LABELS = {
    4: "User",
    8: "SPO",
    9: "SPO",
    10: "DM",
    11: "SPO",
}
MICRO_DASHBOARD_SCOPE_LABELS = (
    ("district", "District"),
    ("upazila", "Upazila"),
    ("union", "Union"),
    ("ward", "Ward"),
    ("village", "Village"),
    ("multi_village", "Multi Village"),
    ("unassigned", "Unassigned"),
)
MICROSTATIFICATION_DOWNLOAD_TICKET_SALT = "microstatification-download-ticket"
MICROSTATIFICATION_DOWNLOAD_TICKET_MAX_AGE_SECONDS = 300
FIELD_USER_VILLAGE_EDITABLE_FIELDS = {
    "name",
    "name_bn",
    "village_code",
    "latitude",
    "longitude",
    "population",
    "ward_no",
    "sk_shw_name",
    "ss_name",
    "mmw_hp_chwc_name",
    "distance_from_upazila_office_km",
    "bordering_country_name",
    "other_activities",
}


class XLSXBinaryRenderer(renderers.BaseRenderer):
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    format = "xlsx"
    charset = None
    render_style = "binary"

    def render(self, data, accepted_media_type=None, renderer_context=None):
        if data is None:
            return b""
        if isinstance(data, (bytes, bytearray)):
            return bytes(data)
        return data


class JSONFormatXLSXRenderer(renderers.JSONRenderer):
    # Allow using query param `format=xlsx` without triggering DRF 404 on JSON endpoints.
    format = "xlsx"


class JSONFormatCSVRenderer(renderers.JSONRenderer):
    # Allow using query param `format=csv` without triggering DRF 404 on JSON endpoints.
    format = "csv"


def _serialize_user(user):
    profile = getattr(user, "profile", None)
    micro_role = getattr(profile, "micro_role", "") if profile else ""
    profile_data = {
        "user_id": user.id,
        "full_name": build_full_name(user),
        "email": user.email or user.username,
        "micro_role": micro_role or None,
    }
    role = get_malaria_role(user)
    return {
        "id": user.id,
        "email": user.email or "",
        "username": user.username,
        "profile": profile_data,
        "role": role,
    }


def _build_full_name_from_row(row):
    full_name = " ".join(part for part in [row.get("first_name"), row.get("last_name")] if part).strip()
    return full_name or row.get("email") or row.get("username") or ""


def _serialize_user_from_row(row, role):
    profile_data = {
        "user_id": row["id"],
        "full_name": _build_full_name_from_row(row),
        "email": row.get("email") or row.get("username") or "",
    }
    return {
        "id": row["id"],
        "email": row.get("email") or "",
        "username": row.get("username") or "",
        "profile": profile_data,
        "role": role,
    }


def _dm_spo_user_scope_q(user):
    did = get_dm_district_id(user)
    if not did:
        return Q(pk__in=[])
    district_scope = (
        Q(profile__micro_district_id=did)
        | Q(profile__micro_upazila__district_id=did)
        | Q(profile__micro_union__upazila__district_id=did)
        | Q(profile__micro_village__union__upazila__district_id=did)
        | Q(profile__micro_villages__union__upazila__district_id=did)
        | Q(created_by=user)
    )
    return Q(role__in={8, 9, 11}) & district_scope


def _apply_filters(queryset, request, exact_fields, in_fields=None):
    in_fields = in_fields or []
    params = request.query_params

    for field in exact_fields:
        value = params.get(field)
        if value not in (None, ""):
            queryset = queryset.filter(**{field: value})

    for field in in_fields:
        value = params.get(f"{field}__in")
        if value:
            items = [item.strip() for item in value.split(",") if item.strip()]
            queryset = queryset.filter(**{f"{field}__in": items})

    order_field = params.get("_order") or params.get("order")
    ascending = (params.get("_ascending") or "true").lower() not in {"0", "false", "no"}
    if order_field:
        queryset = queryset.order_by(order_field if ascending else f"-{order_field}")

    return queryset


def _apply_micro_village_code_sort(queryset):
    return queryset.annotate(
        micro_sort_ward_no=Upper(Coalesce("ward_no", Value(""))),
        micro_sort_village_code=Upper(Coalesce("village_code", Value(""))),
        micro_sort_name=Upper(Coalesce("name", Value(""))),
    ).order_by(
        "union__upazila__district__name",
        "union__upazila__name",
        "union__name",
        "micro_sort_ward_no",
        "micro_sort_village_code",
        "micro_sort_name",
    )


def _apply_micro_local_record_sort(queryset):
    return queryset.annotate(
        micro_sort_union=Upper(Coalesce("village__union__name", Value(""))),
        micro_sort_ward_no=Upper(Coalesce("village__ward_no", Value(""))),
        micro_sort_village_code=Upper(Coalesce("village__village_code", Value(""))),
        micro_sort_name=Upper(Coalesce("village__name", Value(""))),
    ).order_by(
        "village__union__upazila__district__name",
        "village__union__upazila__name",
        "micro_sort_union",
        "micro_sort_ward_no",
        "micro_sort_village_code",
        "micro_sort_name",
    )


def _count_requested(request):
    return (request.query_params.get("count") or "").lower() in COUNT_QUERY_VALUES


def _get_requested_serializer_fields(request):
    raw_fields = (request.query_params.get("_fields") or "").strip()
    if not raw_fields:
        return None

    fields = []
    for item in raw_fields.split(","):
        field_name = item.strip()
        if field_name and REQUESTED_FIELD_PATTERN.fullmatch(field_name):
            fields.append(field_name)

    return fields or None


def _pagination_requested(request):
    params = request.query_params
    paginate_flag = (params.get("paginate") or "").strip().lower()
    if paginate_flag in {"1", "true", "yes", "on"}:
        return True
    return (params.get("page") or "").strip().isdigit()


def _build_page_url(request, page_number):
    params = request.query_params.copy()
    params["paginate"] = "1"
    params["page"] = str(page_number)
    return f"{request.path}?{urlencode(params, doseq=True)}"


def _paginate_queryset_if_requested(request, queryset, default_page_size=100, max_page_size=500):
    if not _pagination_requested(request):
        return None

    raw_page = (request.query_params.get("page") or "1").strip()
    raw_size = (request.query_params.get("page_size") or str(default_page_size)).strip()
    page_number = int(raw_page) if raw_page.isdigit() else 1
    page_size = int(raw_size) if raw_size.isdigit() else default_page_size
    page_size = max(1, min(page_size, max_page_size))

    paginator = Paginator(queryset, page_size)
    try:
        page_obj = paginator.page(page_number)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages if paginator.num_pages > 0 else 1)

    next_url = _build_page_url(request, page_obj.next_page_number()) if page_obj.has_next() else None
    prev_url = _build_page_url(request, page_obj.previous_page_number()) if page_obj.has_previous() else None
    return {
        "count": paginator.count,
        "next": next_url,
        "previous": prev_url,
        "results": page_obj.object_list,
    }


def _build_list_cache_tag(request, queryset, scope_key):
    max_updated = queryset.aggregate(max_updated=Max("updated_at")).get("max_updated")
    max_updated_iso = max_updated.isoformat() if max_updated else ""
    payload = f"{scope_key}|{request.get_full_path()}|{max_updated_iso}|{queryset.count()}"
    return f"\"{hashlib.md5(payload.encode('utf-8')).hexdigest()}\""


def _cache_headers_for_queryset(request, queryset, scope_key):
    etag = _build_list_cache_tag(request, queryset, scope_key)
    max_updated = queryset.aggregate(max_updated=Max("updated_at")).get("max_updated")
    last_modified = max_updated.strftime("%a, %d %b %Y %H:%M:%S GMT") if max_updated else None
    return etag, last_modified


def _set_list_cache_headers(response, etag, last_modified):
    response["Cache-Control"] = "private, max-age=30, stale-while-revalidate=30"
    response["ETag"] = etag
    if last_modified:
        response["Last-Modified"] = last_modified


def _current_dhaka_month():
    return timezone.now().astimezone(ZoneInfo("Asia/Dhaka")).month


def _current_dhaka_year():
    return timezone.now().astimezone(ZoneInfo("Asia/Dhaka")).year


def _current_dhaka_date():
    return timezone.now().astimezone(ZoneInfo("Asia/Dhaka")).date()


def _month_start_date(reporting_year, month_number):
    return date(reporting_year, month_number, 1)


def _default_month_close_date(reporting_year, month_number):
    return date(reporting_year, month_number, monthrange(reporting_year, month_number)[1])


def _normalize_month_access_close_date(raw_value, reporting_year, month_number):
    if isinstance(raw_value, date):
        return raw_value

    if raw_value in (None, ""):
        return _default_month_close_date(reporting_year, month_number)

    value = str(raw_value).strip()
    if not value:
        return _default_month_close_date(reporting_year, month_number)

    if "T" in value:
        value = value.split("T", 1)[0]

    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(value, pattern).date()
        except ValueError:
            continue

    return _default_month_close_date(reporting_year, month_number)


def _month_field_to_number(field_name):
    try:
        return MONTH_COLUMNS.index(field_name) + 1
    except ValueError:
        return None


def _profile_local_approval_scope_q(profile):
    if not profile:
        return None

    multi_village_ids = list(profile.micro_villages.values_list("id", flat=True)) if hasattr(profile, "micro_villages") else []
    if multi_village_ids:
        return Q(local_record__village_id__in=multi_village_ids)

    if getattr(profile, "micro_village_id", None):
        return Q(local_record__village_id=profile.micro_village_id)
    if getattr(profile, "micro_union_id", None):
        ward_no = (getattr(profile, "micro_ward_no", "") or "").strip()
        if ward_no:
            return Q(
                local_record__village__union_id=profile.micro_union_id,
                local_record__village__ward_no__iexact=ward_no,
            )
        return Q(local_record__village__union_id=profile.micro_union_id)
    if getattr(profile, "micro_upazila_id", None):
        return Q(local_record__village__union__upazila_id=profile.micro_upazila_id)
    if getattr(profile, "micro_district_id", None):
        return Q(local_record__village__union__upazila__district_id=profile.micro_district_id)

    return None


def _get_rejected_month_fields(record_type, instance):
    filters = {
        "reporting_year": instance.reporting_year,
        "status": MonthlyApproval.STATUS_REJECTED,
    }
    if record_type == MonthlyApproval.RECORD_TYPE_NON_LOCAL:
        filters["non_local_record"] = instance
    else:
        filters["local_record"] = instance

    rejected_months = MonthlyApproval.objects.filter(**filters).values_list("month", flat=True)
    return {
        MONTH_COLUMNS[month_number - 1]
        for month_number in rejected_months
        if 1 <= month_number <= len(MONTH_COLUMNS)
    }


def _resolve_month_access_district_id(district_id=None, district_name=None):
    if district_id not in (None, ""):
        try:
            return int(district_id)
        except (TypeError, ValueError):
            return None
    district_name = (district_name or "").strip()
    if not district_name:
        return None
    return District.objects.filter(name__iexact=district_name).values_list("id", flat=True).first()


def _local_record_district_id(instance):
    try:
        return instance.village.union.upazila.district_id
    except AttributeError:
        return (
            Village.objects.filter(pk=getattr(instance, "village_id", None))
            .values_list("union__upazila__district_id", flat=True)
            .first()
        )


def _non_local_record_district_id(instance):
    return _resolve_month_access_district_id(district_name=getattr(instance, "district_or_state", ""))


def _get_month_access_lookup(reporting_year, district_id=None, district_name=None):
    today = _current_dhaka_date()
    resolved_district_id = _resolve_month_access_district_id(district_id=district_id, district_name=district_name)
    setting_filter = Q(district__isnull=True)
    if resolved_district_id:
        setting_filter |= Q(district_id=resolved_district_id)
    settings_by_month = {
        (setting.month, setting.district_id): setting
        for setting in MonthAccessSetting.objects.filter(
            setting_filter,
            reporting_year=reporting_year,
        ).only(
            "month",
            "close_date",
            "district_id",
        )
    }
    lookup = {}

    for month_number in range(1, len(MONTH_COLUMNS) + 1):
        month_start = _month_start_date(reporting_year, month_number)
        setting = (
            settings_by_month.get((month_number, resolved_district_id))
            if resolved_district_id
            else None
        ) or settings_by_month.get((month_number, None))
        close_date = setting.close_date if setting and setting.close_date else _default_month_close_date(
            reporting_year,
            month_number,
        )
        lookup[month_number] = month_start <= today <= close_date

    return lookup


def _get_open_month_fields(reporting_year, district_id=None, district_name=None):
    access_lookup = _get_month_access_lookup(
        reporting_year,
        district_id=district_id,
        district_name=district_name,
    )
    return {
        MONTH_COLUMNS[month_number - 1]
        for month_number, is_open in access_lookup.items()
        if is_open and 1 <= month_number <= len(MONTH_COLUMNS)
    }


def _sync_monthly_approvals_for_user_submission(record_type, instance, changed_month_fields):
    for field in changed_month_fields:
        month_number = _month_field_to_number(field)
        if month_number is None:
            continue

        month_value = getattr(instance, field, 0) or 0
        filters = {
            "reporting_year": instance.reporting_year,
            "month": month_number,
        }
        if record_type == MonthlyApproval.RECORD_TYPE_NON_LOCAL:
            filters["non_local_record"] = instance
        else:
            filters["local_record"] = instance

        if month_value > 0:
            MonthlyApproval.objects.update_or_create(
                defaults={
                    "status": MonthlyApproval.STATUS_PENDING,
                    "approved_by": None,
                    "approved_at": None,
                },
                **filters,
            )
        else:
            MonthlyApproval.objects.filter(**filters).delete()


def _build_dashboard_payload():
    role_rows = MalariaUserRole.objects.filter(role=MalariaUserRole.ROLE_SPO).count()
    village_count = Village.objects.count()
    assignment_count = LocalRecord.objects.count()
    approvals = MonthlyApproval.objects.filter(status=MonthlyApproval.STATUS_APPROVED)
    approved_months = approvals.count()

    approved_lookup = {
        (approval.local_record_id, approval.month)
        for approval in approvals.filter(local_record__isnull=False).only("local_record_id", "month")
    }
    unapproved_with_data = 0
    for row in LocalRecord.objects.only("id", *MONTH_COLUMNS):
        for index, column in enumerate(MONTH_COLUMNS, start=1):
            if getattr(row, column) > 0 and (row.id, index) not in approved_lookup:
                unapproved_with_data += 1

    return {
        "totalSKs": role_rows,
        "totalVillages": village_count,
        "totalAssignments": assignment_count,
        "approvedMonths": approved_months,
        "unapprovedWithData": unapproved_with_data,
    }


def _get_user_profile(user):
    try:
        return user.profile
    except Exception:
        return None


def _get_micro_scope(profile, multi_villages):
    if not profile:
        return "unassigned"
    if multi_villages:
        return "multi_village"
    if getattr(profile, "micro_village_id", None):
        return "village"
    if getattr(profile, "micro_union_id", None):
        ward_no = (getattr(profile, "micro_ward_no", "") or "").strip()
        if ward_no:
            return "ward"
        return "union"
    if getattr(profile, "micro_upazila_id", None):
        return "upazila"
    if getattr(profile, "micro_district_id", None):
        return "district"
    return "unassigned"


def _collect_profile_district_names(profile, multi_villages):
    names = set()
    if not profile:
        return names

    if getattr(profile, "micro_district", None):
        names.add(profile.micro_district.name)
    if getattr(profile, "micro_upazila", None) and getattr(profile.micro_upazila, "district", None):
        names.add(profile.micro_upazila.district.name)
    if getattr(profile, "micro_union", None):
        upazila = getattr(profile.micro_union, "upazila", None)
        district = getattr(upazila, "district", None)
        if district:
            names.add(district.name)
    if getattr(profile, "micro_village", None):
        union = getattr(profile.micro_village, "union", None)
        upazila = getattr(union, "upazila", None)
        district = getattr(upazila, "district", None)
        if district:
            names.add(district.name)

    for village in multi_villages:
        union = getattr(village, "union", None)
        upazila = getattr(union, "upazila", None)
        district = getattr(upazila, "district", None)
        if district:
            names.add(district.name)

    return names


def _normalize_micro_header(value):
    return " ".join(str(value or "").replace("\n", " ").split()).strip().lower()


def _normalize_micro_key(value):
    return " ".join(str(value or "").split()).strip().lower()


def _normalize_micro_export_format(value):
    normalized = (value or "xlsx").strip().lower()
    if normalized in MICROSTATIFICATION_EXPORT_FORMATS:
        return normalized
    return None


def _get_micro_request_export_format(request):
    return _normalize_micro_export_format(
        request.query_params.get("export_format")
        or request.query_params.get("format")
    )


def _normalize_micro_district_name(value):
    normalized = (value or "").strip()
    if not normalized:
        return ""

    if normalized.lower() in MICROSTATIFICATION_ALL_DISTRICTS_VALUES:
        return MICROSTATIFICATION_ALL_DISTRICTS_TOKEN

    for district_name in MICROSTATIFICATION_TEMPLATE_FILES:
        if district_name.lower() == normalized.lower():
            return district_name

    return normalized


def _resolve_micro_template_path(district_name):
    template_name = MICROSTATIFICATION_TEMPLATE_FILES.get(district_name)
    if not template_name:
        return None
    return MICROSTATIFICATION_TEMPLATE_DIR / template_name


def _build_micro_designation_lookup():
    designation_lookup = {}
    role_label_map = {8: "SPO", 9: "SPO", 11: "SPO"}
    users = (
        User.objects.filter(role__in=role_label_map.keys())
        .select_related("profile")
    )

    for user in users:
        designation = role_label_map.get(getattr(user, "role", None), "")
        if not designation:
            continue

        profile = getattr(user, "profile", None)
        candidate_names = {
            user.username,
            build_full_name(user),
            getattr(profile, "micro_sk_shw_name", "") if profile else "",
        }
        for candidate in candidate_names:
            normalized = _normalize_micro_key(candidate)
            if normalized:
                designation_lookup[normalized] = designation

    return designation_lookup


def _get_micro_designation(name, designation_lookup):
    normalized = _normalize_micro_key(name)
    if not normalized:
        return None
    return designation_lookup.get(normalized) or None


def _copy_micro_template_row_style(ws, template_row_idx, target_row_idx):
    source_dimension = ws.row_dimensions.get(template_row_idx)
    if source_dimension and source_dimension.height is not None:
        ws.row_dimensions[target_row_idx].height = source_dimension.height

    for column_idx in range(1, ws.max_column + 1):
        source_cell = ws.cell(template_row_idx, column_idx)
        target_cell = ws.cell(target_row_idx, column_idx)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def _clear_micro_template_sheet_rows(ws):
    for row in ws.iter_rows(
        min_row=MICROSTATIFICATION_TEMPLATE_DATA_START_ROW,
        max_row=ws.max_row,
        max_col=ws.max_column,
    ):
        for cell in row:
            cell.value = None
            cell._value = None
            cell._hyperlink = None
            cell.comment = None


def _normalize_micro_cell_value(value):
    if value is None:
        return None

    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None

    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        return value

    if hasattr(value, "is_finite"):
        try:
            return value if value.is_finite() else None
        except TypeError:
            return value

    return value


def _get_micro_record_value(record, field_name):
    if not record:
        return None
    value = getattr(record, field_name, None)
    return _normalize_micro_cell_value(value)


def _get_micro_template_cell_value(header, sequence, village, local_record, designation_lookup):
    normalized = _normalize_micro_header(header)
    union = village.union
    upazila = union.upazila if union else None
    district = upazila.district if upazila else None
    district_name = district.name if district else ""
    upazila_name = upazila.name if upazila else ""
    assignment = get_local_record_display_details(
        local_record,
        village=village,
        designation_lookup=designation_lookup,
    )

    if normalized == "sl":
        return sequence
    if normalized == "sq no":
        return sequence
    if normalized == "country":
        return "Bangladesh"
    if normalized == "division":
        return MICROSTATIFICATION_DIVISION_BY_DISTRICT.get(district_name)
    if normalized == "district":
        return _normalize_micro_cell_value(district_name)
    if normalized == "upazila":
        return _normalize_micro_cell_value(upazila_name)
    if normalized == "union":
        return _normalize_micro_cell_value(union.name if union else None)
    if normalized == "ward no":
        return _normalize_micro_cell_value(village.ward_no)
    if normalized == "name of sk/shw":
        return _normalize_micro_cell_value(assignment["name"])
    if normalized == "desig.":
        return _normalize_micro_cell_value(assignment["designation"])
    if normalized == "name of ss":
        return _normalize_micro_cell_value(assignment["ss_name"])
    if normalized == "village name (english)":
        return _normalize_micro_cell_value(village.name)
    if normalized == "village name (bangla)":
        return _normalize_micro_cell_value(village.name_bn)
    if normalized == "village code":
        return _normalize_micro_cell_value(village.village_code)
    if normalized == "latitude":
        return _normalize_micro_cell_value(
            float(village.latitude) if village.latitude is not None else None
        )
    if normalized == "longitute":
        return _normalize_micro_cell_value(
            float(village.longitude) if village.longitude is not None else None
        )
    if normalized == "population":
        if local_record and local_record.population is not None:
            return _normalize_micro_cell_value(local_record.population)
        return _normalize_micro_cell_value(village.population)
    if normalized == "hh number":
        return _get_micro_record_value(local_record, "hh")
    if "active llins" in normalized:
        for field_name in ITN_FIELDS:
            year = field_name.split("_")[-1]
            if year in normalized:
                return _get_micro_record_value(local_record, field_name)
        return None
    month_lookup = {
        "january": "jan_cases",
        "february": "feb_cases",
        "march": "mar_cases",
        "april": "apr_cases",
        "may": "may_cases",
        "june": "jun_cases",
        "july": "jul_cases",
        "august": "aug_cases",
        "september": "sep_cases",
        "october": "oct_cases",
        "november": "nov_cases",
        "december": "dec_cases",
    }
    if normalized in month_lookup:
        return _get_micro_record_value(local_record, month_lookup[normalized])
    if "name of mmw" in normalized:
        return _normalize_micro_cell_value(village.mmw_hp_chwc_name)
    if "distance" in normalized and "upazila office" in normalized:
        return _normalize_micro_cell_value(
            float(village.distance_from_upazila_office_km)
            if village.distance_from_upazila_office_km is not None
            else None
        )
    if "name of border" in normalized:
        return _normalize_micro_cell_value(village.bordering_country_name)
    if normalized.startswith("others activities"):
        return _normalize_micro_cell_value(village.other_activities)
    return None


def _populate_micro_template_sheet(ws, villages, local_record_lookup, designation_lookup):
    original_max_row = ws.max_row
    _clear_micro_template_sheet_rows(ws)
    headers = [
        ws.cell(MICROSTATIFICATION_TEMPLATE_HEADER_ROW, column_idx).value
        for column_idx in range(1, ws.max_column + 1)
    ]

    for sequence, village in enumerate(villages, start=1):
        target_row = MICROSTATIFICATION_TEMPLATE_DATA_START_ROW + sequence - 1
        if target_row > original_max_row:
            _copy_micro_template_row_style(
                ws,
                MICROSTATIFICATION_TEMPLATE_DATA_START_ROW,
                target_row,
            )

        local_record = local_record_lookup.get(village.id)
        for column_idx, header in enumerate(headers, start=1):
            ws.cell(target_row, column_idx).value = _get_micro_template_cell_value(
                header,
                sequence,
                village,
                local_record,
                designation_lookup,
            )

    last_data_row = MICROSTATIFICATION_TEMPLATE_DATA_START_ROW + max(len(villages) - 1, 0)
    formula_range_pattern = re.compile(r"([A-Z]+)(\d+):([A-Z]+)(\d+)")
    for row in ws.iter_rows(min_row=1, max_row=MICROSTATIFICATION_TEMPLATE_HEADER_ROW - 1):
        for cell in row:
            if cell.data_type != "f" or not isinstance(cell.value, str):
                continue

            def _replace_formula_range(match):
                start_col, start_row, end_col, _end_row = match.groups()
                if int(start_row) < MICROSTATIFICATION_TEMPLATE_DATA_START_ROW:
                    return match.group(0)
                return f"{start_col}{start_row}:{end_col}{last_data_row}"

            cell.value = formula_range_pattern.sub(_replace_formula_range, cell.value)


def _populate_micro_template_workbook(workbook, district_upazila_rows, local_record_lookup, designation_lookup):
    normalized_rows = {
        _normalize_micro_key(upazila_name): villages
        for upazila_name, villages in district_upazila_rows.items()
    }
    sheet_lookup = {
        _normalize_micro_key(sheet.title): sheet
        for sheet in workbook.worksheets
    }

    for sheet in workbook.worksheets:
        villages = normalized_rows.get(_normalize_micro_key(sheet.title), [])
        _populate_micro_template_sheet(sheet, villages, local_record_lookup, designation_lookup)

    for upazila_name, villages in district_upazila_rows.items():
        if _normalize_micro_key(upazila_name) in sheet_lookup:
            continue
        sheet = workbook.copy_worksheet(workbook.worksheets[0])
        sheet.title = _unique_micro_sheet_title(workbook, upazila_name)
        _populate_micro_template_sheet(sheet, villages, local_record_lookup, designation_lookup)

    return workbook


def _get_micro_export_filename(district_name, export_format="xlsx"):
    normalized_format = _normalize_micro_export_format(export_format) or "xlsx"
    extension = "csv" if normalized_format == "csv" else "xlsx"
    return f"microstatification_data_{slugify(district_name)}.{extension}"


def _store_micro_export_debug_copy(filename, export_bytes):
    try:
        MICROSTATIFICATION_EXPORT_DEBUG_DIR.mkdir(parents=True, exist_ok=True)
        export_path = MICROSTATIFICATION_EXPORT_DEBUG_DIR / filename
        if export_path.suffix.lower() == ".xlsx":
            district_slug = Path(filename).stem.removeprefix("microstatification_data_")
            for stale_path in MICROSTATIFICATION_EXPORT_DEBUG_DIR.glob(
                f"microstatification_data_{district_slug}_*.xlsx"
            ):
                stale_path.unlink(missing_ok=True)

        temp_path = export_path.with_suffix(".tmp")
        temp_path.write_bytes(export_bytes)
        temp_path.replace(export_path)
        return export_path
    except OSError:
        # Debug copy is optional; never fail user downloads if filesystem permissions block it.
        return None


def _get_micro_export_cache_path(district_name, export_format="xlsx"):
    return MICROSTATIFICATION_EXPORT_DEBUG_DIR / _get_micro_export_filename(
        district_name,
        export_format=export_format,
    )


def _get_micro_export_source_updated_at(district_name):
    timestamps = []

    template_path = _resolve_micro_template_path(district_name)
    if template_path and template_path.exists():
        timestamps.append(
            datetime.fromtimestamp(template_path.stat().st_mtime, tz=ZoneInfo("UTC"))
        )

    village_updated_at = Village.objects.filter(
        union__upazila__district__name=district_name
    ).aggregate(last=Max("updated_at"))["last"]
    if village_updated_at is not None:
        timestamps.append(village_updated_at)

    local_record_updated_at = LocalRecord.objects.filter(
        village__union__upazila__district__name=district_name
    ).aggregate(last=Max("updated_at"))["last"]
    if local_record_updated_at is not None:
        timestamps.append(local_record_updated_at)

    non_local_record_updated_at = NonLocalRecord.objects.filter(
        district_or_state__iexact=district_name
    ).aggregate(last=Max("updated_at"))["last"]
    if non_local_record_updated_at is not None:
        timestamps.append(non_local_record_updated_at)

    upload_updated_at = MicrostatificationDataUpload.objects.filter(
        district=district_name
    ).aggregate(last=Max("updated_at"))["last"]
    if upload_updated_at is not None:
        timestamps.append(upload_updated_at)

    if not timestamps:
        return None

    return max(timestamps)


def _load_cached_micro_export_payload(
    district_name,
    export_format="xlsx",
    ignore_freshness=False,
):
    normalized_format = _normalize_micro_export_format(export_format)
    if not normalized_format:
        return None

    cache_path = _get_micro_export_cache_path(
        district_name,
        export_format=normalized_format,
    )
    if not cache_path.exists():
        return None

    try:
        cached_at = datetime.fromtimestamp(cache_path.stat().st_mtime, tz=ZoneInfo("UTC"))
    except OSError:
        return None

    if not ignore_freshness:
        max_age_seconds = MICROSTATIFICATION_EXPORT_CACHE_TTL_SECONDS
        if max_age_seconds > 0:
            cache_age = (timezone.now().astimezone(ZoneInfo("UTC")) - cached_at).total_seconds()
            if cache_age > max_age_seconds:
                return None

        source_updated_at = _get_micro_export_source_updated_at(district_name)
        if source_updated_at is not None and cached_at < source_updated_at:
            return None

    try:
        export_bytes = cache_path.read_bytes()
    except OSError:
        return None

    if normalized_format == "xlsx" and not export_bytes.startswith(b"PK\x03\x04"):
        return None

    return {
        "filename": cache_path.name,
        "bytes": export_bytes,
        "debug_path": str(cache_path),
        "content_type": (
            "text/csv; charset=utf-8"
            if normalized_format == "csv"
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        "export_format": normalized_format,
        "cache_hit": True,
    }


def _build_micro_download_ticket(user_id, district_name, export_format="xlsx"):
    return signing.dumps(
        {
            "user_id": user_id,
            "district": district_name,
            "format": export_format,
        },
        salt=MICROSTATIFICATION_DOWNLOAD_TICKET_SALT,
    )


def _load_micro_download_ticket(ticket):
    return signing.loads(
        ticket,
        salt=MICROSTATIFICATION_DOWNLOAD_TICKET_SALT,
        max_age=MICROSTATIFICATION_DOWNLOAD_TICKET_MAX_AGE_SECONDS,
    )


def _micro_download_error_response(detail, status_code):
    payload = json.dumps({"detail": detail}, ensure_ascii=False)
    return HttpResponse(
        payload,
        status=status_code,
        content_type="application/json",
    )


def _normalize_micro_xlsx_bytes(workbook_bytes, template_path=None):
    output = BytesIO()
    typed_blank_cell_pattern = re.compile(r'(<c\b[^>]*?)\s+t="[^"]+"([^>]*)/>')
    empty_value_pattern = re.compile(r"<v\s*/>")
    sheet_namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    relationship_namespace = "http://schemas.openxmlformats.org/package/2006/relationships"
    content_type_namespace = "http://schemas.openxmlformats.org/package/2006/content-types"
    shared_string_type = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings"
    )
    shared_string_content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"
    )
    ET.register_namespace("", sheet_namespace)
    ET.register_namespace("", relationship_namespace)
    ET.register_namespace("", content_type_namespace)

    namespace_map = {
        "main": sheet_namespace,
        "rel": relationship_namespace,
        "ct": content_type_namespace,
    }

    def _serialize_xml(root, namespace=None):
        content = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        if namespace:
            content = content.replace(
                f'xmlns:ns0="{namespace}"'.encode("utf-8"),
                f'xmlns="{namespace}"'.encode("utf-8"),
                1,
            )
            content = content.replace(b"<ns0:", b"<")
            content = content.replace(b"</ns0:", b"</")
        return content

    shared_string_lookup = {}
    shared_string_values = []
    shared_string_count = 0
    member_contents = {}
    member_order = []

    with zipfile.ZipFile(BytesIO(workbook_bytes), "r") as source_zip:
        source_members = source_zip.infolist()
        for member in source_members:
            member_order.append(member.filename)
            content = source_zip.read(member.filename)

            if member.filename.startswith("xl/worksheets/sheet") and member.filename.endswith(".xml"):
                text = content.decode("utf-8")
                text = typed_blank_cell_pattern.sub(r"\1\2/>", text)
                text = empty_value_pattern.sub("", text)
                root = ET.fromstring(text)

                for cell in root.findall(".//main:c", namespace_map):
                    if cell.get("t") != "inlineStr":
                        continue

                    inline_text = "".join(
                        text_part or ""
                        for text_part in cell.itertext()
                    )
                    shared_string_count += 1
                    if inline_text not in shared_string_lookup:
                        shared_string_lookup[inline_text] = len(shared_string_values)
                        shared_string_values.append(inline_text)

                    cell.set("t", "s")
                    for child in list(cell):
                        cell.remove(child)

                    value_node = ET.SubElement(cell, f"{{{sheet_namespace}}}v")
                    value_node.text = str(shared_string_lookup[inline_text])

                content = _serialize_xml(root, namespace=sheet_namespace)

            member_contents[member.filename] = (member, content)

    if shared_string_values and "xl/_rels/workbook.xml.rels" in member_contents:
        member, content = member_contents["xl/_rels/workbook.xml.rels"]
        root = ET.fromstring(content)
        for rel in root.findall("rel:Relationship", namespace_map):
            target = rel.get("Target", "")
            if target.startswith("/xl/"):
                rel.set("Target", target.removeprefix("/xl/"))
        has_shared_strings = any(
            rel.get("Type") == shared_string_type
            for rel in root.findall("rel:Relationship", namespace_map)
        )
        if not has_shared_strings:
            relationship_ids = [
                int(match.group(1))
                for rel in root.findall("rel:Relationship", namespace_map)
                for match in [re.match(r"rId(\d+)$", rel.get("Id", ""))]
                if match
            ]
            next_id = max(relationship_ids, default=0) + 1
            ET.SubElement(
                root,
                f"{{{relationship_namespace}}}Relationship",
                {
                    "Type": shared_string_type,
                    "Target": "sharedStrings.xml",
                    "Id": f"rId{next_id}",
                },
            )
        member_contents["xl/_rels/workbook.xml.rels"] = (
            member,
            _serialize_xml(root, namespace=relationship_namespace),
        )

    if shared_string_values and "[Content_Types].xml" in member_contents:
        member, content = member_contents["[Content_Types].xml"]
        root = ET.fromstring(content)
        has_shared_strings_override = any(
            override.get("PartName") == "/xl/sharedStrings.xml"
            for override in root.findall("ct:Override", namespace_map)
        )
        if not has_shared_strings_override:
            ET.SubElement(
                root,
                f"{{{content_type_namespace}}}Override",
                {
                    "PartName": "/xl/sharedStrings.xml",
                    "ContentType": shared_string_content_type,
                },
            )
        member_contents["[Content_Types].xml"] = (
            member,
            _serialize_xml(root, namespace=content_type_namespace),
        )

    if template_path and Path(template_path).exists():
        with zipfile.ZipFile(template_path, "r") as template_zip:
            for template_member_name in (
                "_rels/.rels",
                "docProps/app.xml",
                "xl/workbook.xml",
                "xl/_rels/workbook.xml.rels",
            ):
                if template_member_name not in member_contents:
                    continue
                member, _content = member_contents[template_member_name]
                member_contents[template_member_name] = (
                    member,
                    template_zip.read(template_member_name),
                )

    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target_zip:
        for filename in member_order:
            member, content = member_contents[filename]
            target_zip.writestr(member, content)

        if shared_string_values:
            shared_root = ET.Element(
                f"{{{sheet_namespace}}}sst",
                {
                    "count": str(shared_string_count),
                    "uniqueCount": str(len(shared_string_values)),
                },
            )
            for value in shared_string_values:
                si_node = ET.SubElement(shared_root, f"{{{sheet_namespace}}}si")
                text_node = ET.SubElement(si_node, f"{{{sheet_namespace}}}t")
                if value != value.strip():
                    text_node.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
                text_node.text = value

            target_zip.writestr(
                "xl/sharedStrings.xml",
                _serialize_xml(shared_root, namespace=sheet_namespace),
            )

    return output.getvalue()


def _unique_micro_sheet_title(workbook, desired_title):
    base_title = desired_title[:31] or "Sheet"
    if base_title not in workbook.sheetnames:
        return base_title

    suffix = 2
    while True:
        suffix_text = f" ({suffix})"
        candidate = f"{base_title[:31 - len(suffix_text)]}{suffix_text}"
        if candidate not in workbook.sheetnames:
            return candidate
        suffix += 1


def _copy_micro_template_sheet_to_workbook(source_ws, target_workbook):
    target_ws = target_workbook.create_sheet(
        title=_unique_micro_sheet_title(target_workbook, source_ws.title)
    )

    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws.cell(
                row=source_cell.row,
                column=source_cell.column,
                value=source_cell.value,
            )
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale
    target_ws.sheet_format.defaultRowHeight = source_ws.sheet_format.defaultRowHeight
    target_ws.sheet_properties = copy(source_ws.sheet_properties)
    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.page_setup = copy(source_ws.page_setup)
    target_ws.print_options = copy(source_ws.print_options)

    for row_idx, row_dimension in source_ws.row_dimensions.items():
        target_row_dimension = target_ws.row_dimensions[row_idx]
        target_row_dimension.height = row_dimension.height
        target_row_dimension.hidden = row_dimension.hidden

    for key, column_dimension in source_ws.column_dimensions.items():
        target_column_dimension = target_ws.column_dimensions[key]
        target_column_dimension.width = column_dimension.width
        target_column_dimension.hidden = column_dimension.hidden
        target_column_dimension.bestFit = column_dimension.bestFit

    return target_ws


def _build_microstatification_dashboard_payload(request_user, district_id=None):
    district_name = None
    if district_id:
        district_name = District.objects.filter(pk=district_id).values_list("name", flat=True).first()

    managed_user_filter = (
        Q(role__in={8, 9, 10, 11})
        | Q(role=4, created_by=request_user)
    )
    if is_malaria_dm(request_user):
        managed_user_filter = _dm_spo_user_scope_q(request_user)
    elif district_id:
        managed_user_filter &= Q(profile__micro_district_id=district_id)
    managed_users = list(
        User.objects.filter(managed_user_filter)
        .select_related(
            "profile",
            "profile__micro_district",
            "profile__micro_upazila__district",
            "profile__micro_union__upazila__district",
            "profile__micro_village__union__upazila__district",
        )
        .prefetch_related("profile__micro_villages__union__upazila__district")
        .order_by("username")
    )

    role_counts = {label: 0 for label in MICRO_DASHBOARD_ROLE_LABELS.values()}
    scope_counts = {key: 0 for key, _ in MICRO_DASHBOARD_SCOPE_LABELS}
    district_user_counts = defaultdict(int)
    assigned_users = 0
    workflow_users = 0

    for user in managed_users:
        role_label = MICRO_DASHBOARD_ROLE_LABELS.get(getattr(user, "role", 4), "User")
        role_counts[role_label] += 1

        profile = _get_user_profile(user)
        multi_villages = list(profile.micro_villages.all()) if profile else []
        scope_key = _get_micro_scope(profile, multi_villages)
        scope_counts[scope_key] += 1

        if scope_key != "unassigned":
            assigned_users += 1

        if profile and getattr(profile, "data_collection_type", "") == "microstatification":
            workflow_users += 1

        for district_name in _collect_profile_district_names(profile, multi_villages):
            district_user_counts[district_name] += 1

    district_stats = []
    village_queryset = Village.objects
    if district_id:
        village_queryset = village_queryset.filter(union__upazila__district_id=district_id)
    village_rows = (
        village_queryset.values(
            "union__upazila__district_id",
            "union__upazila__district__name",
        )
        .annotate(
            villages=Count("id"),
            population=Coalesce(Sum("population"), 0),
            last_updated=Max("updated_at"),
        )
        .order_by("-villages", "union__upazila__district__name")
    )

    upload_queryset = MicrostatificationDataUpload.objects
    if district_name:
        upload_queryset = upload_queryset.filter(district__iexact=district_name)
    upload_rows = {
        row["district"]: row
        for row in upload_queryset.values("district")
        .annotate(
            uploads=Count("id"),
            villages_created=Coalesce(Sum("villages_created"), 0),
            villages_updated=Coalesce(Sum("villages_updated"), 0),
            last_upload_at=Max("created_at"),
        )
        .order_by("district")
    }

    total_population = 0
    total_villages = 0
    latest_village_update = None
    for row in village_rows:
        district_name = row["union__upazila__district__name"]
        upload_data = upload_rows.get(district_name, {})
        population = int(row["population"] or 0)
        villages = int(row["villages"] or 0)
        total_population += population
        total_villages += villages

        last_updated = row.get("last_updated")
        if last_updated and (latest_village_update is None or last_updated > latest_village_update):
            latest_village_update = last_updated

        district_stats.append(
            {
                "id": row["union__upazila__district_id"],
                "name": district_name,
                "villages": villages,
                "population": population,
                "assigned_users": district_user_counts.get(district_name, 0),
                "uploads": int(upload_data.get("uploads") or 0),
                "villages_touched": int(upload_data.get("villages_created") or 0)
                + int(upload_data.get("villages_updated") or 0),
                "last_upload_at": upload_data.get("last_upload_at"),
                "last_updated": last_updated,
            }
        )

    recent_upload_query = MicrostatificationDataUpload.objects.select_related("uploaded_by")
    if district_name:
        recent_upload_query = recent_upload_query.filter(district__iexact=district_name)
    recent_upload_qs = list(recent_upload_query.order_by("-created_at")[:6])
    recent_uploads = MicrostatificationDataUploadSerializer(recent_upload_qs, many=True).data
    upload_trend = [
        {
            "label": timezone.localtime(upload.created_at).strftime("%d %b"),
            "district": upload.district,
            "villages_touched": int(upload.villages_created or 0) + int(upload.villages_updated or 0),
            "uploads": 1,
            "created_at": upload.created_at,
        }
        for upload in reversed(recent_upload_qs)
    ]

    role_breakdown = [
        {
            "key": label.lower().replace(" ", "_"),
            "label": label,
            "count": role_counts[label],
        }
        for label in dict.fromkeys(MICRO_DASHBOARD_ROLE_LABELS.values())
    ]
    scope_breakdown = [
        {
            "key": key,
            "label": label,
            "count": scope_counts[key],
        }
        for key, label in MICRO_DASHBOARD_SCOPE_LABELS
    ]

    total_uploads = sum(item["uploads"] for item in district_stats)
    districts_with_users = sum(1 for item in district_stats if item["assigned_users"] > 0)
    last_upload_at = recent_upload_qs[0].created_at if recent_upload_qs else None
    assignment_coverage = round((assigned_users / len(managed_users)) * 100, 1) if managed_users else 0
    reporting_year = _current_dhaka_year()
    local_record_queryset = LocalRecord.objects.filter(reporting_year=reporting_year)
    non_local_record_queryset = NonLocalRecord.objects.filter(reporting_year=reporting_year)
    if district_id:
        local_record_queryset = local_record_queryset.filter(village__union__upazila__district_id=district_id)
        if district_name:
            non_local_record_queryset = non_local_record_queryset.filter(district_or_state__iexact=district_name)
        else:
            non_local_record_queryset = non_local_record_queryset.none()
    local_case_totals = local_record_queryset.aggregate(
        **{field: Coalesce(Sum(field), 0) for field in MONTH_COLUMNS}
    )
    non_local_case_totals = non_local_record_queryset.aggregate(
        **{field: Coalesce(Sum(field), 0) for field in MONTH_COLUMNS}
    )
    case_presence_q = Q()
    for field in MONTH_COLUMNS:
        case_presence_q |= Q(**{f"{field}__gt": 0})
    reported_case_rows = (
        local_record_queryset
        .filter(case_presence_q)
        .count()
    )
    total_reported_cases = (
        sum(int(local_case_totals.get(field) or 0) for field in MONTH_COLUMNS)
        + sum(int(non_local_case_totals.get(field) or 0) for field in MONTH_COLUMNS)
    )
    monthly_district_updates = [
        {
            "key": field_name,
            "label": label,
            "short_label": short_label,
            "count": int(local_case_totals.get(field_name) or 0),
        }
        for field_name, label, short_label in MONTH_COLUMN_LABELS
    ]

    return {
        "generated_at": timezone.now(),
        "totals": {
            "managed_users": len(managed_users),
            "workflow_users": workflow_users,
            "assigned_users": assigned_users,
            "pending_users": max(len(managed_users) - assigned_users, 0),
            "assignment_coverage": assignment_coverage,
            "districts": len(district_stats),
            "districts_with_users": districts_with_users,
            "villages": total_villages,
            "population": total_population,
            "reported_cases": total_reported_cases,
            "reported_case_rows": reported_case_rows,
            "monthly_district_updates": monthly_district_updates,
            "reporting_year": reporting_year,
            "uploads": total_uploads,
            "last_upload_at": last_upload_at,
            "last_village_update": latest_village_update,
        },
        "role_breakdown": role_breakdown,
        "scope_breakdown": scope_breakdown,
        "district_stats": district_stats,
        "upload_trend": upload_trend,
        "recent_uploads": recent_uploads,
    }


def _profile_local_scope_q(profile):
    if not profile:
        return None

    multi_village_ids = list(profile.micro_villages.values_list("id", flat=True)) if hasattr(profile, "micro_villages") else []
    if multi_village_ids:
        return Q(village_id__in=multi_village_ids)

    if getattr(profile, "micro_village_id", None):
        return Q(village_id=profile.micro_village_id)
    if getattr(profile, "micro_union_id", None):
        ward_no = (getattr(profile, "micro_ward_no", "") or "").strip()
        if ward_no:
            return Q(village__union_id=profile.micro_union_id, village__ward_no__iexact=ward_no)
        return Q(village__union_id=profile.micro_union_id)
    if getattr(profile, "micro_upazila_id", None):
        return Q(village__union__upazila_id=profile.micro_upazila_id)
    if getattr(profile, "micro_district_id", None):
        return Q(village__union__upazila__district_id=profile.micro_district_id)

    return None


def _record_within_profile_scope(profile, instance):
    if not profile or not instance:
        return False

    if hasattr(profile, "micro_villages") and profile.micro_villages.filter(id=instance.village_id).exists():
        return True

    if getattr(profile, "micro_village_id", None):
        return instance.village_id == profile.micro_village_id

    if getattr(profile, "micro_union_id", None):
        if getattr(instance.village, "union_id", None) != profile.micro_union_id:
            return False
        ward_no = (getattr(profile, "micro_ward_no", "") or "").strip()
        if ward_no:
            instance_ward = (getattr(instance.village, "ward_no", "") or "").strip()
            return instance_ward.lower() == ward_no.lower()
        return True

    if getattr(profile, "micro_upazila_id", None):
        upazila_id = getattr(getattr(instance.village, "union", None), "upazila_id", None)
        return upazila_id == profile.micro_upazila_id

    if getattr(profile, "micro_district_id", None):
        district_id = getattr(getattr(getattr(instance.village, "union", None), "upazila", None), "district_id", None)
        return district_id == profile.micro_district_id

    return False


def _village_within_profile_scope(profile, instance):
    if not profile or not instance:
        return False

    if hasattr(profile, "micro_villages") and profile.micro_villages.filter(id=instance.id).exists():
        return True

    if getattr(profile, "micro_village_id", None):
        return instance.id == profile.micro_village_id

    if getattr(profile, "micro_union_id", None):
        if getattr(instance, "union_id", None) != profile.micro_union_id:
            return False
        ward_no = (getattr(profile, "micro_ward_no", "") or "").strip()
        if ward_no:
            instance_ward = (getattr(instance, "ward_no", "") or "").strip()
            return instance_ward.lower() == ward_no.lower()
        return True

    if getattr(profile, "micro_upazila_id", None):
        upazila_id = getattr(getattr(instance, "union", None), "upazila_id", None)
        return upazila_id == profile.micro_upazila_id

    if getattr(profile, "micro_district_id", None):
        district_id = getattr(getattr(getattr(instance, "union", None), "upazila", None), "district_id", None)
        return district_id == profile.micro_district_id

    return False


def _union_within_profile_scope(profile, union):
    """True if a new village may be created under this Union for the given UserProfile scope."""
    if not profile or not union:
        return False

    multi_village_ids = (
        list(profile.micro_villages.values_list("id", flat=True)) if hasattr(profile, "micro_villages") else []
    )
    if multi_village_ids:
        union_ids = set(
            Village.objects.filter(id__in=multi_village_ids).values_list("union_id", flat=True).distinct()
        )
        return union.id in union_ids

    if getattr(profile, "micro_village_id", None):
        try:
            anchor = Village.objects.only("union_id").get(pk=profile.micro_village_id)
        except Village.DoesNotExist:
            return False
        return anchor.union_id == union.id

    if getattr(profile, "micro_union_id", None):
        return union.id == profile.micro_union_id

    if getattr(profile, "micro_upazila_id", None):
        return getattr(union, "upazila_id", None) == profile.micro_upazila_id

    if getattr(profile, "micro_district_id", None):
        district_id = getattr(getattr(union, "upazila", None), "district_id", None)
        return district_id == profile.micro_district_id

    return False


def _village_allowed_for_profile(profile, village):
    if not profile or not village:
        return False
    return _village_within_profile_scope(profile, village)


def _local_record_in_dm_district(user, instance):
    did = get_dm_district_id(user)
    if not did or not instance:
        return False
    village = getattr(instance, "village", None)
    if not village:
        return False
    union = getattr(village, "union", None)
    upazila = getattr(union, "upazila", None) if union else None
    district_id = getattr(upazila, "district_id", None) if upazila else None
    return district_id == did


def _non_local_in_dm_district(user, instance):
    did = get_dm_district_id(user)
    if not did or not instance:
        return False
    try:
        dname = District.objects.only("name").get(pk=did).name
    except District.DoesNotExist:
        return False
    country = (getattr(instance, "country", "") or "").strip().lower()
    dist = (getattr(instance, "district_or_state", "") or "").strip().lower()
    return country == "bangladesh" and dist == (dname or "").strip().lower()


def _village_in_dm_district(user, village):
    did = get_dm_district_id(user)
    if not did or not village:
        return False
    union = getattr(village, "union", None)
    upazila = getattr(union, "upazila", None) if union else None
    district_id = getattr(upazila, "district_id", None) if upazila else None
    return district_id == did


def _user_managed_in_dm_district(actor, target_user):
    """True if target SPO is assigned inside the DM district or was created by the DM."""
    if not is_malaria_dm(actor) or not target_user:
        return True
    did = get_dm_district_id(actor)
    if not did:
        return False
    if get_malaria_role(target_user) != "spo":
        return False
    if getattr(target_user, "created_by_id", None) == getattr(actor, "id", None):
        return True
    profile = getattr(target_user, "profile", None)
    if not profile:
        return False
    if getattr(profile, "micro_district_id", None) == did:
        return True
    if getattr(profile, "micro_upazila_id", None):
        return Upazila.objects.filter(pk=profile.micro_upazila_id, district_id=did).exists()
    if getattr(profile, "micro_union_id", None):
        return Union.objects.filter(pk=profile.micro_union_id, upazila__district_id=did).exists()
    if getattr(profile, "micro_village_id", None):
        return Village.objects.filter(pk=profile.micro_village_id, union__upazila__district_id=did).exists()
    if hasattr(profile, "micro_villages") and profile.micro_villages.filter(union__upazila__district_id=did).exists():
        return True
    return False


def _assert_dm_nonlocal_sk_user(actor, sk_user):
    """
    District managers may only bind non-local rows to SPO users in their district.
    """
    if not is_malaria_dm(actor):
        return
    if not sk_user:
        raise PermissionError(
            "Specify sk_user: an SPO in your district must own this non-local record."
        )
    if not _user_managed_in_dm_district(actor, sk_user):
        raise PermissionError(
            "You can only assign non-local records to users in your district."
        )
    if get_malaria_role(sk_user) != "spo":
        raise PermissionError(
            "You can only assign non-local records to SPO users in your district."
        )


def _apply_privileged_new_user_profile(actor, user, malaria_role):
    """When a DM creates a malaria user, pin them to the DM's district for SPO."""
    if not is_malaria_dm(actor) or malaria_role != MalariaUserRole.ROLE_SPO:
        return
    did = get_dm_district_id(actor)
    if not did:
        return
    profile = getattr(user, "profile", None)
    if not profile:
        return
    profile.micro_district_id = did
    profile.micro_role = "spo"
    profile.save(update_fields=["micro_district", "micro_role"])


def _assert_privileged_role_assignment(actor, target_user, new_role):
    """Raise PermissionError if DM tries to assign outside allowed roles or users."""
    if is_malaria_global_admin(actor):
        return
    if not is_malaria_dm(actor):
        raise PermissionError("Admin access required.")
    if new_role not in {MalariaUserRole.ROLE_SPO}:
        raise PermissionError("District managers may only assign the SPO role.")
    if not _user_managed_in_dm_district(actor, target_user):
        raise PermissionError("You can only manage users in your district.")


def _assert_privileged_user_create(actor, malaria_role):
    if is_malaria_global_admin(actor):
        return
    if not is_malaria_dm(actor):
        raise PermissionError("Admin access required.")
    if malaria_role != MalariaUserRole.ROLE_SPO:
        raise PermissionError("District managers may only create SPO users.")
    if not get_dm_district_id(actor):
        raise PermissionError("Your account must have an assigned district to create users.")


def _ensure_local_record_editable(user, instance, validated_data):
    if is_malaria_global_admin(user):
        return
    if is_malaria_dm(user):
        if _local_record_in_dm_district(user, instance):
            return
        raise PermissionError("You can only edit local records in your district.")

    profile = getattr(user, "profile", None)
    can_access = instance.sk_user_id == user.id or _record_within_profile_scope(profile, instance)
    if not can_access:
        raise PermissionError("You can only edit local records within your assigned scope.")

    rejected_month_fields = _get_rejected_month_fields(MonthlyApproval.RECORD_TYPE_LOCAL, instance)
    changed_month_fields = [
        field
        for field in MONTH_COLUMNS
        if field in validated_data and validated_data[field] != getattr(instance, field)
    ]
    allowed_month_fields = (
        _get_open_month_fields(
            instance.reporting_year,
            district_id=_local_record_district_id(instance),
        )
        | rejected_month_fields
    )
    disallowed = [field for field in changed_month_fields if field not in allowed_month_fields]
    if disallowed:
        raise ValueError("You can only edit months that are still open or months returned as not approved.")


def _ensure_non_local_record_editable(user, instance, validated_data):
    if is_malaria_global_admin(user):
        return
    if is_malaria_dm(user):
        if not _non_local_in_dm_district(user, instance):
            raise PermissionError("You can only edit non-local records in your district.")
        new_sk = validated_data.get("sk_user")
        if new_sk is not None and new_sk.pk != instance.sk_user_id:
            _assert_dm_nonlocal_sk_user(user, new_sk)
        return

    if instance.sk_user_id != user.id:
        raise PermissionError("You can only edit your own non-local records.")

    rejected_month_fields = _get_rejected_month_fields(MonthlyApproval.RECORD_TYPE_NON_LOCAL, instance)
    changed_month_fields = [
        field
        for field in MONTH_COLUMNS
        if field in validated_data and validated_data[field] != getattr(instance, field)
    ]
    disallowed = [
        field
        for field in changed_month_fields
        if field not in (
            _get_open_month_fields(
                instance.reporting_year,
                district_id=_non_local_record_district_id(instance),
            )
            | rejected_month_fields
        )
    ]
    if disallowed:
        raise ValueError("You can only edit months that are still open or months returned as not approved.")


class MalariaSessionView(APIView):
    permission_classes = [IsAuthenticated, HasMalariaAccess]

    def get(self, request):
        role = get_malaria_role(request.user)
        profile = getattr(request.user, "profile", None)
        micro_role = getattr(profile, "micro_role", "") if profile else ""
        micro_villages = list(profile.micro_villages.values_list("id", flat=True)) if profile and hasattr(profile, "micro_villages") else []
        payload = {
            "user": _serialize_user(request.user),
            "profile": {
                "user_id": request.user.id,
                "full_name": build_full_name(request.user),
                "email": request.user.email or request.user.username,
                "micro_role": micro_role or None,
                "micro_district": getattr(profile, "micro_district_id", None) if profile else None,
                "micro_upazila": getattr(profile, "micro_upazila_id", None) if profile else None,
                "micro_union": getattr(profile, "micro_union_id", None) if profile else None,
                "micro_village": getattr(profile, "micro_village_id", None) if profile else None,
                "micro_villages": micro_villages,
                "micro_ward_no": getattr(profile, "micro_ward_no", "") if profile else "",
                "micro_sk_shw_name": getattr(profile, "micro_sk_shw_name", "") if profile else "",
                "micro_designation": getattr(profile, "micro_designation", "") if profile else "",
                "micro_ss_name": getattr(profile, "micro_ss_name", "") if profile else "",
            },
            "role": role,
        }
        return Response(MalariaSessionSerializer(payload).data)


class MalariaMeView(MalariaSessionView):
    pass


MALARIA_GRID_LAYOUT_KEYS = frozenset(
    {
        MalariaGridColumnLayout.GRID_LOCAL_RECORDS,
        MalariaGridColumnLayout.GRID_NON_LOCAL_RECORDS,
    }
)


class MalariaGridColumnLayoutView(APIView):
    """GET: shared column layout for all malaria users. PUT: malaria admin publishes layout."""

    def get_permissions(self):
        if self.request.method == "PUT":
            return [IsAuthenticated(), IsMalariaAdmin()]
        return [IsAuthenticated(), HasMalariaAccess()]

    def get(self, request, grid_key):
        if grid_key not in MALARIA_GRID_LAYOUT_KEYS:
            return Response({"detail": "Unknown grid_key."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            obj = MalariaGridColumnLayout.objects.get(grid_key=grid_key)
        except MalariaGridColumnLayout.DoesNotExist:
            return Response(
                {
                    "grid_key": grid_key,
                    "column_widths": {},
                    "is_expanded_to_header_width": False,
                    "updated_at": None,
                }
            )
        return Response(MalariaGridColumnLayoutResponseSerializer(obj).data)

    def put(self, request, grid_key):
        if grid_key not in MALARIA_GRID_LAYOUT_KEYS:
            return Response({"detail": "Unknown grid_key."}, status=status.HTTP_400_BAD_REQUEST)
        serializer = MalariaGridColumnLayoutPayloadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        obj, _ = MalariaGridColumnLayout.objects.update_or_create(
            grid_key=grid_key,
            defaults={
                "column_widths": serializer.validated_data["column_widths"],
                "is_expanded_to_header_width": serializer.validated_data["is_expanded_to_header_width"],
            },
        )
        return Response(MalariaGridColumnLayoutResponseSerializer(obj).data)


class MalariaLogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        return Response({"message": "Logged out"})


class ProfilesView(APIView):
    permission_classes = [IsAuthenticated, HasMalariaAccess]

    def get(self, request):
        users = User.objects.filter(Q(malaria_role__isnull=False) | Q(role__in=(8, 9, 10, 11)))
        if is_malaria_dm(request.user):
            users = users.filter(_dm_spo_user_scope_q(request.user))
        users = (
            users.distinct()
            .values("id", "first_name", "last_name", "email", "username")
            .order_by("username")
        )
        users = _apply_filters(users, request, exact_fields=["id"], in_fields=["id"])

        single_user_id = request.query_params.get("user_id")
        if single_user_id:
            users = users.filter(id=single_user_id)

        many_user_ids = request.query_params.get("user_id__in")
        if many_user_ids:
            users = users.filter(id__in=[item for item in many_user_ids.split(",") if item])

        malaria_role = request.query_params.get("malaria_role")
        if malaria_role:
            users = users.filter(malaria_role__role=malaria_role)

        data = [
            {
                "user_id": user["id"],
                "full_name": _build_full_name_from_row(user),
                "email": user["email"] or user["username"],
            }
            for user in users
        ]
        serializer = ProfileSerializer(data, many=True)
        return Response(serializer.data)


class UserRolesView(APIView):
    permission_classes = [IsAuthenticated, HasMalariaAccess]

    def get(self, request):
        explicit_roles = list(MalariaUserRole.objects.values("user_id", "role", "created_at", "updated_at").order_by("user__username"))
        payload = [
            {
                "user_id": role["user_id"],
                "role": role["role"],
                "created_at": role["created_at"],
                "updated_at": role["updated_at"],
            }
            for role in explicit_roles
        ]

        role_filter = request.query_params.get("role")
        if role_filter:
            payload = [row for row in payload if row["role"] == role_filter]

        user_id = request.query_params.get("user_id")
        if user_id not in (None, ""):
            payload = [row for row in payload if str(row["user_id"]) == user_id]

        user_ids = request.query_params.get("user_id__in")
        if user_ids:
            allowed_ids = {item.strip() for item in user_ids.split(",") if item.strip()}
            payload = [row for row in payload if str(row["user_id"]) in allowed_ids]

        if is_malaria_dm(request.user):
            allowed_ids = set(
                User.objects.filter(_dm_spo_user_scope_q(request.user)).values_list("id", flat=True)
            )
            payload = [row for row in payload if row["user_id"] in allowed_ids]

        if _count_requested(request):
            return Response({"data": payload, "count": len(payload)})
        return Response(payload)

    def post(self, request):
        if not has_malaria_privileged_access(request.user):
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        user_id = request.data.get("user_id")
        role = request.data.get("role")
        if not user_id or not role:
            return Response({"detail": "user_id and role are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(pk=user_id)
        except User.DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

        try:
            _assert_privileged_role_assignment(request.user, user, role)
        except PermissionError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)

        role_obj, _ = MalariaUserRole.objects.update_or_create(user=user, defaults={"role": role})
        return Response(MalariaUserRoleSerializer(role_obj).data, status=status.HTTP_201_CREATED)


class MalariaUserCreateView(APIView):
    permission_classes = [IsAuthenticated, IsMalariaPrivileged]

    def post(self, request):
        serializer = MalariaUserCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated = serializer.validated_data

        try:
            _assert_privileged_user_create(request.user, validated["role"])
        except PermissionError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)

        email = validated["email"].strip().lower()
        username = email
        if User.objects.filter(username=username).exists():
            return Response({"detail": "A user with this email already exists."}, status=status.HTTP_400_BAD_REQUEST)

        full_name = validated["full_name"].strip()
        name_parts = [part for part in full_name.split(" ", 1) if part]
        first_name = name_parts[0] if name_parts else ""
        last_name = name_parts[1] if len(name_parts) > 1 else ""

        user = User(
            username=username,
            email=email,
            first_name=first_name,
            last_name=last_name,
            role=11 if validated["role"] == MalariaUserRole.ROLE_SPO else 4,
            is_staff=False,
            created_by=request.user,
        )
        user.set_password(validated["password"])
        user.save()
        MalariaUserRole.objects.update_or_create(user=user, defaults={"role": validated["role"]})
        _apply_privileged_new_user_profile(request.user, user, validated["role"])

        return Response({"user": _serialize_user(user)}, status=status.HTTP_201_CREATED)


class MalariaUsersView(APIView):
    permission_classes = [IsAuthenticated, IsMalariaPrivileged]

    def get(self, request):
        roles = MalariaUserRole.objects.select_related("user").order_by("user__username")
        role_filter = request.query_params.get("role")
        if role_filter in {"admin", "dm", "spo"}:
            roles = roles.filter(role=role_filter)

        user_id = request.query_params.get("user_id")
        if user_id not in (None, ""):
            roles = roles.filter(user_id=user_id)

        if is_malaria_dm(request.user):
            roles = roles.filter(user_id__in=User.objects.filter(_dm_spo_user_scope_q(request.user)).values("id"))

        payload = []
        for role_obj in roles:
            user = role_obj.user
            payload.append(
                {
                    "id": user.id,
                    "username": user.username,
                    "email": user.email or "",
                    "full_name": build_full_name(user),
                    "role": role_obj.role,
                    "created_at": role_obj.created_at,
                    "updated_at": role_obj.updated_at,
                }
            )

        if _count_requested(request):
            return Response({"data": payload, "count": len(payload)})
        return Response(payload)

    def post(self, request):
        serializer = MalariaUserCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated = serializer.validated_data

        try:
            _assert_privileged_user_create(request.user, validated["role"])
        except PermissionError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)

        email = validated["email"].strip().lower()
        username = email
        if User.objects.filter(username=username).exists():
            return Response({"detail": "A user with this email already exists."}, status=status.HTTP_400_BAD_REQUEST)

        full_name = validated["full_name"].strip()
        name_parts = [part for part in full_name.split(" ", 1) if part]
        first_name = name_parts[0] if name_parts else ""
        last_name = name_parts[1] if len(name_parts) > 1 else ""

        user = User(
            username=username,
            email=email,
            first_name=first_name,
            last_name=last_name,
            role=11 if validated["role"] == MalariaUserRole.ROLE_SPO else 4,
            is_staff=False,
            created_by=request.user,
        )
        user.set_password(validated["password"])
        user.save()
        role_obj, _ = MalariaUserRole.objects.update_or_create(user=user, defaults={"role": validated["role"]})
        _apply_privileged_new_user_profile(request.user, user, validated["role"])
        return Response(
            {
                "id": user.id,
                "username": user.username,
                "email": user.email or "",
                "full_name": build_full_name(user),
                "role": role_obj.role,
                "created_at": role_obj.created_at,
                "updated_at": role_obj.updated_at,
            },
            status=status.HTTP_201_CREATED,
        )


class MalariaMasterDataView(APIView):
    permission_classes = [IsAuthenticated, HasMalariaAccess]

    def get(self, request):
        include_villages = str(request.query_params.get("include_villages", "1")).strip().lower() not in {"0", "false", "no"}
        did = get_dm_district_id(request.user) if is_malaria_dm(request.user) else None
        districts_qs = District.objects.all()
        upazilas_qs = Upazila.objects.select_related("district").all()
        unions_qs = Union.objects.select_related("upazila", "upazila__district").all()
        villages_qs = Village.objects.select_related("union", "union__upazila", "union__upazila__district").all()
        if did:
            districts_qs = districts_qs.filter(pk=did)
            upazilas_qs = upazilas_qs.filter(district_id=did)
            unions_qs = unions_qs.filter(upazila__district_id=did)
            villages_qs = villages_qs.filter(union__upazila__district_id=did)
        payload = {
            "districts": DistrictSerializer(districts_qs, many=True).data,
            "upazilas": UpazilaSerializer(upazilas_qs, many=True).data,
            "unions": UnionSerializer(unions_qs, many=True).data,
            "villages": (
                VillageSerializer(
                    villages_qs,
                    many=True,
                ).data
                if include_villages
                else []
            ),
        }
        return Response(payload)


class MalariaAdminWriteMixin:
    def get_permissions(self):
        if self.action in {"update", "partial_update", "destroy"}:
            return [IsAuthenticated(), IsMalariaPrivileged()]
        if self.action == "create":
            return [IsAuthenticated(), HasMalariaAccess()]
        return [IsAuthenticated(), HasMalariaAccess()]


class RequestedFieldsViewMixin:
    def get_serializer(self, *args, **kwargs):
        requested_fields = _get_requested_serializer_fields(self.request)
        if requested_fields and "fields" not in kwargs:
            kwargs["fields"] = requested_fields
        return super().get_serializer(*args, **kwargs)


class DistrictViewSet(RequestedFieldsViewMixin, MalariaAdminWriteMixin, viewsets.ModelViewSet):
    queryset = District.objects.all()
    serializer_class = DistrictSerializer
    pagination_class = None

    def get_permissions(self):
        if self.action == "destroy":
            return [IsAuthenticated(), IsMalariaGlobalAdmin()]
        return super().get_permissions()

    def get_queryset(self):
        queryset = super().get_queryset()
        if is_malaria_dm(self.request.user):
            did = get_dm_district_id(self.request.user)
            if did:
                queryset = queryset.filter(pk=did)
            else:
                queryset = queryset.none()
        return _apply_filters(queryset, self.request, exact_fields=["id", "name"], in_fields=["id"])


class UpazilaViewSet(RequestedFieldsViewMixin, MalariaAdminWriteMixin, viewsets.ModelViewSet):
    queryset = Upazila.objects.select_related("district").all()
    serializer_class = UpazilaSerializer
    pagination_class = None

    def get_permissions(self):
        if self.action == "destroy":
            return [IsAuthenticated(), IsMalariaGlobalAdmin()]
        return super().get_permissions()

    def get_queryset(self):
        queryset = super().get_queryset()
        if is_malaria_dm(self.request.user):
            did = get_dm_district_id(self.request.user)
            if did:
                queryset = queryset.filter(district_id=did)
            else:
                queryset = queryset.none()
        return _apply_filters(queryset, self.request, exact_fields=["id", "district_id"], in_fields=["id"])


class UnionViewSet(RequestedFieldsViewMixin, MalariaAdminWriteMixin, viewsets.ModelViewSet):
    queryset = Union.objects.select_related("upazila", "upazila__district").all()
    serializer_class = UnionSerializer
    pagination_class = None

    def get_permissions(self):
        if self.action == "destroy":
            return [IsAuthenticated(), IsMalariaGlobalAdmin()]
        return super().get_permissions()

    def get_queryset(self):
        queryset = super().get_queryset()
        if is_malaria_dm(self.request.user):
            did = get_dm_district_id(self.request.user)
            if did:
                queryset = queryset.filter(upazila__district_id=did)
            else:
                queryset = queryset.none()
        return _apply_filters(queryset, self.request, exact_fields=["id", "upazila_id"], in_fields=["id"])


class VillageViewSet(RequestedFieldsViewMixin, MalariaAdminWriteMixin, viewsets.ModelViewSet):
    queryset = Village.objects.select_related("union", "union__upazila", "union__upazila__district").all()
    serializer_class = VillageSerializer
    pagination_class = None

    def get_permissions(self):
        if self.action == "destroy":
            return [IsAuthenticated(), IsMalariaPrivileged()]
        return [IsAuthenticated(), HasMalariaAccess()]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        if not has_malaria_privileged_access(request.user):
            profile = getattr(request.user, "profile", None)
            union = serializer.validated_data.get("union")
            if union is None:
                return Response(
                    {"detail": "Union is required to create a village."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if not _union_within_profile_scope(profile, union):
                return Response(
                    {"detail": "You can only create villages within your assigned scope."},
                    status=status.HTTP_403_FORBIDDEN,
                )
            ward_no_profile = (getattr(profile, "micro_ward_no", "") or "").strip()
            if getattr(profile, "micro_union_id", None) and ward_no_profile:
                incoming_ward = (serializer.validated_data.get("ward_no") or "").strip()
                if incoming_ward.lower() != ward_no_profile.lower():
                    return Response(
                        {"detail": "You can only create villages within your assigned ward."},
                        status=status.HTTP_403_FORBIDDEN,
                    )
        elif is_malaria_dm(request.user):
            profile = getattr(request.user, "profile", None)
            union = serializer.validated_data.get("union")
            if union is None or not _union_within_profile_scope(profile, union):
                return Response(
                    {"detail": "You can only create villages within your assigned district."},
                    status=status.HTTP_403_FORBIDDEN,
                )
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = _apply_filters(
            queryset,
            self.request,
            exact_fields=["id", "union_id", "ward_no"],
            in_fields=["id", "union_id", "ward_no"],
        )

        query = (self.request.query_params.get("q") or "").strip()
        if query:
            queryset = queryset.filter(
                Q(name__icontains=query)
                | Q(name_bn__icontains=query)
                | Q(village_code__icontains=query)
                | Q(ward_no__icontains=query)
                | Q(sk_shw_name__icontains=query)
                | Q(ss_name__icontains=query)
            )

        requested_order = (
            self.request.query_params.get("_order")
            or self.request.query_params.get("order")
            or ""
        ).strip()
        has_scope_filters = any(
            (self.request.query_params.get(key) or "").strip()
            for key in ("union_id", "union_id__in", "ward_no", "ward_no__in")
        )
        if not requested_order and has_scope_filters:
            queryset = _apply_micro_village_code_sort(queryset)

        limit = (self.request.query_params.get("limit") or "").strip()
        if limit.isdigit():
            size = max(1, min(int(limit), 200))
            queryset = queryset[:size]

        if is_malaria_dm(self.request.user):
            did = get_dm_district_id(self.request.user)
            if did:
                queryset = queryset.filter(union__upazila__district_id=did)
            else:
                queryset = queryset.none()

        return queryset

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        if is_malaria_dm(request.user) and not _village_in_dm_district(request.user, instance):
            return Response(
                {"detail": "You can only edit villages in your district."},
                status=status.HTTP_403_FORBIDDEN,
            )

        if not has_malaria_privileged_access(request.user):
            profile = getattr(request.user, "profile", None)
            can_access = instance.local_records.filter(sk_user=request.user).exists() or _village_within_profile_scope(profile, instance)
            if not can_access:
                return Response(
                    {"detail": "You can only edit villages within your assigned scope."},
                    status=status.HTTP_403_FORBIDDEN,
                )

            disallowed_fields = sorted(
                set(serializer.validated_data.keys()) - FIELD_USER_VILLAGE_EDITABLE_FIELDS,
            )
            if disallowed_fields:
                return Response(
                    {
                        "detail": (
                            "You can only edit Name of MMW, Health post & CHW(C), "
                            "Village Distance from upazila office (KM), Name of Border with others country, "
                            "and Others Activities (TDA/Dev care)."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        self.perform_update(serializer)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if is_malaria_dm(request.user) and not _village_in_dm_district(request.user, instance):
            return Response(
                {"detail": "You can only delete villages in your district."},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)


class LocalRecordViewSet(RequestedFieldsViewMixin, viewsets.ModelViewSet):
    serializer_class = LocalRecordSerializer
    permission_classes = [IsAuthenticated, HasMalariaAccess]
    pagination_class = None

    def get_queryset(self):
        queryset = LocalRecord.objects.select_related(
            "sk_user",
            "sk_user__profile",
            "village",
            "village__union",
            "village__union__upazila",
            "village__union__upazila__district",
        ).all()

        if has_malaria_privileged_access(self.request.user):
            if is_malaria_dm(self.request.user):
                did = get_dm_district_id(self.request.user)
                if did:
                    queryset = queryset.filter(village__union__upazila__district_id=did)
                else:
                    queryset = queryset.none()
        else:
            profile = getattr(self.request.user, "profile", None)
            scope_q = _profile_local_scope_q(profile)
            if scope_q is not None:
                queryset = queryset.filter(Q(sk_user=self.request.user) | scope_q)
            else:
                queryset = queryset.filter(sk_user=self.request.user)

        reporting_year_value = (self.request.query_params.get("reporting_year") or "").strip().lower()
        exact_fields = ["id", "reporting_year", "sk_user_id", "village_id", "metadata_approval_status"]
        if reporting_year_value == "latest":
            latest_year = queryset.aggregate(max_year=Max("reporting_year")).get("max_year")
            if latest_year is not None:
                queryset = queryset.filter(reporting_year=latest_year)
            exact_fields = ["id", "sk_user_id", "village_id", "metadata_approval_status"]

        queryset = _apply_filters(
            queryset,
            self.request,
            exact_fields=exact_fields,
            in_fields=["id", "sk_user_id"],
        )

        district_id = self.request.query_params.get("district_id")
        if district_id not in (None, ""):
            queryset = queryset.filter(village__union__upazila__district_id=district_id)

        district_name = (self.request.query_params.get("district_name") or "").strip()
        if district_name:
            queryset = queryset.filter(village__union__upazila__district__name__iexact=district_name)

        requested_order = (
            self.request.query_params.get("_order")
            or self.request.query_params.get("order")
            or ""
        ).strip()
        if not requested_order:
            queryset = _apply_micro_local_record_sort(queryset)

        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        etag, last_modified = _cache_headers_for_queryset(request, queryset, "local-records")
        if request.headers.get("If-None-Match") == etag:
            response = Response(status=status.HTTP_304_NOT_MODIFIED)
            _set_list_cache_headers(response, etag, last_modified)
            return response

        page_payload = _paginate_queryset_if_requested(request, queryset)
        if page_payload is not None:
            serializer = self.get_serializer(page_payload["results"], many=True)
            response = Response(
                {
                    "count": page_payload["count"],
                    "next": page_payload["next"],
                    "previous": page_payload["previous"],
                    "results": serializer.data,
                }
            )
            _set_list_cache_headers(response, etag, last_modified)
            return response

        serializer = self.get_serializer(queryset, many=True)
        response = Response(serializer.data)
        _set_list_cache_headers(response, etag, last_modified)
        return response

    def create(self, request, *args, **kwargs):
        payload = request.data.copy()
        if not has_malaria_privileged_access(request.user):
            payload["sk_user"] = request.user.id

        serializer = self.get_serializer(data=payload)
        serializer.is_valid(raise_exception=True)
        validated = serializer.validated_data
        reporting_year = validated.get("reporting_year", timezone.now().year)

        if not has_malaria_privileged_access(request.user):
            profile = getattr(request.user, "profile", None)
            village_obj = validated.get("village")
            if village_obj is not None and _profile_local_scope_q(profile) is not None:
                village_scoped = (
                    Village.objects.select_related("union", "union__upazila", "union__upazila__district")
                    .filter(pk=village_obj.pk)
                    .first()
                )
                if not village_scoped or not _village_allowed_for_profile(profile, village_scoped):
                    return Response(
                        {
                            "detail": "You can only create local records for villages within your assigned scope.",
                        },
                        status=status.HTTP_403_FORBIDDEN,
                    )

        village = validated["village"]
        defaults = validated.copy()
        defaults.setdefault("sk_user", request.user)
        record, created = LocalRecord.objects.update_or_create(
            village=village,
            reporting_year=reporting_year,
            defaults=defaults,
        )
        if not has_malaria_privileged_access(request.user):
            submitted_month_fields = [
                field
                for field in MONTH_COLUMNS
                if validated.get(field, 0)
            ]
            month_fields_for_approval = (
                submitted_month_fields
                if submitted_month_fields
                else sorted(
                    _get_open_month_fields(
                        reporting_year,
                        district_id=_local_record_district_id(record),
                    )
                )
            )
            _sync_monthly_approvals_for_user_submission(
                MonthlyApproval.RECORD_TYPE_LOCAL,
                record,
                month_fields_for_approval,
            )
            meta_sub = parse_local_metadata_from_request(payload)
            record.metadata_approval_status = METADATA_APPROVAL_PENDING
            if local_metadata_submission_nonempty(meta_sub):
                record.metadata_pending = merge_local_metadata_pending(record.metadata_pending, meta_sub)
            record.save(
                update_fields=[
                    "metadata_approval_status",
                    "metadata_pending",
                    "updated_at",
                ]
            )
        out = self.get_serializer(record)
        return Response(out.data, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        month_field_set = set(MONTH_COLUMNS)
        non_month_changes = [
            field
            for field, value in serializer.validated_data.items()
            if field not in month_field_set and value != getattr(instance, field, None)
        ]
        changed_month_fields = [
            field
            for field in MONTH_COLUMNS
            if field in serializer.validated_data and serializer.validated_data[field] != getattr(instance, field)
        ]

        try:
            _ensure_local_record_editable(request.user, instance, serializer.validated_data)
        except PermissionError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        self.perform_update(serializer)
        instance = serializer.instance
        if has_malaria_privileged_access(request.user):
            if "sk_user_designation" in request.data:
                designation = str(request.data.get("sk_user_designation") or "").strip()
                profile = getattr(instance.sk_user, "profile", None)
                if profile is not None:
                    profile.micro_designation = designation
                    profile.save(update_fields=["micro_designation"])
        else:
            meta_sub = parse_local_metadata_from_request(request.data)
            if local_metadata_submission_nonempty(meta_sub):
                instance.metadata_pending = merge_local_metadata_pending(instance.metadata_pending, meta_sub)
                instance.metadata_approval_status = METADATA_APPROVAL_PENDING
                instance.metadata_rejection_note = ""
                instance.save(
                    update_fields=[
                        "metadata_pending",
                        "metadata_approval_status",
                        "metadata_rejection_note",
                        "updated_at",
                    ]
                )
        if not has_malaria_privileged_access(request.user) and (changed_month_fields or non_month_changes):
            month_fields_for_approval = (
                changed_month_fields
                if changed_month_fields
                else sorted(
                    _get_open_month_fields(
                        serializer.instance.reporting_year,
                        district_id=_local_record_district_id(serializer.instance),
                    )
                )
            )
            _sync_monthly_approvals_for_user_submission(
                MonthlyApproval.RECORD_TYPE_LOCAL,
                serializer.instance,
                month_fields_for_approval,
            )
        instance.refresh_from_db()
        return Response(self.get_serializer(instance).data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if is_malaria_global_admin(request.user):
            return super().destroy(request, *args, **kwargs)
        if is_malaria_dm(request.user):
            if _local_record_in_dm_district(request.user, instance):
                return super().destroy(request, *args, **kwargs)
            return Response(
                {"detail": "You can only delete local records in your district."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if instance.sk_user_id == request.user.id:
            return super().destroy(request, *args, **kwargs)
        return Response(
            {"detail": "You can only delete your own local records."},
            status=status.HTTP_403_FORBIDDEN,
        )

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated, IsMalariaPrivileged])
    def approve_metadata(self, request, pk=None):
        instance = self.get_object()
        pending = instance.metadata_pending if isinstance(instance.metadata_pending, dict) else {}
        has_payload = bool(pending.get("village")) or bool(pending.get("profile"))

        # Idempotent: already approved with nothing left to merge (e.g. double-click).
        if instance.metadata_approval_status == METADATA_APPROVAL_APPROVED and not has_payload:
            return Response(self.get_serializer(instance).data)

        if instance.metadata_approval_status == METADATA_APPROVAL_REJECTED:
            return Response(
                {
                    "detail": (
                        "Metadata was rejected. The SPO must save again to submit a new version before you can approve."
                    ),
                    "metadata_approval_status": instance.metadata_approval_status,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # PENDING (with or without payload), or APPROVED with orphaned pending JSON — apply.
        if instance.metadata_approval_status == METADATA_APPROVAL_PENDING or has_payload:
            apply_local_metadata_pending_to_db(instance, request.user)
            instance.refresh_from_db()
            return Response(self.get_serializer(instance).data)

        return Response(
            {
                "detail": (
                    "No pending metadata to approve. Save metadata changes as SPO first, "
                    "or refresh the table."
                ),
                "metadata_approval_status": instance.metadata_approval_status,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated, IsMalariaPrivileged])
    def reject_metadata(self, request, pk=None):
        instance = self.get_object()
        if instance.metadata_approval_status != METADATA_APPROVAL_PENDING:
            return Response(
                {"detail": "Record is not pending metadata approval."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        note = str(request.data.get("note") or "")
        reject_local_metadata(instance, request.user, note)
        instance.refresh_from_db()
        return Response(self.get_serializer(instance).data)


class NonLocalRecordViewSet(RequestedFieldsViewMixin, viewsets.ModelViewSet):
    serializer_class = NonLocalRecordSerializer
    permission_classes = [IsAuthenticated, HasMalariaAccess]
    pagination_class = None

    def get_queryset(self):
        queryset = NonLocalRecord.objects.select_related("sk_user").all()
        if has_malaria_privileged_access(self.request.user):
            if is_malaria_dm(self.request.user):
                did = get_dm_district_id(self.request.user)
                if did:
                    try:
                        dname = District.objects.only("name").get(pk=did).name
                    except District.DoesNotExist:
                        dname = None
                    if dname:
                        queryset = queryset.filter(
                            country__iexact="Bangladesh",
                            district_or_state__iexact=dname,
                        )
                    else:
                        queryset = queryset.none()
                else:
                    queryset = queryset.none()
        else:
            queryset = queryset.filter(sk_user=self.request.user)
        reporting_year_value = (self.request.query_params.get("reporting_year") or "").strip().lower()
        exact_fields = [
            "id",
            "reporting_year",
            "sk_user_id",
            "country",
            "district_or_state",
            "metadata_approval_status",
        ]
        if reporting_year_value == "latest":
            latest_year = queryset.aggregate(max_year=Max("reporting_year")).get("max_year")
            if latest_year is not None:
                queryset = queryset.filter(reporting_year=latest_year)
            exact_fields = ["id", "sk_user_id", "country", "district_or_state", "metadata_approval_status"]
        queryset = _apply_filters(
            queryset,
            self.request,
            exact_fields=exact_fields,
            in_fields=["id", "sk_user_id"],
        )
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        etag, last_modified = _cache_headers_for_queryset(request, queryset, "non-local-records")
        if request.headers.get("If-None-Match") == etag:
            response = Response(status=status.HTTP_304_NOT_MODIFIED)
            _set_list_cache_headers(response, etag, last_modified)
            return response

        page_payload = _paginate_queryset_if_requested(request, queryset)
        if page_payload is not None:
            serializer = self.get_serializer(page_payload["results"], many=True)
            response = Response(
                {
                    "count": page_payload["count"],
                    "next": page_payload["next"],
                    "previous": page_payload["previous"],
                    "results": serializer.data,
                }
            )
            _set_list_cache_headers(response, etag, last_modified)
            return response

        serializer = self.get_serializer(queryset, many=True)
        response = Response(serializer.data)
        _set_list_cache_headers(response, etag, last_modified)
        return response

    def create(self, request, *args, **kwargs):
        payload = request.data.copy()
        # Ensure sk_user is always set to avoid DB NOT NULL failures.
        # Non-admin users are always bound to themselves.
        if not payload.get("sk_user"):
            if has_malaria_privileged_access(request.user) and is_malaria_dm(request.user):
                return Response(
                    {
                        "detail": (
                            "Specify sk_user: an SPO in your district must own this non-local record."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            payload["sk_user"] = request.user.id
        if not has_malaria_privileged_access(request.user):
            payload["sk_user"] = request.user.id
        serializer = self.get_serializer(data=payload)
        serializer.is_valid(raise_exception=True)

        if is_malaria_dm(request.user):
            try:
                _assert_dm_nonlocal_sk_user(
                    request.user,
                    serializer.validated_data.get("sk_user"),
                )
            except PermissionError as exc:
                return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)

        self.perform_create(serializer)
        instance = serializer.instance
        if has_malaria_privileged_access(request.user):
            gm = payload.get("grid_metadata")
            if isinstance(gm, dict):
                instance.grid_metadata = {**(instance.grid_metadata or {}), **gm}
                instance.save(update_fields=["grid_metadata", "updated_at"])
        else:
            submitted_month_fields = [
                field
                for field in MONTH_COLUMNS
                if serializer.validated_data.get(field, 0)
            ]
            month_fields_for_approval = (
                submitted_month_fields
                if submitted_month_fields
                else sorted(
                    _get_open_month_fields(
                        instance.reporting_year,
                        district_id=_non_local_record_district_id(instance),
                    )
                )
            )
            _sync_monthly_approvals_for_user_submission(
                MonthlyApproval.RECORD_TYPE_NON_LOCAL,
                instance,
                month_fields_for_approval,
            )
            meta_sub = filter_nonlocal_metadata_submission(payload.get("metadata_submission"))
            instance.metadata_approval_status = METADATA_APPROVAL_PENDING
            if meta_sub:
                instance.metadata_pending = merge_nonlocal_metadata_pending(instance.metadata_pending, meta_sub)
            instance.save(
                update_fields=[
                    "metadata_approval_status",
                    "metadata_pending",
                    "updated_at",
                ]
            )
        instance.refresh_from_db()
        return Response(self.get_serializer(instance).data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        partial = kwargs.pop("partial", False)
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        month_field_set = set(MONTH_COLUMNS)
        non_month_changes = [
            field
            for field, value in serializer.validated_data.items()
            if field not in month_field_set and value != getattr(instance, field, None)
        ]
        changed_month_fields = [
            field
            for field in MONTH_COLUMNS
            if field in serializer.validated_data and serializer.validated_data[field] != getattr(instance, field)
        ]

        try:
            _ensure_non_local_record_editable(request.user, instance, serializer.validated_data)
        except PermissionError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        self.perform_update(serializer)
        instance = serializer.instance
        if has_malaria_privileged_access(request.user):
            gm = request.data.get("grid_metadata")
            if isinstance(gm, dict):
                instance.grid_metadata = {**(instance.grid_metadata or {}), **gm}
                instance.save(update_fields=["grid_metadata", "updated_at"])
        else:
            meta_sub = filter_nonlocal_metadata_submission(request.data.get("metadata_submission"))
            if meta_sub:
                instance.metadata_pending = merge_nonlocal_metadata_pending(instance.metadata_pending, meta_sub)
                instance.metadata_approval_status = METADATA_APPROVAL_PENDING
                instance.metadata_rejection_note = ""
                instance.save(
                    update_fields=[
                        "metadata_pending",
                        "metadata_approval_status",
                        "metadata_rejection_note",
                        "updated_at",
                    ]
                )
        if not has_malaria_privileged_access(request.user) and (changed_month_fields or non_month_changes):
            month_fields_for_approval = (
                changed_month_fields
                if changed_month_fields
                else sorted(
                    _get_open_month_fields(
                        serializer.instance.reporting_year,
                        district_id=_non_local_record_district_id(serializer.instance),
                    )
                )
            )
            _sync_monthly_approvals_for_user_submission(
                MonthlyApproval.RECORD_TYPE_NON_LOCAL,
                serializer.instance,
                month_fields_for_approval,
            )
        instance.refresh_from_db()
        return Response(self.get_serializer(instance).data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if is_malaria_global_admin(request.user):
            return super().destroy(request, *args, **kwargs)
        if is_malaria_dm(request.user):
            if _non_local_in_dm_district(request.user, instance):
                return super().destroy(request, *args, **kwargs)
            return Response(
                {"detail": "You can only delete non-local records in your district."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if instance.sk_user_id == request.user.id:
            return super().destroy(request, *args, **kwargs)
        return Response(
            {"detail": "You can only delete your own non-local records."},
            status=status.HTTP_403_FORBIDDEN,
        )

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated, IsMalariaPrivileged])
    def approve_metadata(self, request, pk=None):
        instance = self.get_object()
        pending = instance.metadata_pending if isinstance(instance.metadata_pending, dict) else {}
        has_payload = bool(pending)

        if instance.metadata_approval_status == METADATA_APPROVAL_APPROVED and not has_payload:
            return Response(self.get_serializer(instance).data)

        if instance.metadata_approval_status == METADATA_APPROVAL_REJECTED:
            return Response(
                {
                    "detail": (
                        "Metadata was rejected. Wait for the SPO to save again before approving."
                    ),
                    "metadata_approval_status": instance.metadata_approval_status,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if instance.metadata_approval_status == METADATA_APPROVAL_PENDING or has_payload:
            apply_nonlocal_metadata_pending_to_db(instance, request.user)
            instance.refresh_from_db()
            return Response(self.get_serializer(instance).data)

        return Response(
            {
                "detail": "No pending metadata to approve. Refresh the table.",
                "metadata_approval_status": instance.metadata_approval_status,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated, IsMalariaPrivileged])
    def reject_metadata(self, request, pk=None):
        instance = self.get_object()
        if instance.metadata_approval_status != METADATA_APPROVAL_PENDING:
            return Response(
                {"detail": "Record is not pending metadata approval."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        note = str(request.data.get("note") or "")
        reject_nonlocal_metadata(instance, request.user, note)
        instance.refresh_from_db()
        return Response(self.get_serializer(instance).data)


class MonthAccessSettingViewSet(RequestedFieldsViewMixin, viewsets.ModelViewSet):
    serializer_class = MonthAccessSettingSerializer
    pagination_class = None

    def get_permissions(self):
        if self.action in {"list", "retrieve"}:
            return [IsAuthenticated(), HasMalariaAccess()]
        return [IsAuthenticated(), IsMalariaPrivileged()]

    def get_queryset(self):
        queryset = MonthAccessSetting.objects.all()
        if is_malaria_dm(self.request.user):
            did = get_dm_district_id(self.request.user)
            queryset = queryset.filter(district_id=did) if did else queryset.none()
        elif self.request.query_params.get("district_id") in (None, ""):
            queryset = queryset.filter(district__isnull=True)
        return _apply_filters(
            queryset,
            self.request,
            exact_fields=["id", "reporting_year", "month", "district_id"],
            in_fields=["id", "month"],
        )

    def create(self, request, *args, **kwargs):
        payload = request.data.copy()
        district_id = payload.get("district") or payload.get("district_id")

        if is_malaria_dm(request.user):
            district_id = get_dm_district_id(request.user)
            if not district_id:
                return Response(
                    {"detail": "Your account must have an assigned district to change month access."},
                    status=status.HTTP_403_FORBIDDEN,
                )
            payload["district"] = district_id
        elif district_id not in (None, ""):
            payload["district"] = district_id
        else:
            payload["district"] = None

        try:
            reporting_year = int(payload.get("reporting_year"))
            month_number = int(payload.get("month"))
            payload["close_date"] = _normalize_month_access_close_date(
                payload.get("close_date"),
                reporting_year,
                month_number,
            ).isoformat()
            payload["is_open"] = True
        except (TypeError, ValueError):
            pass

        existing_instance = None
        if "reporting_year" in payload and "month" in payload:
            existing_instance = MonthAccessSetting.objects.filter(
                reporting_year=payload.get("reporting_year"),
                month=payload.get("month"),
                district_id=payload.get("district"),
            ).first()

        serializer = self.get_serializer(existing_instance, data=payload)
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        response_serializer = self.get_serializer(instance)
        return Response(
            response_serializer.data,
            status=status.HTTP_200_OK if existing_instance else status.HTTP_201_CREATED,
        )


class MonthlyApprovalViewSet(RequestedFieldsViewMixin, viewsets.ModelViewSet):
    serializer_class = MonthlyApprovalSerializer
    pagination_class = None

    def get_permissions(self):
        if self.action in {"list", "retrieve"}:
            return [IsAuthenticated(), HasMalariaAccess()]
        return [IsAuthenticated(), IsMalariaPrivileged()]

    def get_queryset(self):
        queryset = MonthlyApproval.objects.all()
        if is_malaria_global_admin(self.request.user):
            pass
        elif is_malaria_dm(self.request.user):
            did = get_dm_district_id(self.request.user)
            if did:
                try:
                    dname = District.objects.only("name").get(pk=did).name
                except District.DoesNotExist:
                    dname = None
                if dname:
                    queryset = queryset.filter(
                        Q(local_record__village__union__upazila__district_id=did)
                        | Q(
                            non_local_record__country__iexact="Bangladesh",
                            non_local_record__district_or_state__iexact=dname,
                        )
                    )
                else:
                    queryset = queryset.none()
            else:
                queryset = queryset.none()
        else:
            profile = getattr(self.request.user, "profile", None)
            local_scope_q = _profile_local_approval_scope_q(profile)
            visible_local_q = Q(local_record__sk_user=self.request.user)
            if local_scope_q is not None:
                visible_local_q |= local_scope_q
            queryset = queryset.filter(visible_local_q | Q(non_local_record__sk_user=self.request.user))

        record_type = self.request.query_params.get("record_type")
        if record_type == MonthlyApproval.RECORD_TYPE_LOCAL:
            queryset = queryset.filter(local_record__isnull=False)
        elif record_type == MonthlyApproval.RECORD_TYPE_NON_LOCAL:
            queryset = queryset.filter(non_local_record__isnull=False)
        queryset = _apply_filters(queryset, self.request, exact_fields=["reporting_year", "month", "status"], in_fields=[])

        record_id = self.request.query_params.get("record_id")
        if record_id:
            if record_type == MonthlyApproval.RECORD_TYPE_NON_LOCAL:
                queryset = queryset.filter(non_local_record_id=record_id)
            else:
                queryset = queryset.filter(local_record_id=record_id)
        return queryset

    def create(self, request, *args, **kwargs):
        payload = request.data if isinstance(request.data, list) else [request.data]
        approvals = []

        for item in payload:
            record_type = item.get("record_type", MonthlyApproval.RECORD_TYPE_LOCAL)
            record_id = item.get("record_id")
            month = item.get("month")
            reporting_year = item.get("reporting_year", timezone.now().year)
            status_value = item.get("status", MonthlyApproval.STATUS_APPROVED)

            if not record_id or not month:
                return Response({"detail": "record_id and month are required."}, status=status.HTTP_400_BAD_REQUEST)

            lookup = {
                "reporting_year": reporting_year,
                "month": month,
            }
            defaults = {
                "status": status_value,
                "approved_by": request.user if status_value != MonthlyApproval.STATUS_PENDING else None,
                "approved_at": (
                    item.get("approved_at") or timezone.now()
                    if status_value != MonthlyApproval.STATUS_PENDING
                    else None
                ),
            }

            if record_type == MonthlyApproval.RECORD_TYPE_NON_LOCAL:
                lookup["non_local_record_id"] = record_id
            else:
                lookup["local_record_id"] = record_id

            approval, _ = MonthlyApproval.objects.update_or_create(defaults=defaults, **lookup)
            approvals.append(approval)

        serializer = self.get_serializer(approvals, many=True)
        if len(serializer.data) == 1:
            return Response(serializer.data[0], status=status.HTTP_201_CREATED)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class MalariaDashboardSummaryView(APIView):
    permission_classes = [IsAuthenticated, HasMalariaAccess]

    def get(self, request):
        return Response(_build_dashboard_payload())


class MalariaDashboardView(APIView):
    permission_classes = [IsAuthenticated, HasMalariaAccess]

    def get(self, request):
        return Response(_build_dashboard_payload())


class MicrostatificationDashboardSummaryView(APIView):
    permission_classes = [IsAuthenticated, IsMalariaPrivileged]

    def get(self, request):
        district_id = None
        if is_malaria_dm(request.user):
            district_id = get_dm_district_id(request.user)
            if not district_id:
                return Response(
                    {"detail": "Your account must have an assigned district to view this dashboard."},
                    status=status.HTTP_403_FORBIDDEN,
                )
        return Response(_build_microstatification_dashboard_payload(request.user, district_id=district_id))


class MicrostatificationDataUploadView(APIView):
    """Upload and process microstatification Excel files"""
    permission_classes = [IsAuthenticated, IsMalariaAdmin]

    def get(self, request):
        uploads = MicrostatificationDataUpload.objects.select_related("uploaded_by").all()
        serializer = MicrostatificationDataUploadSerializer(uploads, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = MicrostatificationDataUploadFileSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        excel_file = serializer.validated_data['excel_file']
        district_name = serializer.validated_data['district']
        
        try:
            import openpyxl
            import io
            
            # Read Excel file
            file_content = excel_file.read()
            excel_file.seek(0)
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            
            # Create upload record
            upload = MicrostatificationDataUpload.objects.create(
                district=district_name,
                excel_file=excel_file,
                uploaded_by=request.user,
                parsed_data={},
                upload_note="Processing..."
            )

            sync_result = sync_microstatification_workbook(
                workbook,
                district_name=district_name,
                prune_stale=True,
            )

            upload.parsed_data = sync_result.parsed_data
            upload.districts_created = sync_result.districts_created
            upload.upazilas_created = sync_result.upazilas_created
            upload.unions_created = sync_result.unions_created
            upload.villages_created = sync_result.villages_created
            upload.villages_updated = sync_result.villages_updated
            upload.upload_note = (
                "✓ Upload completed successfully\n"
                f"Districts: {sync_result.districts_created}, "
                f"Upazilas created: {sync_result.upazilas_created}, "
                f"Unions created: {sync_result.unions_created}, "
                f"Villages created: {sync_result.villages_created}, "
                f"Villages updated: {sync_result.villages_updated}, "
                f"Villages deleted: {sync_result.villages_deleted}, "
                f"Local records deleted: {sync_result.local_records_deleted}"
            )
            upload.save()
            
            serializer_out = MicrostatificationDataUploadSerializer(upload)
            return Response(serializer_out.data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {"detail": f"Error processing file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )


class MicrostatificationDataDownloadView(APIView):
    permission_classes = [IsAuthenticated, IsMalariaAdmin]
    renderer_classes = [
        XLSXBinaryRenderer,
        renderers.JSONRenderer,
        JSONFormatCSVRenderer,
    ]

    def _build_district_export_context(self, district_name, designation_lookup=None):
        villages = list(
            _apply_micro_village_code_sort(
                Village.objects.select_related(
                    "union",
                    "union__upazila",
                    "union__upazila__district",
                ).filter(union__upazila__district__name=district_name)
            )
        )

        district_upazila_rows = defaultdict(list)
        for village in villages:
            upazila = village.union.upazila if village.union else None
            if upazila is None:
                continue
            district_upazila_rows[upazila.name].append(village)

        village_ids = [village.id for village in villages]
        reporting_year = _current_dhaka_year()
        record_queryset = LocalRecord.objects.filter(
            reporting_year=reporting_year,
            village_id__in=village_ids,
        )

        # Prefer current-year records, but fall back to the newest available year
        # so exported downloads stay populated with the latest data.
        if village_ids and not record_queryset.exists():
            latest_reporting_year = LocalRecord.objects.filter(
                village_id__in=village_ids,
            ).aggregate(latest=Max("reporting_year"))["latest"]
            if latest_reporting_year is not None:
                reporting_year = latest_reporting_year
                record_queryset = LocalRecord.objects.filter(
                    reporting_year=reporting_year,
                    village_id__in=village_ids,
                )

        local_record_lookup = {
            record.village_id: record
            for record in record_queryset.select_related("sk_user", "sk_user__profile").only(
                "village_id",
                "sk_user__email",
                "sk_user__first_name",
                "sk_user__last_name",
                "sk_user__role",
                "sk_user__username",
                "sk_user__profile__micro_designation",
                "sk_user__profile__micro_role",
                "sk_user__profile__micro_sk_shw_name",
                "sk_user__profile__micro_ss_name",
                "hh",
                "population",
                *ITN_FIELDS,
                *MONTH_COLUMNS,
            )
        }

        return {
            "district_upazila_rows": district_upazila_rows,
            "local_record_lookup": local_record_lookup,
            "designation_lookup": designation_lookup or _build_micro_designation_lookup(),
        }

    def _build_csv_export_bytes(
        self,
        district_upazila_rows,
        local_record_lookup,
        designation_lookup,
    ):
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(MICROSTATIFICATION_CSV_HEADERS)

        sequence = 1
        for villages in district_upazila_rows.values():
            for village in villages:
                local_record = local_record_lookup.get(village.id)
                row_values = [
                    _get_micro_template_cell_value(
                        header,
                        sequence,
                        village,
                        local_record,
                        designation_lookup,
                    )
                    for header in MICROSTATIFICATION_CSV_HEADERS
                ]
                writer.writerow(["" if value is None else value for value in row_values])
                sequence += 1

        return output.getvalue().encode("utf-8-sig")

    def _build_indigenous_export_rows(
        self,
        district_upazila_rows,
        local_record_lookup,
        designation_lookup,
    ):
        rows = []
        sequence = 1
        for villages in district_upazila_rows.values():
            for village in villages:
                local_record = local_record_lookup.get(village.id)
                row_values = [
                    _get_micro_template_cell_value(
                        header,
                        sequence,
                        village,
                        local_record,
                        designation_lookup,
                    )
                    for header in MICROSTATIFICATION_CSV_HEADERS
                ]
                rows.append(["" if value is None else value for value in row_values])
                sequence += 1
        return rows

    def _build_imported_export_rows(self, district_name):
        reporting_year = _current_dhaka_year()
        queryset = NonLocalRecord.objects.filter(
            reporting_year=reporting_year,
            district_or_state__iexact=district_name,
        )

        if not queryset.exists():
            latest_reporting_year = NonLocalRecord.objects.filter(
                district_or_state__iexact=district_name,
            ).aggregate(latest=Max("reporting_year"))["latest"]
            if latest_reporting_year is not None:
                reporting_year = latest_reporting_year
                queryset = NonLocalRecord.objects.filter(
                    reporting_year=reporting_year,
                    district_or_state__iexact=district_name,
                )

        rows = []
        for sequence, record in enumerate(
            queryset.order_by("country", "upazila_or_township", "union_name", "village_name"),
            start=1,
        ):
            rows.append(
                [
                    sequence,  # SL
                    record.country or "",
                    "",  # Division
                    district_name,
                    record.upazila_or_township or "",
                    record.union_name or "",
                    "",  # Ward No
                    "",  # Name of SK/SHW
                    "",  # Desig.
                    "",  # Name of SS
                    record.village_name or "",  # Village Name (English)
                    "",  # Village Name (Bangla)
                    "",  # Village Code
                    "",  # Latitude
                    "",  # Longitude
                    "",  # Population
                    "",  # HH Number
                    "",  # 2026 (Active LLINs)
                    "",  # 2025 (Active LLINs)
                    "",  # 2024 (Active LLINs)
                    record.jan_cases or "",
                    record.feb_cases or "",
                    record.mar_cases or "",
                    record.apr_cases or "",
                    record.may_cases or "",
                    record.jun_cases or "",
                    record.jul_cases or "",
                    record.aug_cases or "",
                    record.sep_cases or "",
                    record.oct_cases or "",
                    record.nov_cases or "",
                    record.dec_cases or "",
                    "",  # Name of MMW, Health post & CHW(C)
                    "",  # Village Distance from upazila office (KM)
                    "",  # Name of Border with others country
                    "",  # Others Activities (TDA/Dev care)
                ]
            )
        return rows

    def _build_combined_csv_export_bytes(self, indigenous_rows, imported_rows):
        output = StringIO()
        writer = csv.writer(output)
        combined_headers = ["Dataset", *MICROSTATIFICATION_CSV_HEADERS]
        writer.writerow(combined_headers)

        for row_values in indigenous_rows:
            writer.writerow(["Indigenous", *row_values])
        for row_values in imported_rows:
            writer.writerow(["Imported", *row_values])

        return output.getvalue().encode("utf-8-sig")

    def _build_combined_xlsx_export_bytes(self, indigenous_rows, imported_rows):
        import openpyxl

        workbook = openpyxl.Workbook()
        indigenous_sheet = workbook.active
        indigenous_sheet.title = "Indigenous"
        imported_sheet = workbook.create_sheet("Imported")

        indigenous_sheet.append(MICROSTATIFICATION_CSV_HEADERS)
        for row_values in indigenous_rows:
            indigenous_sheet.append(row_values)

        imported_sheet.append(MICROSTATIFICATION_CSV_HEADERS)
        for row_values in imported_rows:
            imported_sheet.append(row_values)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _build_export_file(
        self,
        district_name,
        export_format="xlsx",
        store_debug_copy=True,
        designation_lookup=None,
        use_cache=True,
        validate_xlsx=True,
    ):
        import openpyxl

        normalized_format = _normalize_micro_export_format(export_format)
        if not normalized_format:
            return Response(
                {
                    "detail": (
                        f"Unsupported export format: {export_format}. "
                        "Supported formats are xlsx and csv."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not district_name or district_name == MICROSTATIFICATION_ALL_DISTRICTS_TOKEN:
            return Response(
                {
                    "detail": (
                        "Please select a district. "
                        "Template-formatted microstatification exports are generated one district at a time."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if district_name not in MICROSTATIFICATION_TEMPLATE_FILES:
            return Response(
                {"detail": f"Unsupported district: {district_name}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if use_cache:
            cached_payload = _load_cached_micro_export_payload(
                district_name,
                export_format=normalized_format,
            )
            if cached_payload is not None:
                return cached_payload

        context = self._build_district_export_context(
            district_name,
            designation_lookup=designation_lookup,
        )
        district_upazila_rows = context["district_upazila_rows"]
        local_record_lookup = context["local_record_lookup"]
        resolved_designation_lookup = context["designation_lookup"]

        indigenous_rows = self._build_indigenous_export_rows(
            district_upazila_rows,
            local_record_lookup,
            resolved_designation_lookup,
        )
        imported_rows = self._build_imported_export_rows(district_name)

        if normalized_format == "csv":
            filename = _get_micro_export_filename(
                district_name,
                export_format=normalized_format,
            )
            export_bytes = self._build_combined_csv_export_bytes(
                indigenous_rows,
                imported_rows,
            )
            export_debug_path = None
            if store_debug_copy:
                export_debug_path = _store_micro_export_debug_copy(filename, export_bytes)
            return {
                "filename": filename,
                "bytes": export_bytes,
                "debug_path": str(export_debug_path) if export_debug_path else None,
                "content_type": "text/csv; charset=utf-8",
                "export_format": normalized_format,
            }

        export_bytes = self._build_combined_xlsx_export_bytes(
            indigenous_rows,
            imported_rows,
        )

        if validate_xlsx:
            # Validate generated workbook bytes before returning the file.
            # This prevents partially written or structurally invalid files from being served.
            try:
                openpyxl.load_workbook(BytesIO(export_bytes), data_only=False)
            except Exception as exc:
                return Response(
                    {"detail": f"Failed to generate a valid XLSX export: {exc}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        filename = _get_micro_export_filename(
            district_name,
            export_format=normalized_format,
        )
        export_debug_path = None
        if store_debug_copy:
            export_debug_path = _store_micro_export_debug_copy(filename, export_bytes)

        return {
            "filename": filename,
            "bytes": export_bytes,
            "debug_path": str(export_debug_path) if export_debug_path else None,
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "export_format": normalized_format,
        }

    def _build_all_district_zip_response(self, export_format):
        zip_buffer = TemporaryFile()
        try:
            designation_lookup = _build_micro_designation_lookup()
            zip_compression = (
                zipfile.ZIP_STORED
                if export_format == "xlsx"
                else zipfile.ZIP_DEFLATED
            )
            zip_kwargs = {
                "compression": zip_compression,
                "allowZip64": True,
            }
            if export_format != "xlsx":
                zip_kwargs["compresslevel"] = 1
            with zipfile.ZipFile(
                zip_buffer,
                "w",
                **zip_kwargs,
            ) as archive:
                for district_name in MICROSTATIFICATION_TEMPLATE_FILES:
                    export_payload = _load_cached_micro_export_payload(
                        district_name,
                        export_format=export_format,
                        ignore_freshness=True,
                    )
                    if export_payload is None:
                        export_payload = self._build_export_file(
                            district_name,
                            export_format=export_format,
                            store_debug_copy=True,
                            designation_lookup=designation_lookup,
                            use_cache=True,
                            validate_xlsx=False,
                        )
                    if isinstance(export_payload, Response):
                        detail = "Unable to generate export file."
                        payload = getattr(export_payload, "data", None)
                        if isinstance(payload, dict):
                            detail = payload.get("detail") or detail
                        elif isinstance(payload, str):
                            detail = payload or detail
                        zip_buffer.close()
                        return _micro_download_error_response(
                            f"{district_name}: {detail}",
                            export_payload.status_code,
                        )

                    archive.writestr(
                        export_payload["filename"],
                        export_payload["bytes"],
                    )

            zip_buffer.seek(0)
        except Exception as exc:
            zip_buffer.close()
            return _micro_download_error_response(
                f"Unable to generate all-district ZIP export: {exc}",
                status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        timestamp = timezone.now().astimezone(ZoneInfo("Asia/Dhaka")).strftime("%Y%m%d_%H%M%S")
        filename = f"microstatification_all_districts_{export_format}_{timestamp}.zip"
        response = FileResponse(
            zip_buffer,
            as_attachment=True,
            filename=filename,
            content_type="application/zip",
        )
        response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response["Pragma"] = "no-cache"
        response["Expires"] = "0"
        return response

    def _build_download_response(self, district_name, export_format="xlsx"):
        normalized_format = _normalize_micro_export_format(export_format)
        if not normalized_format:
            return _micro_download_error_response(
                "Unsupported format. Use xlsx or csv.",
                status.HTTP_400_BAD_REQUEST,
            )

        if district_name == MICROSTATIFICATION_ALL_DISTRICTS_TOKEN:
            return self._build_all_district_zip_response(normalized_format)

        export_payload = self._build_export_file(
            district_name,
            export_format=normalized_format,
        )
        if isinstance(export_payload, Response):
            detail = "Unable to generate export file."
            payload = getattr(export_payload, "data", None)
            if isinstance(payload, dict):
                detail = payload.get("detail") or detail
            elif isinstance(payload, str):
                detail = payload or detail
            return _micro_download_error_response(detail, export_payload.status_code)

        export_filename = export_payload.get("filename") or _get_micro_export_filename(district_name)
        export_bytes = export_payload.get("bytes") or b""
        export_debug_path = export_payload.get("debug_path")
        content_type = export_payload.get("content_type") or "application/octet-stream"

        if normalized_format == "xlsx" and not export_bytes.startswith(b"PK\x03\x04"):
            return _micro_download_error_response(
                "Generated export is invalid. Please try again.",
                status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        response = HttpResponse(
            export_bytes,
            content_type=content_type,
        )
        response["Content-Disposition"] = f'attachment; filename="{export_filename}"'
        response["Content-Length"] = str(len(export_bytes))
        response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response["Pragma"] = "no-cache"
        response["Expires"] = "0"
        if export_debug_path:
            response["X-Micro-Export-Path"] = str(export_debug_path)
        return response

    def get(self, request):
        district_name = _normalize_micro_district_name(request.query_params.get("district"))
        if not district_name:
            return _micro_download_error_response(
                "Please select a district or choose All Districts for ZIP export.",
                status.HTTP_400_BAD_REQUEST,
            )

        export_format = _get_micro_request_export_format(request)
        if not export_format:
            return _micro_download_error_response(
                "Unsupported format. Use xlsx or csv.",
                status.HTTP_400_BAD_REQUEST,
            )

        return self._build_download_response(district_name, export_format)


class MicrostatificationDataDownloadLinkView(APIView):
    permission_classes = [IsAuthenticated, IsMalariaPrivileged]
    renderer_classes = [
        renderers.JSONRenderer,
        JSONFormatXLSXRenderer,
        JSONFormatCSVRenderer,
    ]

    def get(self, request):
        district_name = _normalize_micro_district_name(request.query_params.get("district"))
        if not district_name:
            return Response(
                {"detail": "Please select a district or choose All Districts."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if is_malaria_dm(request.user):
            if district_name == MICROSTATIFICATION_ALL_DISTRICTS_TOKEN:
                return Response(
                    {"detail": "District managers cannot download all-district exports."},
                    status=status.HTTP_403_FORBIDDEN,
                )
            did = get_dm_district_id(request.user)
            if not did:
                return Response(
                    {"detail": "Your account must have an assigned district to download exports."},
                    status=status.HTTP_403_FORBIDDEN,
                )
            allowed_name = District.objects.filter(pk=did).values_list("name", flat=True).first()
            if not allowed_name or district_name.strip().lower() != allowed_name.strip().lower():
                return Response(
                    {"detail": "You are not allowed to download this district."},
                    status=status.HTTP_403_FORBIDDEN,
                )

        if (
            district_name != MICROSTATIFICATION_ALL_DISTRICTS_TOKEN
            and district_name not in MICROSTATIFICATION_TEMPLATE_FILES
        ):
            return Response(
                {"detail": f"Unsupported district: {district_name}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        export_format = _get_micro_request_export_format(request)
        if not export_format:
            return Response(
                {"detail": "Unsupported format. Use xlsx or csv."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ticket = _build_micro_download_ticket(
            request.user.id,
            district_name,
            export_format=export_format,
        )
        download_path = (
            f"{reverse('malaria-download-microstatification-file')}?ticket={ticket}"
        )
        return Response({"download_url": download_path})


class MicrostatificationDataDirectDownloadView(MicrostatificationDataDownloadView):
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        ticket = (request.query_params.get("ticket") or "").strip()
        if not ticket:
            return _micro_download_error_response(
                "Missing download ticket.",
                status.HTTP_400_BAD_REQUEST,
            )

        try:
            payload = _load_micro_download_ticket(ticket)
        except SignatureExpired:
            return _micro_download_error_response(
                "Download link expired. Please try again.",
                status.HTTP_400_BAD_REQUEST,
            )
        except BadSignature:
            return _micro_download_error_response(
                "Invalid download link.",
                status.HTTP_400_BAD_REQUEST,
            )

        user_id = payload.get("user_id")
        district_name = _normalize_micro_district_name(payload.get("district"))
        export_format = _normalize_micro_export_format(payload.get("format"))
        if not district_name or not export_format:
            return _micro_download_error_response(
                "Invalid download link.",
                status.HTTP_400_BAD_REQUEST,
            )

        user = User.objects.filter(id=user_id, is_active=True).first()
        if user is None or not has_malaria_privileged_access(user):
            return _micro_download_error_response(
                "You are not allowed to download this file.",
                status.HTTP_403_FORBIDDEN,
            )

        if is_malaria_dm(user):
            if district_name == MICROSTATIFICATION_ALL_DISTRICTS_TOKEN:
                return _micro_download_error_response(
                    "District managers cannot download all-district exports.",
                    status.HTTP_403_FORBIDDEN,
                )
            did = get_dm_district_id(user)
            if not did:
                return _micro_download_error_response(
                    "You are not allowed to download this file.",
                    status.HTTP_403_FORBIDDEN,
                )
            try:
                d_allowed = (District.objects.only("name").get(pk=did).name or "").strip().lower()
            except District.DoesNotExist:
                d_allowed = ""
            if not d_allowed or district_name.strip().lower() != d_allowed:
                return _micro_download_error_response(
                    "You are not allowed to download this file.",
                    status.HTTP_403_FORBIDDEN,
                )

        return self._build_download_response(district_name, export_format)
